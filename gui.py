#!/usr/bin/env python3
"""
FANZA同人向け 低コスト脚本生成パイプライン - GUI版
Claude API直接対応
Skills: prompt_compactor → low_cost_pipeline → script_quality_supervisor
UI: Material Design 3 inspired
"""

import sys
import types as _types

# v9.0-hotfix: platform.system()がこの環境でハングするため静的値に差し替え
# （anthropic SDKの_build_headers()が毎回呼ぶ）
import platform as _platform
_platform.system = lambda: "Windows"
_platform.platform = lambda: "Windows-11"
_platform.machine = lambda: "AMD64"
_platform.release = lambda: "11"
_platform.python_version = lambda: ".".join(str(x) for x in sys.version_info[:3])

# v9.0-hotfix: darkdetectがこの環境でハングするためwinregで代替
_dd = _types.ModuleType("darkdetect")
try:
    import winreg as _winreg
    _key = _winreg.OpenKey(
        _winreg.HKEY_CURRENT_USER,
        r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
    _val, _ = _winreg.QueryValueEx(_key, "AppsUseLightTheme")
    _winreg.CloseKey(_key)
    _dd_theme = "Light" if _val else "Dark"
except Exception:
    _dd_theme = "Dark"
_dd.theme = lambda: _dd_theme
_dd.isDark = lambda: _dd_theme == "Dark"
_dd.isLight = lambda: _dd_theme == "Light"
_dd.listener = lambda callback: None
sys.modules["darkdetect"] = _dd

import json
import csv
import time
import random
import threading
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed

import tkinter as tk
import customtkinter as ctk

# Excel出力用（オプション）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PNG Info読み取り用（オプション）
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# ドラッグ&ドロップ用（オプション）
try:
    import windnd
    WINDND_AVAILABLE = True
except ImportError:
    WINDND_AVAILABLE = False

try:
    import anthropic
except ImportError:
    print("Error: anthropic library is required. Run: pip install anthropic")
    sys.exit(1)

from char_builder import (
    AGE_OPTIONS, RELATIONSHIP_OPTIONS, ARCHETYPE_OPTIONS,
    FIRST_PERSON_OPTIONS, SPEECH_STYLE_OPTIONS,
    HAIR_COLOR_OPTIONS, HAIR_STYLE_OPTIONS,
    BODY_TYPE_OPTIONS, CHEST_OPTIONS, CLOTHING_OPTIONS,
    SHYNESS_OPTIONS, build_custom_character_data
)
from schema_validator import (
    validate_context, validate_outline, validate_scene, validate_results
)
from concept_presets import CONCEPT_PRESETS

# === Font Awesome 6 アイコンフォント ===
FONTS_DIR = Path(__file__).parent / "fonts"
_FA_FONT_PATH = FONTS_DIR / "fa-solid-900.ttf"
if _FA_FONT_PATH.exists():
    ctk.FontManager.load_font(str(_FA_FONT_PATH))

# フォント定数
FONT_JP = "Noto Sans JP"
FONT_ICON = "Font Awesome 6 Free Solid"
FONT_MONO = "Consolas"


class Icons:
    """Font Awesome 6 Solid ユニコードコードポイント"""
    FILM = "\uf008"
    LOCK = "\uf023"
    FOLDER = "\uf07b"
    FOLDER_OPEN = "\uf07c"
    USER = "\uf007"
    BOOK = "\uf02d"
    GEAR = "\uf013"
    COINS = "\uf51d"
    LIST = "\uf022"
    CLOCK = "\uf017"
    PLAY = "\uf04b"
    SAVE = "\uf0c7"
    STOP = "\uf04d"
    WAND = "\uf0d0"
    XMARK = "\uf057"
    WARNING = "\uf071"
    CHART = "\uf080"
    CHECK = "\uf058"
    CHEVRON_UP = "\uf077"
    CHEVRON_DOWN = "\uf078"
    DOWNLOAD = "\uf019"
    FILE_EXPORT = "\uf56e"
    PALETTE = "\uf53f"
    IMAGE = "\uf03e"


def icon_text_label(parent, icon: str, text: str, icon_size: int = 14, text_size: int = 14,
                    text_weight: str = "bold", text_color=None, fg_color="transparent"):
    """FAアイコン + テキストを横並びで配置するヘルパー"""
    if text_color is None:
        text_color = MaterialColors.ON_SURFACE
    frame = ctk.CTkFrame(parent, fg_color=fg_color)
    ctk.CTkLabel(
        frame, text=icon,
        font=ctk.CTkFont(family=FONT_ICON, size=icon_size),
        text_color=text_color
    ).pack(side="left", padx=(0, 8))
    ctk.CTkLabel(
        frame, text=text,
        font=ctk.CTkFont(family=FONT_JP, size=text_size, weight=text_weight),
        text_color=text_color
    ).pack(side="left")
    return frame


# === Material Design 3 カラーパレット ===
class MaterialColors:
    """
    Material You / M3 Dynamic Color System
    Neutral Gray + Blue Accent palette
    Clean, minimal, professional (Notion/Linear/Figma inspired)
    """

    # === M3 Tonal Palette (Neutral Gray seed + Blue accent) ===
    # Primary — Clean blue for primary actions
    PRIMARY = "#3B82F6"           # Blue-500
    PRIMARY_CONTAINER = "#EFF6FF" # Blue-50
    ON_PRIMARY = "#FFFFFF"
    ON_PRIMARY_CONTAINER = "#1A1A1A"  # Near-black

    # Secondary — Neutral gray
    SECONDARY = "#5C5C5C"         # Medium gray
    SECONDARY_CONTAINER = "#E0E0E0"   # Light gray
    ON_SECONDARY = "#FFFFFF"
    ON_SECONDARY_CONTAINER = "#1C1C1C"  # Near-black

    # Tertiary — Teal accent for variety
    TERTIARY = "#0D9488"          # Teal-600
    TERTIARY_CONTAINER = "#CCFBF1"    # Teal-100

    # Error
    ERROR = "#DC2626"             # Red-600
    ERROR_CONTAINER = "#FEE2E2"   # Red-100
    ON_ERROR = "#FFFFFF"

    # Success (Extended)
    SUCCESS = "#16A34A"           # Green-600
    SUCCESS_CONTAINER = "#DCFCE7" # Green-100

    # === Surface Tones (Pure Neutral R=G=B — zero color cast) ===
    BACKGROUND = "#FAFAFA"        # Near-white warm neutral
    SURFACE = "#FAFAFA"           # Match background
    SURFACE_DIM = "#C2C2C2"       # Dimmed surface
    SURFACE_BRIGHT = "#FAFAFA"
    SURFACE_CONTAINER_LOWEST = "#FFFFFF"   # Pure white — cards, dialogs
    SURFACE_CONTAINER_LOW = "#F5F5F5"      # Subtle elevation step
    SURFACE_CONTAINER = "#EEEEEE"          # Medium elevation — input bg
    SURFACE_CONTAINER_HIGH = "#E0E0E0"     # Higher elevation
    SURFACE_CONTAINER_HIGHEST = "#D6D6D6"  # Highest — section headers

    # On Surface — High contrast text
    ON_BACKGROUND = "#1A1A1A"     # Near black
    ON_SURFACE = "#1A1A1A"        # Near black for readability
    ON_SURFACE_VARIANT = "#525252"    # Secondary/hint/placeholder (7.8:1 on white)

    # Outline — WCAG 1.4.11 non-text contrast compliant
    OUTLINE = "#757575"           # Input borders (4.6:1 on white)
    OUTLINE_VARIANT = "#8E8E8E"   # Card borders, dividers (3.3:1 on white)

    # Inverse
    INVERSE_SURFACE = "#2D2D2D"   # Dark bg for snackbar, tooltips
    INVERSE_ON_SURFACE = "#F0F0F0"
    INVERSE_PRIMARY = "#93C5FD"   # Blue-300 for dark surface accent

    # Scrim & Shadow
    SCRIM = "#000000"
    SHADOW = "#000000"

    # === Legacy aliases for compatibility ===
    SURFACE_VARIANT = SURFACE_CONTAINER
    PRIMARY_VARIANT = "#2563EB"   # Blue-600 (hover)
    PRIMARY_LIGHT = INVERSE_PRIMARY
    ACCENT = TERTIARY
    ACCENT_VARIANT = "#5EEAD4"    # Teal-300
    ACCENT_DARK = "#0F766E"       # Teal-700
    WARNING = "#D97706"           # Amber-600
    SURFACE_DARK = INVERSE_SURFACE
    ON_DARK = INVERSE_ON_SURFACE
    ON_ACCENT = ON_PRIMARY


# === 設定 ===
MAX_RETRIES = 3
MAX_RETRIES_OVERLOADED = 6  # 529 Overloaded専用（長時間待機）
RETRY_DELAY = 2
RETRY_DELAY_OVERLOADED = 15  # 529 Overloaded初回待機秒数
CONCURRENT_BATCH_SIZE = 2       # Wave内同時生成数（v8.7: 5→2 ストーリー一貫性向上）
CONCURRENT_MIN_SCENES = 13      # 並列化の最小シーン数
CONCURRENT_WAVE_COOLDOWN = 2.0  # Wave間クールダウン(秒)

# プロバイダー設定
PROVIDER_CLAUDE = "claude"

OUTPUT_DIR = Path(__file__).parent
SKILLS_DIR = OUTPUT_DIR / "skills"
JAILBREAK_FILE = OUTPUT_DIR / "jailbreak.md"
DANBOORU_TAGS_JSON = OUTPUT_DIR / "danbooru_tags.json"
CONFIG_FILE = OUTPUT_DIR / "config.json"
LOG_FILE = OUTPUT_DIR / "log.txt"
CONTEXT_DIR = OUTPUT_DIR / "context"
DRAFTS_DIR = OUTPUT_DIR / "drafts"
FINAL_DIR = OUTPUT_DIR / "final"
EXPORTS_DIR = OUTPUT_DIR / "exports"
SOURCES_DIR = OUTPUT_DIR / "sources"
CHARACTERS_DIR = OUTPUT_DIR / "characters"
CHAR_SKILLS_DIR = SKILLS_DIR / "characters"
PROFILES_DIR = OUTPUT_DIR / "profiles"

# プリセットキャラクター
PRESETS_DIR = Path(__file__).parent / "presets"
PRESET_CHARS_DIR = PRESETS_DIR / "characters"
PRESET_INDEX_FILE = PRESETS_DIR / "preset_index.json"

# ディレクトリ作成
for d in [CONTEXT_DIR, DRAFTS_DIR, FINAL_DIR, EXPORTS_DIR, SOURCES_DIR, CHARACTERS_DIR, CHAR_SKILLS_DIR, PROFILES_DIR]:
    d.mkdir(exist_ok=True, parents=True)

# モデル設定
MODELS = {
    "haiku": "claude-haiku-4-5-20251001",        # 高品質（複雑タスク用）
    "haiku_fast": "claude-3-haiku-20240307",      # 低コスト（シンプルタスク用: 4x安い）
    "sonnet": "claude-sonnet-4-20250514",         # プレミアム（最重要シーン用）
    "opus": "claude-opus-4-5-20250924",           # 最高品質（クライマックス清書用）
}

# コスト（USD per 1M tokens）
COSTS = {
    "claude-3-haiku-20240307": {"input": 0.25, "output": 1.25},
    "claude-haiku-4-5-20251001": {"input": 1.00, "output": 5.00},
    "claude-sonnet-4-20250514": {"input": 3.00, "output": 15.00},
    "claude-opus-4-5-20250924": {"input": 5.00, "output": 25.00},
}

# テーマ選択肢
THEME_OPTIONS = {
    "指定なし": "",
    "凌辱・屈辱": "humiliation",
    "強制・無理やり": "forced",
    "純愛・ラブラブ": "love",
    "寝取られ・NTR": "netorare",
    "和姦・合意": "vanilla",
    "堕ち・調教": "corruption",
    "痴漢・公共": "chikan",
    "上司・OL": "office",
    "先生・生徒": "teacher_student",
    "メイド・ご主人様": "maid",
    "催眠・洗脳": "hypnosis",
    "異種姦・モンスター": "monster",
    "時間停止": "time_stop",
    "ハーレム": "harem",
    "女性優位・痴女": "femdom",
    "近親相姦": "incest",
    "異世界・ファンタジー": "isekai",
    "温泉・お風呂": "onsen",
    "睡眠・夜這い": "sleep",
    "輪姦・集団": "gangbang",
    "医療・診察": "medical",
    "水着・プール": "swimsuit",
    "スポーツ・部活": "sports",
    "アイドル・芸能": "idol",
    "隣人・日常": "neighbor",
    "風俗・パパ活": "prostitution",
    "露出・盗撮": "voyeur",
    "触手": "tentacle",
    "逆レイプ": "reverse_rape",
    "コスプレ・撮影": "cosplay",
}
_THEME_KEY_TO_JP = {v: k for k, v in THEME_OPTIONS.items() if v}

# 男性キャラクター プリセット
MALE_PRESETS = {
    "おまかせ": "",
    "筋肉質の青年": "muscular_male, young_man, short_hair, toned_body, 1boy",
    "小太りの中年": "fat_man, middle_aged, chubby_male, ugly_man, 1boy",
    "イケメン青年": "handsome, bishounen, young_man, slim, 1boy",
    "強面のおじさん": "old_man, muscular_male, scar, stubble, intimidating, 1boy",
    "ガリガリの陰キャ": "skinny_male, messy_hair, glasses, otaku, 1boy",
    "黒人マッチョ": "dark-skinned_male, muscular_male, tall_male, abs, 1boy",
    "太ったおじさん": "fat_man, old_man, hairy_male, bald, ugly_man, 1boy",
    "普通体型の少年": "young_man, average_build, 1boy",
}

# 男性キャラ 髪型オプション
MALE_HAIR_STYLE_OPTIONS = {
    "おまかせ": "",
    "短髪": "short_hair",
    "長髪": "long_hair",
    "はげ": "bald",
    "坊主": "buzz_cut",
    "オールバック": "slicked_back_hair",
    "ボサボサ": "messy_hair",
}

# 男性キャラ 髪色オプション
MALE_HAIR_COLOR_OPTIONS = {
    "おまかせ": "",
    "黒髪": "black_hair",
    "茶髪": "brown_hair",
    "金髪": "blonde_hair",
    "白髪": "grey_hair",
    "赤髪": "red_hair",
}

# 男性キャラ 肌色オプション
MALE_SKIN_COLOR_OPTIONS = {
    "おまかせ": "",
    "色白": "pale_skin",
    "普通": "",
    "日焼け": "tanned, dark_skin",
    "褐色": "dark-skinned_male",
    "黒肌": "dark-skinned_male, very_dark_skin",
}

# 時間帯オプション
TIME_OF_DAY_OPTIONS = {
    "おまかせ": "",
    "朝": "morning, sunrise, morning_light, warm_lighting",
    "昼": "daytime, sunlight, bright, afternoon",
    "夕方": "evening, sunset, orange_sky, golden_hour",
    "夜": "night, moonlight, dark, dim_lighting",
}

# 場所タイプオプション
LOCATION_TYPE_OPTIONS = {
    "おまかせ": "",
    "屋内": "indoors",
    "屋外": "outdoors",
}

# 男性外見 日本語→SDタグ変換マップ
MALE_TAG_MAP = {
    "はげ": "bald",
    "禿": "bald",
    "坊主": "buzz_cut",
    "小太り": "chubby_male",
    "太った": "fat_man",
    "デブ": "fat_man",
    "肥満": "fat_man, obese",
    "中年": "middle_aged",
    "おじさん": "old_man",
    "おっさん": "old_man",
    "老人": "elderly",
    "おじいちゃん": "old_man, grey_hair",
    "筋肉": "muscular_male",
    "マッチョ": "muscular_male, abs",
    "ガリガリ": "skinny_male",
    "痩せ": "skinny_male",
    "長身": "tall_male",
    "低身長": "short_male",
    "イケメン": "handsome, bishounen",
    "ブサイク": "ugly_man",
    "キモい": "ugly_man",
    "メガネ": "glasses",
    "眼鏡": "glasses",
    "ひげ": "beard",
    "髭": "beard",
    "無精ひげ": "stubble",
    "黒人": "dark-skinned_male",
    "白髪": "grey_hair",
    "金髪": "blonde_hair",
    "短髪": "short_hair",
    "長髪": "long_hair",
    "オールバック": "slicked_back_hair",
    "スーツ": "business_suit",
    "作業着": "work_clothes",
    "裸": "nude_male, shirtless",
    "タンクトップ": "tank_top",
    "サラリーマン": "salaryman, business_suit",
    "ヤンキー": "delinquent, pompadour",
    "体毛": "hairy_male",
    "傷": "scar",
    "強面": "intimidating",
    "優しい": "gentle_expression",
    "怖い": "intimidating, menacing",
    "少年": "young_man, shota",
    "青年": "young_man",
    "紳士": "gentleman, formal",
}


def parse_male_description(text: str) -> str:
    """ユーザーの自由入力テキストからSDタグを生成する"""
    if not text or not text.strip():
        return ""
    tags = set()
    tags.add("1boy")
    for keyword, sd_tags in MALE_TAG_MAP.items():
        if keyword in text:
            for t in sd_tags.split(", "):
                tags.add(t.strip())
    # マッチしなかった場合はデフォルト
    if len(tags) <= 1:  # 1boyのみ
        tags.add("1boy")
    return ", ".join(sorted(tags))


# ストーリー構成プリセット（プロローグ/本編/エピローグ %）
STRUCTURE_PRESETS = {
    "標準バランス (10/80/10)": {"prologue": 10, "epilogue": 10},
    "エロ重視 (5/90/5)": {"prologue": 5, "epilogue": 5},
    "ストーリー重視 (20/70/10)": {"prologue": 20, "epilogue": 10},
    "じっくり展開 (15/75/10)": {"prologue": 15, "epilogue": 10},
    "カスタム": None,
}

# テーマ別ストーリー・演出ガイド
THEME_GUIDES = {
    "netorare": {
        "name": "寝取られ・NTR",
        "story_arc": "日常→接近→裏切り→堕ち→完堕ち",
        "key_emotions": ["背徳感", "罪悪感", "快楽への抗えなさ", "比較（彼氏より...）"],
        "story_elements": [
            "彼氏/夫がいる設定を明確に",
            "最初は抵抗・罪悪感",
            "徐々に快楽に負ける",
            "「彼氏には言えない」「こんなの初めて」",
            "最終的に寝取り男を求める"
        ],
        "dialogue_tone": """罪悪感と快感の葛藤、比較表現、堕ちていく過程。
■核心: 「彼氏より感じる自分」への罪悪感→快楽で塗り潰される過程を丁寧に描け
■比較表現の書き方: thought内で「○○くんとは…違う…」「こんなの…初めて…」のように彼氏を想起させる
■speech変化: 序盤「やめて…彼氏がいるの」→中盤「ごめんなさい…ごめん…」→終盤「もっと…♡」
■NG: 彼氏の存在を忘れた普通のエロ。NTRは「裏切っている意識」が常にあってこそ成立する""",
        "use_heart": False,  # ♡は使わない
        "sd_tags": "netorare, cheating, corruption, guilt, unfaithful, stolen",
        "sd_expressions": "conflicted, guilty_pleasure, ahegao, mindbreak",
        # NTR谷あり型: 一度落ちてから再上昇。中盤で罪悪感の谷を作る
        "intensity_curve": "valley",  # 1→3→2→4→5→4
        "foreplay_ratio": 0.25,  # 前戯25%（心理描写重視）
        "intro_ratio": 0.08,  # 導入8%（日常→接近）
    },
    "humiliation": {
        "name": "凌辱・屈辱",
        "story_arc": "支配→抵抗→屈服→快楽堕ち",
        "key_emotions": ["屈辱", "恐怖", "抵抗", "やがて快感に負ける"],
        "story_elements": [
            "力関係の差を明確に",
            "抵抗するが徐々に体が反応",
            "「やめて」「嫌」から変化",
            "屈辱的な状況設定"
        ],
        "dialogue_tone": """抵抗、懇願、屈辱感、やがて快感を認める。
■核心: プライドが砕かれる過程。抵抗→身体の裏切り→屈辱の中の快楽→完全屈服
■speech変化: 「ふざけないで」→「…っ…見るな…」→「嫌…なのに…♡」→「…はい…♡」
■男性speech: 嘲笑・蔑み・命令で屈辱を増幅。「ほら、もう濡れてるぞ」「認めろよ」
■NG: いきなりの快楽堕ち。屈辱される過程の「心が折れる瞬間」を丁寧に描け""",
        "use_heart": False,
        "sd_tags": "humiliation, forced, reluctant, crying, tears",
        "sd_expressions": "crying, fearful, reluctant, trembling, broken",
        "intensity_curve": "ascending",  # 1→2→3→4→5（右肩上がり）
        "foreplay_ratio": 0.15,
        "intro_ratio": 0.05,  # 導入最小（即行為開始）
    },
    "forced": {
        "name": "強制・無理やり",
        "story_arc": "襲われる→抵抗→屈服→（オプション：快楽堕ち）",
        "key_emotions": ["恐怖", "抵抗", "絶望", "やがて諦め/快感"],
        "story_elements": [
            "逃げられない状況",
            "必死の抵抗",
            "力で押さえつけられる",
            "「やめて」「助けて」"
        ],
        "dialogue_tone": """懇願、抵抗、絶望、諦め。
■核心: 恐怖→抵抗→身体の裏切り→心が壊れる。暴力ではなく「快楽で壊される」恐怖を描け
■speech変化: 「助けて！」→「やめ…て…」→（言葉にならない）→「…もう…いい…」
■thought活用: speechで拒否しつつ、thoughtで「なんで…感じてる…」の矛盾を描け
■NG: 終始同じトーンの悲鳴。段階的に「壊れていく」グラデーションが命""",
        "use_heart": False,
        "sd_tags": "forced, rape, struggling, restrained, pinned_down",
        "sd_expressions": "crying, screaming, fearful, defeated",
        "intensity_curve": "ascending",
        "foreplay_ratio": 0.10,  # 前戯短め（強制的に進行）
        "intro_ratio": 0.05,
    },
    "love": {
        "name": "純愛・ラブラブ",
        "story_arc": "告白→初々しさ→情熱→幸福",
        "key_emotions": ["恥じらい", "愛情", "幸福感", "一体感"],
        "story_elements": [
            "両想いの確認",
            "初々しい恥じらい",
            "愛情表現",
            "「好き」「愛してる」"
        ],
        "dialogue_tone": """甘い、恥ずかしがり、愛情たっぷり。
■核心: 好きだからこそ恥ずかしい。信頼と恥じらいの共存。二人の感情が呼応する描写
■speech特徴: 「好き」を直接言えない段階→言える段階の変化。「…見ないで…」「…嬉しい…♡」
■男性speech: 命令ではなく確認と愛情。「怖くない？」「綺麗だよ」「もっと近くに来て」
■thought: 「嬉しい…でも恥ずかしい…」の二律背反。幸福感を軸に描け
■NG: 強制系と同じspeechパターン。純愛は「相互の愛情」が前提""",
        "use_heart": True,  # ♡OK
        "sd_tags": "romantic, loving, gentle, passionate, consensual",
        "sd_expressions": "blushing, happy, loving, content, peaceful",
        "intensity_curve": "wave",  # 焦らし+感情描写の波型
        "foreplay_ratio": 0.30,  # 前戯30%（丁寧に描写）
        "intro_ratio": 0.10,
    },
    "vanilla": {
        "name": "和姦・合意",
        "story_arc": "ムード→合意→行為→満足",
        "key_emotions": ["期待", "興奮", "快感", "満足"],
        "story_elements": [
            "自然な流れ",
            "お互いの同意",
            "楽しむ雰囲気"
        ],
        "dialogue_tone": "自然、楽しそう、気持ちいい",
        "use_heart": True,
        "sd_tags": "consensual, enjoying, willing, happy_sex",
        "sd_expressions": "happy, enjoying, moaning, satisfied",
        "intensity_curve": "ascending",
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.08,
    },
    "corruption": {
        "name": "堕ち・調教",
        "story_arc": "純粋→揺らぎ→堕落→完堕ち",
        "key_emotions": ["戸惑い", "背徳感", "快楽への目覚め", "依存"],
        "story_elements": [
            "最初は純粋・清楚",
            "徐々に快楽を覚える",
            "「こんなの知らなかった」",
            "最終的に求めるように"
        ],
        "dialogue_tone": """戸惑いから快楽への変化、堕ちていく過程。
■核心: 無垢な状態から快楽を教え込まれ、自ら求めるようになる変化。「教育」の過程を描け
■speech変化: 「なにこれ…」→「もう少し…だけ…」→「もっと教えて♡」→「壊して♡」
■thought: 「知らなかった…こんなの」→「だめ…でも止められない」→「もっと…♡」
■NG: 最初から積極的。堕ちの「段階」が命。知らない→知る→求めるの3段階を丁寧に""",
        "use_heart": False,
        "sd_tags": "corruption, training, breaking, mindbreak",
        "sd_expressions": "confused, awakening, addicted, broken, ahegao",
        # 調教型: 段階的にエスカレート、最後に完堕ち
        "intensity_curve": "staircase",  # 1→2→3→3→4→4→5
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.08,
    },
    "chikan": {
        "name": "痴漢・公共",
        "story_arc": "被害→抵抗できない→感じてしまう",
        "key_emotions": ["恐怖", "羞恥", "声が出せない", "感じてしまう罪悪感"],
        "story_elements": [
            "公共の場（電車など）",
            "周りにバレられない",
            "声を出せない状況",
            "体が勝手に反応"
        ],
        "dialogue_tone": """小声、我慢、羞恥。
■核心: 公共の場で声を出せない→それが逆に感覚を鋭敏にする。「バレたら終わり」の緊張感
■speech: 全て小声・途切れがち。「やめ…ここ…電車…」「声…出ちゃ…」「人が…いるのに…」
■thought: 「バレたら…」「こんなとこで…感じちゃだめ…」「…また来るの…待ってる…♡」
■NG: 大声の喘ぎ。痴漢は「声を押し殺す」のが基本。moanも小声表現にすること""",
        "use_heart": False,
        "sd_tags": "chikan, groping, public, train, crowded, molested",
        "sd_expressions": "embarrassed, trying_not_to_moan, biting_lip, conflicted",
        "intensity_curve": "wave",
        "foreplay_ratio": 0.20,  # 痴漢: 触る→戸惑い→感じる の波を作る
        "intro_ratio": 0.05,
    },
    "office": {
        "name": "上司・OL",
        "story_arc": "職場→関係発展→密会→背徳",
        "key_emotions": ["緊張", "背徳感", "禁断の興奮", "秘密"],
        "story_elements": [
            "上下関係",
            "バレてはいけない",
            "仕事中の緊張感",
            "オフィスでの密会"
        ],
        "dialogue_tone": "敬語混じり、緊張、背徳感",
        "use_heart": False,
        "sd_tags": "office, office_lady, suit, desk, workplace, secret",
        "sd_expressions": "nervous, secretive, professional_facade",
        "intensity_curve": "valley",  # 緊張→行為→中断リスク→再開
        "foreplay_ratio": 0.25,
        "intro_ratio": 0.10,
    },
    "teacher_student": {
        "name": "先生・生徒",
        "story_arc": "禁断→誘惑/誘われ→一線を越える→背徳",
        "key_emotions": ["禁断", "背徳感", "支配/被支配", "秘密"],
        "story_elements": [
            "立場の差",
            "禁じられた関係",
            "教室/保健室などの場所",
            "バレたら終わり"
        ],
        "dialogue_tone": """敬語と砕けた表現の混在、禁断感。
■核心: 敬語の崩壊=関係性の崩壊。「先生」という呼称がそのままエロさを増幅する
■speech変化: 「先生…だめです…」→「先生…もう…」→「せんせ…♡」（敬語が崩壊していく）
■thought: 「先生と生徒なのに…」「バレたら…退学…」「でも…止められない…」
■場所効果: 学校の教室・準備室・体育倉庫=日常空間でのエロが背徳感を増幅
■NG: 最初から恋人同士のような口調。「先生と生徒」の立場差が常にあること""",
        "use_heart": False,
        "sd_tags": "teacher, student, classroom, forbidden, taboo",
        "sd_expressions": "nervous, forbidden_pleasure, secretive",
        "intensity_curve": "valley",  # 禁断→一線→罪悪感→再犯
        "foreplay_ratio": 0.25,
        "intro_ratio": 0.10,
    },
    "maid": {
        "name": "メイド・ご主人様",
        "story_arc": "奉仕→親密→特別な奉仕",
        "key_emotions": ["忠誠", "奉仕", "主従関係", "愛情"],
        "story_elements": [
            "主従関係",
            "「ご主人様」呼び",
            "奉仕の延長",
            "命令への従順"
        ],
        "dialogue_tone": "丁寧語、奉仕精神、従順",
        "use_heart": True,
        "sd_tags": "maid, maid_uniform, master, servant, obedient",
        "sd_expressions": "devoted, obedient, eager_to_please",
        "intensity_curve": "staircase",  # 奉仕→命令→特別奉仕の段階
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.08,
    },
    "hypnosis": {
        "name": "催眠・洗脳",
        "story_arc": "暗示→無意識→操作→覚醒しても体が覚えている",
        "key_emotions": ["ぼんやり", "抵抗できない", "無意識の快感", "自分じゃない感覚"],
        "story_elements": [
            "催眠術や暗示のきっかけ",
            "意識がぼやける描写",
            "命令に逆らえない体",
            "「なぜ体が勝手に...」という混乱",
            "覚醒後も体が反応してしまう"
        ],
        "dialogue_tone": """ぼんやりした口調、命令への無抵抗、覚醒時の混乱と羞恥。
■核心: 「自分の意思で感じている」と思わせるのが催眠の本質。無理やりではなく「自発的に見える」
■speech変化: 「あれ…なんか…」→「気持ちいい…から…いいかな…」→「もっとして…♡」
■thought: 「おかしい…でも…自然なこと…」「催眠じゃない…好きだから…」
■覚醒時: 「え…なんで…裸…？」「嘘…私…何を…」の混乱が重要
■NG: ロボットのような無感情応答。催眠中も「自然に見える」感情表現を入れること""",
        "use_heart": False,
        "sd_tags": "hypnosis, mind_control, blank_eyes, spiral_eyes, trance",
        "sd_expressions": "empty_eyes, dazed, vacant, drooling, mindless, confused",
        # 催眠型: 段階的に深化、覚醒で一時ダウン→再度堕ちる
        "intensity_curve": "valley",
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.08,
    },
    "monster": {
        "name": "異種姦・モンスター",
        "story_arc": "遭遇→捕獲→異種交配→快楽堕ち",
        "key_emotions": ["恐怖", "嫌悪", "異物感", "未知の快感に溺れる"],
        "story_elements": [
            "人外の存在との遭遇",
            "逃げられない状況",
            "人間にはない刺激",
            "「人間じゃないのに...」という背徳感",
            "触手や異形の描写"
        ],
        "dialogue_tone": "恐怖と驚き、徐々に快感に変わる声、人間離れした行為への反応",
        "use_heart": False,
        "sd_tags": "monster, tentacles, interspecies, creature, non-human",
        "sd_expressions": "scared, disgusted, surprised, overwhelmed, ahegao",
        "intensity_curve": "two_stage",  # 1体目→2体目/形態変化の2段階
        "foreplay_ratio": 0.10,
        "intro_ratio": 0.05,
    },
    "time_stop": {
        "name": "時間停止",
        "story_arc": "停止→観察→いたずら→解除の瞬間",
        "key_emotions": ["無防備", "知らないうちに", "解除後の混乱", "証拠に気づく恥辱"],
        "story_elements": [
            "時間が止まるきっかけ",
            "止まった世界での自由行動",
            "好きなポーズに変えられる",
            "解除後の「何かされた？」感覚",
            "体に残る痕跡"
        ],
        "dialogue_tone": "停止中は無言（ナレーション中心）、解除後は混乱と気づきの描写",
        "use_heart": False,
        "sd_tags": "time_stop, frozen, mannequin_pose, unconscious, sleeping",
        "sd_expressions": "frozen, blank_expression, sleeping, confused, shocked",
        "intensity_curve": "plateau",  # 停止中は一定intensity→解除で急変
        "foreplay_ratio": 0.15,
        "intro_ratio": 0.08,
        "world_rules": [
            "【時間停止の対象キャラ】意識はあるが身体は完全に動かない（操り人形状態）。表情も固定される",
            "【セリフ制約】停止中の対象キャラのセリフはthought（心の声）のみ。speechは禁止（声が出せないため）。moanも禁止（声が出せないため、身体反応の描写で代替）",
            "【男性の行動】時間停止の使い手は自由に動ける。セリフはspeechで出力してよい",
            "【時間停止解除後】対象キャラは混乱・違和感・身体の痕跡に気づく。speechとthought両方使用可",
            "【身体反応】停止中でも発汗・充血・濡れ等の生理現象は発生する（不随意反応）。ただし声は出ない",
        ],
    },
    "harem": {
        "name": "ハーレム",
        "story_arc": "出会い→好意集中→争奪→全員で奉仕",
        "key_emotions": ["独占欲", "嫉妬", "競争心", "共有の快楽"],
        "story_elements": [
            "複数ヒロインが主人公を取り合う",
            "嫉妬や競争の描写",
            "「私の方が上手」的な比較",
            "最終的に全員でのシーン",
            "各キャラの個性が際立つ"
        ],
        "dialogue_tone": "各キャラが個性的に競い合う、嫉妬と甘え、協力と競争",
        "use_heart": True,
        "sd_tags": "harem, multiple_girls, group, jealous, competitive",
        "sd_expressions": "jealous, competitive, eager, cooperative, blush",
        "intensity_curve": "wave",  # キャラ交代時にintensity下がる波型
        "foreplay_ratio": 0.15,
        "intro_ratio": 0.08,
    },
    "femdom": {
        "name": "女性優位・痴女",
        "story_arc": "主導権掌握→翻弄→支配→ご褒美",
        "key_emotions": ["支配欲", "優越感", "相手をからかう楽しさ", "征服感"],
        "story_elements": [
            "女性がリードする関係",
            "男性を翻弄する",
            "「こんなに感じてるの？」的なからかい",
            "騎乗位や言葉責め",
            "主導権は常に女性側"
        ],
        "dialogue_tone": "上から目線、からかい、余裕のある態度、小悪魔的",
        "use_heart": True,
        "sd_tags": "femdom, dominatrix, female_domination, sitting_on_face, riding",
        "sd_expressions": "smirk, confident, teasing, dominant, looking_down",
        "intensity_curve": "staircase",  # 焦らし→ご褒美のリズム
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.08,
    },
    "incest": {
        "name": "近親相姦",
        "story_arc": "家族の日常→意識→禁断→堕ちる",
        "key_emotions": ["背徳感", "罪悪感", "家族への愛と欲望の混同", "秘密"],
        "story_elements": [
            "家族設定を明確に（兄妹/姉弟/母子など）",
            "普段の家族関係からの逸脱",
            "「家族なのに...」という葛藤",
            "二人だけの秘密",
            "他の家族にバレない緊張感"
        ],
        "dialogue_tone": "普段の呼び方（お兄ちゃん、お姉ちゃん等）と背徳感、家族の呼称が興奮を増す",
        "use_heart": False,
        "sd_tags": "incest, siblings, family, forbidden_love, taboo, secret",
        "sd_expressions": "guilty, conflicted, forbidden_pleasure, secretive",
        "intensity_curve": "valley",  # 日常→禁断→罪悪感で一旦引く→再び
        "foreplay_ratio": 0.25,
        "intro_ratio": 0.10,
    },
    "isekai": {
        "name": "異世界・ファンタジー",
        "story_arc": "召喚/転移→出会い→異文化交流→異世界の夜",
        "key_emotions": ["驚き", "好奇心", "異文化への戸惑い", "特別な絆"],
        "story_elements": [
            "異世界の独自ルールや文化",
            "種族差（エルフ/獣人/魔族など）",
            "魔法や特殊能力の存在",
            "「この世界では普通」という価値観の違い"
        ],
        "dialogue_tone": "ファンタジー風の言い回し、異文化のギャップ、冒険と興奮",
        "use_heart": True,
        "sd_tags": "fantasy, isekai, elf, magic, castle, medieval, adventurer, demon_girl",
        "sd_expressions": "curious, amazed, blushing, excited, fantasy_glow",
        "intensity_curve": "staircase",  # 異文化交流→段階的にエスカレート
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.10,
    },
    "onsen": {
        "name": "温泉・お風呂",
        "story_arc": "入浴→のぼせ/接近→肌の触れ合い→湯けむりの中で",
        "key_emotions": ["リラックス", "羞恥", "開放感", "肌の密着感"],
        "story_elements": [
            "温泉や浴場のシチュエーション",
            "裸の状態からのスタート",
            "湯気や水音の演出",
            "のぼせて判断力低下"
        ],
        "dialogue_tone": "リラックスした口調、恥じらい、湯に溶ける感覚",
        "use_heart": True,
        "sd_tags": "onsen, hot_spring, bathing, wet, steam, towel, nude, water",
        "sd_expressions": "relaxed, flushed, steamy, embarrassed, dreamy",
        "intensity_curve": "wave",  # リラックス→興奮の波
        "foreplay_ratio": 0.25,
        "intro_ratio": 0.10,
    },
    "sleep": {
        "name": "睡眠・夜這い",
        "story_arc": "就寝→忍び寄る→まどろみの中で→覚醒",
        "key_emotions": ["無防備", "まどろみ", "夢か現実かの曖昧さ", "気づいた時の羞恥"],
        "story_elements": [
            "寝ている状態からの開始",
            "夢うつつの曖昧な意識",
            "「夢かと思った」反応",
            "声を殺す緊張感"
        ],
        "dialogue_tone": "寝ぼけた声、小声、曖昧な反応、覚醒後の動揺",
        "use_heart": False,
        "sd_tags": "sleeping, night_crawling, bed, nightgown, pajamas, dark_room, blanket",
        "sd_expressions": "sleeping, half_asleep, drowsy, confused, surprised",
        "intensity_curve": "staircase",  # 寝ている間→半覚醒→完全覚醒の段階
        "foreplay_ratio": 0.30,
        "intro_ratio": 0.05,
    },
    "gangbang": {
        "name": "輪姦・集団",
        "story_arc": "囲まれる→抵抗→連続→意識朦朧",
        "key_emotions": ["恐怖", "圧倒", "逃げ場のなさ", "快感の波状攻撃"],
        "story_elements": [
            "複数の男性による同時攻め",
            "一人では対処できない状況",
            "交代で休みなく続く",
            "意識が飛びそうになる"
        ],
        "dialogue_tone": "懇願、悲鳴、もう無理という限界、複数からの声",
        "use_heart": False,
        "sd_tags": "gangbang, group_sex, multiple_boys, surrounded, double_penetration",
        "sd_expressions": "overwhelmed, crying, exhausted, ahegao, drooling",
        "intensity_curve": "two_stage",  # 1回目→休息→2回目の2段階
        "foreplay_ratio": 0.10,
        "intro_ratio": 0.05,
    },
    "medical": {
        "name": "医療・診察",
        "story_arc": "受診→診察→特別な処置→堕ちる",
        "key_emotions": ["不安", "羞恥", "信頼の裏切り", "逆らえない立場"],
        "story_elements": [
            "医師/看護師と患者の関係",
            "診察の名目で触れる",
            "医療器具の利用",
            "「治療のためだから」という口実"
        ],
        "dialogue_tone": "専門用語混じり、冷静な口調と羞恥、命令的な指示",
        "use_heart": False,
        "sd_tags": "medical, doctor, nurse, hospital, examination, stethoscope, lab_coat",
        "sd_expressions": "nervous, embarrassed, clinical, vulnerable, exposed",
        "intensity_curve": "staircase",  # 検査→触診→処置の段階的エスカレーション
        "foreplay_ratio": 0.25,
        "intro_ratio": 0.10,
    },
    "swimsuit": {
        "name": "水着・プール",
        "story_arc": "水遊び→接近→水着のまま→開放的に",
        "key_emotions": ["開放感", "眩しさ", "日差しと熱さ", "夏の高揚感"],
        "story_elements": [
            "プールや海辺のシチュエーション",
            "水着姿の露出度",
            "水に濡れた肌の描写",
            "更衣室やシャワールームへの移動"
        ],
        "dialogue_tone": "明るく開放的、恥ずかしがり、夏の雰囲気",
        "use_heart": True,
        "sd_tags": "swimsuit, bikini, pool, beach, wet, sunlight, summer, ocean",
        "sd_expressions": "cheerful, embarrassed, wet, flushed, playful",
        "intensity_curve": "ascending",  # 水遊び→徐々にエスカレート
        "foreplay_ratio": 0.25,
        "intro_ratio": 0.10,
    },
    "sports": {
        "name": "スポーツ・部活",
        "story_arc": "練習→二人きり→汗だく→体育倉庫で",
        "key_emotions": ["競争心", "汗の匂い", "仲間意識からの逸脱", "若さの衝動"],
        "story_elements": [
            "部活や運動後のシチュエーション",
            "汗だくの体の描写",
            "体育倉庫/更衣室/部室",
            "先輩後輩の上下関係"
        ],
        "dialogue_tone": "元気で体育会系、先輩後輩の口調、汗と興奮",
        "use_heart": True,
        "sd_tags": "sports, gym_uniform, bloomers, sweat, locker_room, sporty, athletic",
        "sd_expressions": "sweaty, energetic, competitive, embarrassed, panting",
        "intensity_curve": "ascending",  # 運動の延長→そのまま上昇
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.10,
    },
    "idol": {
        "name": "アイドル・芸能",
        "story_arc": "華やかな表舞台→裏の顔→秘密の関係→堕ちる",
        "key_emotions": ["表と裏のギャップ", "禁止された恋愛", "スキャンダルの恐怖", "本当の自分"],
        "story_elements": [
            "アイドルや芸能人の設定",
            "恋愛禁止のルール",
            "マネージャーやプロデューサーとの関係",
            "バレたら終わりの緊張感"
        ],
        "dialogue_tone": "表向きのアイドル口調と本音、ギャップ萌え、秘密の関係",
        "use_heart": True,
        "sd_tags": "idol, stage, microphone, costume, backstage, dressing_room, celebrity",
        "sd_expressions": "stage_smile, nervous, secretive, gap_moe, blushing",
        "intensity_curve": "valley",  # 表舞台→裏→葛藤で一旦引く→本音解放
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.10,
    },
    "neighbor": {
        "name": "隣人・日常",
        "story_arc": "日常の接点→距離が縮まる→一線を越える→秘密の関係",
        "key_emotions": ["親しみ", "日常の延長", "近すぎる距離", "ご近所バレの恐怖"],
        "story_elements": [
            "隣の部屋/隣の家の住人",
            "日常的な接点（回覧板、ゴミ出し等）",
            "壁が薄い/声が漏れる",
            "顔を合わせる気まずさ"
        ],
        "dialogue_tone": "日常会話からの逸脱、親しみ、声を抑える、ご近所トーク",
        "use_heart": True,
        "sd_tags": "neighbor, apartment, casual_clothes, domestic, everyday, next_door",
        "sd_expressions": "friendly, nervous, secretive, blushing, trying_to_be_quiet",
        "intensity_curve": "wave",  # 日常→興奮→冷静→再燃の波
        "foreplay_ratio": 0.25,
        "intro_ratio": 0.10,
    },
    "prostitution": {
        "name": "風俗・パパ活",
        "story_arc": "交渉→サービス→本番→リピーター",
        "key_emotions": ["割り切り", "プロ意識", "本気と演技の境目", "金銭と快楽"],
        "story_elements": [
            "金銭的な関係が前提",
            "サービスのプロフェッショナル感",
            "「仕事だから」と割り切る態度",
            "徐々に本気になる/させる"
        ],
        "dialogue_tone": "営業トーク、割り切った態度、徐々に本音が漏れる",
        "use_heart": False,
        "sd_tags": "prostitution, escort, hotel_room, lingerie, money, transaction",
        "sd_expressions": "professional, seductive, calculating, genuine_pleasure",
        "intensity_curve": "ascending",  # サービス→本番→本気の上昇
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.08,
    },
    "voyeur": {
        "name": "露出・盗撮",
        "story_arc": "覗き/露出→バレそうな瞬間→興奮→暴走",
        "key_emotions": ["スリル", "背徳感", "バレる恐怖と興奮", "見られる快感"],
        "story_elements": [
            "人目を気にする状況",
            "覗き/盗撮/露出のシチュエーション",
            "バレそうでバレないギリギリ",
            "見られている興奮"
        ],
        "dialogue_tone": "小声、ドキドキ、「誰かに見られたら」、スリルと興奮",
        "use_heart": False,
        "sd_tags": "voyeurism, exhibitionism, peeping, hidden_camera, outdoors, public",
        "sd_expressions": "thrilled, nervous, excited, peeking, exposed",
        "intensity_curve": "wave",  # スリル→安堵→再びスリルの波
        "foreplay_ratio": 0.20,
        "intro_ratio": 0.08,
    },
    "tentacle": {
        "name": "触手",
        "story_arc": "遭遇→拘束→同時多点攻め→快楽堕ち",
        "key_emotions": ["恐怖", "嫌悪感", "抵抗できない", "人外の快感"],
        "story_elements": [
            "触手による全身拘束",
            "複数箇所同時の刺激",
            "粘液や吸盤の異物感",
            "「こんなもので感じるなんて」という屈辱"
        ],
        "dialogue_tone": "恐怖と嫌悪、拘束への抵抗、異物感からの快楽、堕ちていく絶望",
        "use_heart": False,
        "sd_tags": "tentacles, tentacle_sex, restrained, slime, monster, bondage, multiple_insertions",
        "sd_expressions": "scared, disgusted, struggling, overwhelmed, ahegao, crying",
        "intensity_curve": "staircase",  # 拘束→挿入→増殖の段階
        "foreplay_ratio": 0.10,
        "intro_ratio": 0.05,
    },
    "reverse_rape": {
        "name": "逆レイプ",
        "story_arc": "出会い→主導権奪取→男を押し倒す→搾り取る",
        "key_emotions": ["支配欲", "攻撃的な欲望", "男の困惑", "容赦ない快楽"],
        "story_elements": [
            "女性が完全に主導権を握る",
            "男性の抵抗を無視/楽しむ",
            "一方的に押し倒す/跨る",
            "「逃がさない」「もっと出して」"
        ],
        "dialogue_tone": "攻撃的、命令口調、男の抵抗を嘲笑う、容赦ない搾取",
        "use_heart": False,
        "sd_tags": "reverse_rape, femdom, aggressive, pinning_down, cowgirl_position, straddling",
        "sd_expressions": "aggressive, predatory, smirk, wild, dominant, crazed",
        "intensity_curve": "ascending",  # 一方的にエスカレート
        "foreplay_ratio": 0.15,
        "intro_ratio": 0.08,
    },
    "cosplay": {
        "name": "コスプレ・撮影",
        "story_arc": "撮影会→ポーズ指示→過激に→そのままの衣装で",
        "key_emotions": ["なりきり", "露出の恥じらい", "カメラへの意識", "キャラと自分の境界"],
        "story_elements": [
            "コスプレ衣装を着た状態",
            "カメラマンとの関係",
            "ポーズ指示が徐々にエスカレート",
            "衣装のまま/半脱ぎの興奮"
        ],
        "dialogue_tone": "キャラなりきり口調、恥ずかしがり、カメラを意識した言動",
        "use_heart": True,
        "sd_tags": "cosplay, costume, photoshoot, camera, posing, dressing_up, roleplay",
        "sd_expressions": "posing, embarrassed, in_character, blushing, camera_aware",
        "intensity_curve": "ascending",  # ポーズ→過激に→そのまま行為へ
        "foreplay_ratio": 0.25,
        "intro_ratio": 0.10,
    },
}

# v8.9: テーマ別時間軸スパン設定
# "single_event": 一晩/単発の出来事（翌日すら不要）
# "few_days": 数日間（2-3日、複数回の接触が物語の本質）
# "flexible": テーマに応じて自由（ただし翌週の乱用禁止）
_THEME_TIME_SPAN = {
    # single_event: 単発テーマ（1回の出来事で完結）
    "humiliation": "single_event",
    "forced": "single_event",
    "chikan": "single_event",
    "time_stop": "single_event",
    "gangbang": "single_event",
    "sleep": "single_event",
    "monster": "single_event",
    "tentacle": "single_event",
    "reverse_rape": "single_event",
    # few_days: 複数日テーマ（心理変化・関係発展に時間が必要）
    "netorare": "few_days",
    "corruption": "few_days",
    "office": "few_days",
    "teacher_student": "few_days",
    "neighbor": "few_days",
    "prostitution": "few_days",
    "idol": "few_days",
    "incest": "few_days",
    # flexible: 自由テーマ（シチュエーション次第）
    "love": "flexible",
    "vanilla": "flexible",
    "maid": "flexible",
    "hypnosis": "flexible",
    "harem": "flexible",
    "femdom": "flexible",
    "isekai": "flexible",
    "onsen": "flexible",
    "medical": "flexible",
    "swimsuit": "flexible",
    "sports": "flexible",
    "voyeur": "flexible",
    "cosplay": "flexible",
}

def _get_time_axis_instruction(theme: str, act1: int) -> str:
    """テーマ別の時間軸プロンプト指示を生成"""
    span = _THEME_TIME_SPAN.get(theme, "flexible")
    if span == "single_event":
        return (
            "\n## ⚠️⚠️ 時間軸ルール（厳守）\n"
            "- **全シーンは「同一イベント内」の出来事**。「翌日」「翌週」「数日後」は全て禁止\n"
            "- 導入から本番まで連続した1回の出来事として書くこと\n"
            "- story_flowに時間経過表現を書くな。「直後」「その場で」のみ使用可\n"
            "- CG集は「一瞬〜数時間の出来事」として構成すること\n"
        )
    elif span == "few_days":
        return (
            "\n## ⚠️⚠️ 時間軸ルール（厳守）\n"
            f"- **全シーンは「最大3日間」に収まること**。「翌週」「数週間後」「翌月」は禁止\n"
            f"- 導入{act1}シーンは「1日目」に収めよ。行為開始〜本番は「1-2日目」。余韻は「最終日」\n"
            "- 許可する時間表現: 「翌日」「翌朝」「その夜」「数時間後」\n"
            "- 禁止する時間表現: 「翌週」「数日後」「翌々週」「後日」「数週間後」\n"
            "- 1シーンごとに日が変わるのは禁止。同じ日に最低3-5シーンを配置すること\n"
        )
    else:  # flexible
        return (
            "\n## ⚠️⚠️ 時間軸ルール（厳守）\n"
            f"- 導入{act1}シーンは「同日中」に収めること。行為開始後は連続した出来事として書け\n"
            "- 「翌週」「数週間後」「翌月」の大きな時間ジャンプは禁止\n"
            "- 許可: 「翌日」「翌朝」「その夜」「数時間後」「少し休んで」\n"
            "- story_flowで毎シーン日付が変わるパターンは禁止\n"
        )

# テーマ自動推定用キーワードマップ
THEME_KEYWORD_MAP = {
    "netorare": ["寝取", "NTR", "ntr", "彼氏持ち", "人妻", "不倫", "浮気", "他の男"],
    "humiliation": ["凌辱", "屈辱", "辱め", "恥辱", "陵辱"],
    "forced": ["強制", "無理やり", "レイプ", "rape", "暴行", "襲", "犯す"],
    "love": ["純愛", "ラブラブ", "恋人", "恋愛", "両想い", "甘々", "いちゃいちゃ"],
    "vanilla": ["和姦", "合意", "普通", "ノーマル"],
    "corruption": ["堕ち", "調教", "堕落", "快楽堕ち", "洗脳堕ち", "開発"],
    "chikan": ["痴漢", "電車", "満員", "公共", "バス"],
    "office": ["上司", "OL", "オフィス", "職場", "部下", "同僚", "社長"],
    "teacher_student": ["先生", "教師", "生徒", "教え子", "塾", "家庭教師", "教授"],
    "maid": ["メイド", "ご主人様", "執事", "使用人", "お嬢様"],
    "hypnosis": ["催眠", "洗脳", "暗示", "マインドコントロール", "操り"],
    "monster": ["異種", "モンスター", "オーク", "ゴブリン", "魔物", "獣"],
    "time_stop": ["時間停止", "時止め", "時を止め", "フリーズ"],
    "harem": ["ハーレム", "複数", "全員", "3P", "4P"],
    "femdom": ["痴女", "女性優位", "逆転", "女王様", "M男", "ドS"],
    "incest": ["近親", "兄妹", "姉弟", "母子", "父娘", "義妹", "義姉", "義母"],
    "isekai": ["異世界", "ファンタジー", "勇者", "魔王", "エルフ", "転生", "召喚"],
    "onsen": ["温泉", "お風呂", "入浴", "混浴", "銭湯", "露天風呂"],
    "sleep": ["夜這い", "寝て", "睡眠", "眠って", "就寝", "添い寝", "寝込み"],
    "gangbang": ["輪姦", "集団", "輪姦す", "大勢", "ぶっかけ", "連続"],
    "medical": ["医者", "診察", "病院", "ナース", "看護師", "検診", "治療"],
    "swimsuit": ["水着", "プール", "海", "ビーチ", "ビキニ", "スク水"],
    "sports": ["部活", "スポーツ", "体操", "ブルマ", "運動", "体育", "マネージャー"],
    "idol": ["アイドル", "芸能", "ライブ", "プロデューサー", "地下アイドル", "配信者"],
    "neighbor": ["隣人", "隣の", "お隣", "マンション", "アパート", "ご近所"],
    "prostitution": ["風俗", "パパ活", "援助", "デリヘル", "ソープ", "キャバ", "ホテヘル"],
    "voyeur": ["露出", "盗撮", "覗き", "見せつけ", "野外", "見られ"],
    "tentacle": ["触手", "tentacle"],
    "reverse_rape": ["逆レイプ", "逆レ", "押し倒す", "搾り取", "逆襲"],
    "cosplay": ["コスプレ", "撮影会", "レイヤー", "コス", "衣装"],
}


def _infer_theme_from_concept(concept: str) -> str:
    """コンセプト文からテーマを自動推定する。

    THEME_KEYWORD_MAPのキーワードとコンセプトを照合し、
    最もマッチ数が多いテーマを返す。マッチなしなら空文字。
    """
    if not concept:
        return ""
    scores = {}
    for theme_key, keywords in THEME_KEYWORD_MAP.items():
        score = sum(1 for kw in keywords if kw in concept)
        if score > 0:
            scores[theme_key] = score
    if not scores:
        return ""
    return max(scores, key=scores.get)


def _build_dynamic_theme_guide(concept: str) -> dict:
    """テーマガイドが見つからない場合に、コンセプトからミニマルなガイドを生成する。

    コンセプトのトーンを検出して合理的なデフォルト値を返す。
    """
    # トーン検出
    dark_keywords = ["強制", "凌辱", "暴行", "レイプ", "堕", "調教", "犯", "脅"]
    sweet_keywords = ["恋", "愛", "甘", "ラブ", "好き", "告白", "デート"]
    comedic_keywords = ["ギャグ", "コメディ", "ドジ", "ラッキースケベ", "間違い"]

    dark_score = sum(1 for kw in dark_keywords if kw in concept)
    sweet_score = sum(1 for kw in sweet_keywords if kw in concept)
    comedic_score = sum(1 for kw in comedic_keywords if kw in concept)

    if dark_score > sweet_score and dark_score > comedic_score:
        return {
            "name": "ダーク系",
            "story_arc": "導入→支配→エスカレート→堕ちる",
            "key_emotions": ["恐怖", "抵抗", "屈服", "快楽堕ち"],
            "story_elements": ["力関係の差", "抵抗と屈服", "徐々に体が反応"],
            "dialogue_tone": "抵抗、懇願、やがて快感に負ける",
            "use_heart": False,
            "sd_tags": "forced, reluctant, dark",
            "sd_expressions": "crying, fearful, reluctant, ahegao",
            "intensity_curve": "ascending",
            "foreplay_ratio": 0.15,
            "intro_ratio": 0.05,
        }
    elif sweet_score > dark_score:
        return {
            "name": "甘々系",
            "story_arc": "出会い→親密→告白→結ばれる",
            "key_emotions": ["恥じらい", "愛情", "幸福感", "一体感"],
            "story_elements": ["両想い", "恥じらい", "愛情表現"],
            "dialogue_tone": "甘い、恥ずかしがり、愛情たっぷり",
            "use_heart": True,
            "sd_tags": "romantic, loving, gentle, consensual",
            "sd_expressions": "blushing, happy, loving, content",
            "intensity_curve": "wave",
            "foreplay_ratio": 0.30,
            "intro_ratio": 0.10,
        }
    elif comedic_score > 0:
        return {
            "name": "コミカル系",
            "story_arc": "ハプニング→巻き込まれ→流されて→楽しむ",
            "key_emotions": ["驚き", "戸惑い", "照れ", "楽しさ"],
            "story_elements": ["偶然の状況", "ドジやハプニング", "ノリと勢い"],
            "dialogue_tone": "ツッコミ、驚き、流される、楽しむ",
            "use_heart": True,
            "sd_tags": "comedic, surprised, accidental, happy",
            "sd_expressions": "surprised, embarrassed, laughing, happy",
            "intensity_curve": "ascending",
            "foreplay_ratio": 0.20,
            "intro_ratio": 0.10,
        }
    else:
        return {
            "name": "汎用",
            "story_arc": "出会い→接近→行為→余韻",
            "key_emotions": ["期待", "興奮", "快感", "満足"],
            "story_elements": ["自然な展開", "段階的なエスカレーション"],
            "dialogue_tone": "自然、感情豊か",
            "use_heart": True,
            "sd_tags": "consensual, enjoying",
            "sd_expressions": "blushing, moaning, satisfied",
            "intensity_curve": "ascending",
            "foreplay_ratio": 0.20,
            "intro_ratio": 0.08,
        }


QUALITY_POSITIVE_TAGS = "(masterpiece, best_quality:1.2)"
QUALITY_TAGS_DISABLED = "__DISABLED__"  # カスタムモードで空欄→quality tags無し

# 体位タグリスト（体位重複防止システム用）
POSITION_TAGS = {
    "missionary", "missionary_position", "cowgirl_position", "reverse_cowgirl",
    "doggy_style", "from_behind", "standing_sex", "standing",
    "sitting", "sitting_on_lap", "straddling", "spooning",
    "prone_bone", "mating_press", "suspended_congress",
    "leg_lock", "legs_up", "legs_over_head",
    "face_sitting", "sixty_nine", "paizuri", "fellatio",
    "cunnilingus", "handjob", "against_wall", "bent_over",
    "on_side", "spread_legs", "all_fours", "on_back",
    "on_stomach", "kneeling", "squatting", "lotus_position",
}

# 体位代替マップ（同一体位検出時のフォールバック）
POSITION_FALLBACKS = {
    "missionary": ["cowgirl_position", "spooning", "from_behind"],
    "missionary_position": ["cowgirl_position", "spooning", "from_behind"],
    "doggy_style": ["prone_bone", "standing_sex", "mating_press"],
    "cowgirl_position": ["reverse_cowgirl", "sitting_on_lap", "missionary"],
    "reverse_cowgirl": ["cowgirl_position", "prone_bone", "on_side"],
    "from_behind": ["prone_bone", "against_wall", "standing_sex"],
    "standing_sex": ["against_wall", "suspended_congress", "from_behind"],
    "prone_bone": ["doggy_style", "mating_press", "on_stomach"],
    "mating_press": ["missionary", "legs_up", "prone_bone"],
    "spooning": ["on_side", "from_behind", "missionary"],
    "paizuri": ["fellatio", "handjob", "cowgirl_position"],
    "fellatio": ["paizuri", "handjob", "kneeling"],
    "all_fours": ["doggy_style", "prone_bone", "on_back"],
    "sitting_on_lap": ["straddling", "cowgirl_position", "face_sitting"],
    "against_wall": ["standing_sex", "from_behind", "bent_over"],
    "bent_over": ["against_wall", "doggy_style", "standing_sex"],
    "face_sitting": ["sixty_nine", "cowgirl_position", "sitting_on_lap"],
    "legs_up": ["mating_press", "missionary", "legs_over_head"],
    "spread_legs": ["missionary", "legs_up", "on_back"],
    "standing": ["against_wall", "standing_sex", "from_behind"],
    "sitting": ["sitting_on_lap", "straddling", "lotus_position"],
    "straddling": ["cowgirl_position", "sitting_on_lap", "face_sitting"],
    "on_back": ["missionary", "spread_legs", "legs_up"],
    "on_side": ["spooning", "from_behind", "prone_bone"],
    "on_stomach": ["prone_bone", "doggy_style", "all_fours"],
    "kneeling": ["all_fours", "doggy_style", "fellatio"],
    "squatting": ["cowgirl_position", "sitting_on_lap", "standing_sex"],
    "lotus_position": ["sitting_on_lap", "straddling", "cowgirl_position"],
    "suspended_congress": ["standing_sex", "against_wall", "mating_press"],
    "legs_over_head": ["legs_up", "mating_press", "missionary"],
    "sixty_nine": ["face_sitting", "on_back", "on_side"],
    "cunnilingus": ["face_sitting", "sixty_nine", "spread_legs"],
    "handjob": ["fellatio", "paizuri", "kneeling"],
    "leg_lock": ["missionary", "cowgirl_position", "mating_press"],
}

# intensity別 体位優先度（高intensityではより激しい体位を優先）
_POSITION_INTENSITY_PREFERENCE = {
    5: {"mating_press", "prone_bone", "suspended_congress", "legs_over_head",
        "standing_sex", "against_wall", "all_fours"},
    4: {"doggy_style", "cowgirl_position", "reverse_cowgirl", "from_behind",
        "bent_over", "standing_sex"},
    3: {"missionary", "spooning", "sitting_on_lap", "on_side"},
}

# アングル代替マップ（同一アングル連続時のフォールバック）
ANGLE_FALLBACKS = {
    "from_above": ["from_side", "pov", "dutch_angle"],
    "from_below": ["from_side", "straight-on", "dutch_angle"],
    "from_behind": ["from_side", "from_above", "pov"],
    "from_side": ["from_above", "pov", "straight-on"],
    "pov": ["from_above", "from_side", "dutch_angle"],
    "straight-on": ["from_side", "from_above", "pov"],
    "dutch_angle": ["from_side", "from_above", "pov"],
}

# アングルタグ管理（v9.0: アングル多様性）
_ANGLE_TAGS = {
    "close-up", "full_body", "upper_body", "from_behind",
    "from_above", "from_below", "pov", "dutch_angle", "side_view",
    "cowboy_shot", "portrait", "wide_shot", "from_side", "straight-on",
    "between_legs",
}

# intensity別 推奨アングルプール
_INTENSITY_ANGLE_MAP = {
    1: ["portrait", "full_body", "upper_body", "cowboy_shot"],
    2: ["upper_body", "close-up", "cowboy_shot", "side_view"],
    3: ["close-up", "from_above", "upper_body", "pov", "side_view"],
    4: ["pov", "from_behind", "from_above", "from_below", "close-up", "dutch_angle"],
    5: ["close-up", "pov", "from_below", "dutch_angle", "from_above"],
}

# SDプロンプトから除去すべき品質/スタイルタグ（v9.0: シーン固有タグのみ出力）
_QUALITY_TAGS_TO_REMOVE = {
    "masterpiece", "best_quality", "highest_quality", "absurdres",
    "highres", "very_detailed", "intricate_details", "detailed",
    "score_9", "score_8_up", "score_7_up", "score_6_up",
    "source_anime", "source_pony", "rating_explicit", "rating_questionable",
    "best quality", "high quality", "ultra detailed",
    "amazing_quality", "very_aesthetic", "newest",
}

# ── 物理状態トラッキング用タグリスト (v9.0) ──────────────────────────
_CLOTHING_STATE_TAGS = {
    "nude", "topless", "bottomless", "panties_aside", "skirt_lift",
    "shirt_lift", "bra_pull", "torn_clothes", "undressing",
    "clothes_removed", "dress_lift", "swimsuit_aside",
    "open_shirt", "unbuttoned", "no_bra", "no_panties",
    "partially_undressed", "naked_shirt", "panties_only",
    "stockings_only", "completely_nude", "naked",
}

_FLUID_STATE_TAGS = {
    "cum", "cum_on_body", "cum_on_face", "cum_in_pussy", "cum_overflow",
    "cum_string", "cum_pool", "pussy_juice", "sweat", "drooling",
    "tears", "wet", "saliva", "body_fluids", "cum_drip",
    "love_juice", "saliva_trail", "mixed_fluids", "excessive_cum",
}

_EXPRESSION_STATE_TAGS = {
    "ahegao", "blush", "crying", "tears", "panting", "drooling",
    "tongue_out", "rolling_eyes", "trembling", "heart_pupils",
    "torogao", "half-closed_eyes", "clenched_teeth", "open_mouth",
}

# ── ナラティブ→SDタグ意味連携マップ (Phase1) ──────────────────────────
# description/mood/direction/character_feelings の日本語キーワード → Danbooruタグ
_NARRATIVE_SD_MAP = {
    "position": {
        "正常位": ["missionary", "on_back"],
        "バック": ["doggy_style", "from_behind"],
        "後ろから": ["doggy_style", "from_behind"],
        "背面": ["from_behind"],
        "騎乗位": ["cowgirl_position", "girl_on_top"],
        "跨": ["cowgirl_position", "straddling"],
        "四つん這い": ["all_fours"],
        "仰向け": ["on_back", "lying"],
        "うつ伏せ": ["on_stomach", "prone"],
        "寝バック": ["prone_bone", "on_stomach"],
        "立ちバック": ["standing_sex", "from_behind"],
        "立位": ["standing_sex", "standing"],
        "座位": ["sitting", "sitting_on_lap"],
        "膝立ち": ["kneeling"],
        "開脚": ["spread_legs"],
        "足を持ち上げ": ["legs_up"],
        "抱え上げ": ["carrying", "suspended"],
        "側臥位": ["spooning", "on_side"],
        "横向き": ["on_side"],
        "駅弁": ["carrying", "legs_around_waist"],
        "対面座位": ["sitting_on_lap", "face_to_face"],
        "寝そべ": ["lying", "on_back"],
        "馬乗り": ["straddling", "girl_on_top"],
        "押し倒": ["pinned_down"],
        "圧迫": ["mating_press", "pinned_down"],
    },
    "action": {
        "挿入": ["vaginal", "penetration"],
        "突い": ["thrusting", "sex"],
        "突か": ["thrusting", "sex"],
        "腰を振": ["hip_thrust", "sex"],
        "ピストン": ["thrusting", "sex"],
        "愛撫": ["caressing"],
        "胸を揉": ["breast_grab", "groping"],
        "乳首": ["nipple_tweak"],
        "フェラ": ["fellatio", "oral"],
        "咥え": ["fellatio", "oral"],
        "舐め": ["licking"],
        "クンニ": ["cunnilingus", "oral"],
        "手マン": ["fingering"],
        "指を入": ["fingering", "insertion"],
        "パイズリ": ["paizuri"],
        "足コキ": ["footjob"],
        "素股": ["grinding", "intercrural"],
        "中出し": ["cum_in_pussy", "creampie"],
        "射精": ["ejaculation", "cum"],
        "イク": ["orgasm"],
        "絶頂": ["orgasm"],
        "アナル": ["anal"],
    },
    "expression": {
        "恥ずかし": ["embarrassed", "blush", "looking_away"],
        "照れ": ["blush", "embarrassed"],
        "涙": ["tears", "crying"],
        "泣い": ["tears", "crying"],
        "泣き": ["tears", "crying"],
        "嗚咽": ["crying", "sobbing"],
        "絶頂": ["orgasm", "ahegao"],
        "イッ": ["orgasm", "ahegao"],
        "アヘ": ["ahegao", "rolling_eyes"],
        "白目": ["rolling_eyes"],
        "トロ": ["torogao", "half-closed_eyes"],
        "恍惚": ["dazed", "half-closed_eyes"],
        "虚ろ": ["empty_eyes", "dazed"],
        "放心": ["dazed", "empty_eyes"],
        "怯え": ["scared", "trembling"],
        "怖": ["fearful", "trembling"],
        "微笑": ["smile"],
        "笑顔": ["smile", "happy"],
        "喘い": ["panting", "open_mouth"],
        "喘ぎ": ["panting", "open_mouth"],
        "息を荒": ["heavy_breathing", "panting"],
        "苦悶": ["pained_expression", "clenched_teeth"],
        "歯を食いしば": ["clenched_teeth"],
        "唇を噛": ["lip_biting"],
        "舌を出": ["tongue_out"],
        "目を見開": ["wide-eyed"],
        "目を閉じ": ["closed_eyes"],
        "目を逸ら": ["looking_away"],
        "赤面": ["blush", "embarrassed"],
        "快楽": ["pleasure", "blush"],
    },
    "body_state": {
        "汗だく": ["sweaty_body", "sweat_drops"],
        "汗": ["sweat"],
        "震え": ["trembling"],
        "ビクビク": ["trembling"],
        "痙攣": ["trembling", "convulsion"],
        "髪が乱れ": ["messy_hair"],
        "髪を振り乱": ["messy_hair"],
        "体を反ら": ["arched_back"],
        "仰け反": ["arched_back", "head_back"],
        "力が抜け": ["limp", "exhausted"],
        "ぐったり": ["exhausted", "limp"],
        "脱力": ["exhausted", "limp"],
        "肌が紅潮": ["blush", "flushed"],
        "全身が熱": ["flushed", "sweat"],
        "よだれ": ["drooling"],
    },
    "hand_action": {
        "シーツを掴": ["gripping_sheets"],
        "シーツを握": ["gripping_sheets"],
        "腰を掴": ["grabbing_hips"],
        "腰に手": ["hand_on_hip"],
        "髪を掴": ["hair_pulling"],
        "髪を引": ["hair_pulling"],
        "首に手": ["hand_on_neck"],
        "手を繋": ["holding_hands"],
        "抱き締": ["hugging"],
        "しがみつ": ["clinging"],
        "胸を掴": ["breast_grab"],
        "壁に手": ["hand_on_wall"],
    },
    "mood": {
        "背徳": ["dark_atmosphere"],
        "甘い": ["romantic"],
        "甘美": ["romantic"],
        "激しい": ["intense"],
        "激情": ["intense"],
        "切ない": ["melancholy"],
        "優しい": ["gentle"],
        "緊張": ["nervous"],
        "恐怖": ["dark_atmosphere"],
        "興奮": ["excited"],
        "羞恥": ["embarrassed"],
        "屈辱": ["humiliation"],
        "陶酔": ["dazed"],
        "官能": ["sensual"],
        "退廃": ["dark_atmosphere"],
    },
}

# mood↔表情矛盾検出・修正マップ
_MOOD_EXPRESSION_CONFLICTS = {
    # (mood keywords, conflicting SD tag, replacement SD tag)
    "negative_smile": {
        "mood_kw": ["恐怖", "屈辱", "苦痛", "嫌悪", "絶望", "怯え", "恐れ", "怖い"],
        "conflict_tag": "smile",
        "replace_with": ["fearful"],
    },
    "positive_crying": {
        "mood_kw": ["幸福", "愛情", "甘い", "喜び", "幸せ", "嬉し", "ラブラブ"],
        "conflict_tag": "crying",
        "replace_with": ["happy"],
    },
    "shame_confident": {
        "mood_kw": ["羞恥", "恥ずかし", "恥辱", "照れ"],
        "conflict_tag": "confident",
        "replace_with": ["blush", "looking_away"],
    },
}

# description↔SDタグ不整合チェック (validate_script用)
_CRITICAL_NARRATIVE_CHECKS = [
    # (日本語キーワード, 必須SDタグ候補(いずれか1つあればOK), チェック名)
    ("仰向け", {"on_back", "lying", "missionary"}, "仰向け→on_back系タグ不足"),
    ("四つん這い", {"all_fours", "doggy_style"}, "四つん這い→all_fours系タグ不足"),
    ("涙", {"tears", "crying"}, "涙→tears/cryingタグ不足"),
    ("震え", {"trembling"}, "震え→tremblingタグ不足"),
    ("フェラ", {"fellatio", "oral"}, "フェラ→fellatio/oralタグ不足"),
    ("騎乗位", {"cowgirl_position", "girl_on_top"}, "騎乗位→cowgirl系タグ不足"),
    ("正常位", {"missionary", "on_back"}, "正常位→missionary系タグ不足"),
    ("バック", {"doggy_style", "from_behind"}, "バック→doggy_style系タグ不足"),
]

# ── Danbooruタグ正規化マップ (Phase5) ──────────────────────────
# 非Danbooruタグ/誤タグ → 正規Danbooruタグ or None(除去)
_TAG_ALIAS_MAP = {
    # 非視覚的タグ → 除去
    "passionate": None,
    "intense_sex": "rough_sex",
    "gentle_sex": "sex",
    "emotional": None,
    "loving": None,
    "aggressive": None,
    "desperate": None,
    "tender": None,
    "intimate": None,
    "sensual": None,
    "erotic": None,
    # 誤タグ → 正規タグ
    "cowgirl": "cowgirl_position",
    "reverse_cowgirl": "reverse_cowgirl",
    "heart_pupils": "heart-shaped_pupils",
    "cum_overflow": "overflow",
    "gentle_smile": "smile",
    "crying_face": "crying",
    "scared_expression": "scared",
    "laying_down": "lying",
    "laying": "lying",
    "cum_inside": "cum_in_pussy",
    "vaginal_sex": "vaginal",
    "anal_sex": "anal",
    "breast_squeeze": "breast_grab",
    "breast_press": "breast_grab",
    "eye_contact": "looking_at_viewer",
    "teary_eyes": "watery_eyes",
    "blushing": "blush",
    "shaking": "trembling",
    "clothed_female_nude_male": "cfnm",
    "nude_male_clothed_female": "cfnm",
    "doggy": "doggy_style",
    "missionary_position": "missionary",
}

# 非Danbooruタグ検出用セット（validate_script用）
_NON_DANBOORU_TAGS = {
    "passionate", "intense_sex", "gentle_sex", "emotional", "loving",
    "aggressive", "desperate", "tender", "intimate", "sensual", "erotic",
    "cowgirl", "heart_pupils", "cum_overflow", "gentle_smile",
    "crying_face", "scared_expression", "laying_down", "laying",
    "cum_inside", "vaginal_sex", "anal_sex", "breast_squeeze",
    "breast_press", "eye_contact", "teary_eyes", "blushing",
    "shaking", "doggy", "missionary_position",
}

# ── キャラ特性永続化 (Phase6) ──────────────────────────
# 性格別表情バイアス（表情タグ注入時に性格に合わせた補正）
_PERSONALITY_EXPRESSION_BIAS = {
    "tsundere": {"boost": ["blush", "looking_away", "furrowed_brow"], "suppress": {"smile"}},
    "kuudere": {"boost": ["expressionless", "half-closed_eyes"], "suppress": {"smile", "blush"}},
    "dandere": {"boost": ["looking_down", "shy", "fidgeting"], "suppress": {"embarrassed"}},
    "yandere": {"boost": ["crazy_eyes", "smile", "dark_expression"], "suppress": {"embarrassed", "looking_away"}},
    "submissive": {"boost": ["downcast_eyes", "obedient"], "suppress": {"confident", "angry"}},
    "cheerful": {"boost": ["smile", "happy"], "suppress": {"expressionless"}},
    "serious": {"boost": ["determined", "furrowed_brow"], "suppress": set()},
    "shy": {"boost": ["blush", "looking_away", "covering_face"], "suppress": {"confident"}},
    "bold": {"boost": ["confident", "smirk"], "suppress": {"shy", "looking_away"}},
}

# アクセサリ永続化対象タグ
_ACCESSORY_PERSISTENT_TAGS = {
    "glasses", "hair_ribbon", "hair_ornament", "hairclip", "hairband",
    "hat", "beret", "headband", "choker", "necklace", "earrings",
    "hair_flower", "bow", "hair_bow", "scrunchie", "twintails",
    "ponytail", "braid", "twin_braids", "side_ponytail",
}


# ── 場所タグ動的解決 (v9.3再実装) ──────────────────────────
# 場所タググループ: 排他的な場所カテゴリ（同時に2つの場所にはいられない）
_LOCATION_TAG_GROUPS = {
    "train":     {"train_interior", "train", "subway", "crowded_train", "packed_train"},
    "toilet":    {"toilet", "restroom", "public_restroom", "bathroom_stall", "toilet_stall"},
    "classroom": {"classroom", "school", "chalkboard", "school_desk"},
    "office":    {"office", "desk", "cubicle", "meeting_room", "conference_room"},
    "bedroom":   {"bedroom", "bed", "on_bed", "pillow", "blanket", "futon"},
    "bathroom":  {"bathroom", "shower", "bathtub", "shower_room", "changing_room"},
    "outdoor":   {"park", "alley", "street", "rooftop", "beach", "forest"},
    "car":       {"car_interior", "car", "back_seat", "backseat"},
    "gym":       {"gym", "gym_storeroom", "locker_room", "pool", "swimming_pool"},
}

# シーンのlocation文字列→場所グループのキーワードマッピング
_LOC_GROUP_KEYWORD_MAP = {
    # train
    "電車": "train", "車内": "train", "通勤": "train", "train": "train",
    "地下鉄": "train", "subway": "train", "満員": "train",
    # toilet
    "トイレ": "toilet", "toilet": "toilet", "restroom": "toilet",
    "お手洗い": "toilet", "化粧室": "toilet",
    # classroom
    "教室": "classroom", "classroom": "classroom", "学校": "classroom",
    "school": "classroom", "放課後": "classroom",
    # office
    "オフィス": "office", "office": "office", "事務所": "office",
    "会議室": "office", "meeting": "office", "職場": "office",
    # bedroom
    "寝室": "bedroom", "bedroom": "bedroom", "ベッド": "bedroom",
    "布団": "bedroom", "自室": "bedroom", "部屋": "bedroom",
    # bathroom
    "風呂": "bathroom", "浴室": "bathroom", "bath": "bathroom",
    "shower": "bathroom", "シャワー": "bathroom", "脱衣": "bathroom",
    # outdoor
    "公園": "outdoor", "路地": "outdoor", "屋上": "outdoor",
    "rooftop": "outdoor", "alley": "outdoor", "outside": "outdoor",
    # car
    "車": "car", "car": "car", "後部座席": "car", "backseat": "car",
    # gym
    "体育": "gym", "gym": "gym", "プール": "gym", "pool": "gym",
    "更衣室": "gym", "locker": "gym",
}

# テーマ別: 場所ごとの追加タグ
_THEME_LOCATION_DETAIL_TAGS = {
    "chikan": {
        "train":    ["train_interior", "crowded", "standing", "public"],
        "toilet":   ["public_restroom", "toilet_stall", "indoors", "against_wall"],
        "outdoor":  ["alley", "public", "against_wall"],
        "default":  ["public", "crowded"],
    },
    "teacher_student": {
        "classroom": ["classroom", "chalkboard", "school_desk", "indoors"],
        "office":    ["office", "desk", "indoors"],
        "gym":       ["gym_storeroom", "indoors"],
        "bathroom":  ["shower_room", "indoors", "wet"],
        "default":   ["classroom", "indoors"],
    },
    "office": {
        "office":   ["office", "desk", "indoors", "workplace"],
        "toilet":   ["restroom", "indoors", "against_wall"],
        "default":  ["office", "indoors"],
    },
}


def _resolve_scene_location_group(scene: dict) -> str:
    """シーンのlocationフィールドから場所グループを推定"""
    loc = scene.get("location", "")
    if not loc:
        return ""
    loc_lower = loc.lower()
    for kw, group in _LOC_GROUP_KEYWORD_MAP.items():
        if kw in loc_lower:
            return group
    return ""


# ── 体位サポートタグマップ (Phase2) ──────────────────────────
# 体位タグに必須のサポートタグ（視覚的整合性のため）
_POSITION_SUPPORT_TAGS = {
    "cowgirl_position": ["girl_on_top", "straddling"],
    "reverse_cowgirl": ["girl_on_top", "straddling"],
    "missionary": ["on_back", "lying"],
    "doggy_style": ["all_fours", "from_behind"],
    "mating_press": ["on_back", "legs_up"],
    "prone_bone": ["on_stomach", "prone"],
    "fellatio": ["kneeling", "oral"],
    "standing_sex": ["standing"],
    "spooning": ["on_side", "lying"],
    "suspended": ["carrying", "legs_around_waist"],
    "lotus_position": ["sitting", "face_to_face"],
    "reverse_suspended": ["carrying", "from_behind"],
    "piledriver": ["legs_up", "upside-down"],
    "amazon_position": ["girl_on_top", "sitting_on_person"],
    "sixty-nine": ["oral", "lying"],
}


# ── タグ順序最適化グループ (Phase3) ──────────────────────────
# SDモデルは先頭のタグほど影響が強い。意味グループ別に最適順序で並べ替え
_TAG_ORDER_GROUPS = {
    # group_name: (priority, tag_set)  ※priority数値が小さいほど先頭
    "subject": (0, {"1girl", "2girls", "1boy", "2boys", "multiple_girls", "multiple_boys", "solo"}),
    "action": (2, {
        "sex", "vaginal", "anal", "penetration", "fellatio", "oral", "paizuri",
        "cunnilingus", "handjob", "footjob", "thrusting", "ejaculation", "cum",
        "creampie", "orgasm", "fingering", "grinding", "intercrural",
        "deep_penetration", "cum_in_pussy", "double_penetration",
    }),
    "position": (3, {
        "missionary", "doggy_style", "cowgirl_position", "reverse_cowgirl",
        "mating_press", "prone_bone", "standing_sex", "spooning",
        "all_fours", "on_back", "on_stomach", "sitting_on_lap",
        "spread_legs", "legs_up", "straddling", "girl_on_top",
        "from_behind", "lying", "kneeling", "standing", "prone",
        "piledriver", "lotus_position", "suspended",
    }),
    "expression": (4, {
        "ahegao", "blush", "crying", "tears", "panting", "drooling",
        "tongue_out", "rolling_eyes", "trembling", "heart_pupils",
        "torogao", "half-closed_eyes", "clenched_teeth", "open_mouth",
        "smile", "embarrassed", "scared", "fearful", "dazed",
        "looking_away", "closed_eyes", "wide-eyed", "lip_biting",
        "empty_eyes",
    }),
    "clothing": (5, {
        "nude", "naked", "completely_nude", "topless", "bottomless",
        "panties_aside", "skirt_lift", "shirt_lift", "undressing",
        "open_shirt", "no_bra", "no_panties", "partially_undressed",
        "school_uniform", "sailor_uniform", "serafuku", "blazer",
        "bikini", "swimsuit", "lingerie", "maid", "nurse",
    }),
    "body_state": (6, {
        "sweat", "sweaty_body", "sweat_drops", "cum_on_body", "cum_on_face",
        "cum_overflow", "cum_string", "cum_pool", "cum_drip",
        "pussy_juice", "body_fluids", "flushed", "arched_back",
        "exhausted", "limp",
    }),
    "hand_limb": (7, {
        "gripping_sheets", "grabbing_hips", "hair_pulling",
        "holding_hands", "hugging", "clinging", "breast_grab",
        "hand_on_wall", "hand_on_hip", "hand_on_neck",
        "caressing", "groping",
    }),
    "angle": (8, {
        "pov", "from_above", "from_below", "from_side",
        "dutch_angle", "close-up", "full_body", "cowboy_shot",
        "upper_body", "lower_body", "portrait", "wide_shot",
        "depth_of_field",
    }),
    "background": (9, {
        "classroom", "bedroom", "bathroom", "kitchen", "living_room",
        "office", "hotel_room", "onsen", "park", "forest", "beach",
        "car_interior", "train_interior", "indoors", "outdoors",
        "alley", "rooftop", "pool", "shrine", "dungeon",
    }),
    "lighting": (10, {
        "dim_lighting", "soft_lighting", "warm_lighting", "natural_lighting",
        "moonlight", "sunlight", "candlelight", "backlight",
        "neon", "golden_hour", "light_rays",
    }),
    "atmosphere": (11, {
        "romantic", "dark_atmosphere", "intense", "melancholy",
        "gentle", "sensual",
    }),
    "male": (12, {
        "faceless_male", "muscular_male", "veiny_arms",
    }),
}

# intensity連動ウェイト強化マップ (Phase3)
_INTENSITY_WEIGHT_BOOST = {
    5: {"orgasm": 1.4, "ahegao": 1.4, "cum": 1.3, "cum_in_pussy": 1.3, "rolling_eyes": 1.3},
    4: {"sex": 1.3, "penetration": 1.3, "panting": 1.2, "sweaty_body": 1.2},
    3: {"blush": 1.2, "undressing": 1.2, "open_mouth": 1.1},
}


def _reorder_sd_tags(tags: list, intensity: int = 3, char_danbooru: list = None) -> list:
    """SDタグを意味グループ別に最適順序で並べ替え。
    SDモデルは先頭のタグほど影響が強いため、重要タグを前方に配置。"""
    import re as _re_order

    # タグ→(正規化名, 元タグ) のマッピング
    tag_entries = []
    for tag in tags:
        m = _re_order.match(r'^\(([^:]+):([\d.]+)\)$', tag.strip())
        if m:
            norm = m.group(1).strip().lower().replace(" ", "_")
        else:
            norm = tag.strip().lower().replace(" ", "_")
        tag_entries.append((norm, tag))

    # キャラタグセット（priority 1）
    _char_set = set()
    if char_danbooru:
        _char_set = {c.lower().replace(" ", "_") for c in char_danbooru}

    # 各タグのグループとpriority割当
    def _get_priority(norm_tag):
        if norm_tag in _char_set:
            return 1  # character tags
        for _gname, (_prio, _gset) in _TAG_ORDER_GROUPS.items():
            if norm_tag in _gset:
                return _prio
        return 8  # デフォルト: angle付近（未分類タグ）

    # 安定ソート（同一priority内は元の順序維持）
    sorted_entries = sorted(tag_entries, key=lambda e: _get_priority(e[0]))
    return [entry[1] for entry in sorted_entries]


def _extract_narrative_sd_tags(scene: dict) -> list:
    """description/mood/direction/character_feelingsから視覚的SDタグを抽出。
    最大8タグまで。"""
    text_parts = []
    for key in ("description", "mood", "direction", "character_feelings"):
        val = scene.get(key, "")
        if val:
            text_parts.append(str(val))
    combined = " ".join(text_parts)
    if not combined:
        return []

    extracted = []
    seen = set()
    for _category, mappings in _NARRATIVE_SD_MAP.items():
        for keyword, sd_tags in mappings.items():
            if keyword in combined:
                for tag in sd_tags:
                    if tag not in seen:
                        seen.add(tag)
                        extracted.append(tag)
    # 上限8タグ
    return extracted[:8]


def _generate_negative_prompt(scene: dict, theme: str = "") -> str:
    """シーン固有のネガティブプロンプトを生成（APIコスト不要）。"""
    # ベース共通ネガティブ
    base = [
        "worst_quality", "low_quality", "bad_anatomy", "bad_hands",
        "missing_fingers", "extra_digits", "fewer_digits",
        "text", "signature", "watermark", "username",
        "blurry", "jpeg_artifacts", "cropped",
    ]

    intensity = scene.get("intensity", 3)
    sd = scene.get("sd_prompt", "")
    _sd_tags = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}
    mood = scene.get("mood", "")

    # intensity連動: 低intensity時はexplicitコンテンツ抑制
    if intensity <= 2:
        base.extend(["explicit", "nsfw", "sex", "penetration"])

    # 衣装連動: nudeなら着衣をネガティブに
    if _sd_tags & {"nude", "naked", "completely_nude"}:
        base.extend(["clothed", "dressed", "uniform"])

    # テーマ連動
    _theme_neg = {
        "humiliation": ["happy", "willing", "smile", "romantic"],
        "love": ["forced", "unwilling", "scared"],
        "sleep": ["awake", "standing", "walking"],
        "bondage": ["free_hands", "standing_freely"],
    }
    if theme in _theme_neg:
        base.extend(_theme_neg[theme])

    # 表情連動: 現在の表情と矛盾する表情をネガティブに
    if _sd_tags & {"ahegao", "rolling_eyes"}:
        base.extend(["calm", "composed", "serious"])
    if _sd_tags & {"crying", "tears"}:
        base.extend(["smile", "happy", "laughing"])
    if _sd_tags & {"smile", "happy"}:
        base.extend(["crying", "tears", "sad"])

    # 重複排除
    seen = set()
    result = []
    for tag in base:
        if tag not in seen:
            seen.add(tag)
            result.append(tag)
    return ", ".join(result)


def deduplicate_sd_tags(prompt: str) -> str:
    """SDプロンプトのタグを重複排除（順序保持）"""
    import re as _re
    tags = [t.strip() for t in prompt.split(",") if t.strip()]
    seen = set()
    result = []
    for tag in tags:
        # (tag:weight) 形式のタグからタグ名を正しく抽出
        m = _re.match(r'^\(([^:]+):([\d.]+)\)$', tag.strip())
        if m:
            normalized = m.group(1).strip().lower().replace(" ", "_")
        else:
            normalized = _re.sub(r'\([^)]*:[\d.]+\)', '', tag).strip().lower().replace(" ", "_")
        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(tag)
    return ", ".join(result)


def validate_script(results: list, theme: str = "", char_profiles: list = None) -> dict:
    """FANZA CG集基準で生成済み台本を自動検証（APIコスト不要）。

    Returns:
        dict with score, scene_issues, repeated_moans, repeated_onomatopoeia, total_issues, summary
    """
    import re as _re

    heroine_names = set()
    if char_profiles:
        for cp in char_profiles:
            name = cp.get("character_name", "")
            if name:
                heroine_names.add(name)

    scene_issues = {}
    all_moan_texts = []   # [(scene_id, text)]
    all_speech_texts = [] # [(scene_id, text)]
    all_thought_texts = [] # [(scene_id, text)]
    all_onom_sets = []    # [(scene_id, frozenset)]
    prev_angle_tags = set()
    prev_position_tags = set()
    _val_angle_history = []  # v9.0: 3連続同一アングル検出用

    for i, scene in enumerate(results):
        scene_id = scene.get("scene_id", i + 1)
        if scene.get("mood") == "エラー":
            continue
        problems = []

        # --- bubbles ---
        bubbles = scene.get("bubbles", [])

        # dialogue形式（旧フォーマット）からbubblesへのfallback変換
        if not bubbles and scene.get("dialogue"):
            bubbles = []
            for d in scene["dialogue"]:
                line_text = d.get("line", "")
                speaker = d.get("speaker", "")
                # emotionから推定: 喘ぎ系emotionならmoan、それ以外はspeech
                emotion = d.get("emotion", "")
                _moan_emotions = {"快感", "絶頂", "陶酔", "悶え", "昂り", "高潮", "恍惚"}
                btype = "moan" if emotion in _moan_emotions else "speech"
                bubbles.append({"type": btype, "speaker": speaker, "text": line_text})

        # 吹き出し数（1-3個: 主人公1-2 + 男性0-1）
        if len(bubbles) > 3:
            problems.append(f"吹き出し{len(bubbles)}個（上限3個）")
        elif len(bubbles) == 0:
            problems.append("吹き出しが0個")

        # 男セリフ数（≤1/ページ）
        male_speech_count = 0
        for b in bubbles:
            if b.get("type") == "speech":
                speaker = b.get("speaker", "")
                if speaker and heroine_names and speaker not in heroine_names:
                    male_speech_count += 1
        if male_speech_count > 1:
            problems.append(f"男性セリフ{male_speech_count}個（推奨1個以下）")

        # 男セリフ内容チェック（♡含有・喘ぎ・甘え語尾）
        for b in bubbles:
            speaker = b.get("speaker", "")
            is_male = speaker and heroine_names and speaker not in heroine_names
            if is_male:
                txt = b.get("text", "")
                btype = b.get("type", "")
                if "♡" in txt or "♥" in txt:
                    problems.append(f"男性「{speaker}」のセリフに♡: 「{txt}」")
                if btype == "moan":
                    problems.append(f"男性「{speaker}」がmoan(喘ぎ)タイプ: 「{txt}」")
                if any(k in txt for k in ["ぃ", "ぉ", "きもち", "もっとぉ", "すきぃ"]):
                    problems.append(f"男性「{speaker}」に甘え語尾: 「{txt}」")
                # 男性セリフ観察実況チェック（「～だな」系）
                import re as _re_val
                _clean_txt = txt.rstrip("…♡♥")
                if _re_val.search(r".{4,}(?:だな|してるな|だろうな)$", _clean_txt):
                    problems.append(f"男性「{speaker}」観察型: 「{txt}」")

        # thought部位ラベル冒頭チェック
        _BODY_PART_CHECK = ["胸…", "太もも…", "お尻…", "首筋…", "耳…", "唇…",
                            "舌…", "脚…", "背中…", "髪…", "うなじ…", "乳首…",
                            "おっぱい…", "ふともも…", "おしり…"]
        for b in bubbles:
            if b.get("type") == "thought":
                txt = b.get("text", "")
                if any(txt.startswith(bp) for bp in _BODY_PART_CHECK):
                    problems.append(f"thought部位ラベル冒頭: 「{txt}」")

        # 設備名混入チェック（location_detailの設備名がセリフに漏れている）
        _FIXTURE_WORDS = ["便器", "便座", "手洗い台", "手洗い鏡", "トイレットペーパー"]
        for b in bubbles:
            txt = b.get("text", "")
            if not txt:
                continue
            # 「肉便器」はエロ漫画スラングなので除外
            check_txt = txt.replace("肉便器", "")
            for fw in _FIXTURE_WORDS:
                if fw in check_txt:
                    problems.append(f"設備名混入「{fw}」検出: 「{txt}」")
                    break

        # 不自然表現チェック（書き言葉・医学用語・過剰敬語の検出）
        _UNNATURAL_WORDS = [
            "信じられない", "考えられない", "受け入れてしまう", "感じてしまう",
            "何も考えられない", "体温が上がる", "抗えない", "もう我慢できない",
            "壊れてしまいそう", "心臓が高鳴る", "全身が痺れるような",
            "理性が飛びそう", "快感が走る", "抵抗する力がなくなる",
            "体が反応してしまう", "頭が真っ白になる",
        ]
        _MEDICAL_WORDS = ["性器", "挿入", "射精", "絶頂", "愛液", "勃起", "膣内"]
        _POLITE_WORDS = [
            "してもよろしいですか", "感じてしまいます", "見ないでください",
            "触らないでください", "行ってしまいます", "出てしまいます",
            "止められません", "お願いします", "ありがとうございます",
            "気持ちいいです", "嬉しい気持ちです", "大丈夫です",
        ]
        for b in bubbles:
            txt = b.get("text", "")
            if not txt:
                continue
            for uw in _UNNATURAL_WORDS:
                if uw in txt:
                    problems.append(f"不自然表現「{uw}」検出: 「{txt}」")
                    break
            for mw in _MEDICAL_WORDS:
                if mw in txt:
                    problems.append(f"医学用語「{mw}」検出: 「{txt}」")
                    break
            for pw in _POLITE_WORDS:
                if pw in txt:
                    problems.append(f"過剰敬語「{pw}」検出: 「{txt}」")
                    break

        # moanタイプ内容検証（3段階: 漢字/助詞/非喘ぎ語彙チェック）
        # 根拠: MOAN_POOL全400エントリは100%仮名+装飾(♡…っー゛)。
        #   1. 漢字含有 → 非喘ぎ確定
        #   2. 文末助詞 → 会話文が混入
        #   3. 非喘ぎ語彙 → AFTERMATH_POOL等の身体状況報告が混入
        _NON_MOAN_WORDS = frozenset([
            "ぼーっと", "ぐったり", "ふわふわ", "ごめん", "どしよ",
            "なにこれ", "もうむり", "もう…むり", "なにこれ…", "ごめん…",
            "どしよ…", "ぼーっと…", "ぐったり…", "ふわふわ…",
        ])
        for b in bubbles:
            if b.get("type") == "moan":
                txt = b.get("text", "")
                if not txt:
                    continue
                has_kanji = bool(_re.search(r'[\u4e00-\u9faf\u3400-\u4dbf]', txt))
                has_sentence_ending = bool(_re.search(
                    r'(だ|です|ます|ない|ない…|ている|てる|する|される|して|した|しい)$', txt))
                is_non_moan_word = txt.rstrip("…♡♡♡") in _NON_MOAN_WORDS or txt in _NON_MOAN_WORDS
                if has_kanji or has_sentence_ending or is_non_moan_word:
                    problems.append(f"moanタイプに非喘ぎテキスト: 「{txt}」")

        # speechタイプ身体状況報告チェック（intensity>=3のアクションシーン）
        # 根拠: CG集のspeechは感情的反応。「汗すごい」「指先痺れ」等の
        #        身体状態の客観的報告はナレーション/ト書きであり、セリフとして不自然。
        _BODY_REPORT_KW = [
            "涙が", "汗すごい", "汗が", "声出ない", "息できない",
            "力入んない", "頭まっしろ", "目が回る", "指先痺れ",
            "全身痺れ", "まだ震えて", "震えてる", "動けない",
            "立てない", "からだ重い", "呼吸が", "ぼーっと",
            "ぐったり", "ふわふわ", "思考が", "意識が",
        ]
        if scene.get("intensity", 0) >= 3:
            for b in bubbles:
                if b.get("type") in ("speech", "thought"):
                    txt = b.get("text", "")
                    for kw in _BODY_REPORT_KW:
                        if kw in txt:
                            problems.append(f"身体状況報告セリフ: 「{txt}」（{b.get('type')}）")
                            break

        # 同一シーン内テキスト重複チェック
        bubble_texts_in_scene = [b.get("text", "") for b in bubbles if b.get("text")]
        if len(bubble_texts_in_scene) != len(set(bubble_texts_in_scene)):
            problems.append(f"同一シーン内にテキスト重複あり")

        # moan・speech・thought追跡（クロスシーン重複検出用）
        for b in bubbles:
            if b.get("type") == "moan":
                all_moan_texts.append((scene_id, b.get("text", "")))
            elif b.get("type") == "speech":
                all_speech_texts.append((scene_id, b.get("text", "")))
            elif b.get("type") == "thought":
                all_thought_texts.append((scene_id, b.get("text", "")))

        # --- onomatopoeia ---
        onom = scene.get("onomatopoeia", [])
        all_onom_sets.append((scene_id, frozenset(onom) if onom else frozenset()))

        # --- 必須フィールド ---
        for field in ("title", "description", "mood", "sd_prompt", "direction"):
            if not scene.get(field):
                problems.append(f"「{field}」が空")

        # --- description品質 ---
        desc = scene.get("description", "")
        intensity = scene.get("intensity", 0)
        if 0 < len(desc) < 30:
            problems.append(f"description短すぎ（{len(desc)}文字）")
        if intensity >= 4 and desc:
            # 具体性キーワード: これらが1つでもあれば具体的と判定
            _CONCRETE_DESC_KW = [
                # 体位・行為
                "正常位", "後背位", "騎乗位", "背面", "立ち", "座位",
                "バック", "対面", "側位", "寝バック", "駅弁",
                "挿入", "ピストン", "腰を", "突き", "押し当て",
                "咥え", "舐め", "吸い", "しゃぶ", "フェラ", "パイズリ",
                "手コキ", "指を", "弄", "愛撫し",
                # 身体反応
                "汗", "涙", "震え", "痙攣", "力が抜け", "仰け反",
                "ビクビク", "ガクガク", "びくっ", "跳ね",
                # 具体的な動き
                "掴み", "押さえ", "引き寄せ", "しがみつ", "抱き",
                "開かせ", "持ち上げ", "覆いかぶさ", "跨", "乗り",
                "四つん這い", "うつ伏せ", "仰向け", "膝立ち",
                # 体の部位（具体的描写の指標）
                "胸を", "腰を", "脚を", "太もも", "尻を", "首筋",
            ]
            if not any(kw in desc for kw in _CONCRETE_DESC_KW):
                problems.append("descriptionが抽象的（具体的な体位・行為を記述すべき）")

        # --- sd_prompt: 日本語混入 ---
        sd = scene.get("sd_prompt", "")
        jp_chars = _re.findall(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]+', sd)
        if jp_chars:
            problems.append(f"sd_promptに日本語: {', '.join(jp_chars[:3])}")

        # --- sd_prompt: 連続同一アングル ---
        angle_kw = {"from_above", "from_below", "from_behind", "from_side",
                    "pov", "straight-on", "dutch_angle", "close-up",
                    "full_body", "upper_body", "cowboy_shot", "portrait",
                    "wide_shot", "side_view", "between_legs"}
        cur_angles = {kw for kw in angle_kw if kw in sd.lower()}
        if cur_angles and cur_angles == prev_angle_tags:
            problems.append(f"前シーンと同一アングル: {', '.join(cur_angles)}")
        # v9.0: 3連続同一アングル検出
        _cur_angle_key = frozenset(cur_angles) if cur_angles else None
        _val_angle_history.append(_cur_angle_key)
        if (len(_val_angle_history) >= 3
                and _cur_angle_key is not None
                and _val_angle_history[-2] == _cur_angle_key
                and _val_angle_history[-3] == _cur_angle_key):
            problems.append(f"3連続同一アングル: {', '.join(cur_angles)}")
        prev_angle_tags = cur_angles

        # --- sd_prompt: 連続同一体位 ---
        sd_tags_norm = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}
        cur_positions = sd_tags_norm & POSITION_TAGS
        if cur_positions and cur_positions == prev_position_tags:
            problems.append(f"前シーンと同一体位: {', '.join(cur_positions)}")
        prev_position_tags = cur_positions

        # --- sd_prompt: 室内外タグ矛盾 ---
        sd_low = sd.lower()
        outdoor_markers = {"outdoors", "park", "forest", "beach", "poolside", "rooftop", "garden"}
        indoor_markers = {"indoors", "classroom", "bedroom", "bathroom", "kitchen", "elevator",
                          "office", "living_room", "train_interior", "car_interior"}
        indoor_only_tags = {"ceiling", "fluorescent_light", "wallpaper", "chandelier",
                            "carpet", "wooden_floor", "tile_floor", "ceiling_fan"}
        outdoor_only_tags = {"sky", "cloud", "horizon", "grass", "trees", "ocean", "sun"}
        sd_tags_set = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}
        has_outdoor = bool(sd_tags_set & outdoor_markers)
        has_indoor = bool(sd_tags_set & indoor_markers)
        has_window = "window" in sd_low
        if has_outdoor:
            bad = sd_tags_set & indoor_only_tags
            if bad:
                problems.append(f"室内外矛盾: outdoor+{','.join(list(bad)[:3])}")
        if has_indoor and not has_window:
            bad = sd_tags_set & outdoor_only_tags
            if bad and "open_air_bath" not in sd_low:
                problems.append(f"室内外矛盾: indoor+{','.join(list(bad)[:3])}(window無し)")

        # --- sd_prompt: 照明-時間帯整合性 ---
        morning_kw = {"morning", "sunrise", "daytime", "afternoon"}
        night_kw = {"night", "midnight", "late_night"}
        night_light_bad = {"sunlight", "bright_daylight", "blue_sky", "morning_light"}
        morning_light_bad = {"moonlight", "darkness", "night_sky", "starlight"}
        time_in_sd = sd_tags_set & (morning_kw | night_kw)
        if time_in_sd & morning_kw:
            bad = sd_tags_set & morning_light_bad
            if bad:
                problems.append(f"照明矛盾: 朝昼+{','.join(list(bad)[:2])}")
        if time_in_sd & night_kw:
            bad = sd_tags_set & night_light_bad
            if bad:
                problems.append(f"照明矛盾: 夜+{','.join(list(bad)[:2])}")

        # --- sd_prompt: 背景タグ存在確認 ---
        bg_tags = {
            # 基本
            "outdoors", "indoors",
            # 学校
            "classroom", "library", "gym", "hallway", "stairwell",
            "locker_room", "infirmary", "rooftop", "club_room",
            "storage_room", "school",
            # 住居
            "bedroom", "bathroom", "kitchen", "living_room",
            "japanese_room", "balcony", "basement", "study",
            "entrance", "closet", "garage",
            # 商業・仕事
            "office", "elevator", "warehouse", "factory",
            "convenience_store", "store",
            # 宿泊
            "hotel_room", "ryokan_room", "inn_room", "cabin",
            # 飲食
            "cafe", "restaurant", "izakaya", "bar", "cafeteria",
            # 交通
            "car_interior", "train_interior", "bus_interior",
            "airplane_interior", "ship_interior", "train_station",
            # 娯楽
            "karaoke_room", "internet_cafe", "arcade", "theater",
            "studio",
            # 屋外・自然
            "park", "forest", "beach", "mountain", "river", "lake",
            "garden", "alley", "bridge", "riverbank", "field",
            "grassland", "cliff", "cave",
            # 風呂・温泉
            "onsen", "bath", "pool", "open_air_bath", "bathhouse",
            "sauna",
            # 宗教
            "shrine", "temple", "church", "graveyard",
            # ファンタジー
            "dungeon", "castle", "tower", "prison", "tavern",
            "throne_room",
            # SF
            "spaceship_interior", "laboratory", "cockpit",
            # 日本建築
            "engawa", "storehouse", "barn",
        }
        if sd and not (sd_tags_set & bg_tags):
            problems.append("sd_promptに背景/場所タグが無い")

        # --- sd_prompt: v9.0 品質/スタイルタグ混入検出 ---
        if sd:
            _found_quality = sd_tags_set & _QUALITY_TAGS_TO_REMOVE
            if _found_quality:
                problems.append(f"sd_promptに品質タグ混入: {', '.join(list(_found_quality)[:3])}")
            if "<lora:" in sd.lower():
                problems.append("sd_promptにLoRAタグ混入")

        if problems:
            scene_issues[scene_id] = problems

    # --- クロスシーン: story_flow重複チェック（完全一致 + 高類似度） ---
    seen_flows = {}  # flow_text -> scene_id
    for i, scene in enumerate(results):
        flow = scene.get("story_flow", "")
        if not flow or len(flow) < 10:
            continue
        scene_id = scene.get("scene_id", i + 1)
        # 完全一致チェック
        if flow in seen_flows:
            scene_issues.setdefault(scene_id, []).append(
                f"story_flow重複（シーン{seen_flows[flow]}と完全同一）")
        else:
            # 高類似度チェック（先頭20文字一致 = ほぼコピペ）
            flow_prefix = flow[:20]
            for prev_flow, prev_sid in seen_flows.items():
                if prev_flow[:20] == flow_prefix and prev_sid != scene_id:
                    scene_issues.setdefault(scene_id, []).append(
                        f"story_flow類似（シーン{prev_sid}と先頭20字一致）")
                    break
            seen_flows[flow] = scene_id

    # --- v8.2: クロスシーン: story_flow構造テンプレ検出 ---
    _SF_NORMALIZE_RE = _re.compile(r'(挿入|中出し|絶頂|愛撫|フェラ|座位|正常位|バック|騎乗位|側位|対面)')
    def _sf_skeleton(text):
        return _SF_NORMALIZE_RE.sub("\u25c6", text[:30])
    sf_skeletons = {}
    for r in results:
        sf = r.get("story_flow", "")
        if sf:
            sk = _sf_skeleton(sf)
            sf_skeletons[sk] = sf_skeletons.get(sk, 0) + 1
    threshold_sf = max(3, len(results) // 8)
    for sk, cnt in sf_skeletons.items():
        if cnt >= threshold_sf:
            scene_issues.setdefault("global", []).append(
                f"story_flow構造反復: 「{sk[:20]}…」が{cnt}回")

    # --- クロスシーン: description類似チェック（先頭30文字一致=コピペ） ---
    seen_descs = {}  # desc_prefix -> scene_id
    for i, scene in enumerate(results):
        desc = scene.get("description", "")
        if not desc or len(desc) < 15:
            continue
        scene_id = scene.get("scene_id", i + 1)
        desc_prefix = desc[:15]  # v8.2根本修正: 30字→15字に短縮
        if desc_prefix in seen_descs:
            scene_issues.setdefault(scene_id, []).append(
                f"description類似（シーン{seen_descs[desc_prefix]}と先頭15字一致）")
        else:
            seen_descs[desc_prefix] = scene_id

    # --- v8.2: クロスシーン: description冒頭10字反復チェック ---
    desc_prefix10 = {}
    for r in results:
        d = r.get("description", "")[:10]
        if d:
            desc_prefix10[d] = desc_prefix10.get(d, 0) + 1
    threshold_dp = max(3, len(results) // 6)
    for d, cnt in desc_prefix10.items():
        if cnt >= threshold_dp:
            scene_issues.setdefault("global", []).append(
                f"description冒頭「{d}」が{cnt}回反復")

    # --- v8.2根本修正: ストーリー膠着検出 ---
    # title重複（同一title3回以上 = 同じイベントの繰り返し）
    _title_counter_v = {}
    for r in results:
        t = r.get("title", "")
        if t:
            _title_counter_v[t] = _title_counter_v.get(t, 0) + 1
    for t, cnt in _title_counter_v.items():
        if cnt >= 3:
            scene_issues.setdefault("global", []).append(
                f"ストーリー膠着: title「{t[:15]}」が{cnt}回反復（同一イベント繰り返し）")
    # situation先頭20字が3シーン以上で一致
    _sit_prefix = {}
    for r in results:
        s = r.get("situation", r.get("description", ""))[:20]
        if s:
            _sit_prefix[s] = _sit_prefix.get(s, 0) + 1
    for s, cnt in _sit_prefix.items():
        if cnt >= 3:
            scene_issues.setdefault("global", []).append(
                f"ストーリー膠着: 状況「{s}」が{cnt}回反復")

    # --- v8.2: クロスシーン: mood反復チェック ---
    mood_counter = {}
    for r in results:
        m = r.get("mood", "")
        if m:
            mood_counter[m] = mood_counter.get(m, 0) + 1
    threshold_mood = max(3, len(results) // 5)
    for m, cnt in mood_counter.items():
        if cnt >= threshold_mood:
            scene_issues.setdefault("global", []).append(
                f"mood「{m[:15]}」が{cnt}回反復（{cnt}シーンで同一mood）")
    # mood先頭6字反復
    mood_prefix6 = {}
    for r in results:
        m = r.get("mood", "")[:6]
        if m:
            mood_prefix6[m] = mood_prefix6.get(m, 0) + 1
    threshold_mp = max(4, len(results) // 4)
    for m, cnt in mood_prefix6.items():
        if cnt >= threshold_mp:
            scene_issues.setdefault("global", []).append(
                f"mood接頭辞「{m}」が{cnt}回反復")

    # --- クロスシーン: title長さチェック ---
    for i, scene in enumerate(results):
        title = scene.get("title", "")
        scene_id = scene.get("scene_id", i + 1)
        if len(title) > 25:
            scene_issues.setdefault(scene_id, []).append(
                f"title長すぎ({len(title)}字): 「{title[:30]}...」")

    # --- クロスシーン: title品質チェック（句点混入・location混入） ---
    _VALIDATE_LEAK_WORDS = ["タイル", "白い壁", "天井", "床", "ベンチ", "洗面台",
                            "カーテン", "ドア", "窓", "机", "排水", "蛇口"]
    _TITLE_LOCATION_KW = [
        "トイレ", "個室", "便所", "教室", "部室", "保健室", "屋上", "体育館",
        "プール", "更衣室", "シャワー室", "ベッドルーム", "リビング", "キッチン",
        "浴室", "風呂", "脱衣所", "車内", "電車内", "バス内", "駐車場",
        "エレベーター", "階段", "廊下", "倉庫", "物置", "地下室", "ホテル",
        "旅館", "カラオケ", "ネカフェ", "漫喫", "オフィス", "会議室",
    ]
    for i, scene in enumerate(results):
        title = scene.get("title", "")
        scene_id = scene.get("scene_id", i + 1)
        if not title:
            continue
        # 句点チェック（タイトルに「。」は不適）
        if "。" in title:
            scene_issues.setdefault(scene_id, []).append(
                f"titleに句点混入: 「{title}」")
        # location混入チェック（場所名がtitleに含まれる）
        loc = scene.get("location_detail", scene.get("location", ""))
        if loc and len(loc) >= 4 and loc in title:
            scene_issues.setdefault(scene_id, []).append(
                f"titleにlocation混入: 「{title}」（location: {loc}）")
        # 場所キーワードが支配的なタイトル（場所KWが2つ以上）
        loc_kw_count = sum(1 for kw in _TITLE_LOCATION_KW if kw in title)
        if loc_kw_count >= 2:
            scene_issues.setdefault(scene_id, []).append(
                f"titleが場所名の羅列: 「{title}」（場所KW{loc_kw_count}個）")
        # location leak語チェック（建材/設備名がtitleに混入）
        for lw in _VALIDATE_LEAK_WORDS:
            if lw in title:
                scene_issues.setdefault(scene_id, []).append(
                    f"title location leak: 「{title}」（{lw}混入）")
                break
        # 末尾切断チェック（1文字助詞で終わる不自然なタイトル）
        if len(title) >= 3 and title[-1] in "新のとがをにでへは":
            scene_issues.setdefault(scene_id, []).append(
                f"title末尾切断: 「{title}」（「{title[-1]}」で終了）")
        # description断片混入チェック（助詞「の」で始まる/名詞で途切れる不完全title）
        desc = scene.get("description", "")
        if desc and len(title) >= 6:
            # titleがdescriptionの部分文字列（10文字以上の断片）
            if len(title) >= 10 and title in desc:
                scene_issues.setdefault(scene_id, []).append(
                    f"titleにdescription断片混入: 「{title}」")
            # 「の」で始まるtitle（文の途中から切り取られた形跡）
            elif title.startswith("の") or title.startswith("と"):
                scene_issues.setdefault(scene_id, []).append(
                    f"title不完全（助詞で開始）: 「{title}」")
            # 「張」「れ」等の送り仮名で終わる（文の途中で途切れた形跡）
            elif len(title) >= 8 and title[-1] in "張貼掛掲載映写":
                scene_issues.setdefault(scene_id, []).append(
                    f"title不完全（途中で途切れ）: 「{title}」")

    # --- v8.2: クロスシーン: title接頭辞2字反復チェック ---
    title_prefix2 = {}
    for r in results:
        t = r.get("title", "")[:2]
        if t:
            title_prefix2[t] = title_prefix2.get(t, 0) + 1
    threshold_tp = max(4, len(results) // 6)
    for t, cnt in title_prefix2.items():
        if cnt >= threshold_tp:
            scene_issues.setdefault("global", []).append(
                f"title接頭辞「{t}」が{cnt}回反復")

    # --- クロスシーン: title重複チェック ---
    seen_titles = {}  # title -> scene_id
    for i, scene in enumerate(results):
        title = scene.get("title", "")
        if not title:
            continue
        scene_id = scene.get("scene_id", i + 1)
        if title in seen_titles:
            scene_issues.setdefault(scene_id, []).append(
                f"title重複「{title}」（シーン{seen_titles[title]}と同一）")
        else:
            seen_titles[title] = scene_id

    # --- クロスシーン: titleキーワード過剰使用チェック ---
    _title_kw_count = {}
    _TITLE_CHECK_KW = ["膣奥", "膣", "理性", "崩壊", "限界", "快感", "堕ち", "抵抗",
                        "連続", "激突", "責め", "声", "最後", "扉", "壁", "視線",
                        "奥", "腰", "廊下"]
    for scene in results:
        title = scene.get("title", "")
        for kw in _TITLE_CHECK_KW:
            if kw in title:
                _title_kw_count[kw] = _title_kw_count.get(kw, 0) + 1
    _total_scenes = len(results)
    _title_kw_threshold = max(3, _total_scenes // 10)  # 10シーンにつき1回まで許容
    for kw, cnt in _title_kw_count.items():
        if cnt >= _title_kw_threshold:
            scene_issues.setdefault("global", []).append(
                f"titleキーワード過剰: 「{kw}」が{cnt}回使用（推奨2回以下）")

    # --- クロスシーン: description連続類似チェック（3連続で同一行為キーワード） ---
    _DESC_ACT_KW = ["膣奥", "突かれ", "責められ", "腰を振", "ピストン",
                     "挿入", "フェラ", "パイズリ", "騎乗", "バック",
                     "正常位", "四つん這い"]
    _desc_kw_per_scene = []
    for scene in results:
        desc = scene.get("description", "")
        kws = frozenset(kw for kw in _DESC_ACT_KW if kw in desc)
        _desc_kw_per_scene.append(kws)
    for k in range(2, len(_desc_kw_per_scene)):
        common = _desc_kw_per_scene[k] & _desc_kw_per_scene[k-1] & _desc_kw_per_scene[k-2]
        if len(common) >= 2:  # 2キーワード以上一致で類似判定（1つだけなら正常）
            sid = results[k].get("scene_id", k + 1)
            scene_issues.setdefault(sid, []).append(
                f"description3連続類似（行為キーワード同一: {common}）")

    # --- クロスシーン: character_feelings類似チェック ---
    seen_feelings = {}  # feelings_str -> scene_id
    for i, scene in enumerate(results):
        feelings = scene.get("character_feelings", {})
        if not feelings or not isinstance(feelings, dict):
            continue
        scene_id = scene.get("scene_id", i + 1)
        feelings_str = str(sorted(feelings.values()))
        if len(feelings_str) < 15:
            continue
        if feelings_str in seen_feelings:
            scene_issues.setdefault(scene_id, []).append(
                f"character_feelings重複（シーン{seen_feelings[feelings_str]}と同一）")
        else:
            seen_feelings[feelings_str] = scene_id

    # --- クロスシーン: scene_id重複チェック ---
    scene_ids = [s.get("scene_id", i+1) for i, s in enumerate(results)]
    if len(scene_ids) != len(set(scene_ids)):
        dupes = [sid for sid in scene_ids if scene_ids.count(sid) > 1]
        for sid in set(dupes):
            scene_issues.setdefault(sid, []).append(f"scene_id {sid} が重複している")

    # --- クロスシーン: bubble完全重複チェック ---
    prev_bubble_set = set()
    for i, scene in enumerate(results):
        scene_id = scene.get("scene_id", i + 1)
        bubbles = scene.get("bubbles", [])
        curr_bubble_set = frozenset(b.get("text", "") for b in bubbles if b.get("text"))
        if curr_bubble_set and curr_bubble_set == prev_bubble_set:
            scene_issues.setdefault(scene_id, []).append("前シーンとbubbleが完全同一（重複）")
        prev_bubble_set = curr_bubble_set

    # --- クロスシーン: 喘ぎ重複（類似マッチ含む） ---
    moan_map = {}
    for sid, text in all_moan_texts:
        moan_map.setdefault(text, []).append(sid)
    repeated_moans = {t: sids for t, sids in moan_map.items() if len(sids) > 1}
    # 類似喘ぎ検出（正規化ハッシュ + 先頭4文字バケット → O(n)）
    moan_norm_map = {}  # normalized_text -> [(sid, original_text)]
    moan_prefix_map = {}  # prefix4 -> [(sid, original_text)]
    for sid, text in all_moan_texts:
        norm = _normalize_bubble_text(text)
        moan_norm_map.setdefault(norm, []).append((sid, text))
        if len(norm) >= 4:
            prefix = norm[:4]
            moan_prefix_map.setdefault(prefix, []).append((sid, text))
    # 正規化一致で重複検出
    for norm, entries in moan_norm_map.items():
        if len(entries) > 1:
            unique_texts = {}
            for sid, text in entries:
                unique_texts.setdefault(text, []).append(sid)
            texts = list(unique_texts.keys())
            for i in range(len(texts)):
                for j in range(i + 1, len(texts)):
                    key = f"{texts[i]}≈{texts[j]}"
                    if key not in repeated_moans:
                        repeated_moans[key] = [unique_texts[texts[i]][0], unique_texts[texts[j]][0]]
    # 先頭4文字一致で類似検出
    for prefix, entries in moan_prefix_map.items():
        if len(entries) > 1:
            unique_texts = {}
            for sid, text in entries:
                unique_texts.setdefault(text, []).append(sid)
            texts = list(unique_texts.keys())
            for i in range(len(texts)):
                for j in range(i + 1, len(texts)):
                    key = f"{texts[i]}≈{texts[j]}"
                    if key not in repeated_moans:
                        repeated_moans[key] = [unique_texts[texts[i]][0], unique_texts[texts[j]][0]]

    # --- クロスシーン: speech重複チェック ---
    speech_map = {}
    for sid, text in all_speech_texts:
        speech_map.setdefault(text, []).append(sid)
    repeated_speech = {t: sids for t, sids in speech_map.items() if len(sids) > 1}
    for text, sids in repeated_speech.items():
        for sid in sids[1:]:
            scene_issues.setdefault(sid, []).append(f"speech重複「{text}」（シーン{sids[0]}と同一）")

    # --- クロスシーン: thought先頭パターン反復チェック ---
    # 先頭パターンが同じthoughtが4回以上出現した場合に警告（「だめ…」パターン等）
    thought_prefix_counter = {}  # prefix -> [(scene_id, full_text)]
    for sid, text in all_thought_texts:
        if text and len(text) >= 2:
            # 最初の「…」より前を取得（「だめ…声が…」→「だめ」）
            first_part = text.split("\u2026")[0].replace("\u2665", "").replace("\u3063", "").strip()
            if len(first_part) >= 2:
                prefix = first_part[:3]
            else:
                clean = text.replace("\u2026", "").replace("\u2665", "").replace("\u3063", "").strip()
                prefix = clean[:2] if len(clean) >= 2 else ""
            if prefix:
                thought_prefix_counter.setdefault(prefix, []).append((sid, text))
    for prefix, entries in thought_prefix_counter.items():
        if len(entries) >= 4:
            scene_ids_str = ",".join(str(e[0]) for e in entries[:6])
            scene_issues.setdefault("global", []).append(
                f"thought先頭「{prefix}」が{len(entries)}回反復（シーン{scene_ids_str}）")

    # --- クロスシーン: thoughtテキスト内キーワード頻度チェック ---
    _THOUGHT_CONTENT_KW = [
        "だめ", "声", "やめて", "おく", "なか", "廊下", "聞こえ",
        # v8.2追加: 感情サイクリング検出
        "こわい", "きもち", "いや", "すき", "もう", "ほしい",
        "おかしく", "とまら", "しんじ", "たすけ", "はずか", "にげ",
    ]
    _thought_kw_count = {}
    for sid, text in all_thought_texts:
        for kw in _THOUGHT_CONTENT_KW:
            if kw in text:
                _thought_kw_count[kw] = _thought_kw_count.get(kw, 0) + 1
    _thought_total = max(1, len(all_thought_texts))
    for kw, cnt in _thought_kw_count.items():
        # 全thoughtの25%以上で同一キーワードが出現したら警告
        if cnt >= max(4, _thought_total // 4):
            scene_issues.setdefault("global", []).append(
                f"thoughtキーワード過剰: 「{kw}」が{cnt}/{_thought_total}回出現（25%超過）")

    # --- クロスシーン: 男性セリフ長文チェック（15文字超え） ---
    for i, scene in enumerate(results):
        scene_id = scene.get("scene_id", i + 1)
        for b in scene.get("bubbles", []):
            speaker = b.get("speaker", "")
            is_male = speaker and heroine_names and speaker not in heroine_names
            if is_male and b.get("type") == "speech":
                txt = b.get("text", "")
                # ♡…っ等の装飾を除いた実質文字数
                core = txt.replace("…", "").replace("♡", "").replace("っ", "").replace("♥", "").strip()
                if len(core) > 15:
                    scene_issues.setdefault(scene_id, []).append(
                        f"男性セリフ長文({len(core)}字): 「{txt}」")

    # --- クロスシーン: 男性セリフ末尾フレーズ反復チェック（「最高だ」等が3回以上） ---
    _male_speech_suffix_counter = {}  # suffix -> [(scene_id, full_text)]
    _MALE_SUFFIX_CHECK_LEN = 4  # 末尾N文字で比較
    _MALE_SUFFIX_THRESHOLD = max(3, len(results) // 15)  # 15シーンに1回まで許容
    for i, scene in enumerate(results):
        scene_id = scene.get("scene_id", i + 1)
        for b in scene.get("bubbles", []):
            speaker = b.get("speaker", "")
            is_male = speaker and heroine_names and speaker not in heroine_names
            if is_male and b.get("type") == "speech":
                txt = b.get("text", "")
                if not txt or len(txt) < 3:
                    continue
                # 装飾除去して末尾フレーズを抽出
                core = txt.replace("…", "").replace("♡", "").replace("っ", "").replace("♥", "").strip()
                if len(core) < 3:
                    continue
                suffix = core[-_MALE_SUFFIX_CHECK_LEN:] if len(core) >= _MALE_SUFFIX_CHECK_LEN else core
                _male_speech_suffix_counter.setdefault(suffix, []).append((scene_id, txt))
    for suffix, entries in _male_speech_suffix_counter.items():
        if len(entries) >= _MALE_SUFFIX_THRESHOLD:
            scene_ids_str = ",".join(str(e[0]) for e in entries[:6])
            scene_issues.setdefault("global", []).append(
                f"男性セリフ末尾「{suffix}」が{len(entries)}回反復（シーン{scene_ids_str}）")

    # --- クロスシーン: 男性セリフ多様性スコア（ユニーク率50%未満で警告） ---
    _all_male_speeches = []
    for scene in results:
        for b in scene.get("bubbles", []):
            speaker = b.get("speaker", "")
            is_male = speaker and heroine_names and speaker not in heroine_names
            if is_male and b.get("type") == "speech":
                txt = b.get("text", "")
                if txt:
                    _all_male_speeches.append(txt)
    if len(_all_male_speeches) >= 5:
        _unique_male = len(set(_all_male_speeches))
        _diversity_pct = _unique_male * 100 // len(_all_male_speeches)
        if _diversity_pct < 50:
            scene_issues.setdefault("global", []).append(
                f"男性セリフ多様性低: {_unique_male}/{len(_all_male_speeches)}({_diversity_pct}%ユニーク)")

    # --- クロスシーン: 不自然表現チェック（「らめ」「括弧」「一人称ブレ」） ---
    for i, scene in enumerate(results):
        scene_id = scene.get("scene_id", i + 1)
        for b in scene.get("bubbles", []):
            txt = b.get("text", "")
            if "らめ" in txt:
                scene_issues.setdefault(scene_id, []).append(
                    f"不自然表現「らめ」: 「{txt}」")
            if "」" in txt or "「" in txt:
                scene_issues.setdefault(scene_id, []).append(
                    f"括弧混入: 「{txt}」")

    # --- クロスシーン: 吹き出し内ナレーション混入検出 ---
    for i, scene in enumerate(results):
        scene_id = scene.get("scene_id", i + 1)
        for b in scene.get("bubbles", []):
            txt = b.get("text", "")
            btype = b.get("type", "")
            if not txt:
                continue
            # 句号（。）は吹き出しに不適（ナレーション混入の兆候）
            if "。" in txt:
                scene_issues.setdefault(scene_id, []).append(
                    f"吹き出しに句号: 「{txt[:25]}」")
            # moanに説明文・会話文が混入（漢字3文字以上連続 = 喘ぎではない）
            if btype == "moan" and _re.search(r'[\u4e00-\u9faf]{3,}', txt):
                scene_issues.setdefault(scene_id, []).append(
                    f"moanに説明文混入: 「{txt[:25]}」")

    # --- クロスシーン: description外見反復検出 ---
    _desc_appearance_seqs = []
    for scene in results:
        desc = scene.get("description", "")
        # 先頭30文字から外見キーワードを抽出
        _desc_appearance_seqs.append(desc[:30] if desc else "")
    for k in range(2, len(_desc_appearance_seqs)):
        d0, d1, d2 = _desc_appearance_seqs[k-2], _desc_appearance_seqs[k-1], _desc_appearance_seqs[k]
        if d0 and d1 and d2 and d0 == d1 == d2:
            sid = results[k].get("scene_id", k + 1)
            scene_issues.setdefault(sid, []).append(
                f"description先頭3連続同一: 「{d0[:20]}…」")

    # --- クロスシーン: オノマトペ近接重複（3シーン以内） ---
    repeated_onom = []
    for k in range(1, len(all_onom_sets)):
        cur_sid, cur_set = all_onom_sets[k]
        if not cur_set:
            continue
        for j in range(max(0, k - 3), k):
            _, prev_set = all_onom_sets[j]
            if prev_set and cur_set == prev_set:
                repeated_onom.append((all_onom_sets[j][0], cur_sid))
                break

    # --- クロスシーン: 3シーン連続同一location ---
    locations_list = []
    for scene in results:
        loc = scene.get("location_detail", scene.get("location", ""))
        locations_list.append(loc.strip().lower() if loc else "")
    for k in range(2, len(locations_list)):
        if locations_list[k] and locations_list[k] == locations_list[k-1] == locations_list[k-2]:
            sid = results[k].get("scene_id", k + 1)
            scene_issues.setdefault(sid, []).append(
                f"3シーン連続同一location: {locations_list[k]}")

    # --- クロスシーン: アングル全体分布偏り ---
    angle_counter = {}
    for scene in results:
        sd_text = scene.get("sd_prompt", "").lower()
        for akw in ("from_above", "from_below", "from_behind", "from_side",
                     "pov", "straight-on", "dutch_angle"):
            if akw in sd_text:
                angle_counter[akw] = angle_counter.get(akw, 0) + 1
    total_scenes = len(results)
    if total_scenes >= 5:
        for akw, cnt in angle_counter.items():
            if cnt / total_scenes >= 0.4:
                scene_issues.setdefault("global", []).append(
                    f"アングル偏り: {akw}が{cnt}/{total_scenes}シーン({cnt*100//total_scenes}%)")

    # --- クロスシーン: 体位全体分布偏り ---
    position_counter = {}
    for scene in results:
        sd_text = scene.get("sd_prompt", "")
        _sd_tags = {t.strip().lower().replace(" ", "_") for t in sd_text.split(",") if t.strip()}
        for ptag in _sd_tags & POSITION_TAGS:
            position_counter[ptag] = position_counter.get(ptag, 0) + 1
    if total_scenes >= 5:
        for ptag, cnt in position_counter.items():
            if cnt / total_scenes >= 0.4:
                scene_issues.setdefault("global", []).append(
                    f"体位偏り: {ptag}が{cnt}/{total_scenes}シーン({cnt*100//total_scenes}%)")

    # --- 体位バリエーション統計 ---
    unique_positions = set(position_counter.keys())
    position_variety = {
        "unique_count": len(unique_positions),
        "positions_used": sorted(unique_positions),
        "distribution": position_counter,
    }

    # --- v7.1: intensity不一致検出 ---
    _LOW_I_MOANS = _re.compile(r"[あんはぁ]っ|んほ|あへ|んっ|あぁ|はぁ")
    for i, scene in enumerate(results):
        sid = scene.get("scene_id", i + 1)
        intensity = scene.get("intensity", 3)
        for bubble in scene.get("bubbles", []):
            txt = bubble.get("text", "")
            btype = bubble.get("type", "")
            if not txt:
                continue
            if intensity <= 2:
                if "♡" in txt:
                    scene_issues.setdefault(sid, []).append(
                        f"intensity{intensity}に♡: 「{txt[:20]}」")
                if btype in ("speech", "thought") and _LOW_I_MOANS.search(txt):
                    scene_issues.setdefault(sid, []).append(
                        f"intensity{intensity}に喘ぎ表現: 「{txt[:20]}」")
            if intensity >= 4:
                if btype in ("speech", "thought"):
                    for polite in ("です", "ます", "ください"):
                        if polite in txt:
                            scene_issues.setdefault(sid, []).append(
                                f"intensity{intensity}に丁寧語「{polite}」: 「{txt[:20]}」")
                            break

    # --- v7.1: 語尾パターン反復検出（3連続以上） ---
    _suffix_seq = []  # [(scene_id, suffix)]
    for i, scene in enumerate(results):
        sid = scene.get("scene_id", i + 1)
        for bubble in scene.get("bubbles", []):
            txt = bubble.get("text", "").rstrip()
            if len(txt) >= 3:
                _suffix_seq.append((sid, txt[-3:]))
    for idx in range(len(_suffix_seq) - 2):
        s1 = _suffix_seq[idx][1]
        if s1 == _suffix_seq[idx + 1][1] == _suffix_seq[idx + 2][1]:
            if idx > 0 and _suffix_seq[idx - 1][1] == s1:
                continue
            sid = _suffix_seq[idx][0]
            scene_issues.setdefault(sid, []).append(
                f"語尾パターン3連続「{s1}」")

    # --- v7.1: thought長さチェック（20文字超=ナレーション化） ---
    for i, scene in enumerate(results):
        sid = scene.get("scene_id", i + 1)
        for bubble in scene.get("bubbles", []):
            if bubble.get("type") == "thought":
                txt = bubble.get("text", "")
                if len(txt) > 20:
                    scene_issues.setdefault(sid, []).append(
                        f"thought長すぎ({len(txt)}文字): 「{txt[:25]}…」")

    # --- クロスシーン: ストーリーリセット検出（最後の10%でi≤2はリセットの兆候） ---
    _total = len(results)
    if _total >= 20:
        _epilogue_start = max(1, _total - max(5, _total // 10))
        _reset_scenes = []
        for i in range(_epilogue_start, _total):
            si = results[i].get("intensity", 3)
            sid = results[i].get("scene_id", i + 1)
            if si <= 2:
                _reset_scenes.append(sid)
        if _reset_scenes:
            scene_issues.setdefault("global", []).append(
                f"ストーリーリセット疑い: 終盤シーン{_reset_scenes}がi≤2（導入の繰り返し）")

    # --- クロスシーン: i=4連続過多検出 ---
    _max_consecutive_4 = 0
    _curr_run = 0
    for s in results:
        if s.get("intensity", 3) == 4:
            _curr_run += 1
            _max_consecutive_4 = max(_max_consecutive_4, _curr_run)
        else:
            _curr_run = 0
    if _max_consecutive_4 > 6:
        scene_issues.setdefault("global", []).append(
            f"i=4連続{_max_consecutive_4}シーン（上限4-6・i=3ブレイク不足）")

    # --- クロスシーン: i≤2連続過多検出（テンポ停滞） ---
    _max_consecutive_low = 0
    _curr_low_run = 0
    for s in results:
        if s.get("intensity", 3) <= 2:
            _curr_low_run += 1
            _max_consecutive_low = max(_max_consecutive_low, _curr_low_run)
        else:
            _curr_low_run = 0
    if _max_consecutive_low > 6:
        scene_issues.setdefault("global", []).append(
            f"i≤2連続{_max_consecutive_low}シーン（上限6・テンポ停滞）")

    # --- THOUGHT↔SPEECH感情矛盾チェック ---
    _ct_positive_th = ["幸せ", "嬉しい", "好き", "大好き", "気持ちいい", "もっと", "欲しい", "♡"]
    _ct_negative_sp = ["やめて", "嫌", "離して", "痛い", "やだ", "助けて", "来ないで"]
    _ct_negative_th = ["怖い", "嫌だ", "逃げ", "助けて", "痛い", "無理", "嫌い"]
    _ct_positive_sp = ["もっと", "気持ちいい", "好き", "♡", "嬉しい", "幸せ", "ちょうだい"]
    _ct_exempt = any(k in (theme or "").lower() for k in ["forced", "reluctant", "陵辱", "強制"])
    _ct_count = 0
    for s in results:
        _i = s.get("intensity", 3)
        if _ct_exempt and 3 <= _i <= 4:
            continue
        _thoughts = [b.get("text", "") for b in s.get("bubbles", []) if b.get("type") == "thought"]
        _speeches = [b.get("text", "") for b in s.get("bubbles", []) if b.get("type") == "speech"]
        for _th in _thoughts:
            for _sp in _speeches:
                if (any(kw in _th for kw in _ct_positive_th) and any(kw in _sp for kw in _ct_negative_sp)):
                    _ct_count += 1
                elif (any(kw in _th for kw in _ct_negative_th) and any(kw in _sp for kw in _ct_positive_sp)):
                    _ct_count += 1
    if _ct_count > 0:
        scene_issues.setdefault("global", []).append(f"THOUGHT↔SPEECH感情矛盾: {_ct_count}件")

    # --- N-gram語彙多様性チェック（4文字以上の繰り返し表現検出）---
    from collections import Counter as _Counter
    _ngram_counter = _Counter()
    for s in results:
        for b in s.get("bubbles", []):
            txt = b.get("text", "")
            if len(txt) >= 4:
                for _ng_start in range(len(txt) - 3):
                    _ngram_counter[txt[_ng_start:_ng_start + 4]] += 1
    _repeated_ngrams = [(ng, cnt) for ng, cnt in _ngram_counter.most_common(20) if cnt > 5]
    if _repeated_ngrams:
        _ngram_report = ", ".join(f"「{ng}」×{cnt}" for ng, cnt in _repeated_ngrams[:5])
        scene_issues.setdefault("global", []).append(
            f"N-gram反復: {_ngram_report}（計{len(_repeated_ngrams)}パターン）")

    # v8.9: メタ参照description検出（「シーンXXの場面では」等のAPI生成アーティファクト）
    import re as _re_val
    _META_REF_PATTERN_V = _re_val.compile(r'シーン\d+')
    _meta_ref_count = 0
    for i, scene in enumerate(results):
        desc = scene.get("description", "")
        if desc and _META_REF_PATTERN_V.search(desc):
            sid = scene.get("scene_id", i + 1)
            scene_issues.setdefault(f"S{sid}", []).append(
                f"メタ参照: descriptionに「シーンXX」パターン検出（API生成アーティファクト）")
            _meta_ref_count += 1
    if _meta_ref_count > 0:
        scene_issues.setdefault("global", []).append(
            f"メタ参照description: {_meta_ref_count}シーンで「シーンXX」パターン検出")

    # v8.9: 時間軸バリデーション（「翌週」「数日後」等の時間ジャンプ検出）
    # エピローグ（最終10%）は時間ジャンプ許可
    # テーマ別: single_eventは「翌日」も検出 / few_days/flexibleは「翌週」以上のみ
    _TIME_JUMP_KW = ["翌週", "翌々週", "数日後", "一週間後", "数週間後", "翌月", "数ヶ月後",
                     "1週間後", "２週間後", "次の週", "来週", "後日"]
    _time_span_v = _THEME_TIME_SPAN.get(theme, "flexible")
    if _time_span_v == "single_event":
        _TIME_JUMP_KW.extend(["翌日", "翌朝"])
    _time_jump_count = 0
    _n_results = len(results)
    _epilogue_start_v = max(1, _n_results - max(1, _n_results // 10))
    for i, scene in enumerate(results):
        if i >= _epilogue_start_v:
            continue  # エピローグは時間ジャンプ許可
        sid = scene.get("scene_id", i + 1)
        desc = scene.get("description", "")
        sflow = scene.get("story_flow", "")
        for kw in _TIME_JUMP_KW:
            if kw in desc:
                scene_issues.setdefault(f"S{sid}", []).append(
                    f"時間軸ジャンプ禁止: descriptionに「{kw}」（同日内に圧縮すべき）")
                _time_jump_count += 1
                break
            if kw in sflow:
                scene_issues.setdefault(f"S{sid}", []).append(
                    f"時間軸ジャンプ禁止: story_flowに「{kw}」（同日内に圧縮すべき）")
                _time_jump_count += 1
                break
    if _time_jump_count > 0:
        scene_issues.setdefault("global", []).append(
            f"時間軸ジャンプ検出: {_time_jump_count}シーンで「翌週」等の大きな時間経過（CG集は基本同日の出来事）")

    # v8.8: テーマ別バリデーション（time_stop: 停止中の女性speech/moan検出）
    if theme == "time_stop":
        _TS_ACTIVE_KW = ["時間停止", "身動き", "止まった", "停止した", "動けない"]
        _TS_RELEASED_KW = ["時間再開", "再び動き", "解除", "現実に戻", "混乱して"]
        for i, scene in enumerate(results):
            desc = scene.get("description", "")
            is_frozen = any(kw in desc for kw in _TS_ACTIVE_KW)
            is_released = any(kw in desc for kw in _TS_RELEASED_KW)
            if is_frozen and not is_released:
                sid = scene.get("scene_id", i + 1)
                for b in scene.get("bubbles", []):
                    speaker = b.get("speaker", "")
                    btype = b.get("type", "")
                    if not _is_male_speaker(speaker):
                        if btype == "speech":
                            scene_issues.setdefault(f"S{sid}", []).append(
                                f"時間停止中の女性speech禁止: 「{b.get('text', '')[:20]}」")
                        elif btype == "moan":
                            scene_issues.setdefault(f"S{sid}", []).append(
                                f"時間停止中のmoan禁止（声が出せない）: 「{b.get('text', '')[:20]}」")
    # v8.8: mood品質チェック（テーマkey_emotionsがそのままmoodに使われている）
    _theme_guide_v = THEME_GUIDES.get(theme, {})
    _theme_key_emotions_v = set(_theme_guide_v.get("key_emotions", []))
    if _theme_key_emotions_v:
        for i, scene in enumerate(results):
            m = scene.get("mood", "")
            if m and m in _theme_key_emotions_v:
                sid = scene.get("scene_id", i + 1)
                scene_issues.setdefault(f"S{sid}", []).append(
                    f"moodがテーマ感情そのまま: 「{m}」→具体的なmoodにすべき")

    # v9.0: 物理状態一貫性チェック（服装復活・体液消失検出）
    _val_max_undress = 0
    _val_had_cum = False
    for i, scene in enumerate(results):
        sid = scene.get("scene_id", i + 1)
        sd = scene.get("sd_prompt", "")
        _sd_tags_v = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}

        # 脱衣レベル検出
        _cur_lv = 0
        for _tag_v in _sd_tags_v:
            if _tag_v in ("nude", "naked", "completely_nude"):
                _cur_lv = max(_cur_lv, 5)
            elif _tag_v in ("topless", "bottomless", "panties_only", "naked_shirt", "stockings_only"):
                _cur_lv = max(_cur_lv, 4)
            elif _tag_v in ("panties_aside", "open_shirt", "bra_removed", "torn_clothes", "no_bra", "no_panties"):
                _cur_lv = max(_cur_lv, 3)

        # 服装復活検出（脱衣レベルが2段階以上逆行）
        if _val_max_undress >= 4 and _cur_lv <= 1 and i > 0:
            scene_issues.setdefault(f"S{sid}", []).append(
                f"服装復活検出: 前シーンで脱衣レベル{_val_max_undress}→現シーン{_cur_lv}（sd_promptに脱衣タグ不足）")
        _val_max_undress = max(_val_max_undress, _cur_lv)

        # 体液消失検出（射精後なのにcum系タグなし）
        _has_cum_tags = bool(_sd_tags_v & {"cum", "cum_on_body", "cum_on_face", "cum_in_pussy",
                                            "cum_overflow", "cum_string", "cum_pool", "cum_drip"})
        if _val_had_cum and not _has_cum_tags and scene.get("intensity", 3) >= 3:
            scene_issues.setdefault(f"S{sid}", []).append(
                "体液消失検出: 前シーンで射精があったがsd_promptにcum系タグなし")
        if _has_cum_tags:
            _val_had_cum = True

    # Phase1: ナラティブ↔SDタグ不整合チェック
    for i, scene in enumerate(results):
        sid = scene.get("scene_id", i + 1)
        desc = scene.get("description", "")
        sd = scene.get("sd_prompt", "")
        if not desc or not sd:
            continue
        _sd_tags_nc = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}
        # ウェイト付きタグも正規化
        import re as _re_nc
        _sd_tags_nc = {_re_nc.sub(r'[()]', '', t).split(":")[0].strip() for t in _sd_tags_nc}
        for _kw, _required_tags, _check_name in _CRITICAL_NARRATIVE_CHECKS:
            if _kw in desc and not (_sd_tags_nc & _required_tags):
                scene_issues.setdefault(f"S{sid}", []).append(
                    f"SD不整合: {_check_name}")

    # Phase5: 非Danbooruタグ検出
    for i, scene in enumerate(results):
        sid = scene.get("scene_id", i + 1)
        sd = scene.get("sd_prompt", "")
        if not sd:
            continue
        import re as _re_p5
        for t in sd.split(","):
            _t_norm_p5 = t.strip().lower().replace(" ", "_")
            _t_inner_p5 = _re_p5.sub(r'[()]', '', _t_norm_p5).split(":")[0].strip()
            if _t_inner_p5 in _NON_DANBOORU_TAGS:
                scene_issues.setdefault(f"S{sid}", []).append(
                    f"非Danbooruタグ: {_t_inner_p5}")

    # Phase7: 構図偏りチェック
    import re as _re_p7
    _closeup_check = {"close-up", "portrait", "upper_body", "face_focus"}
    _closeup_v_count = 0
    n_scenes_v = len(results)
    if n_scenes_v >= 8:
        for scene in results:
            sd = scene.get("sd_prompt", "")
            if not sd:
                continue
            _p7_tags = {_re_p7.sub(r'[()]', '', t.strip().lower().replace(" ", "_")).split(":")[0].strip()
                        for t in sd.split(",") if t.strip()}
            if _p7_tags & _closeup_check:
                _closeup_v_count += 1
        if n_scenes_v > 0 and _closeup_v_count / n_scenes_v > 0.60:
            scene_issues.setdefault("全体", []).append(
                f"構図偏り: close-up系が{_closeup_v_count}/{n_scenes_v}({_closeup_v_count*100//n_scenes_v}%)で60%超過")

    # 体位サポートタグ完全性チェック
    import re as _re_pos
    for i, scene in enumerate(results):
        sid = scene.get("scene_id", i + 1)
        sd = scene.get("sd_prompt", "")
        if not sd:
            continue
        _pos_tags_v = {_re_pos.sub(r'[()]', '', t.strip().lower().replace(" ", "_")).split(":")[0].strip()
                       for t in sd.split(",") if t.strip()}
        for _pos_k, _pos_supports in _POSITION_SUPPORT_TAGS.items():
            if _pos_k in _pos_tags_v:
                _missing_sup = [s for s in _pos_supports[:2] if s not in _pos_tags_v]
                if _missing_sup:
                    scene_issues.setdefault(f"S{sid}", []).append(
                        f"体位サポート不足: {_pos_k}に{','.join(_missing_sup)}がない")

    # アクセサリ消失チェック（初回検出後に消失したシーンを検出）
    _acc_first_seen = {}  # tag -> first scene index
    for i, scene in enumerate(results):
        sd = scene.get("sd_prompt", "")
        if not sd:
            continue
        _acc_tags_v = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}
        for _acc_tag in _ACCESSORY_PERSISTENT_TAGS:
            if _acc_tag in _acc_tags_v and _acc_tag not in _acc_first_seen:
                _acc_first_seen[_acc_tag] = i
    # 初回検出後、5シーン以上連続で消失していたら警告
    if _acc_first_seen and len(results) >= 10:
        for _acc_tag, _first_idx in _acc_first_seen.items():
            _consecutive_missing = 0
            for i in range(_first_idx + 1, len(results)):
                sd = results[i].get("sd_prompt", "")
                if not sd:
                    continue
                _acc_v = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}
                if _acc_tag not in _acc_v:
                    _consecutive_missing += 1
                else:
                    _consecutive_missing = 0
                if _consecutive_missing >= 5:
                    sid = results[i].get("scene_id", i + 1)
                    scene_issues.setdefault(f"S{sid}", []).append(
                        f"アクセサリ消失: {_acc_tag}が5シーン以上連続で欠落")
                    break  # 1タグにつき1回だけ警告

    n_issues = sum(len(v) for v in scene_issues.values()) + len(repeated_moans) + len(repeated_onom)
    # スコア計算: シーン数で正規化（大規模シーンでもscore=0にならないように）
    # ≤20シーン: 従来通り n_issues * 5 で減点
    # >20シーン: issues_per_scene ベースで減点（1 issue/scene = -33点）
    n_scenes = max(1, len(results))
    if n_scenes <= 20:
        score = max(0, 100 - n_issues * 5)
    else:
        issues_per_scene = n_issues / n_scenes
        score = max(0, 100 - int(issues_per_scene * 33))

    return {
        "score": score,
        "scene_issues": scene_issues,
        "repeated_moans": repeated_moans,
        "repeated_onomatopoeia": repeated_onom,
        "position_variety": position_variety,
        "total_issues": n_issues,
        "summary": f"品質スコア: {score}/100（{n_issues}件の問題検出）"
    }




def _fix_character_name(name: str, correct_names: list) -> str:
    """キャラ名の表記ブレを修正"""
    if not name or not correct_names:
        return name
    # 完全一致ならそのまま
    if name in correct_names:
        return name
    # 部分一致: 正しい名前がnameに含まれる or nameが正しい名前に含まれる
    for cn in correct_names:
        if cn in name or name in cn:
            return cn
    # 姓が一致するパターン（中野三子→中野三玖）
    for cn in correct_names:
        if len(cn) >= 3 and len(name) >= 3:
            # 姓（先頭2文字）が一致し、名が異なる場合
            if cn[:2] == name[:2] and cn != name:
                return cn
            # 先頭1文字が一致し残り文字数が同じ場合
            if cn[0] == name[0] and len(cn) == len(name) and cn != name:
                return cn
    return name


def _fix_names_in_text(text: str, correct_names: list) -> str:
    """テキスト内のキャラ名表記ブレを修正"""
    import re
    if not text or not correct_names:
        return text
    for correct in correct_names:
        if len(correct) < 3:
            continue
        # 姓（先頭2文字）+ 名（残り）のパターンで検索
        family = correct[:2]
        given = correct[2:]
        if not given:
            continue
        # family + given長と同じ文字数の漢字/ひらがな/カタカナを検索
        # 正しい名前以外をマッチさせる
        pattern = re.escape(family) + r'([\u4e00-\u9faf\u3040-\u309f\u30a0-\u30ff]{' + str(len(given)) + '})'
        for m in re.finditer(pattern, text):
            found_given = m.group(1)
            if found_given != given:
                wrong_name = family + found_given
                text = text.replace(wrong_name, correct)
                log_message(f"  名前修正: {wrong_name}→{correct}")
    return text


def _normalize_bubble_text(text: str) -> str:
    """セリフテキストを正規化して類似判定に使用。
    装飾除去+濁点/半濁点除去+カタカナ→ひらがな。
    例: 「あ゛あ゛っ♡♡」→「ああ」, 「ああっ♡」→「ああ」 → 同系統と判定可能
    """
    # 装飾文字除去
    t = text.replace("♡", "").replace("♥", "").replace("…", "").replace("っ", "").replace("ー", "").strip()
    # 濁点・半濁点除去（漫画的な「あ゛」「お゛」表現の正規化）
    # U+309B ゛, U+309C ゜, U+3099 結合濁点, U+309A 結合半濁点
    t = t.replace("\u309B", "").replace("\u309C", "").replace("\u3099", "").replace("\u309A", "")
    # 全角濁点的な表記も除去
    t = t.replace("゛", "").replace("゜", "")
    # カタカナ→ひらがな変換
    result = []
    for ch in t:
        cp = ord(ch)
        if 0x30A1 <= cp <= 0x30F6:
            result.append(chr(cp - 0x60))
        else:
            result.append(ch)
    return "".join(result)

def _is_similar_bubble(text1: str, text2: str, strict: bool = False) -> bool:
    """2つのセリフが類似しているか判定。
    strict=False（デフォルト）: 完全一致 or 正規化一致 or 先頭一致
    strict=True: 完全一致 or 正規化完全一致のみ（短い喘ぎ声向け）
    """
    if text1 == text2:
        return True
    n1 = _normalize_bubble_text(text1)
    n2 = _normalize_bubble_text(text2)
    if n1 == n2:
        return True
    if strict:
        return False
    # 長めのテキスト（4文字以上の正規化結果）のみ先頭一致チェック
    if len(n1) >= 4 and len(n2) >= 4 and n1[:4] == n2[:4]:
        return True
    return False

def _analyze_scene_context(scene: dict) -> str:
    """シーンのdescription/title/moodからコンテキストタイプを判定。
    Returns: 'non_sexual' | 'foreplay' | 'sexual' | 'climax' | 'aftermath'"""
    desc = (scene.get("description", "") + " " + scene.get("title", "")
            + " " + scene.get("mood", "")).lower()
    intensity = scene.get("intensity", 3)

    # 事後シーン
    aftermath_kw = ["事後", "余韻", "虚脱", "罪悪感", "後悔", "戻って", "帰る",
                    "眠り", "崩れ落ち", "終えた", "身繕い", "動けない", "虚ろ",
                    "後片付け", "放心", "脱力", "ぐったり", "呆然", "立てない",
                    "意識が戻", "我に返", "現実に戻"]
    if any(k in desc for k in aftermath_kw):
        return "aftermath"

    # 非エロシーン（歩き・日常・会話のみ）
    non_sexual_kw = ["歩く", "歩き", "歩いて", "通りを", "散歩", "食事", "食堂",
                     "休む", "休憩", "眺め", "待つ", "待って", "帰省", "到着",
                     "村に着", "自室で", "くつろ", "話しかけ", "説明を受",
                     "呼び止め", "誘われ", "連れ", "囲まれて", "聞き入",
                     "聞いて", "話を聞", "習慣", "近づき", "語りかけ",
                     "声をかけ"]
    # 非エロキーワードに該当し、かつ性行為キーワードが無ければnon_sexual
    sex_act_kw = ["挿入", "突き", "突かれ", "犯さ", "抱かれ", "愛撫", "舐", "咥",
                  "胸を", "乳首", "腰を振", "ピストン", "フェラ", "クンニ",
                  "手マン", "正常位", "騎乗", "バック", "結合", "肉棒"]
    has_non_sexual = any(k in desc for k in non_sexual_kw)
    has_sex = any(k in desc for k in sex_act_kw)

    if has_non_sexual and not has_sex:
        return "non_sexual"

    # 絶頂シーン
    if intensity >= 5 or any(k in desc for k in ["絶頂", "イク", "果て", "限界",
                                                   "痙攣", "理性崩壊", "アヘ"]):
        return "climax"

    # 前戯（触れる・撫でる・キス・服を脱がせる等、性行為未満の接触）
    foreplay_kw = ["触れ", "触って", "撫で", "キス", "抱きしめ", "脱がせ", "脱がさ",
                   "裸に"]
    if intensity <= 2 or (not has_sex and any(k in desc for k in foreplay_kw)):
        return "foreplay"

    return "sexual"


def _deduplicate_across_scenes(results: list, theme: str = "",
                                heroine_names: list = None,
                                char_profiles: list = None,
                                concept: str = "") -> None:
    """シーン間の同一・類似セリフを検出し、プールから代替セリフに置換。
    - 文脈判定: descriptionを解析し、非エロシーンにエロセリフを入れない
    - 重複保護: 同一セリフが検出された場合、プールから代替セリフに置換
    - ヒロイン名リスト以外のspeakerは全て男性と判定
    - テーマ/intensityに応じてプールカテゴリを絞り込み
    - 性格タイプに応じてプール混合比率を調整
    - v8.7: concept引数追加でkey_linesをプールに統合"""
    try:
        from ero_dialogue_pool import (
            get_moan_pool, get_speech_pool, pick_replacement, SPEECH_MALE_POOL,
            SPEECH_FEMALE_POOL, THOUGHT_POOL, NEUTRAL_POOL, AFTERMATH_POOL,
            get_male_speech_pool, get_female_speech_pool, get_pattern_key_lines
        )
        has_pool = True
    except ImportError:
        try:
            from ero_dialogue_pool import (
                get_moan_pool, get_speech_pool, pick_replacement, SPEECH_MALE_POOL,
                SPEECH_FEMALE_POOL, THOUGHT_POOL, get_male_speech_pool,
                get_female_speech_pool
            )
            has_pool = True
            NEUTRAL_POOL = None
            AFTERMATH_POOL = None
            get_pattern_key_lines = None
        except ImportError:
            has_pool = False
            NEUTRAL_POOL = None
            AFTERMATH_POOL = None
            get_pattern_key_lines = None
            log_message("ero_dialogue_pool.py未検出、重複は除去のみ（置換なし）")

    # ヒロイン名セット構築（これ以外のspeakerは全て男性扱い）
    _heroine_set = set()
    if heroine_names:
        for n in heroine_names:
            if n:
                _heroine_set.add(n)
                if len(n) >= 2:
                    _heroine_set.add(n[:2])
                    if len(n) >= 3:
                        _heroine_set.add(n[2:])

    # 性格タイプ判定（性格別プール混合用）
    _personality_type = _detect_personality_type(char_profiles) if char_profiles else ""
    _pool_mix = _PERSONALITY_POOL_MIX.get(_personality_type, {}) if _personality_type else {}

    # キャラ固有プール読み込み
    _dedup_char_pool = {}
    if char_profiles:
        _cp0 = char_profiles[0]
        _cp_id = generate_char_id(_cp0.get("work_title", ""), _cp0.get("character_name", ""))
        _dedup_char_pool = load_character_pool(_cp_id)

    def _is_male_speaker(speaker: str) -> bool:
        if not speaker:
            return False
        for h in _heroine_set:
            if h in speaker:
                return False
        return True

    def _get_male_pool_for_theme(theme_str: str, intensity: int) -> list:
        t = theme_str.lower() if theme_str else ""
        pool = []
        if any(k in t for k in ["痴漢", "chikan", "公共", "public", "電車", "train", "トイレ"]):
            pool.extend(SPEECH_MALE_POOL.get("chikan", []))
            pool.extend(SPEECH_MALE_POOL.get("taunt", []))
            pool.extend(SPEECH_MALE_POOL.get("public", []))
            pool.extend(SPEECH_MALE_POOL.get("command", []))
        elif any(k in t for k in ["ntr", "寝取", "夜這", "村", "レイプ", "陵辱", "調教", "奴隷"]):
            pool.extend(SPEECH_MALE_POOL.get("command", []))
            pool.extend(SPEECH_MALE_POOL.get("dirty", []))
        elif any(k in t for k in ["純愛", "ラブ", "恋人", "カップル"]):
            pool.extend(SPEECH_MALE_POOL.get("gentle", []))
            pool.extend(SPEECH_MALE_POOL.get("praise", []))
        else:
            if intensity >= 4:
                pool.extend(SPEECH_MALE_POOL.get("command", []))
                pool.extend(SPEECH_MALE_POOL.get("dirty", []))
            else:
                pool.extend(SPEECH_MALE_POOL.get("dirty", []))
                pool.extend(SPEECH_MALE_POOL.get("praise", []))
        return pool if pool else [v for sp in SPEECH_MALE_POOL.values() for v in sp]

    def _get_pool_for_context(ctx: str, intensity: int, is_male: bool,
                              btype: str) -> list:
        """文脈・intensity・性別・タイプに応じた最適プールを選択"""
        # 非エロシーン → 中立プールのみ（エロ混入防止の最重要ガード）
        if ctx == "non_sexual":
            if NEUTRAL_POOL:
                return NEUTRAL_POOL.get("male" if is_male else "female", [])
            # フォールバック: 内蔵の中立セリフ
            if is_male:
                return ["ああ", "そうか", "来い", "行くぞ", "どうした"]
            return ["うん…", "え…", "あ、はい", "そう…", "ん…"]

        # 事後シーン
        if ctx == "aftermath":
            if AFTERMATH_POOL:
                return AFTERMATH_POOL.get("male" if is_male else "female", [])
            if is_male:
                return ["もういいぞ", "帰れ", "次もな"]
            return ["もう…むり", "動けない…", "ごめん…", "ぼーっと…"]

        # エロシーンの通常ロジック
        if btype == "moan":
            pool = []
            if _dedup_char_pool and "moan" in _dedup_char_pool:
                pool.extend(_dedup_char_pool["moan"].get(str(intensity), []))
            pool.extend(get_moan_pool(intensity))
            return pool
        elif btype == "thought":
            pool = []
            # v8.7: key_lines統合（thought用、先頭配置で選択確率UP）
            if concept and get_pattern_key_lines:
                kl_phase = "early" if intensity <= 2 else ("late" if intensity >= 5 else "mid")
                kl = get_pattern_key_lines(theme, concept, kl_phase)
                if kl:
                    pool.extend(kl)
            if _dedup_char_pool and "thought" in _dedup_char_pool:
                # フェーズ推定で適切なthoughtを混合
                try:
                    from ero_dialogue_pool import infer_phase as _inf_ph
                    _ph = _inf_ph(intensity, 0, 1)  # 推定（dedup時はscene_index不定）
                    pool.extend(_dedup_char_pool["thought"].get(_ph, []))
                except (ImportError, Exception):
                    pass
            pool.extend(get_speech_pool("thought", theme, intensity))
            return pool
        elif is_male:
            return _get_male_pool_for_theme(theme, intensity)
        else:
            # 女性speech: v8.7 key_lines → キャラ固有プール → 性格タイプ → intensity連動フォールバック
            pool = []
            # v8.7: key_lines統合（speech用、先頭配置で選択確率UP）
            if concept and get_pattern_key_lines:
                kl_phase = "early" if intensity <= 2 else ("late" if intensity >= 5 else "mid")
                kl = get_pattern_key_lines(theme, concept, kl_phase)
                if kl:
                    pool.extend(kl)
            if _dedup_char_pool and "speech" in _dedup_char_pool:
                try:
                    from ero_dialogue_pool import infer_phase as _inf_ph2
                    _ph2 = _inf_ph2(intensity, 0, 1)
                    pool.extend(_dedup_char_pool["speech"].get(_ph2, []))
                except (ImportError, Exception):
                    pass
            if _pool_mix:
                # 性格別プール混合: primary(2倍) + secondary(1倍)
                for cat in _pool_mix.get("primary", []):
                    entries = SPEECH_FEMALE_POOL.get(cat, [])
                    pool.extend(entries)
                    pool.extend(entries)  # 2倍ウェイト
                for cat in _pool_mix.get("secondary", []):
                    pool.extend(SPEECH_FEMALE_POOL.get(cat, []))
                # intensity補正: 高intensityではecstasy/plea追加
                if intensity >= 4 and "ecstasy" not in _pool_mix.get("primary", []):
                    pool.extend(SPEECH_FEMALE_POOL.get("ecstasy", []))
                if intensity >= 3 and "plea" not in _pool_mix.get("primary", []):
                    pool.extend(SPEECH_FEMALE_POOL.get("plea", []))
            else:
                # デフォルト: intensity連動
                if intensity <= 2:
                    pool.extend(SPEECH_FEMALE_POOL.get("denial", []))
                    pool.extend(SPEECH_FEMALE_POOL.get("embarrassed", []))
                elif intensity == 3:
                    pool.extend(SPEECH_FEMALE_POOL.get("plea", []))
                    pool.extend(SPEECH_FEMALE_POOL.get("acceptance", []))
                    pool.extend(SPEECH_FEMALE_POOL.get("embarrassed", []))
                elif intensity == 4:
                    pool.extend(SPEECH_FEMALE_POOL.get("acceptance", []))
                    pool.extend(SPEECH_FEMALE_POOL.get("plea", []))
                    pool.extend(SPEECH_FEMALE_POOL.get("ecstasy", []))
                else:
                    pool.extend(SPEECH_FEMALE_POOL.get("ecstasy", []))
                    pool.extend(SPEECH_FEMALE_POOL.get("plea", []))
                    pool.extend(SPEECH_FEMALE_POOL.get("submissive", []))
            return pool if pool else [v for sp in SPEECH_FEMALE_POOL.values() for v in sp]

    # v8.8: 使用済みテキスト追跡（speech/thoughtは完全一致のみ。moanのみprefix3類似チェック復活）
    used_moan_raw = set()
    used_moan_texts = set()
    used_thought_raw = set()
    used_thought_texts = set()
    used_speech_raw = set()
    used_speech_texts = set()
    # v8.8: moan限定の軽量類似チェック（speech/thoughtは完全一致のみ維持）
    _moan_prefix3_counter = {}  # 先頭3字カウンター
    _MOAN_PREFIX_LIMIT = max(3, len(results) // 8)  # 20→3, 50→6, 100→12
    # v8.6: 語尾構造パターン追跡（「~し…」「~る♡」等の3連続防止）— 有用なので維持
    _SUFFIX_STRUCT_RE = re.compile(r'(し…|る♡|の…|て…|く…|で…|に…|だし|よ…|か…|ない|った|って)$')
    _recent_suffix_structs = []  # 直近の語尾構造パターン（スライディングウィンドウ）

    replace_count = 0

    for scene in results:
        # dialogue形式（旧フォーマット）からbubblesへの変換
        if "bubbles" not in scene and scene.get("dialogue"):
            _moan_emotions = {"快感", "絶頂", "陶酔", "悶え", "昂り", "高潮", "恍惚"}
            scene["bubbles"] = []
            for d in scene["dialogue"]:
                emotion = d.get("emotion", "")
                btype = "moan" if emotion in _moan_emotions else "speech"
                scene["bubbles"].append({
                    "type": btype,
                    "speaker": d.get("speaker", ""),
                    "text": d.get("line", ""),
                })
        if "bubbles" not in scene:
            continue
        cleaned_bubbles = []
        sid = scene.get("scene_id", "?")
        intensity = scene.get("intensity", 3)
        ctx = _analyze_scene_context(scene)

        for b in scene["bubbles"]:
            text = b.get("text", "")
            btype = b.get("type", "")
            speaker = b.get("speaker", "")
            is_male = _is_male_speaker(speaker)

            if not text:
                cleaned_bubbles.append(b)
                continue

            # === 置換判定: 重複 or 長文 or パターン過多 or 文脈不整合 ===
            need_replace = False
            reason = ""

            if btype == "moan":
                norm = _normalize_bubble_text(text)
                # v8.8: 完全一致 + prefix3類似チェック（speech/thoughtは完全一致のみ維持）
                if (text in used_moan_raw) or (norm in used_moan_texts):
                    need_replace = True
                    reason = "重複"
                elif len(text) >= 3:
                    prefix3 = text[:3]
                    cnt = _moan_prefix3_counter.get(prefix3, 0)
                    if cnt >= _MOAN_PREFIX_LIMIT:
                        need_replace = True
                        reason = "喘ぎ類似"
                # 非エロシーンで喘ぎは文脈不整合
                if ctx == "non_sexual":
                    need_replace = True
                    reason = "非エロ文脈で喘ぎ"

            elif btype == "thought":
                norm = _normalize_bubble_text(text)
                # v8.7: 完全一致のみ置換（先頭4字類似/キーワード上限/パターン上限を撤廃）
                if (text in used_thought_raw) or (norm in used_thought_texts):
                    need_replace = True
                    reason = "重複"
                # v8.6: 語尾構造パターン3連続チェック（thought）— これは有用なので維持
                if not need_replace and len(text) >= 3:
                    _sm = _SUFFIX_STRUCT_RE.search(text.rstrip())
                    if _sm:
                        _cur_pat = _sm.group(1)
                        if _recent_suffix_structs[-2:].count(_cur_pat) >= 2:
                            need_replace = True
                            reason = f"語尾構造反復({_cur_pat})"

            elif btype == "speech":
                norm = _normalize_bubble_text(text)
                # v8.7: 完全一致のみ置換（先頭4字類似/末尾5字部分一致を撤廃）
                if (text in used_speech_raw) or (norm in used_speech_texts):
                    need_replace = True
                    reason = "重複"
                # v8.6: 語尾構造パターン3連続チェック — 有用なので維持
                if not need_replace and len(text) >= 3:
                    _sm = _SUFFIX_STRUCT_RE.search(text.rstrip())
                    if _sm:
                        _cur_pat = _sm.group(1)
                        if _recent_suffix_structs[-2:].count(_cur_pat) >= 2:
                            need_replace = True
                            reason = f"語尾構造反復({_cur_pat})"
                # 非エロシーンで♡付きセリフは文脈不整合
                if ctx == "non_sexual" and ("♡" in text or "♥" in text):
                    need_replace = True
                    reason = "非エロ文脈で♡"
                # 非エロシーンで明らかなエロセリフは不整合
                if ctx == "non_sexual":
                    erotic_kw = ["感じ", "奥", "イく", "イっ", "出る", "入って",
                                 "締まる", "濡れ", "とろ"]
                    if any(k in text for k in erotic_kw):
                        need_replace = True
                        reason = "非エロ文脈でエロ語"

            # 長文判定は廃止（セリフの長さを制限しない）

            # 事後シーンの文脈不整合チェック（エロ語・戦闘語が混入していたら置換）
            if ctx == "aftermath" and not need_replace:
                _aftermath_bad_kw = ["逃げ", "戦", "攻撃", "殺", "来るな", "来んな",
                                     "奥", "イく", "イっ", "締まる", "入って", "もっと",
                                     "感じ", "濡れ", "♡", "♥"]
                if any(k in text for k in _aftermath_bad_kw):
                    need_replace = True
                    reason = "事後文脈で不適切セリフ"

            # === 置換実行 ===
            if need_replace and has_pool:
                pool = _get_pool_for_context(ctx, intensity, is_male, btype)
                used_set = (used_moan_raw if btype == "moan"
                            else used_thought_raw if btype == "thought"
                            else used_speech_raw)
                norm_fn = _normalize_bubble_text
                # v8.6: 直近の語尾パターンを避ける + intensity考慮
                _avoid = ""
                if btype in ("speech", "thought") and _recent_suffix_structs:
                    _avoid = _recent_suffix_structs[-1]
                replacement = pick_replacement(pool, used_set, norm_fn,
                                               avoid_suffix=_avoid, intensity=intensity)
                if replacement:
                    log_message(f"  S{sid}: {reason}→置換「{text}」→「{replacement}」")
                    b["text"] = replacement
                    replace_count += 1
            elif need_replace and not has_pool:
                # プールがない場合は重複除去のみ（バブルをスキップ）
                if reason == "重複":
                    continue

            # v8.8: 使用済み登録（moanはprefix3カウンターも更新）
            final_text = b.get("text", "")
            final_norm = _normalize_bubble_text(final_text)
            if btype == "moan":
                used_moan_raw.add(final_text)
                used_moan_texts.add(final_norm)
                if len(final_text) >= 3:
                    p3 = final_text[:3]
                    _moan_prefix3_counter[p3] = _moan_prefix3_counter.get(p3, 0) + 1
            elif btype == "thought":
                used_thought_raw.add(final_text)
                used_thought_texts.add(final_norm)
            elif btype == "speech":
                used_speech_raw.add(final_text)
                used_speech_texts.add(final_norm)

            # v8.6: 語尾構造パターンをスライディングウィンドウに記録
            if btype in ("speech", "thought") and len(final_text) >= 3:
                _sm_reg = _SUFFIX_STRUCT_RE.search(final_text.rstrip())
                if _sm_reg:
                    _recent_suffix_structs.append(_sm_reg.group(1))
                    # ウィンドウサイズ6で制限（直近6件のみ追跡）
                    if len(_recent_suffix_structs) > 6:
                        _recent_suffix_structs.pop(0)

            cleaned_bubbles.append(b)

        if cleaned_bubbles:
            scene["bubbles"] = cleaned_bubbles

    if replace_count > 0:
        log_message(f"  重複セリフ計{replace_count}件を置換完了")

    # オノマトペ: 3シーン以内に同じ組み合わせがあれば除去
    for i in range(1, len(results)):
        curr_se = set(results[i].get("onomatopoeia", []))
        if not curr_se:
            continue
        for j in range(max(0, i - 3), i):
            prev_se = set(results[j].get("onomatopoeia", []))
            if prev_se and curr_se == prev_se:
                results[i]["onomatopoeia"] = []
                log_message(f"  S{results[i].get('scene_id', '?')}: S{results[j].get('scene_id', '?')}と同一SE除去")
                break

    # sd_prompt内の体位タグ: 前シーンとの連続重複を検出し代替置換
    import re as _re_dedup
    _prev_pos = set()
    for scene in results:
        sd = scene.get("sd_prompt", "")
        if not sd:
            _prev_pos = set()
            continue
        tags = [t.strip() for t in sd.split(",") if t.strip()]
        _cur_pos = set()
        new_tags = []
        changed = False
        for tag in tags:
            _inner = _re_dedup.sub(r'[()]', '', tag).split(":")[0].strip().lower().replace(" ", "_")
            if _inner in POSITION_TAGS:
                _cur_pos.add(_inner)
                if _inner in _prev_pos:
                    fallbacks = POSITION_FALLBACKS.get(_inner, [])
                    replacement = None
                    for fb in fallbacks:
                        if fb not in _prev_pos:
                            replacement = fb
                            break
                    if replacement:
                        _cur_pos.discard(_inner)
                        _cur_pos.add(replacement)
                        new_tags.append(replacement)
                        changed = True
                        log_message(f"  S{scene.get('scene_id', '?')}: 体位重複置換 {_inner}→{replacement}")
                        continue
            new_tags.append(tag)
        if changed:
            scene["sd_prompt"] = ", ".join(new_tags)
        _prev_pos = _cur_pos

    # sd_prompt内のアングルタグ: 前シーンとの連続重複を検出し代替置換
    _angle_kw = {"from_above", "from_below", "from_behind", "from_side",
                 "pov", "straight-on", "dutch_angle"}
    _prev_angles = set()
    for scene in results:
        sd = scene.get("sd_prompt", "")
        if not sd:
            _prev_angles = set()
            continue
        tags = [t.strip() for t in sd.split(",") if t.strip()]
        _cur_angles = set()
        new_tags = []
        changed = False
        for tag in tags:
            _inner = _re_dedup.sub(r'[()]', '', tag).split(":")[0].strip().lower().replace(" ", "_")
            if _inner in _angle_kw:
                _cur_angles.add(_inner)
                if _inner in _prev_angles:
                    fallbacks = ANGLE_FALLBACKS.get(_inner, [])
                    replacement = None
                    for fb in fallbacks:
                        if fb not in _prev_angles:
                            replacement = fb
                            break
                    if replacement:
                        _cur_angles.discard(_inner)
                        _cur_angles.add(replacement)
                        new_tags.append(replacement)
                        changed = True
                        log_message(f"  S{scene.get('scene_id', '?')}: アングル重複置換 {_inner}→{replacement}")
                        continue
            new_tags.append(tag)
        if changed:
            scene["sd_prompt"] = ", ".join(new_tags)
        _prev_angles = _cur_angles

def auto_fix_script(results: list, char_profiles: list = None, theme: str = "",
                    callback: Optional[Callable] = None, concept: str = "") -> list:
    """生成結果の自動修正（APIコスト不要のローカル後処理）"""
    import re

    _total_scenes = len(results)

    def _progress(step_name: str):
        """auto_fix内の進捗報告 + 停止チェック"""
        if callback:
            callback(f"🔧 自動修正: {step_name}（{_total_scenes}シーン）")
        log_message(f"  auto_fix: {step_name}")

    # === キャラ名の正規化マップ構築 ===
    correct_names = []  # [(correct_full_name, family, given)]
    if char_profiles:
        for cp in char_profiles:
            name = cp.get("character_name", "")
            if not name or len(name) < 2:
                continue
            correct_names.append(name)

    # === キャラ固有セリフプールの読み込み ===
    _char_pool = {}
    if char_profiles:
        _cp0 = char_profiles[0]
        _cp_id = generate_char_id(_cp0.get("work_title", ""), _cp0.get("character_name", ""))
        _char_pool = load_character_pool(_cp_id)
        if _char_pool:
            log_message(f"キャラ固有プール読み込み: {_cp_id}")

    # テキストフィールド一覧
    text_fields = ["description", "location_detail", "direction", "story_flow", "title"]

    _progress("Step 1-4 基本修正")
    for scene in results:
        # 1. "(XX字)" マーカーの除去
        for field in text_fields + ["mood"]:
            if field in scene and isinstance(scene[field], str):
                scene[field] = re.sub(r'[（(]\d+字[以内程度上]*[）)]', '', scene[field]).strip()

        if "character_feelings" in scene and isinstance(scene["character_feelings"], dict):
            scene["character_feelings"] = {
                k: re.sub(r'[（(]\d+字[以内程度上]*[）)]', '', v).strip()
                for k, v in scene["character_feelings"].items()
            }

        # 1.5. 「……」→「…」全フィールド統一（二重三点リーダ修正）
        for field in text_fields + ["mood"]:
            if field in scene and isinstance(scene[field], str):
                while "……" in scene[field]:
                    scene[field] = scene[field].replace("……", "…")
        if "character_feelings" in scene and isinstance(scene["character_feelings"], dict):
            for k, v in scene["character_feelings"].items():
                while "……" in v:
                    v = v.replace("……", "…")
                scene["character_feelings"][k] = v
        if "bubbles" in scene and isinstance(scene["bubbles"], list):
            for bubble in scene["bubbles"]:
                if not isinstance(bubble, dict):
                    continue
                txt = bubble.get("text", "")
                if isinstance(txt, str) and "……" in txt:
                    while "……" in txt:
                        txt = txt.replace("……", "…")
                    bubble["text"] = txt

        # 2. キャラ名の修正（全フィールド対象）
        if correct_names:
            # 2a. character_feelingsのキー修正
            if "character_feelings" in scene and isinstance(scene["character_feelings"], dict):
                new_feelings = {}
                for key, val in scene["character_feelings"].items():
                    corrected_key = _fix_character_name(key, correct_names)
                    new_feelings[corrected_key] = val
                scene["character_feelings"] = new_feelings

            # 2b. bubblesのspeaker修正
            if "bubbles" in scene:
                for bubble in scene["bubbles"]:
                    speaker = bubble.get("speaker", "")
                    if speaker:
                        bubble["speaker"] = _fix_character_name(speaker, correct_names)

            # 2c. テキストフィールド内のキャラ名修正
            for field in text_fields:
                text = scene.get(field, "")
                if text:
                    scene[field] = _fix_names_in_text(text, correct_names)

        # 3. SDプロンプトのquality括弧修正
        if "sd_prompt" in scene and scene["sd_prompt"]:
            sd = scene["sd_prompt"]
            quality_tags = {"masterpiece", "best_quality", "best quality", "high_quality", "highres", "absurdres"}
            match = re.match(r'^\(([^)]+)\)', sd)
            if match:
                inner_tags = [t.strip() for t in match.group(1).split(",")]
                quality_only = []
                non_quality = []
                for tag in inner_tags:
                    base_tag = re.sub(r':[\d.]+$', '', tag).strip().lower()
                    if base_tag in quality_tags:
                        quality_only.append(tag)
                    else:
                        non_quality.append(tag)
                if non_quality:
                    new_quality = "(" + ", ".join(quality_only) + ")"
                    rest = sd[match.end():].lstrip(", ")
                    non_quality_str = ", ".join(non_quality)
                    scene["sd_prompt"] = f"{new_quality}, {non_quality_str}, {rest}".rstrip(", ")

    # 4. scene_id連番修正（1, 2, 3, ... に強制リナンバー）
    for i, scene in enumerate(results):
        scene["scene_id"] = i + 1

    _progress("Step 4.5-4.7 セリフ・表現修正")
    # 4.5. 男性セリフ自動修正（♡除去、moan→speech変換、長文短縮）
    heroine_name_set = set(correct_names) if correct_names else set()
    # 男性speaker名フォールバック（heroine_name_set空時に使用）
    _MALE_SPEAKER_NAMES = frozenset([
        "男", "男性", "おじさん", "おっさん", "先生", "上司", "彼",
        "男の子", "少年", "青年", "紳士", "客", "店長", "先輩", "後輩",
        "兄", "弟", "父", "義父", "義兄", "master", "男A", "男B",
    ])
    def _is_male_by_name(speaker: str) -> bool:
        """heroine_name_set空時のフォールバック男性判定"""
        if not speaker:
            return False
        if heroine_name_set:
            return speaker not in heroine_name_set and not any(h in speaker for h in heroine_name_set)
        return speaker in _MALE_SPEAKER_NAMES or any(m in speaker for m in _MALE_SPEAKER_NAMES)
    for scene in results:
        if "bubbles" in scene:
            for bubble in scene["bubbles"]:
                speaker = bubble.get("speaker", "")
                is_male = speaker and _is_male_by_name(speaker)
                if is_male:
                    # ♡♥を除去
                    txt = bubble.get("text", "")
                    if "♡" in txt or "♥" in txt:
                        txt = txt.replace("♡", "").replace("♥", "").strip()
                        bubble["text"] = txt
                    # moan→speech変換
                    if bubble.get("type") == "moan":
                        bubble["type"] = "speech"

    # 4.52. description内「主人公」→ヒロイン名置換（男性参照時のみ）
    _MALE_ACTION_KEYWORDS = {"挿入", "腰を振", "攻め", "抱き", "犯", "掴", "押し倒", "射精", "突い", "責め", "脱がせ", "握り", "引き寄せ", "覆いかぶさ"}
    _protagonist_fix_count = 0
    heroine_label = correct_names[0] if correct_names else "ヒロイン"
    for scene in results:
        for field in ["description", "story_flow", "direction"]:
            text = scene.get(field, "")
            if not text or "主人公" not in text:
                continue
            # 「主人公が+男性的動作」パターンを検出→「彼」に置換（男性を指す場合）
            for kw in _MALE_ACTION_KEYWORDS:
                if kw in text and "主人公" in text:
                    text = text.replace("主人公", "彼")
                    scene[field] = text
                    _protagonist_fix_count += 1
                    break
            else:
                # 男性動作なし→女性を指している可能性→ヒロイン名に置換
                if "主人公" in text:
                    text = text.replace("主人公", heroine_label)
                    scene[field] = text
                    _protagonist_fix_count += 1
    if _protagonist_fix_count > 0:
        log_message(f"  主人公呼称修正: {_protagonist_fix_count}件")

    # お嬢様口調崩壊マップ（intensity 4-5で適用）
    _OJOUSAMA_BREAKDOWN_MAP = {
        "ですの": "…の…♡",
        "ですわ": "…♡",
        "ですこと": "…♡",
        "でございます": "…♡",
        "なさいませ": "…て…♡",
        "いたしますわ": "ちゃう…♡",
        "くださいませ": "…♡",
        "ましてよ": "…♡",
        "おほほ": "あっ♡",
        "ごきげんよう": "",
        "わたくし": "あたし",
        "いけませんわ": "だめ…♡",
        "よろしくてよ": "いい…♡",
        "存じません": "しらない…",
        "困りますわ": "やだ…♡",
        "お許しください": "ゆるして…♡",
        "なりませんわ": "だめ…♡",
        "ございませんの": "ない…♡",
        "いたしません": "しない…♡",
        "差し上げます": "あげる…♡",
    }

    # 4.52b. お嬢様口調崩壊処理（intensity 4-5で丁寧語を崩す）
    _ojousama_fix_count = 0
    for scene in results:
        intensity = scene.get("intensity", 2)
        if intensity < 4:
            continue
        if "bubbles" not in scene:
            continue
        for bubble in scene["bubbles"]:
            speaker = bubble.get("speaker", "")
            # 男性セリフはスキップ
            if speaker and _is_male_by_name(speaker):
                continue
            txt = bubble.get("text", "")
            if not txt:
                continue
            original = txt
            for formal, broken in _OJOUSAMA_BREAKDOWN_MAP.items():
                if formal in txt:
                    txt = txt.replace(formal, broken)
            # intensity 5: 残った「です」も崩す
            if intensity >= 5 and "です" in txt:
                txt = txt.replace("です", "…♡")
            if txt != original:
                bubble["text"] = txt
                _ojousama_fix_count += 1
                log_message(f"  お嬢様口調崩壊: 「{original}」→「{txt}」")
    if _ojousama_fix_count > 0:
        log_message(f"  お嬢様口調崩壊: {_ojousama_fix_count}件修正")

    # 4.55. 男性セリフ末尾フレーズ反復修正（「最高だ」問題対策）
    # LLMが同一末尾フレーズの男性セリフを繰り返す問題を検出→SPEECH_MALE_POOLから代替置換
    _MALE_SUFFIX_FIX_THRESHOLD = 3  # 同一末尾フレーズ3回以上で修正
    _male_suffix_map = {}  # suffix -> [(scene_idx, bubble_idx, full_text)]
    for si, scene in enumerate(results):
        if "bubbles" not in scene:
            continue
        for bi, bubble in enumerate(scene["bubbles"]):
            speaker = bubble.get("speaker", "")
            is_male = speaker and heroine_name_set and speaker not in heroine_name_set
            if is_male and bubble.get("type") == "speech":
                txt = bubble.get("text", "")
                if not txt or len(txt) < 3:
                    continue
                core = txt.replace("…", "").replace("♡", "").replace("っ", "").replace("♥", "").strip()
                if len(core) < 3:
                    continue
                suffix = core[-4:] if len(core) >= 4 else core
                _male_suffix_map.setdefault(suffix, []).append((si, bi, txt))
    _male_suffix_fix_count = 0
    try:
        from ero_dialogue_pool import SPEECH_MALE_POOL as _MALE_POOL_CHECK
        _has_male_pool = True
    except ImportError:
        _has_male_pool = False
    if _has_male_pool:
        _used_male_texts = set()
        # 先に全ての男性セリフを使用済みとして登録
        for scene in results:
            for b in scene.get("bubbles", []):
                sp = b.get("speaker", "")
                if sp and heroine_name_set and sp not in heroine_name_set and b.get("type") == "speech":
                    _used_male_texts.add(b.get("text", ""))
        for suffix, entries in _male_suffix_map.items():
            if len(entries) < _MALE_SUFFIX_FIX_THRESHOLD:
                continue
            # 最初の2件は残し、3件目以降を置換
            for si, bi, old_txt in entries[2:]:
                intensity = results[si].get("intensity", 3)
                # テーマに応じた男性プールを取得
                try:
                    from ero_dialogue_pool import get_male_speech_pool_for_theme, SPEECH_MALE_POOL as _MSP
                    _pool = get_male_speech_pool_for_theme(theme, intensity)
                except (ImportError, AttributeError):
                    _pool = []
                    try:
                        from ero_dialogue_pool import SPEECH_MALE_POOL as _MSP
                        for cat_entries in _MSP.values():
                            _pool.extend(cat_entries)
                    except ImportError:
                        break
                # 使用済みテキストを除外して代替を選択
                candidates = [t for t in _pool if t not in _used_male_texts and t != old_txt]
                if candidates:
                    import random as _rng_male
                    replacement = _rng_male.choice(candidates)
                    results[si]["bubbles"][bi]["text"] = replacement
                    _used_male_texts.add(replacement)
                    _male_suffix_fix_count += 1
                    log_message(f"  S{results[si].get('scene_id','?')}: 男性セリフ反復修正「{old_txt}」→「{replacement}」")
    if _male_suffix_fix_count > 0:
        log_message(f"  男性セリフ反復修正: {_male_suffix_fix_count}件")

    # 4.56. 男性セリフ末尾フレーズ反復修正（ハードコード代替辞書版 - プール置換漏れの補完）
    _MALE_SUFFIX_ALTERNATIVES = {
        "最高だ": ["たまんねえ", "いい反応だ", "堕ちたな", "もう戻れねえ", "感じてんだろ",
                    "締まるな", "素直だな", "いい顔だな", "我慢すんなよ", "ほら声出せ"],
        "おくだ": ["もっとだ", "逃がさねえ", "そのままだ", "力抜けよ", "震えてんぞ",
                    "声漏れてる", "もう無理だろ", "欲しいんだろ", "体は正直だ", "いいぞ"],
        "もっと": ["まだだ", "止まんな", "そのまま", "いけるだろ", "ほらもう一回"],
    }
    _suffix_counts = {}  # suffix -> count seen so far
    _suffix_fix_count = 0
    for scene in results:
        if "bubbles" not in scene:
            continue
        for bubble in scene["bubbles"]:
            speaker = bubble.get("speaker", "")
            is_male = speaker and heroine_name_set and speaker not in heroine_name_set
            if not is_male:
                continue
            txt = bubble.get("text", "")
            if not txt:
                continue
            for suffix, alts in _MALE_SUFFIX_ALTERNATIVES.items():
                if txt.rstrip("…♡").endswith(suffix) or suffix in txt:
                    _suffix_counts[suffix] = _suffix_counts.get(suffix, 0) + 1
                    if _suffix_counts[suffix] > 2:  # 3回目以降は置換
                        import random as _rng
                        alt = _rng.choice(alts)
                        old_txt = txt
                        bubble["text"] = alt
                        log_message(f"  男性セリフ反復修正: 「{old_txt}」→「{alt}」")
                        _suffix_fix_count += 1
                    break
    if _suffix_fix_count > 0:
        log_message(f"  男性セリフ反復修正(辞書): {_suffix_fix_count}件")

    # 4.57. 男性セリフ均等分配（intensity≥3で男性バブルなしのシーンに補充）
    import random as _rng_47
    _male_inject_count = 0
    _total_scenes_47 = len(results)
    _scenes_needing_male = []
    for idx_47, scene in enumerate(results):
        intensity = scene.get("intensity", 3)
        if intensity < 3:
            continue
        bubbles = scene.get("bubbles", [])
        has_male_bubble = any(
            b.get("speaker", "") and _is_male_by_name(b.get("speaker", ""))
            for b in bubbles
        )
        if not has_male_bubble and 1 <= len(bubbles) < 3:
            _scenes_needing_male.append(idx_47)
    # 40-50%のシーンに男性セリフを注入
    if _scenes_needing_male:
        _inject_target = max(1, int(len(_scenes_needing_male) * 0.45))
        _rng_47.shuffle(_scenes_needing_male)
        _inject_candidates = _scenes_needing_male[:_inject_target]
        try:
            from ero_dialogue_pool import get_male_speech_pool_for_theme as _get_male_47
            _has_male_pool_47 = True
        except ImportError:
            _has_male_pool_47 = False
        if _has_male_pool_47:
            _used_male_47 = set()
            for idx_47 in _inject_candidates:
                scene = results[idx_47]
                intensity = scene.get("intensity", 3)
                _pool_47 = _get_male_47(theme, intensity)
                if not _pool_47:
                    continue
                candidates_47 = [t for t in _pool_47 if t not in _used_male_47]
                if not candidates_47:
                    candidates_47 = _pool_47
                male_text = _rng_47.choice(candidates_47)
                _used_male_47.add(male_text)
                # 男性speakerの名前を推定
                _male_speaker = "男"
                if heroine_name_set:
                    for b in scene.get("bubbles", []):
                        sp = b.get("speaker", "")
                        if sp and sp not in heroine_name_set:
                            _male_speaker = sp
                            break
                new_bubble = {"type": "speech", "speaker": _male_speaker, "text": male_text}
                # バブルの後半に挿入（moanの後、speechの前あたり）
                insert_at = len(scene.get("bubbles", [])) // 2
                scene.setdefault("bubbles", []).insert(insert_at, new_bubble)
                _male_inject_count += 1
    if _male_inject_count > 0:
        log_message(f"  Step 4.57 男性セリフ補充: {_male_inject_count}シーンに注入（i≥3、男性バブルなしの45%に分配）")

    # 4.58. 男性セリフ頻度制限（全体の35%以下に抑制）
    total_scenes = len(results)
    max_male_scenes = max(2, int(total_scenes * 0.35))
    scenes_with_male = []
    for idx, scene in enumerate(results):
        if "bubbles" not in scene:
            continue
        for bubble in scene["bubbles"]:
            speaker = bubble.get("speaker", "")
            if speaker and _is_male_by_name(speaker):
                scenes_with_male.append(idx)
                break
    if len(scenes_with_male) > max_male_scenes:
        # 保持するシーン: intensity 5のシーン + 最初の男性登場シーン + 等間隔
        keep_indices = set()
        # intensity 5は必ず保持
        for idx in scenes_with_male:
            if results[idx].get("intensity", 0) >= 5:
                keep_indices.add(idx)
        # 最初の男性登場も保持
        if scenes_with_male:
            keep_indices.add(scenes_with_male[0])
        # 残り枠を等間隔で配分
        remaining_slots = max_male_scenes - len(keep_indices)
        if remaining_slots > 0:
            candidates = [i for i in scenes_with_male if i not in keep_indices]
            if candidates:
                step = max(1, len(candidates) // remaining_slots)
                for j in range(0, len(candidates), step):
                    if len(keep_indices) >= max_male_scenes:
                        break
                    keep_indices.add(candidates[j])
        # 不要な男性セリフを削除
        _male_freq_removed = 0
        for idx in scenes_with_male:
            if idx in keep_indices:
                continue
            scene = results[idx]
            scene["bubbles"] = [
                b for b in scene["bubbles"]
                if not (b.get("speaker", "") and _is_male_by_name(b.get("speaker", "")))
            ]
            _male_freq_removed += 1
        if _male_freq_removed > 0:
            log_message(f"  男性セリフ頻度制限: {_male_freq_removed}シーンから男性セリフ除去（{len(scenes_with_male)}→{len(scenes_with_male)-_male_freq_removed}シーン）")

    # 4.6. 括弧除去・「らめ」修正・一人称ブレ修正・タイトル長制限
    _46_fix_count = 0
    # 一人称マップ構築（キャラプロファイルから）
    _first_person_map = {}  # character_name -> first_person
    if char_profiles:
        for cp in char_profiles:
            cn = cp.get("character_name", "")
            fp = cp.get("first_person", "")
            if cn and fp:
                _first_person_map[cn] = fp
    for scene in results:
        # タイトル品質修正（長制限 + location混入 + 句点 + 重複）
        title = scene.get("title", "")
        # 句点を含むタイトルは不正（descriptionの混入）
        if "。" in title:
            title = title.split("。")[0].strip()
            scene["title"] = title
            log_message(f"  S{scene.get('scene_id','?')}: タイトル句点除去")
            _46_fix_count += 1
        # location名がタイトルに混入している場合、除去
        _loc_keywords = ["トイレ個室", "の駅の", "教室内", "部屋の中"]
        for lk in _loc_keywords:
            if lk in title and len(title) > 15:
                parts = title.split(lk)
                cleaned = parts[0].strip().rstrip("の")
                if len(cleaned) >= 3:
                    scene["title"] = cleaned
                    title = cleaned
                    log_message(f"  S{scene.get('scene_id','?')}: タイトルlocation除去「{cleaned}」")
                    _46_fix_count += 1
                break
        if len(title) > 25:
            scene["title"] = title[:25].rstrip("。、…")
            log_message(f"  S{scene.get('scene_id','?')}: タイトル短縮")
            _46_fix_count += 1
        if "bubbles" not in scene:
            continue
        for bubble in scene["bubbles"]:
            txt = bubble.get("text", "")
            if not txt:
                continue
            orig = txt
            # 括弧除去
            txt = txt.strip("「」『』""")
            # 「らめ」→「だめ」修正（moanでもspeechでも）
            if "らめ" in txt:
                txt = txt.replace("らめぇぇ", "だめぇ").replace("らめぇん", "だめぇ")
                txt = txt.replace("らめにゃ", "だめぇ").replace("らめらめ", "だめだめ")
                txt = txt.replace("らめなの", "だめなの").replace("らめぇっ", "だめぇっ")
                txt = txt.replace("らめっ", "だめっ").replace("らめぇ", "だめぇ")
                txt = txt.replace("らめ", "だめ")
            # 一人称ブレチェック
            speaker = bubble.get("speaker", "")
            expected_fp = _first_person_map.get(speaker, "")
            if expected_fp and expected_fp != "あたし" and "あたし" in txt:
                txt = txt.replace("あたし", expected_fp)
            if txt != orig:
                bubble["text"] = txt
                _46_fix_count += 1
    if _46_fix_count > 0:
        log_message(f"  括弧/らめ/一人称修正: {_46_fix_count}件")

    # 4.7. 不自然表現の自動修正（書き言葉→話し言葉、句点除去、ひらがな化）
    # v8.7: 2層分離 + 部分置換化（LLM生成セリフの文脈を保持）
    # 層1: 医学用語/敬語/設定用語（必ず部分置換、breakしない→複数適用可）
    _HARD_REPLACEMENTS = {
        # --- 医学用語→俗語 ---
        "性器": "あそこ",
        "挿入して": "いれて",
        "射精して": "だして",
        "絶頂に達": "イっちゃ",
        "愛液が": "ぬるぬる…",
        "勃起": "おっき",
        "口腔内に": "くちのなか…",
        "嚥下する": "ごくん…♡",
        "口内射精": "おくちに…♡",
        # --- 過剰敬語→くだけた表現 ---
        "してもよろしいですか": "して…",
        "感じてしまいます": "感じちゃ…",
        "見ないでください": "みないで…",
        "触らないでください": "さわんな…",
        "行ってしまいます": "イっちゃ…",
        "出てしまいます": "でちゃ…",
        "止められません": "とまんない…",
        "ありがとうございます": "ありがと…",
        "すみません": "ごめん…",
        "分かりました": "うん…",
        "大丈夫です": "だいじょぶ…",
        "嬉しい気持ちです": "うれしい…♡",
        "気持ちいいです": "きもちぃ…♡",
        "お願いします": "おねがい…♡",
        "やめてください": "やめ…",
        "怖いです": "こわい…",
        "痛いです": "いた…",
        "すごいです": "すご…",
        "もう限界です": "もう…むりぃ…♡",
        "声が出てしまいます": "あぁん♡",
        # --- お嬢様口調→CG集 ---
        "でございますの": "…の…♡",
        "いたしますわ": "ちゃう…♡",
        "くださいませ": "…♡",
        "よろしくてよ": "いい…♡",
        # --- 硬い接続詞 ---
        "しかしながら": "でも…",
        "それにもかかわらず": "なのに…",
        "したがって": "…",
        "なぜならば": "…",
        "とはいえ": "…けど…",
        "あるいは": "…",
        "一方で": "…",
        "いわゆる": "…",
        "つまるところ": "…",
        "要するに": "…",
        "察するに": "…",
    }
    # 層2: 文学/小説的表現（短文なら全置換、長文なら部分置換→文脈保持）
    _SOFT_REPLACEMENTS = {
        # --- 長文的表現→短縮 ---
        "信じられない": "うそ…",
        "現実じゃない": "え…うそ…",
        "考えられない": "なんで…",
        "受け入れてしまう": "あ…っ",
        "感じてしまう": "あ…やば…",
        "声が出てしまう": "あ…声…っ",
        "何も考えられない": "まっしろ…",
        "離れたくない": "いかないで…",
        "体温が上がる": "あつい…",
        "ずっと震えてる": "ふるえ…てる",
        "抗えない": "やだ…のに…",
        "本当にいいの？": "いいの…？",
        "もう我慢できない": "むり…♡",
        "恥ずかしい": "はずかし…",
        "どうしよう": "どしよ…",
        # --- 小説的独白→CG集thought ---
        "心臓が高鳴る": "ドキドキ…",
        "体が熱くなってきた": "あつい…",
        "頭が真っ白になる": "なにも…かんがえられ…",
        "全身が痺れるような": "ビリビリ…",
        "理性が飛びそう": "もう…むり…",
        "意識が遠のく": "とお…く…",
        "体の芯が疼く": "うずうず…",
        # --- 説明的表現→感情的表現 ---
        "とても気持ちが良い": "きもちぃ♡",
        "快感が走る": "あっ♡",
        "抵抗する力がなくなる": "ちから…はいんない…",
        "体が反応してしまう": "やだ…かってに…",
        "壊れてしまいそう": "こわれ…ちゃう…♡",
        # 催眠テーマ
        "催眠をかけられ": "ぼーっと…♡",
        "洗脳されて": "あたま…からっぽ…♡",
        "意識が朦朧と": "ふわふわ…♡",
        "催眠状態で": "とろとろ…♡",
        "暗示にかかって": "はい…♡",
        # 痴漢テーマ
        "声を抑えて": "んっ…っ",
        "周りに気づかれない": "バレ…ちゃう…",
        "人目を気にしながら": "だめ…ここ…",
        # 射精受け
        "精液が流れ込む": "あつい…♡♡",
        "子宮に届いた": "おく…♡♡♡",
        "体内に射精される": "中…あつい…♡♡",
        # 2回戦
        "再び挿入される": "また…♡♡",
        "感度が上がって": "さっきより…♡♡",
        # --- 拘束/SM系 ---
        "拘束されている": "にげらんない…",
        "縛られたまま": "うごけない…",
        "目隠しをされ": "みえない…こわい…",
        "自由を奪われ": "からだ…うごかない…",
        # --- 義父/近親系 ---
        "お義父さんに": "パパに…",
        "義理の父に": "パパに…",
        "血のつながりはない": "かぞく…なのに…",
        # --- フェラ/口 ---
        "咥えさせられ": "くわえて…",
        # --- 小説的→CG集 ---
        "彼に抱かれて": "だかれ…♡",
        "快感の波が": "きもち…やば…♡",
        "理性の糸が切れる": "きれ…ちゃう…♡",
        "限界が近い": "もう…むり…♡",
        "喘ぎ声が漏れる": "あ…声…♡",
        "息が荒くなって": "はぁはぁ…",
        "涙が頬を伝う": "なみだ…とまんない…",
        "最後の一線を越え": "こえちゃ…う…♡",
        "快楽に溺れる": "おぼれ…ちゃう…♡",
        "体が震えて": "ふるえてる…",
        "愛おしい気持ちが溢れ": "すき…すき…♡",
        "激しく腰を振られ": "はげし…♡♡",
        # --- 書き言葉→話し言葉 ---
        "胸が苦しい": "くるしい…",
        "涙が止まらない": "なみだ…",
        "全身が震える": "ふるえ…てる…",
        "頭がおかしくなりそう": "おかしく…なる…",
        "逃げ出したい": "にげたい…",
        "声を殺して": "んっ…っ",
        "目を背けたい": "みたくない…",
        "許してほしい": "ゆるして…",
        "もう一度してほしい": "もっかい…♡",
        "気が狂いそう": "おかしく…なっちゃう…",
        "体が言うことを聞かない": "からだ…かって…に…",
        "耐えられない": "むりぃ…♡",
        "苦しいくらいに気持ちいい": "くるし…きもちぃ…♡",
        "何度もイかされて": "またイ…っちゃ…♡♡",
        "お腹の中が熱い": "おなか…あつい…♡",
        "頭の中が真っ白": "しろ…い…",
        "我を忘れて": "もう…なにも…",
        # --- 文語表現→CG集 ---
        "溢れ出す": "あふれて…",
        "身を委ねる": "まかせ…ちゃう…♡",
        "恍惚として": "とろとろ…♡",
        "蕩ける": "とけちゃ…う…♡",
        "嬌声を上げ": "あ…んっ♡",
        "甘い吐息": "はぁ…♡",
        "悦びに": "きもちぃ…",
        "淫らな": "えっちな…",
        # --- 文学的表現→CG集口語 ---
        "心の奥底で": "…こころの…おく…",
        "快楽に支配され": "きもちよすぎ…て…",
        "陶酔に浸り": "とろとろ…\u2665",
        "背徳感に": "いけないこと…",
        "羞恥心が": "はずかし…",
        "嫌悪感を": "いや…",
        "自制心が": "がまん…できな…",
        "抗えない衝動": "とめらんない…",
    }
    # 男性セリフの不自然表現修正（heroine_name_setが必要なので判定付き）
    _MALE_SPEECH_REPLACEMENTS = {
        "可愛いね": "かわいい",
        "気持ちよくしてあげる": "イかせてやる",
        "もっと感じて": "もっと",
        "素直になれよ": "素直にしろ",
        "愛してるよ": "好きだ",
        "気持ちいいだろ？": "いいだろ",
    }
    _HIRAGANA_MAP = {
        "気持ちいい": "きもちぃ",
        "気持ちいぃ": "きもちぃ",
        "気持ち良い": "きもちぃ",
        "大好き": "だいすき",
        "好き": "すき",
        "欲しい": "ほしい",
        "可愛い": "かわいい",
        "怖い": "こわい",
        "嬉しい": "うれしい",
        "凄い": "すごい",
        "駄目": "だめ",
        "嫌": "いや",
        "奥": "おく",
        "中": "なか",
        "熱い": "あつい",
        "深い": "ふかい",
        # v7.0追加
        "痛い": "いたい",
        "汚い": "きたない",
        "苦しい": "くるしい",
        "太い": "ふとい",
        "硬い": "かたい",
        "強い": "つよい",
        "早い": "はやい",
        "速い": "はやい",
        "壊れる": "こわれる",
        "溶ける": "とける",
        "出来ない": "できない",
        # v7.1追加
        "嫌い": "きらい",
        "遅い": "おそい",
        "重い": "おもい",
        "狭い": "せまい",
        # v7.4追加
        "臭い": "くさい",
        "濡れる": "ぬれる",
        "震える": "ふるえる",
        "崩れる": "くずれる",
        "乱れる": "みだれる",
    }
    _fix_count = 0
    for scene in results:
        if "bubbles" not in scene:
            continue
        for bubble in scene["bubbles"]:
            txt = bubble.get("text", "")
            if not txt:
                continue
            original_txt = txt
            # 句点「。」除去
            if "。" in txt:
                txt = txt.replace("。", "…")
            # 男性セリフの不自然表現修正
            speaker = bubble.get("speaker", "")
            is_male = speaker and heroine_name_set and speaker not in heroine_name_set
            if is_male:
                for ng, ok in _MALE_SPEECH_REPLACEMENTS.items():
                    if ng in txt:
                        txt = txt.replace(ng, ok)
            # 設備名混入の部分置換（便器→こんなとこ 等）
            # 「肉便器」はエロ漫画スラングなので保護
            _has_niku = "肉便器" in txt
            _FIXTURE_SUBS = {
                "便器": "こんなとこ",
                "便座": "ここ",
                "手洗い台": "ここ",
                "手洗い鏡": "鏡",
                "トイレットペーパー": "",
            }
            for fw, repl in _FIXTURE_SUBS.items():
                if fw in txt:
                    if fw == "便器" and _has_niku:
                        continue  # 肉便器は保護
                    txt = txt.replace(fw, repl)
            # 置換で生じた空の「…」連続を整理
            while "……" in txt:
                txt = txt.replace("……", "…")
            # v8.7: 層1(HARD) - 医学用語/敬語は常に部分置換（複数適用可）
            for ng, ok in _HARD_REPLACEMENTS.items():
                if ng in txt:
                    txt = txt.replace(ng, ok)
            # v8.7: 層2(SOFT) - 文学表現は短文→全置換、長文→部分置換（文脈保持）
            for ng, ok in _SOFT_REPLACEMENTS.items():
                if ng in txt:
                    if len(txt) <= len(ng) + 5:
                        txt = ok       # 短文→全置換（ほぼ全体がNG表現）
                    else:
                        txt = txt.replace(ng, ok)  # 長文→部分置換（文脈保持）
            # ひらがな化（エロシーン向け）
            for kanji, hira in _HIRAGANA_MAP.items():
                if kanji in txt:
                    txt = txt.replace(kanji, hira)
            bubble["text"] = txt
            if txt != original_txt:
                _fix_count += 1
    if _fix_count > 0:
        log_message(f"  セリフ自動修正: {_fix_count}件の不自然表現を修正")

    # 5. シーン間の同一セリフ・SE重複除去（プールから代替置換）
    #    ※重複セリフをプールから代替置換する
    _progress("Step 5 セリフ重複除去")
    heroine_names = []
    if char_profiles:
        for cp in char_profiles:
            n = cp.get("character_name", "")
            if n:
                heroine_names.append(n)
    try:
        _deduplicate_across_scenes(results, theme=theme, heroine_names=heroine_names,
                                   char_profiles=char_profiles, concept=concept)
    except Exception as _dedup_err:
        log_message(f"  [WARN]セリフ重複除去エラー（スキップ）: {_dedup_err}")
        import traceback
        log_message(traceback.format_exc())

    # 6. 3シーン連続同一locationの自動修正
    try:
        _fix_consecutive_locations(results)
    except Exception as _loc_err:
        log_message(f"  [WARN]location多様化エラー（スキップ）: {_loc_err}")

    # 7. 吹き出し数上限トリミング（3個以下: ヒロイン1-2 + 男性0-1）
    for scene in results:
        bubbles = scene.get("bubbles", [])
        if len(bubbles) > 3:
            # ヒロインセリフ最大2個 + 男性セリフ0-1個に絞る
            heroine_b = [b for b in bubbles if b.get("speaker", "") != "男"]
            male_b = [b for b in bubbles if b.get("speaker", "") == "男"]
            # ヒロイン: moan > thought > speech の優先度で2個選択
            def _bubble_priority(b):
                btype = b.get("type", "")
                if btype == "moan":
                    return 2
                if btype == "thought":
                    return 1
                return 0
            heroine_b.sort(key=_bubble_priority, reverse=True)
            kept = heroine_b[:2]
            if male_b:
                kept.append(male_b[0])
            scene["bubbles"] = kept[:3]

    _progress("Step 8-10 喘ぎ・セリフ品質修正")
    # 8. moanタイプ内容修正（3段階: 漢字/助詞/非喘ぎ語彙 → プールから置換）
    # 根拠: MOAN_POOL全400エントリは仮名+装飾のみ。
    #   漢字・助詞がある=LLMの誤生成。
    #   AFTERMATH_POOL語彙(ぼーっと,ぐったり等)は身体状況報告で喘ぎではない。
    _kanji_re = re.compile(r'[\u4e00-\u9faf\u3400-\u4dbf]')
    _sentence_end_re = re.compile(
        r'(だ|です|ます|ない|ない…|ている|てる|する|される|して|した|しい)$')
    _NON_MOAN_WORDS = frozenset([
        "ぼーっと", "ぐったり", "ふわふわ", "ごめん", "どしよ",
        "なにこれ", "もうむり", "もう…むり", "なにこれ…", "ごめん…",
        "どしよ…", "ぼーっと…", "ぐったり…", "ふわふわ…",
    ])
    try:
        from ero_dialogue_pool import (get_moan_pool, get_speech_pool,
                                       pick_replacement, infer_phase)
        _has_pool = True
    except ImportError:
        _has_pool = False

    def _get_moan_pool_with_char(intensity: int) -> list:
        """get_moan_poolにキャラ固有プールを優先混合"""
        pool = []
        if _char_pool and "moan" in _char_pool:
            char_moans = _char_pool["moan"].get(str(intensity), [])
            pool.extend(char_moans)
        pool.extend(get_moan_pool(intensity))
        return pool

    # 性格タイプ別speech禁止パターン（キャラ個性消失防止）
    _af_personality = ""
    if char_profiles:
        try:
            _af_personality = _detect_personality_type(char_profiles)
        except Exception:
            pass
    _PERSONALITY_SPEECH_EXCLUDE = {
        # _detect_personality_type() の返り値に完全対応
        "seiso": ["もっと！", "すごい！", "最高！", "もっとして！", "ちょうだい♡♡",
                  "壊して", "犯して", "もっと激しく"],  # 清楚=激しい要求NG
        "tsundere": ["好き…♡", "嬉しい♡", "大好き♡", "幸せ♡", "ずっと一緒♡"],  # 低intensityのみ適用
        "kuudere": ["きゃー", "すごーい", "わーい", "もっともっと♡", "いっぱい♡"],
        "ojou": ["ヤバい", "マジ", "ウケる", "まじ", "やべー", "ちょー"],
        "submissive": [],  # 従順=制限なし（何でも受け入れる）
        "sadistic": ["怖い", "やめて", "痛い", "助けて", "許して"],  # S気質=怯えNG
        "gal": ["お願いします", "すみません", "恐れ入り"],  # ギャル=敬語NG
        "genki": ["もう…無理…", "動けない…", "力が…"],  # 元気=脱力NG
        "inkya": ["もっと！", "すごい！", "最高！", "もっとして！", "きゃー"],  # 陰キャ=過度にはしゃぐNG
    }

    # THOUGHT_POOL intensity別フィルタキーワード
    # 低intensity(1-2): 穏やか・戸惑い系 → 激しい快感表現を除外
    # 高intensity(4-5): 快感・崩壊系 → 穏やか・冷静表現を除外
    _THOUGHT_INTENSITY_EXCLUDE = {
        "low": ["♡♡", "壊れ", "溶け", "おかしくなる", "狂", "もっと欲しい", "全部",
                "真っ白", "快感", "気持ちいい", "とまらない", "止まらない", "中毒"],
        "high": ["大丈夫", "平気", "冷静", "落ち着", "普通に", "気にしない", "余裕"],
    }

    def _get_speech_pool_with_char(btype: str, theme_: str, intensity: int,
                                    scene_idx: int = 0, total: int = 1) -> list:
        """get_speech_poolにキャラ固有プールを優先混合 + thought intensity適合フィルタ"""
        phase = ""
        if _has_pool:
            try:
                phase = infer_phase(intensity, scene_idx, total)
            except Exception:
                pass
        pool = []
        if _char_pool and btype in _char_pool:
            if phase and isinstance(_char_pool[btype], dict):
                char_lines = _char_pool[btype].get(phase, [])
                # サブフェーズ→ベースフェーズフォールバック
                if not char_lines and "_" in phase:
                    base_phase = phase.rsplit("_", 1)[0]
                    char_lines = _char_pool[btype].get(base_phase, [])
                pool.extend(char_lines)
        pool.extend(get_speech_pool(btype, theme_, intensity, phase=phase))
        # thought の intensity 適合フィルタ
        if btype == "thought" and pool:
            if intensity <= 2:
                _excl = _THOUGHT_INTENSITY_EXCLUDE["low"]
                _filtered = [p for p in pool if not any(kw in p for kw in _excl)]
                if len(_filtered) >= 10:  # フィルタ後最低10個確保
                    pool = _filtered
            elif intensity >= 4:
                _excl = _THOUGHT_INTENSITY_EXCLUDE["high"]
                _filtered = [p for p in pool if not any(kw in p for kw in _excl)]
                if len(_filtered) >= 10:
                    pool = _filtered
        # 性格タイプ別speech/thoughtフィルタ（キャラ個性消失防止）
        if _af_personality and btype in ("speech", "thought") and pool:
            _p_excl = _PERSONALITY_SPEECH_EXCLUDE.get(_af_personality, [])
            # ツンデレは低intensityのみ素直表現を禁止（高intensityでは堕ちてOK）
            if _af_personality == "tsundere" and intensity >= 4:
                _p_excl = []
            if _p_excl:
                _p_filtered = [p for p in pool if not any(kw in p for kw in _p_excl)]
                if len(_p_filtered) >= 10:
                    pool = _p_filtered
        return pool

    def _get_male_pool_for_theme(theme_str: str, intensity: int) -> list:
        """テーマ・intensity連動で男性セリフプールを返す（auto_fix用）"""
        try:
            from ero_dialogue_pool import SPEECH_MALE_POOL, get_male_speech_pool
        except ImportError:
            return ["もっと", "どうした", "来い", "行くぞ", "いいだろ"]
        t = theme_str.lower() if theme_str else ""
        pool = []
        if any(k in t for k in ["痴漢", "chikan", "公共", "public", "電車", "train", "トイレ"]):
            pool.extend(SPEECH_MALE_POOL.get("chikan", []))
            pool.extend(SPEECH_MALE_POOL.get("taunt", []))
            pool.extend(SPEECH_MALE_POOL.get("public", []))
            pool.extend(SPEECH_MALE_POOL.get("command", []))
        elif any(k in t for k in ["ntr", "寝取", "夜這", "村", "レイプ", "陵辱", "調教", "奴隷"]):
            pool.extend(SPEECH_MALE_POOL.get("command", []))
            pool.extend(SPEECH_MALE_POOL.get("dirty", []))
        elif any(k in t for k in ["純愛", "ラブ", "恋人", "カップル"]):
            pool.extend(SPEECH_MALE_POOL.get("gentle", []))
            pool.extend(SPEECH_MALE_POOL.get("praise", []))
        else:
            if intensity >= 4:
                pool.extend(SPEECH_MALE_POOL.get("command", []))
                pool.extend(SPEECH_MALE_POOL.get("dirty", []))
            else:
                pool.extend(SPEECH_MALE_POOL.get("dirty", []))
                pool.extend(SPEECH_MALE_POOL.get("praise", []))
        return pool if pool else [v for sp in SPEECH_MALE_POOL.values() for v in sp]

    _moan_fix_count = 0
    _used_moan_for_fix = set()
    for scene in results:
        intensity = scene.get("intensity", 3)
        for b in scene.get("bubbles", []):
            if b.get("type") != "moan":
                continue
            txt = b.get("text", "")
            if not txt:
                continue
            stripped = txt.rstrip("…♡♡♡")
            is_non_moan = (bool(_kanji_re.search(txt))
                           or bool(_sentence_end_re.search(txt))
                           or stripped in _NON_MOAN_WORDS
                           or txt in _NON_MOAN_WORDS)
            if is_non_moan and _has_pool:
                pool = _get_moan_pool_with_char(intensity)
                replacement = pick_replacement(pool, _used_moan_for_fix, _normalize_bubble_text)
                if replacement:
                    log_message(f"  moan内容修正: 「{txt}」→「{replacement}」")
                    b["text"] = replacement
                    _used_moan_for_fix.add(replacement)
                    _moan_fix_count += 1
    if _moan_fix_count > 0:
        log_message(f"  moanタイプ内容修正: {_moan_fix_count}件")

    # 8b. MOAN_POOL非含有moanの強制プール置換
    # 根拠: LLMが生成する「ふあひっ」「ん゛お゛っ」等の不自然な濁点組み合わせは
    #   Step 8の漢字/助詞チェックでは検出できないが、MOAN_POOLの400エントリには存在しない。
    #   正規化後にプール内に見つからないmoanは全て強制置換する。
    _moan_pool_fix_count = 0
    try:
        from ero_dialogue_pool import get_all_moan_normalized
        _all_moan_norms = get_all_moan_normalized()
    except ImportError:
        _all_moan_norms = None
    if _all_moan_norms and _has_pool:
        for scene in results:
            intensity = scene.get("intensity", 3)
            for b in scene.get("bubbles", []):
                if b.get("type") != "moan":
                    continue
                txt = b.get("text", "")
                if not txt:
                    continue
                norm = _normalize_bubble_text(txt)
                if norm and norm not in _all_moan_norms:
                    pool = _get_moan_pool_with_char(intensity)
                    replacement = pick_replacement(pool, _used_moan_for_fix, _normalize_bubble_text)
                    if replacement:
                        log_message(f"  moanプール外修正: 「{txt}」→「{replacement}」")
                        b["text"] = replacement
                        _used_moan_for_fix.add(replacement)
                        _moan_pool_fix_count += 1
        if _moan_pool_fix_count > 0:
            log_message(f"  MOANプール外強制置換: {_moan_pool_fix_count}件")

    # 9. speechタイプ身体状況報告修正（intensity>=3のアクションシーン）
    # 根拠: CG集のspeechは感情的反応。身体状態の客観報告はナレーションでありセリフ不適。
    _BODY_REPORT_KW = frozenset([
        "涙が", "汗すごい", "汗が", "声出ない", "息できない",
        "力入んない", "頭まっしろ", "目が回る", "指先痺れ",
        "全身痺れ", "まだ震えて", "震えてる", "動けない",
        "立てない", "からだ重い", "呼吸が", "ぼーっと",
        "ぐったり", "ふわふわ", "思考が", "意識が",
    ])
    _body_fix_count = 0
    _used_speech_for_fix = set()
    if _has_pool:
        if not theme and results:
            # メタデータからテーマ取得（5テーマ自動検出）
            all_desc = " ".join(
                s.get("description", "") + " " + s.get("mood", "")
                for s in results[:5]
            )
            if any(k in all_desc for k in ["寝取", "NTR", "ntr"]):
                theme = "ntr"
            elif any(k in all_desc for k in ["襲", "犯さ", "無理矢理", "暴行", "レイプ", "陵辱", "強制", "痴漢"]):
                theme = "forced"
            elif any(k in all_desc for k in ["催眠", "洗脳", "堕落", "調教", "奴隷"]):
                theme = "corruption"
            elif any(k in all_desc for k in ["嫌がる", "逃げ", "恐怖", "怯え", "抵抗"]):
                theme = "reluctant"
            elif any(k in all_desc for k in ["純愛", "恋人", "カップル", "両想"]):
                theme = "vanilla"
            if theme:
                log_message(f"  テーマ自動検出: {theme}")
        _total_scenes = len(results)
        for _si, scene in enumerate(results):
            intensity = scene.get("intensity", 3)
            if intensity < 3:
                continue
            for b in scene.get("bubbles", []):
                if b.get("type") not in ("speech", "thought"):
                    continue
                txt = b.get("text", "")
                if not txt:
                    continue
                is_body_report = any(kw in txt for kw in _BODY_REPORT_KW)
                if is_body_report:
                    pool = _get_speech_pool_with_char(b["type"], theme, intensity, _si, _total_scenes)
                    # プールから身体状況報告を除外（循環置換防止）
                    pool = [t for t in pool
                            if not any(kw in t for kw in _BODY_REPORT_KW)]
                    replacement = pick_replacement(pool, _used_speech_for_fix,
                                                   _normalize_bubble_text,
                                                   intensity=intensity)
                    if replacement:
                        log_message(f"  身体状況報告修正({b['type']}): 「{txt}」→「{replacement}」")
                        b["text"] = replacement
                        _used_speech_for_fix.add(replacement)
                        _body_fix_count += 1
    if _body_fix_count > 0:
        log_message(f"  身体状況報告修正: {_body_fix_count}件")

    # 9a. 男性セリフ観察実況修正（「～だな」「～してるな」→ 命令/挑発型に置換）
    _MALE_OBSERVATION_PATTERNS = {
        "いい声だな": "もっと鳴け",
        "敏感だな": "もっと感じろ",
        "いい反応だな": "素直になれよ",
        "いい体だな": "脱げ",
        "いい体してんな": "脱げ",
        "エロい体だな": "もっと見せろ",
        "体は正直だな": "隠すなよ",
        "感じやすいな": "もう我慢すんな",
        "素直になったな": "もっとくれ",
        "とろとろだな": "もっと締めろ",
        "ここも敏感だな": "ここも好きだろ",
        "いい顔だな": "もっとくれ",
        "かわいい顔だな": "かわいい",
        "いい表情だな": "もっと見せろ",
        "たまらん顔だな": "たまんねぇ",
        "極上だな": "最高だ",
        "見事だな": "最高だ",
        # v7.4追加
        "反応がいいな": "もっと感じろ",
        "締まりがいいな": "もっと締めろ",
        "いい匂いだな": "いい匂いだ",
        "素直だな": "素直だろ",
        "正直だな": "正直だろ",
    }
    import re as _re_autofix
    # 「～だな」「～してるな」「～だろうな」で終わる観察型パターン検出
    _MALE_OBS_RE = _re_autofix.compile(r".{4,}(?:だな|するな|してるな|だろうな|てるな)$")
    _male_obs_fix_count = 0
    _used_male_obs_fix = set()
    for scene in results:
        for b in scene.get("bubbles", []):
            speaker = b.get("speaker", "")
            if not (speaker and _is_male_by_name(speaker)):
                continue
            if b.get("type") != "speech":
                continue
            txt = b.get("text", "").strip()
            if not txt:
                continue
            # 1. 辞書マッチ
            replaced = False
            for obs_pattern, obs_replacement in _MALE_OBSERVATION_PATTERNS.items():
                if obs_pattern in txt:
                    old_txt = txt
                    b["text"] = obs_replacement
                    log_message(f"  男性観察型修正: 「{old_txt}」→「{obs_replacement}」")
                    _male_obs_fix_count += 1
                    replaced = True
                    break
            if replaced:
                continue
            # 2. 正規表現マッチ（「～だな」系で終わる4文字以上）
            clean = txt.rstrip("…♡♥")
            if _MALE_OBS_RE.match(clean):
                intensity = scene.get("intensity", 3)
                pool = _get_male_pool_for_theme(theme if theme else "", intensity)
                # 観察型を除外
                pool = [p for p in pool if not _MALE_OBS_RE.match(p.rstrip("…♡♥"))]
                replacement = pick_replacement(pool, _used_male_obs_fix, _normalize_bubble_text,
                                               intensity=intensity)
                if replacement:
                    log_message(f"  男性観察型修正(regex): 「{txt}」→「{replacement}」")
                    b["text"] = replacement
                    _used_male_obs_fix.add(replacement)
                    _male_obs_fix_count += 1
    if _male_obs_fix_count > 0:
        log_message(f"  男性観察型セリフ修正: {_male_obs_fix_count}件")

    # 9b. thought部位ラベル冒頭修正（「胸…」「太もも…」等の部位名冒頭を感覚型に置換）
    _BODY_PART_LABELS = [
        "胸", "太もも", "お尻", "首筋", "耳", "唇", "舌", "指", "脚", "腕",
        "背中", "お腹", "腰", "膝", "肩", "足", "髪", "うなじ", "乳首", "クリ",
        "おっぱい", "おしり", "ふともも", "くちびる", "みみ",
    ]
    _BODY_PART_RE = _re_autofix.compile(
        r"^(" + "|".join(_re_autofix.escape(bp) for bp in _BODY_PART_LABELS) + r")…"
    )
    _thought_body_fix_count = 0
    if _has_pool:
        _used_thought_fix = set()
        for _si_tb, scene in enumerate(results):
            intensity = scene.get("intensity", 3)
            for b in scene.get("bubbles", []):
                if b.get("type") != "thought":
                    continue
                txt = b.get("text", "")
                if not txt:
                    continue
                m = _BODY_PART_RE.match(txt)
                if m:
                    pool = _get_speech_pool_with_char("thought", theme if theme else "",
                                                      intensity, _si_tb, len(results))
                    # 部位ラベル冒頭を除外
                    pool = [t for t in pool if not _BODY_PART_RE.match(t)]
                    replacement = pick_replacement(pool, _used_thought_fix, _normalize_bubble_text,
                                                   intensity=intensity)
                    if replacement:
                        log_message(f"  thought部位ラベル修正: 「{txt}」→「{replacement}」")
                        b["text"] = replacement
                        _used_thought_fix.add(replacement)
                        _thought_body_fix_count += 1
    if _thought_body_fix_count > 0:
        log_message(f"  thought部位ラベル冒頭修正: {_thought_body_fix_count}件")

    # 9c. thought 20文字超をトリミング（ナレーション化防止）
    _thought_trim_count = 0
    for scene in results:
        for b in scene.get("bubbles", []):
            if b.get("type") == "thought" and len(b.get("text", "")) > 20:
                txt = b["text"]
                # 「…」で切れ目を探して20文字以内に
                cut = txt[:20].rfind("\u2026")
                if cut > 5:
                    b["text"] = txt[:cut + 1]
                else:
                    b["text"] = txt[:18] + "\u2026"
                _thought_trim_count += 1
    if _thought_trim_count > 0:
        log_message(f"  thought長さトリミング: {_thought_trim_count}件")

    # 10. 同一シーン内テキスト重複修正
    _intra_dup_count = 0
    if _has_pool:
        _total_s10 = len(results)
        for _si10, scene in enumerate(results):
            intensity = scene.get("intensity", 3)
            bubbles = scene.get("bubbles", [])
            seen_texts = set()
            for b in bubbles:
                txt = b.get("text", "")
                if not txt:
                    continue
                if txt in seen_texts:
                    btype = b.get("type", "speech")
                    if btype == "moan":
                        pool = _get_moan_pool_with_char(intensity)
                        repl = pick_replacement(pool, _used_moan_for_fix,
                                                _normalize_bubble_text,
                                                intensity=intensity)
                    else:
                        pool = _get_speech_pool_with_char(btype, theme, intensity, _si10, _total_s10)
                        repl = pick_replacement(pool, _used_speech_for_fix,
                                                _normalize_bubble_text,
                                                intensity=intensity)
                    if repl:
                        log_message(f"  シーン内重複修正: 「{txt}」→「{repl}」")
                        b["text"] = repl
                        _used_speech_for_fix.add(repl)
                        _intra_dup_count += 1
                seen_texts.add(txt)
    if _intra_dup_count > 0:
        log_message(f"  シーン内重複修正: {_intra_dup_count}件")

    # 10b. THOUGHT↔SPEECH感情矛盾修正
    _POSITIVE_THOUGHT_KW = ["幸せ", "嬉しい", "好き", "大好き", "気持ちいい", "もっと", "欲しい", "♡"]
    _NEGATIVE_SPEECH_KW = ["やめて", "嫌", "離して", "痛い", "やだ", "助けて", "来ないで"]
    _NEGATIVE_THOUGHT_KW = ["怖い", "嫌だ", "逃げ", "助けて", "痛い", "無理", "嫌い"]
    _POSITIVE_SPEECH_KW = ["もっと", "気持ちいい", "好き", "♡", "嬉しい", "幸せ", "ちょうだい"]
    # forced/reluctantテーマのi=3-4は矛盾が正常パターン（快楽堕ち）→免除
    _is_contradiction_exempt = any(k in (theme or "").lower()
                                    for k in ["forced", "reluctant", "陵辱", "強制"])
    _contradiction_fix_count = 0
    if _has_pool:
        for _si_ct, scene in enumerate(results):
            intensity = scene.get("intensity", 3)
            if _is_contradiction_exempt and 3 <= intensity <= 4:
                continue
            bubbles = scene.get("bubbles", [])
            _scene_thoughts = [b for b in bubbles if b.get("type") == "thought"]
            _scene_speeches = [b for b in bubbles if b.get("type") == "speech"]
            if not _scene_thoughts or not _scene_speeches:
                continue
            for th_b in _scene_thoughts:
                th_text = th_b.get("text", "")
                is_positive_thought = any(kw in th_text for kw in _POSITIVE_THOUGHT_KW)
                is_negative_thought = any(kw in th_text for kw in _NEGATIVE_THOUGHT_KW)
                for sp_b in _scene_speeches:
                    sp_text = sp_b.get("text", "")
                    is_negative_speech = any(kw in sp_text for kw in _NEGATIVE_SPEECH_KW)
                    is_positive_speech = any(kw in sp_text for kw in _POSITIVE_SPEECH_KW)
                    # ポジティブthought + ネガティブspeech → thoughtを抵抗系に差替え
                    if is_positive_thought and is_negative_speech:
                        pool = _get_speech_pool_with_char("thought", theme, intensity, _si_ct, len(results))
                        _resist_pool = [p for p in pool if any(kw in p for kw in ["でも", "なのに", "…けど", "嫌", "だめ"])]
                        if _resist_pool:
                            repl = pick_replacement(_resist_pool, _used_speech_for_fix, _normalize_bubble_text,
                                                   intensity=intensity)
                            if repl:
                                log_message(f"  感情矛盾修正: シーン{_si_ct+1} thought「{th_text[:12]}」→「{repl}」")
                                th_b["text"] = repl
                                _used_speech_for_fix.add(repl)
                                _contradiction_fix_count += 1
                        break
                    # ネガティブthought + ポジティブspeech → speechを否定系に差替え
                    if is_negative_thought and is_positive_speech:
                        pool = _get_speech_pool_with_char("speech", theme, intensity, _si_ct, len(results))
                        _deny_pool = [p for p in pool if any(kw in p for kw in ["やめ", "だめ", "嫌", "…っ", "痛"])]
                        if _deny_pool:
                            repl = pick_replacement(_deny_pool, _used_speech_for_fix, _normalize_bubble_text,
                                                   intensity=intensity)
                            if repl:
                                log_message(f"  感情矛盾修正: シーン{_si_ct+1} speech「{sp_text[:12]}」→「{repl}」")
                                sp_b["text"] = repl
                                _used_speech_for_fix.add(repl)
                                _contradiction_fix_count += 1
                        break
    if _contradiction_fix_count > 0:
        log_message(f"  感情矛盾修正: {_contradiction_fix_count}件")

    # 10c. シーン間心理状態遷移モデル（2段階以上乖離するセリフを差替え）
    # テーマ別心理遷移定義（各テーマ固有の5段階）
    _THEME_PSYCH_STAGES = {
        "netorare": {
            "stages": ["rejection", "body_betrayal", "guilt", "comparison", "fallen"],
            "speech": {
                "rejection":     ["やめて", "嫌", "離して", "来ないで", "彼氏がいる", "こんなの"],
                "body_betrayal": ["なんで…感じて", "身体が", "嘘…こんな", "おかしい", "反応して"],
                "guilt":         ["ごめん", "ごめんなさい", "許して", "最低", "彼に", "裏切り"],
                "comparison":    ["こんなの初めて", "…違う", "…こんなに", "知らなかった", "負けて"],
                "fallen":        ["もっと", "♡♡", "ちょうだい", "欲しい", "もう…いい…♡", "好き"],
            },
            "thought": {
                "rejection":     ["嫌だ…", "逃げなきゃ", "彼氏に", "こんな男"],
                "body_betrayal": ["なんで…感じてる", "身体が勝手に", "嘘…", "おかしい"],
                "guilt":         ["ごめんね…", "最低だ…私", "彼に顔向け", "裏切ってる"],
                "comparison":    ["…こんなの知らなかった", "彼とは…違う", "こんなに…奥"],
                "fallen":        ["もう…戻れない", "彼よりも…♡", "こっちの方が", "壊れちゃう"],
            },
        },
        "love": {
            "stages": ["shyness", "trust", "passion", "unity", "devotion"],
            "speech": {
                "shyness":   ["恥ずかしい", "見ないで", "明るい", "初めて", "緊張する"],
                "trust":     ["…いいよ", "信じてる", "大丈夫", "怖くない", "あなたなら"],
                "passion":   ["もっと", "近くに", "離さないで", "好き", "感じる"],
                "unity":     ["一緒に", "奥まで", "繋がって", "気持ちいい", "もっと奥"],
                "devotion":  ["大好き", "ずっと", "離れたくない", "幸せ", "また…したい♡"],
            },
            "thought": {
                "shyness":   ["恥ずかしい…", "心臓うるさい", "顔見れない"],
                "trust":     ["怖くない…この人なら", "温かい…", "安心する"],
                "passion":   ["もっと…触れたい", "好き…好き…", "溶けちゃう"],
                "unity":     ["繋がってる…", "同じ気持ち…", "満たされて"],
                "devotion":  ["幸せ…", "ずっと一緒に", "この人だけ…♡"],
            },
        },
        "forced": {
            "stages": ["fear", "resistance", "submission", "pleasure", "dependence"],
            "speech": {
                "fear":       ["やめて", "嫌", "怖い", "助けて", "離して", "痛い", "来ないで"],
                "resistance": ["やだ", "やめ…", "触らないで", "嫌って言ってる", "無理"],
                "submission": ["…もう…いい", "…好きにして", "抵抗…できない", "わかった"],
                "pleasure":   ["あっ…なんで…", "気持ち…", "止まらない", "もっと…♡"],
                "dependence": ["お願い", "もっと", "欲しい♡", "ください", "離さないで♡"],
            },
            "thought": {
                "fear":       ["怖い…", "逃げなきゃ", "誰か…助けて", "嫌だ嫌だ"],
                "resistance": ["負けない…", "こんな奴に…", "感じるもんか"],
                "submission": ["もう…無理…", "力が…入らない", "抵抗できない"],
                "pleasure":   ["なんで…気持ちいい…", "身体が…裏切る", "おかしくなる"],
                "dependence": ["もう…なしじゃ…", "この人がいないと", "壊されちゃった"],
            },
        },
        "corruption": {
            "stages": ["innocence", "curiosity", "temptation", "indulgence", "corruption"],
            "speech": {
                "innocence":  ["これ…なに", "知らない", "変な感じ", "なんで触る", "やめて"],
                "curiosity":  ["…なにこれ", "変…だけど", "もう少しだけ", "気になる"],
                "temptation": ["だめ…なのに", "止められない", "…もっと教えて", "知りたい"],
                "indulgence": ["気持ちいい", "もっと", "教えて♡", "お願い", "すごい"],
                "corruption": ["もっとして♡", "全部♡", "狂っちゃう♡", "壊して♡", "♡♡♡"],
            },
            "thought": {
                "innocence":  ["なにされてるの…", "怖い…", "わからない"],
                "curiosity":  ["…変な感じ", "知りたい…だめ？", "なんだろう…これ"],
                "temptation": ["だめって…わかってるのに", "止まれない…", "もっと…♡"],
                "indulgence": ["気持ちいい…もう…", "こんなの知らなかった", "溺れちゃう"],
                "corruption": ["もう戻れない…♡", "壊れちゃう…♡", "なんでもする…♡"],
            },
        },
        "hypnosis": {
            "stages": ["unaware", "discomfort", "acceptance", "voluntary", "normalized"],
            "speech": {
                "unaware":    ["あれ…", "なんか変", "ぼーっと", "…え？", "なにしてる…？"],
                "discomfort": ["なんで…こんな", "おかしい", "身体が勝手に", "止められない"],
                "acceptance": ["…いいかも", "気持ちいい…から", "自然と…", "求めてる"],
                "voluntary":  ["もっとして", "欲しい", "お願い", "催眠なんかじゃ", "自分から"],
                "normalized": ["当然", "毎日して♡", "ないと困る♡", "もう普通のこと♡"],
            },
            "thought": {
                "unaware":    ["あれ…なにされて…", "頭がぼんやり", "意識が…"],
                "discomfort": ["おかしい…自分の意思じゃ", "身体だけが…", "抗えない"],
                "acceptance": ["…気持ちいいから…いい…", "自然なこと…", "受け入れて"],
                "voluntary":  ["これは…私の意思…", "催眠じゃない…好きだから", "欲しい"],
                "normalized": ["何も…おかしくない…♡", "毎日の日課…♡", "普通…♡"],
            },
        },
        "chikan": {
            "stages": ["confusion", "shame", "hypersensitivity", "surrender", "anticipation"],
            "speech": {
                "confusion":       ["え…", "なに…", "誰…", "ちょっと", "やめ…", "ここ電車"],
                "shame":           ["人が…いるのに", "見られ…", "声…出ちゃ", "恥ずかし"],
                "hypersensitivity":["そこ…だめ…っ", "あっ…敏感…", "やば…", "触らないで…あっ"],
                "surrender":       ["…もう…いい", "好きにして…", "…ん…♡", "止められない…"],
                "anticipation":    ["…また…来る…？", "今日も…♡", "待ってた…♡", "奥まで…♡"],
            },
            "thought": {
                "confusion":       ["え…なに…触られて…", "電車の中で…", "嘘でしょ"],
                "shame":           ["周りにバレたら…", "声出しちゃだめ…", "最低…感じてる"],
                "hypersensitivity":["やば…身体が…", "こんなとこで…感じちゃ…", "敏感すぎ"],
                "surrender":       ["もう…どうでもいい…", "抵抗する力…ない", "このまま…"],
                "anticipation":    ["また…あの人…♡", "期待してる…最低…", "今日も…♡"],
            },
        },
        "humiliation": {
            "stages": ["defiance", "humiliation", "crumbling", "pleasure", "submission"],
            "speech": {
                "defiance":    ["ふざけないで", "こんなの", "絶対に", "屈しない", "離せ"],
                "humiliation": ["…っ", "見るな…", "恥ずかし…", "やめ…て…", "最低…"],
                "crumbling":   ["嫌…なのに…", "なんで…声が…", "身体が…", "…あ…っ"],
                "pleasure":    ["あっ…♡", "だめ…気持ち…", "認めない…けど…♡", "嫌…♡"],
                "submission":  ["…はい…♡", "ごめんなさい…♡", "負けました…♡", "お願い♡♡"],
            },
            "thought": {
                "defiance":    ["こんな奴に…絶対…", "負けない…", "プライドが"],
                "humiliation": ["恥ずかしい…見られてる", "こんな姿…", "屈辱…"],
                "crumbling":   ["なんで…感じてる…", "嫌なはず…なのに", "おかしい"],
                "pleasure":    ["認めたくない…けど…♡", "気持ちいい…嫌…♡", "負けそう"],
                "submission":  ["…負けた…♡", "もう…どうでもいい♡", "この人に…♡"],
            },
        },
        "time_stop": {
            "stages": ["shock", "helplessness", "sensation", "overwhelm", "resignation"],
            "thought": {
                "shock":        ["身体が…動かない…", "なに…されて…", "嘘…止まって…"],
                "helplessness": ["助けて…誰か…", "声も…出せない", "抵抗…できない"],
                "sensation":    ["動けないのに…感じてる", "身体だけ…反応して", "やだ…"],
                "overwhelm":    ["おかしくなる…", "止めて…壊れちゃう", "限界…"],
                "resignation":  ["もう…なんでもいい…", "好きにして…", "慣れちゃった…"],
            },
        },
        "teacher_student": {
            "stages": ["boundary", "forbidden", "secret", "addiction", "devotion"],
            "speech": {
                "boundary":  ["先生…だめです", "生徒ですよ…", "こんなの…いけない", "離れて"],
                "forbidden": ["ここ学校…", "バレたら…", "いけないって…わかってる", "…でも"],
                "secret":    ["二人だけの…秘密", "誰にも言わない", "先生だけ…", "…もっと"],
                "addiction": ["先生…♡", "放課後…待ってる♡", "もっと教えて♡", "好き…♡"],
                "devotion":  ["先生のもの♡", "ずっと…先生だけ♡", "卒業しても♡", "全部♡"],
            },
            "thought": {
                "boundary":  ["先生と生徒なのに…", "いけない…", "こんなの間違って"],
                "forbidden": ["バレたら…退学…", "でも…止められない", "禁断…"],
                "secret":    ["秘密…二人だけの", "背徳感が…", "でも嬉しい"],
                "addiction": ["先生のこと…考えちゃう", "授業中も…♡", "中毒…"],
                "devotion":  ["先生なしじゃ…もう…♡", "全部…先生に♡", "一生…♡"],
            },
        },
        "femdom": {
            "stages": ["dominance", "teasing", "control", "reward", "ownership"],
            "speech": {
                "dominance":  ["ほら…どうしたの？", "跪きなさい", "許可した？", "だめよ"],
                "teasing":    ["こんなになって…♡", "可愛い反応♡", "もっと見せて", "我慢しなさい"],
                "control":    ["私のものでしょ♡", "逃がさない♡", "言うこと聞きなさい♡"],
                "reward":     ["いい子ね♡", "ご褒美あげる♡", "気持ちいい？♡", "もっと♡"],
                "ownership":  ["私だけのもの♡♡", "離さないから♡♡", "永遠に♡♡"],
            },
            "thought": {
                "dominance":  ["この子…可愛い♡", "もっと…支配したい"],
                "teasing":    ["いい反応…♡", "もっと焦らしちゃおう♡", "壊しちゃだめ…まだ"],
                "control":    ["私のもの…♡", "逃がさない♡", "完全に…手の中♡"],
                "reward":     ["頑張ったね…♡", "ご褒美…♡", "気持ちよくしてあげる♡"],
                "ownership":  ["もう離さない…♡♡", "永遠に私の…♡♡"],
            },
        },
        "incest": {
            "stages": ["taboo", "wavering", "crossing", "immersion", "normalization"],
            "speech": {
                "taboo":         ["だめ…家族なのに", "お兄ちゃん…やめて", "いけないこと"],
                "wavering":      ["…だめ…だけど", "家族…なのに…こんな", "嫌…じゃないけど"],
                "crossing":      ["もう…いい…", "家族とか…もう", "…お兄ちゃん…♡"],
                "immersion":     ["お兄ちゃん…もっと♡", "気持ちいい♡", "大好き…♡"],
                "normalization": ["毎晩…来て♡", "お兄ちゃんだけ♡♡", "秘密だよ♡♡"],
            },
            "thought": {
                "taboo":         ["家族なのに…こんなこと", "いけない…", "バレたら"],
                "wavering":      ["だめ…でも…嫌じゃない", "おかしい…家族なのに"],
                "crossing":      ["もう…超えちゃった", "戻れない…", "でも…嬉しい"],
                "immersion":     ["お兄ちゃん…好き…♡", "家族とか…もうどうでも♡"],
                "normalization": ["これが…普通…♡", "もう…離れられない♡"],
            },
        },
    }
    # デフォルト（テーマ別定義がないテーマ用フォールバック）
    _PSYCHOLOGICAL_STAGES = ["resistance", "confusion", "acceptance", "desire", "abandon"]
    _STAGE_SPEECH_PATTERNS = {
        "resistance": ["やめて", "嫌", "離して", "来ないで", "やだ", "助けて", "痛い", "怖い"],
        "confusion":  ["なんで", "わからない", "どうして", "嘘", "おかしい", "信じられない", "混乱"],
        "acceptance": ["…仕方ない", "…もういい", "わかった", "好きに", "…ん…", "いいよ"],
        "desire":     ["もっと", "欲しい", "お願い", "ちょうだい", "止めないで", "気持ちいい"],
        "abandon":    ["壊れ", "なんでもいい", "全部", "おかしく", "どうでも", "♡♡♡", "もう…だめ"],
    }
    _STAGE_THOUGHT_PATTERNS = {
        "resistance": ["逃げ", "嫌だ", "怖い", "助けて", "無理"],
        "confusion":  ["なんで", "わからない", "おかしい", "どうして"],
        "acceptance": ["仕方ない", "受け入れ", "もういい", "諦め"],
        "desire":     ["欲しい", "もっと", "気持ちいい", "♡"],
        "abandon":    ["壊れ", "何も考え", "真っ白", "溶け", "♡♡"],
    }

    # テーマ名→_THEME_PSYCH_STAGES キーのマッピング
    _THEME_PSYCH_KEY_MAP = {
        "netorare": "netorare", "ntr": "netorare", "寝取": "netorare",
        "love": "love", "vanilla": "love", "純愛": "love", "ラブ": "love", "和姦": "love",
        "forced": "forced", "強制": "forced", "無理やり": "forced",
        "corruption": "corruption", "堕ち": "corruption", "調教": "corruption",
        "hypnosis": "hypnosis", "催眠": "hypnosis", "洗脳": "hypnosis",
        "chikan": "chikan", "痴漢": "chikan",
        "humiliation": "humiliation", "凌辱": "humiliation", "屈辱": "humiliation",
        "time_stop": "time_stop", "時間停止": "time_stop",
        "teacher_student": "teacher_student", "先生": "teacher_student",
        "femdom": "femdom", "女性優位": "femdom", "痴女": "femdom",
        "incest": "incest", "近親": "incest",
    }

    def _resolve_theme_psych(theme_str: str) -> tuple:
        """テーマ→(stages_list, speech_patterns, thought_patterns) を解決"""
        _ts = (theme_str or "").lower()
        for _kw, _key in _THEME_PSYCH_KEY_MAP.items():
            if _kw in _ts and _key in _THEME_PSYCH_STAGES:
                _td = _THEME_PSYCH_STAGES[_key]
                return (_td["stages"],
                        _td.get("speech", {}),
                        _td.get("thought", {}))
        return (_PSYCHOLOGICAL_STAGES, _STAGE_SPEECH_PATTERNS, _STAGE_THOUGHT_PATTERNS)

    def _infer_psychological_stage(scene_idx: int, intensity: int, total: int, theme_str: str) -> str:
        """シーン位置/intensity/テーマ→心理段階推定（テーマ別5段階対応）"""
        stages, _, _ = _resolve_theme_psych(theme_str)
        ratio = scene_idx / max(total, 1)
        # 5段階を ratio + intensity で推定
        # 段階0=序盤低i / 段階1=序盤中i / 段階2=中盤 / 段階3=中盤後半高i / 段階4=終盤高i
        if intensity <= 1:
            return stages[0]
        elif intensity == 2:
            return stages[1] if ratio < 0.5 else stages[2]
        elif intensity == 3:
            return stages[2] if ratio < 0.6 else stages[3]
        elif intensity == 4:
            return stages[3]
        else:
            return stages[4]

    def _is_stage_mismatch(text: str, stage: str, patterns_dict: dict, stages_list: list) -> bool:
        """テキストが現在の心理段階から2段階以上離れたキーワードを含むか"""
        stage_idx = stages_list.index(stage) if stage in stages_list else 2
        for other_stage, keywords in patterns_dict.items():
            other_idx = stages_list.index(other_stage) if other_stage in stages_list else 2
            if abs(stage_idx - other_idx) >= 2:
                if any(kw in text for kw in keywords):
                    return True
        return False

    _stage_fix_count = 0
    if _has_pool:
        _total_s_psy = len(results)
        # テーマ別心理パターンを解決（ループ外で1回だけ）
        _psy_stages, _psy_speech, _psy_thought = _resolve_theme_psych(theme)
        for _si_psy, scene in enumerate(results):
            intensity = scene.get("intensity", 3)
            _stage = _infer_psychological_stage(_si_psy, intensity, _total_s_psy, theme)
            for b in scene.get("bubbles", []):
                btype = b.get("type", "speech")
                txt = b.get("text", "")
                if not txt:
                    continue
                if btype == "speech" and _psy_speech:
                    # i=5のspeechはStep 20bで1文字に崩壊するため、ここでの置換は無駄→スキップ
                    if intensity >= 5:
                        continue
                    if _is_stage_mismatch(txt, _stage, _psy_speech, _psy_stages):
                        pool = _get_speech_pool_with_char("speech", theme, intensity, _si_psy, _total_s_psy)
                        # 現在段階のキーワードを含むセリフを優先
                        stage_kw = _psy_speech.get(_stage, [])
                        _stage_pool = [p for p in pool if any(kw in p for kw in stage_kw)]
                        target_pool = _stage_pool if _stage_pool else pool
                        repl = pick_replacement(target_pool, _used_speech_for_fix, _normalize_bubble_text,
                                               intensity=intensity)
                        if repl:
                            log_message(f"  心理遷移修正: シーン{_si_psy+1}({_stage}) speech")
                            b["text"] = repl
                            _used_speech_for_fix.add(repl)
                            _stage_fix_count += 1
                elif btype == "thought" and _psy_thought:
                    if _is_stage_mismatch(txt, _stage, _psy_thought, _psy_stages):
                        pool = _get_speech_pool_with_char("thought", theme, intensity, _si_psy, _total_s_psy)
                        stage_kw = _psy_thought.get(_stage, [])
                        _stage_pool = [p for p in pool if any(kw in p for kw in stage_kw)]
                        target_pool = _stage_pool if _stage_pool else pool
                        repl = pick_replacement(target_pool, _used_speech_for_fix, _normalize_bubble_text,
                                               intensity=intensity)
                        if repl:
                            log_message(f"  心理遷移修正: シーン{_si_psy+1}({_stage}) thought")
                            b["text"] = repl
                            _used_speech_for_fix.add(repl)
                            _stage_fix_count += 1
    if _stage_fix_count > 0:
        log_message(f"  心理遷移修正: {_stage_fix_count}件")

    # 10d. N-gram語彙多様性修正（4文字N-gramが5回超出現→3回目以降をプール代替）
    _ngram_fix_count = 0
    if _has_pool:
        from collections import Counter as _NgramCounter
        # 全バブルからN-gram頻度集計
        _ngram_positions = {}  # ngram -> [(scene_idx, bubble_idx, start_pos)]
        for _si_ng, scene in enumerate(results):
            for _bi_ng, b in enumerate(scene.get("bubbles", [])):
                txt = b.get("text", "")
                if len(txt) >= 4:
                    for _ng_s in range(len(txt) - 3):
                        ng = txt[_ng_s:_ng_s + 4]
                        _ngram_positions.setdefault(ng, []).append((_si_ng, _bi_ng, _ng_s))
        # 5回超のN-gramを持つバブルの3回目以降を置換
        _heavy_ngrams = {ng: positions for ng, positions in _ngram_positions.items()
                         if len(positions) > 5}
        _replaced_bubbles = set()  # (scene_idx, bubble_idx)
        for ng, positions in _heavy_ngrams.items():
            for _occur_idx, (si, bi, _) in enumerate(positions):
                if _occur_idx < 2:
                    continue  # 最初の2回はそのまま
                if (si, bi) in _replaced_bubbles:
                    continue
                scene = results[si]
                b = scene.get("bubbles", [])[bi]
                btype = b.get("type", "speech")
                intensity = scene.get("intensity", 3)
                _total_s_ng = len(results)
                if btype == "moan":
                    pool = _get_moan_pool_with_char(intensity)
                    repl = pick_replacement(pool, _used_moan_for_fix, _normalize_bubble_text,
                                           intensity=intensity)
                else:
                    pool = _get_speech_pool_with_char(btype, theme, intensity, si, _total_s_ng)
                    repl = pick_replacement(pool, _used_speech_for_fix, _normalize_bubble_text,
                                           intensity=intensity)
                if repl:
                    log_message(f"  N-gram反復修正({ng}): シーン{si+1}「{b['text'][:15]}…」→「{repl}」")
                    b["text"] = repl
                    _used_speech_for_fix.add(repl)
                    _replaced_bubbles.add((si, bi))
                    _ngram_fix_count += 1
    if _ngram_fix_count > 0:
        log_message(f"  N-gram反復修正: {_ngram_fix_count}件")

    # 10e. バブル順序ローテーション（同一first-bubble type 3連続防止）
    _BUBBLE_TYPE_ORDER = ["moan", "thought", "speech"]
    _prev_first_type = None
    _consecutive_first = 0
    _bubble_rotate_count = 0
    for scene in results:
        bubbles = scene.get("bubbles", [])
        if not bubbles or len(bubbles) < 2:
            _prev_first_type = None
            _consecutive_first = 0
            continue
        first_type = bubbles[0].get("type", "speech")
        if first_type == _prev_first_type:
            _consecutive_first += 1
        else:
            _consecutive_first = 1
            _prev_first_type = first_type
        if _consecutive_first >= 3:
            intensity = scene.get("intensity", 3)
            # intensity≤2のシーンはmoan-firstにしない
            # 次のタイプを決定
            try:
                cur_idx = _BUBBLE_TYPE_ORDER.index(first_type)
            except ValueError:
                cur_idx = 0
            next_type = _BUBBLE_TYPE_ORDER[(cur_idx + 1) % len(_BUBBLE_TYPE_ORDER)]
            # intensity≤2ではmoan-firstを回避
            if intensity <= 2 and next_type == "moan":
                next_type = _BUBBLE_TYPE_ORDER[(cur_idx + 2) % len(_BUBBLE_TYPE_ORDER)]
            # 該当typeのバブルを先頭に移動
            target_idx = None
            for bi, b in enumerate(bubbles):
                if b.get("type") == next_type and bi > 0:
                    target_idx = bi
                    break
            if target_idx is not None:
                moved = bubbles.pop(target_idx)
                bubbles.insert(0, moved)
                scene["bubbles"] = bubbles
                _prev_first_type = next_type
                _consecutive_first = 1
                _bubble_rotate_count += 1
    if _bubble_rotate_count > 0:
        log_message(f"  バブル順序ローテーション: {_bubble_rotate_count}件（3連続同一first防止）")

    # 11. story_flow重複修正（同一テキストの2回目以降を空にする）
    _seen_flows = {}
    _flow_fix_count = 0
    for scene in results:
        flow = scene.get("story_flow", "")
        if flow and len(flow) >= 10:
            sid = scene.get("scene_id", "?")
            if flow in _seen_flows:
                scene["story_flow"] = ""
                _flow_fix_count += 1
                log_message(f"  S{sid}: story_flow重複削除（S{_seen_flows[flow]}と同一）")
            else:
                _seen_flows[flow] = sid
    if _flow_fix_count > 0:
        log_message(f"  story_flow重複修正: {_flow_fix_count}件")

    # 11b. キャラ名途切れ修復（フルネーム直後に助詞がない場合、姓のみに置換）
    # Haiku 3でフルネーム後のトークン生成が不安定になり助詞+動詞が欠落する問題への対策
    _VALID_AFTER_NAME = set("がをのはにとでもへやより、。）)」』】")
    _name_trunc_count = 0
    if correct_names:
        # フルネーム→短縮名マップを構築
        _name_short_map = {}  # full_name -> short_name
        for full_name in correct_names:
            if not full_name or len(full_name) < 2:
                continue
            # 最初の「・」「 」「＝」「　」の前を姓とする
            short = full_name
            for sep in ["・", " ", "＝", "　"]:
                idx = full_name.find(sep)
                if idx > 0:
                    short = full_name[:idx]
                    break
            if short != full_name:
                _name_short_map[full_name] = short

        if _name_short_map:
            _first_occurrence_done = {}  # full_name -> bool (シーン1でフルネーム初出済みか)
            for i, scene in enumerate(results):
                sid = scene.get("scene_id", i + 1)
                desc = scene.get("description", "")
                if not desc:
                    continue
                for full_name, short_name in _name_short_map.items():
                    if full_name not in desc:
                        continue
                    # シーン1（最初の出現）はフルネームを保持
                    if full_name not in _first_occurrence_done:
                        _first_occurrence_done[full_name] = True
                        # ただしシーン1でも途切れ箇所はチェック
                        # 最初の出現はスキップし、2回目以降のみ修復
                        positions = []
                        start = 0
                        while True:
                            pos = desc.find(full_name, start)
                            if pos < 0:
                                break
                            positions.append(pos)
                            start = pos + len(full_name)
                        if len(positions) <= 1:
                            continue
                        # 2回目以降の出現のみ処理（逆順で置換）
                        new_desc = desc
                        for pos in reversed(positions[1:]):
                            after_pos = pos + len(full_name)
                            if after_pos < len(new_desc):
                                after_char = new_desc[after_pos]
                                if after_char not in _VALID_AFTER_NAME:
                                    new_desc = new_desc[:pos] + short_name + new_desc[after_pos:]
                                    _name_trunc_count += 1
                            else:
                                # 文末にフルネーム→短縮名に
                                new_desc = new_desc[:pos] + short_name + new_desc[after_pos:]
                                _name_trunc_count += 1
                        scene["description"] = new_desc
                    else:
                        # シーン2以降: 全出現を検査
                        positions = []
                        start = 0
                        while True:
                            pos = desc.find(full_name, start)
                            if pos < 0:
                                break
                            positions.append(pos)
                            start = pos + len(full_name)
                        if not positions:
                            continue
                        new_desc = desc
                        for pos in reversed(positions):
                            after_pos = pos + len(full_name)
                            if after_pos < len(new_desc):
                                after_char = new_desc[after_pos]
                                if after_char not in _VALID_AFTER_NAME:
                                    new_desc = new_desc[:pos] + short_name + new_desc[after_pos:]
                                    _name_trunc_count += 1
                            else:
                                new_desc = new_desc[:pos] + short_name + new_desc[after_pos:]
                                _name_trunc_count += 1
                        scene["description"] = new_desc
                        desc = new_desc  # 更新後の値で次のキャラ名をチェック
    if _name_trunc_count > 0:
        log_message(f"  キャラ名途切れ修復: {_name_trunc_count}件（フルネーム→姓に置換）")

    _progress("Step 12-20 description/title/感情修正")
    # 12. description先頭15字重複修正（全既出シーンと比較、最初の句点後に状況挿入）
    # v8.2根本修正: 30字→15字に短縮（「地方出張先のビジネスホテルの一室。」vs「一室、」の差を検出）
    # 方針: 「場所。状況描写...」の「。」の後にvariation文を挿入して先頭を変化させる
    _INTENSITY_DESC_INSERTS = {
        1: [
            "不穏な空気が漂う中、", "緊張感が張り詰める中、", "嫌な予感を覚えながら、",
            "周囲を警戒しつつ、", "息を殺して様子を窺いながら、", "心の準備ができないまま、",
            "逃げ場のない状況で、", "背筋に冷たいものが走る中、",
        ],
        2: [
            "恥ずかしさで体が強張る中、", "心臓の鼓動が速まる中、", "唇を噛みしめながら、",
            "触れられた箇所が熱を持ち始め、", "頬が紅潮していくのを感じながら、",
            "視線を逸らしつつも意識が集中し、", "手足が小刻みに震える中、",
            "初めての感覚に体が跳ねる中、",
        ],
        3: [
            "快感に抗いきれなくなる中、", "甘い痺れが全身に広がり、", "抵抗の力が弱まっていく中、",
            "息が荒くなりながら、", "肌が敏感になっていくのを感じ、", "声を抑えきれなくなりながら、",
            "腰が勝手に動いてしまう中、", "意識が快楽に染まり始める中、",
        ],
        4: [
            "快楽に支配されつつある中、", "もう逃れられないと悟りながら、", "理性が揺らぎ始める中、",
            "全身が敏感に反応する中、", "体の芯から熱が溢れ出す中、", "抵抗の意志が溶けていく中、",
            "自分の声が止められなくなり、", "全身の力が抜けていく中、",
        ],
        5: [
            "理性が完全に崩壊した状態で、", "快楽の波に全身が呑まれ、", "もう何も考えられなくなり、",
            "絶頂の余韻が全身を支配する中、", "白い光に視界が塗りつぶされる中、",
            "意識が飛びそうになりながら、", "体が痙攣を繰り返す中、",
            "自分が誰かも分からなくなり、",
        ],
    }
    _desc_fix_count = 0
    _DESC_PREFIX_LEN = 15  # v8.2: 30→15字に短縮（場所名の微差を検出）
    _seen_desc_prefixes = {}  # prefix -> first scene_id
    for i, scene in enumerate(results):
        desc = scene.get("description", "")
        if not desc or len(desc) < _DESC_PREFIX_LEN:
            sid = scene.get("scene_id", i + 1)
            if desc:
                _seen_desc_prefixes[desc[:_DESC_PREFIX_LEN]] = sid
            continue
        prefix30 = desc[:_DESC_PREFIX_LEN]
        sid = scene.get("scene_id", i + 1)
        if prefix30 in _seen_desc_prefixes:
            intensity = scene.get("intensity", 3)
            inserts = _INTENSITY_DESC_INSERTS.get(intensity, _INTENSITY_DESC_INSERTS[3])
            # 最初の句点を探して挿入位置を決定
            insert_pos = desc.find("。")
            if insert_pos >= 0 and insert_pos < len(desc) - 1:
                insert_pos += 1  # 「。」の直後
            else:
                # 句点なし → 「、」の後に挿入
                insert_pos = desc.find("、")
                if insert_pos >= 0 and insert_pos < len(desc) - 1:
                    insert_pos += 1
                else:
                    insert_pos = 0  # フォールバック: 先頭
            # 未使用のバリエーションを選択（自intensity→隣接intensity→全intensityの順で探索）
            chosen_insert = None
            # 1) 自intensityの全バリエーション
            for try_idx in range(_desc_fix_count, _desc_fix_count + len(inserts)):
                candidate = inserts[try_idx % len(inserts)]
                new_desc = desc[:insert_pos] + candidate + desc[insert_pos:]
                if new_desc[:_DESC_PREFIX_LEN] not in _seen_desc_prefixes:
                    chosen_insert = candidate
                    break
            # 2) 隣接intensity（±1）のバリエーションも試す
            if chosen_insert is None:
                for adj_i in [max(1, intensity - 1), min(5, intensity + 1)]:
                    if adj_i == intensity:
                        continue
                    adj_inserts = _INTENSITY_DESC_INSERTS.get(adj_i, [])
                    for candidate in adj_inserts:
                        new_desc = desc[:insert_pos] + candidate + desc[insert_pos:]
                        if new_desc[:_DESC_PREFIX_LEN] not in _seen_desc_prefixes:
                            chosen_insert = candidate
                            break
                    if chosen_insert:
                        break
            # 3) v8.8修正: 全intensity横断は±2まで制限（i=1のプレフィックスがi=4に入る問題を防止）
            if chosen_insert is None:
                for any_i in [max(1, intensity - 2), min(5, intensity + 2)]:
                    if any_i == intensity or abs(any_i - intensity) <= 1:
                        continue  # ±1は既に試行済み
                    for candidate in _INTENSITY_DESC_INSERTS.get(any_i, []):
                        new_desc = desc[:insert_pos] + candidate + desc[insert_pos:]
                        if new_desc[:_DESC_PREFIX_LEN] not in _seen_desc_prefixes:
                            chosen_insert = candidate
                            break
                    if chosen_insert:
                        break
            # 4) それでも見つからない → 挿入位置を先頭に変更して自intensity±1のみで再試行
            if chosen_insert is None:
                insert_pos = 0
                for any_i in [intensity, max(1, intensity - 1), min(5, intensity + 1)]:
                    for candidate in _INTENSITY_DESC_INSERTS.get(any_i, []):
                        new_desc = candidate + desc
                        if new_desc[:_DESC_PREFIX_LEN] not in _seen_desc_prefixes:
                            chosen_insert = candidate
                            break
                    if chosen_insert:
                        break
            if chosen_insert is None:
                # 最終フォールバック: 全て枯渇 → シーン番号入りで一意性保証
                chosen_insert = f"シーン{sid}の場面では、"
                insert_pos = 0
            new_desc = desc[:insert_pos] + chosen_insert + desc[insert_pos:]
            scene["description"] = new_desc
            _desc_fix_count += 1
            log_message(f"  S{sid}: description重複修正（S{_seen_desc_prefixes[prefix30]}と一致、挿入: {chosen_insert[:15]}...）")
            # 修正後のprefixも登録（二次重複防止）
            new_prefix30 = new_desc[:_DESC_PREFIX_LEN]
            if new_prefix30 not in _seen_desc_prefixes:
                _seen_desc_prefixes[new_prefix30] = sid
        else:
            _seen_desc_prefixes[prefix30] = sid
    if _desc_fix_count > 0:
        log_message(f"  description重複修正: {_desc_fix_count}件")

    # 12a2. description先頭10字prefix二次チェック（"シャワー室の濡れた床で" 等の短い重複をキャッチ）
    _DESC_PREFIX_LEN_SHORT = 10
    _seen_short_prefix = {}  # prefix10 -> [scene_indices]
    for i, scene in enumerate(results):
        desc = scene.get("description", "")
        if not desc or len(desc) < _DESC_PREFIX_LEN_SHORT:
            continue
        short_p = desc[:_DESC_PREFIX_LEN_SHORT]
        _seen_short_prefix.setdefault(short_p, []).append(i)
    _desc_short_fix = 0
    for short_p, indices in _seen_short_prefix.items():
        if len(indices) < 3:
            continue
        # 3回目以降にバリエーション挿入
        for dup_idx in indices[2:]:
            scene = results[dup_idx]
            desc = scene.get("description", "")
            intensity = scene.get("intensity", 3)
            sid = scene.get("scene_id", dup_idx + 1)
            inserts = _INTENSITY_DESC_INSERTS.get(intensity, _INTENSITY_DESC_INSERTS[3])
            candidate = inserts[(_desc_short_fix + dup_idx) % len(inserts)]
            insert_pos = desc.find("。")
            if insert_pos >= 0 and insert_pos < len(desc) - 1:
                insert_pos += 1
            else:
                insert_pos = 0
            new_desc = desc[:insert_pos] + candidate + desc[insert_pos:]
            # 10字prefixが変わったか確認
            if new_desc[:_DESC_PREFIX_LEN_SHORT] != short_p:
                scene["description"] = new_desc
                _desc_short_fix += 1
            else:
                # 先頭挿入で確実に変える
                new_desc = candidate + desc
                scene["description"] = new_desc
                _desc_short_fix += 1
    if _desc_short_fix > 0:
        log_message(f"  description短prefix重複修正: {_desc_short_fix}件（10字prefix 3回以上）")

    # 12b. mood重複修正（同一moodの3回目以降をintensity別バリエーションで置換）
    # v8.9: 6→15個に拡充（100シーンで枯渇防止）
    # テーマ別mood拡張バリアント（テーマ固有の雰囲気をintensity別に）
    _THEME_MOOD_VARIANTS = {
        "netorare": {
            1: ["密かに芽生える禁断の予感", "日常の裏に潜む裏切りの気配"],
            2: ["彼氏の影がちらつく罪悪感", "背徳の甘さに手が伸びる瞬間"],
            3: ["比較が止まらない背徳の沼", "罪悪感を快楽が塗り潰していく"],
            4: ["もう彼の顔を思い出せない堕落", "裏切りの快楽に溺れる暗い恍惚"],
            5: ["完全に堕ちた背徳の極致", "彼氏より深い快楽に染まった証"],
        },
        "love": {
            1: ["二人だけの穏やかな時間の始まり", "心が近づく温かな予感"],
            2: ["恥じらいと信頼が交差する甘い空気", "触れ合うたび深まる想い"],
            3: ["愛する人に身を委ねる幸福な恥じらい", "信頼が肌を通して伝わる温もり"],
            4: ["二つの身体が一つに溶け合う恍惚", "愛情が快楽となって全身を巡る"],
            5: ["魂まで一つになる至福の絶頂", "愛し合う二人だけの永遠の瞬間"],
        },
        "forced": {
            1: ["逃げ場のない恐怖が忍び寄る", "日常が壊れる不吉な予兆"],
            2: ["抵抗する心と従う身体の乖離", "暴力的な手に震える無力感"],
            3: ["屈服しかけた心に残る最後の抵抗", "身体の裏切りに絶望する恥辱"],
            4: ["壊された理性の残骸で感じる快楽", "もう抗えない絶望的な恍惚"],
            5: ["完全に支配された魂の叫び", "壊されて初めて知った快楽の深淵"],
        },
        "corruption": {
            1: ["無垢な日常に忍び込む淫靡な影", "知らない感覚への戸惑い"],
            2: ["好奇心が理性を侵食し始める瞬間", "禁じられた快楽への無自覚な渇望"],
            3: ["堕ちていく自分を止められない背徳感", "快楽を知った身体が求め始める"],
            4: ["もう元には戻れない淫蕩の深み", "堕落を受け入れた解放感と快楽"],
            5: ["完全に書き換わった快楽の価値観", "堕ちきった先に見えた暗い至福"],
        },
        "humiliation": {
            1: ["プライドが試される不穏な空気", "屈辱の予感に身が固まる"],
            2: ["晒される恥辱と消えない自尊心", "見下される視線に震える怒りと羞恥"],
            3: ["誇りが砕ける音を聞きながら感じる快楽", "恥辱の中に芽生える異常な悦び"],
            4: ["プライドの残骸の上で喘ぐ堕落", "屈辱が快楽に変わった暗い恍惚"],
            5: ["全てを差し出して跪く恍惚の底", "屈辱に酔いしれる壊れた心"],
        },
        "chikan": {
            1: ["満員電車の中の異常な気配", "逃げられない空間の圧迫感"],
            2: ["人混みの中で密かに犯される羞恥", "声を出せない状況の絶望的な興奮"],
            3: ["周囲にバレる恐怖と止まらない快感", "公衆の面前で身体が裏切る恥辱"],
            4: ["バレてもいいと思い始める堕落", "人目の中で果てる背徳の絶頂"],
            5: ["衆人環視の中で完全に堕ちた恍惚", "もう隠す気もない公開の悦楽"],
        },
    }
    # テーマキーワード→_THEME_MOOD_VARIANTSキーのマッピング
    _THEME_MOOD_KEY_MAP = {
        "ntr": "netorare", "netorare": "netorare", "寝取": "netorare",
        "love": "love", "vanilla": "love", "純愛": "love", "ラブ": "love",
        "forced": "forced", "強制": "forced", "無理やり": "forced", "レイプ": "forced",
        "corruption": "corruption", "堕ち": "corruption", "調教": "corruption",
        "humiliation": "humiliation", "凌辱": "humiliation", "屈辱": "humiliation",
        "chikan": "chikan", "痴漢": "chikan",
    }
    _MOOD_VARIANTS = {
        1: ["静かな緊張感", "不安と期待が入り混じる空気", "甘い予感が漂う空間",
            "戸惑いと好奇心の狭間", "穏やかだが張りつめた沈黙", "秘めた欲望が滲む雰囲気",
            "微かな胸騒ぎ", "言葉にならない不安の気配", "日常が揺らぐ予兆",
            "纏わりつくような沈黙", "探るような視線の応酬", "薄氷を踏むような空気感",
            "静かに忍び寄る危険の匂い", "何かが始まる直前の静寂", "背筋を伝う冷たい予感"],
        2: ["高まる鼓動と熱気", "抗えない引力に満ちた空気", "肌が触れ合う甘い緊張",
            "理性と欲望がせめぎ合う空間", "息遣いが重なる距離感", "抑えきれない衝動の予感",
            "じわりと滲む背徳の熱", "体温が上がる距離の近さ", "呼吸が乱れ始める瞬間",
            "抗いたいのに逆らえない引力", "思考が鈍くなる甘い空気", "肌が粟立つ接近",
            "逃げ場のない甘い緊張", "拒否と受容の狭間で揺れる心", "指先から伝わる危険な熱"],
        3: ["快楽に溺れる密室", "熱く絡み合う情欲の渦", "理性が崩れていく甘い地獄",
            "汗ばむ肌と乱れる吐息", "止められない快感の連鎖", "貪り合う獣のような熱気",
            "抵抗を忘れる甘美な痺れ", "溶けていく自制心", "肌を這う快楽の余韻",
            "抗うことを諦めた解放感", "熱に浮かされた朦朧とした空気", "羞恥と快楽が混ざり合う瞬間",
            "身体が正直に反応する恥じらい", "言葉にならない喘ぎが漏れる空間", "理性の最後の砦が揺らぐ"],
        4: ["絶頂へ駆け上がる狂熱", "壊れそうなほどの快楽の嵐", "獣じみた情欲が支配する空間",
            "限界を超えた快感の波状攻撃", "理性が完全に溶けた淫靡な世界", "果てしない絶頂の連鎖",
            "思考を奪う圧倒的な快楽", "痙攣が止まらない限界の淵", "獣のように貪り合う激情",
            "全身が快楽に染まった恍惚", "声を上げることしかできない支配", "何度も押し寄せる絶頂の波",
            "自分が誰かも忘れるほどの快楽", "骨の髄まで響く激しい律動", "涙が滲むほどの快感と屈辱"],
        5: ["全てを焼き尽くす最高潮", "意識が飛ぶほどの究極の快楽", "魂ごと蕩ける至福の瞬間",
            "壮絶な絶頂が全身を貫く", "白く染まる意識の果て", "限界を遥かに超えた恍惚",
            "魂が抜けるような壮絶な絶頂", "視界が真っ白に染まる瞬間", "全身の感覚が一点に集約される",
            "崩壊と再生を繰り返す果てしない快楽", "人格が書き換わるほどの衝撃", "もう戻れないと悟る至福",
            "全細胞が悲鳴を上げる絶頂", "存在ごと溶かされる快楽の渦", "永遠に続くかのような恍惚"],
    }
    _mood_fix_count = 0
    _mood_seen_count = {}  # mood_text -> occurrence_count
    _mood_used_variants = set()
    # テーマ別mood辞書を解決（ループ外で1回）
    _resolved_theme_mood = {}
    _tl = (theme or "").lower()
    for _tmk, _tmv in _THEME_MOOD_KEY_MAP.items():
        if _tmk in _tl and _tmv in _THEME_MOOD_VARIANTS:
            _resolved_theme_mood = _THEME_MOOD_VARIANTS[_tmv]
            break
    for scene in results:
        m = scene.get("mood", "")
        if not m:
            continue
        _mood_seen_count[m] = _mood_seen_count.get(m, 0) + 1
        if _mood_seen_count[m] >= 3:  # 3回目以降を置換
            intensity = scene.get("intensity", 3)
            chosen = None
            # テーマ別バリアントを優先
            if _resolved_theme_mood:
                for v in _resolved_theme_mood.get(intensity, []):
                    if v not in _mood_used_variants and v != m:
                        chosen = v
                        break
            # テーマ別で見つからなければ汎用バリアント
            if chosen is None:
                variants = _MOOD_VARIANTS.get(intensity, _MOOD_VARIANTS[3])
                for v in variants:
                    if v not in _mood_used_variants and v != m:
                        chosen = v
                        break
            if chosen is None:
                # 隣接intensityからも探索
                for adj_i in [max(1, intensity - 1), min(5, intensity + 1)]:
                    _pool = _resolved_theme_mood.get(adj_i, []) if _resolved_theme_mood else []
                    _pool = _pool + _MOOD_VARIANTS.get(adj_i, [])
                    for v in _pool:
                        if v not in _mood_used_variants:
                            chosen = v
                            break
                    if chosen:
                        break
            if chosen:
                scene["mood"] = chosen
                _mood_used_variants.add(chosen)
                _mood_fix_count += 1
    if _mood_fix_count > 0:
        log_message(f"  mood重複修正: {_mood_fix_count}件")

    # 12c. v8.8: mood品質チェック（テーマkey_emotionsがそのままmoodに使われている→置換）
    _theme_guide = THEME_GUIDES.get(theme, {})
    _theme_key_emotions = set(_theme_guide.get("key_emotions", []))
    _mood_quality_fix = 0
    if _theme_key_emotions:
        for scene in results:
            m = scene.get("mood", "")
            if m and m in _theme_key_emotions:
                intensity = scene.get("intensity", 3)
                # テーマ別バリアント優先
                _qpool = (_resolved_theme_mood.get(intensity, []) if _resolved_theme_mood else []) + \
                         _MOOD_VARIANTS.get(intensity, _MOOD_VARIANTS[3])
                for v in _qpool:
                    if v not in _mood_used_variants:
                        scene["mood"] = v
                        _mood_used_variants.add(v)
                        _mood_quality_fix += 1
                        break
    if _mood_quality_fix > 0:
        log_message(f"  mood品質修正（テーマ感情→具体mood）: {_mood_quality_fix}件")

    # 12d. v8.8: テーマ別bubble制約（time_stop等、テーマ世界ルールに基づくセリフ型変換）
    _bubble_theme_fix = 0
    if theme == "time_stop":
        # 時間停止中（intensity≥3かつ停止中シーン）: speech→thought変換、moan→thought変換
        _TIME_STOP_ACTIVE_KW = ["時間停止", "身動き", "止まった", "停止した", "動けない", "frozen"]
        _TIME_STOP_RELEASED_KW = ["時間再開", "再び動き", "解除", "現実に戻", "混乱して"]
        for scene in results:
            desc = scene.get("description", "")
            is_frozen = any(kw in desc for kw in _TIME_STOP_ACTIVE_KW)
            is_released = any(kw in desc for kw in _TIME_STOP_RELEASED_KW)
            if not is_frozen or is_released:
                continue
            bubbles = scene.get("bubbles", [])
            for b in bubbles:
                speaker = b.get("speaker", "")
                btype = b.get("type", "")
                if _is_male_speaker(speaker):
                    continue  # 男性（時間停止の使い手）はspeechのまま
                if btype == "speech":
                    b["type"] = "thought"
                    _bubble_theme_fix += 1
                elif btype == "moan":
                    # moanをthoughtに変換（声が出せないため）
                    text = b.get("text", "")
                    clean = text.replace("♡", "").replace("♥", "").strip()
                    if clean:
                        b["type"] = "thought"
                        # 喘ぎの長さで変換パターンを分岐
                        if len(clean) <= 3:
                            b["text"] = f"{clean}…身体が…反応して…"
                        elif clean.endswith("っ") or clean.endswith("ぁ"):
                            b["text"] = f"{clean}…この…感覚…"
                        else:
                            b["text"] = f"{clean}…身体が…どうなって…"
                    _bubble_theme_fix += 1
    if _bubble_theme_fix > 0:
        log_message(f"  テーマ別bubble制約修正（time_stop）: {_bubble_theme_fix}件")

    # 12e. v8.9: 時間軸ジャンプ修正（description/story_flowの「翌週」等を同日表現に置換）
    # エピローグ（最終10%）は時間ジャンプを許可
    # テーマ別: few_days/flexibleなら「翌日」「翌朝」は許可
    _epilogue_start_12e = max(1, len(results) - max(1, len(results) // 10))
    _time_span = _THEME_TIME_SPAN.get(theme, "flexible")
    # single_event: 翌日含めて全禁止 / few_days/flexible: 翌週以上のみ禁止
    _TIME_JUMP_REPLACEMENTS_STRICT = {
        # 「の」付き → 「その直後の」に置換（空文字ではなく文法を保持）
        "翌週の": "その直後の", "翌々週の": "その直後の", "翌月の": "その直後の",
        "数日後の": "その直後の", "一週間後の": "その直後の", "数週間後の": "その直後の",
        "1週間後の": "その直後の", "２週間後の": "その直後の", "次の週の": "その直後の",
        "来週の": "その直後の", "数ヶ月後の": "その直後の",
        # 「、」付き → 時間接続詞に置換
        "翌週、": "その後、", "翌々週、": "その後、",
        "数日後、": "しばらくして、", "一週間後、": "その後、",
        "数週間後、": "その後、", "翌月、": "その後、",
        "来週、": "その後、", "次の週、": "その後、",
        "1週間後、": "その後、", "２週間後、": "その後、",
        "数ヶ月後、": "その後、",
        # 「に」付き
        "翌週に": "その後", "数日後に": "しばらくして",
        "翌々週に": "その後", "一週間後に": "その後",
        "来週に": "その後", "次の週に": "その後",
        "数ヶ月後に": "その後",
        # 「後日」系
        "後日の": "その直後の", "後日、": "その後、", "後日に": "その後",
    }
    # few_days/flexible用: 「翌日」「翌朝」は除外（許可）
    _TIME_JUMP_REPLACEMENTS = dict(_TIME_JUMP_REPLACEMENTS_STRICT)
    if _time_span == "single_event":
        # single_eventでは「翌日」も禁止
        _TIME_JUMP_REPLACEMENTS["翌日の"] = "その直後の"
        _TIME_JUMP_REPLACEMENTS["翌日、"] = "その後、"
        _TIME_JUMP_REPLACEMENTS["翌日に"] = "その後"
        _TIME_JUMP_REPLACEMENTS["翌朝の"] = "その直後の"
        _TIME_JUMP_REPLACEMENTS["翌朝、"] = "その後、"
        _TIME_JUMP_REPLACEMENTS["翌朝に"] = "その後"
    _time_fix_count = 0
    for i, scene in enumerate(results):
        if i >= _epilogue_start_12e:
            continue  # エピローグは時間ジャンプ許可
        for field in ("description", "story_flow"):
            text = scene.get(field, "")
            if not text:
                continue
            new_text = text
            for old_kw, new_kw in _TIME_JUMP_REPLACEMENTS.items():
                if old_kw in new_text:
                    new_text = new_text.replace(old_kw, new_kw)
            if new_text != text:
                scene[field] = new_text
                _time_fix_count += 1
    if _time_fix_count > 0:
        log_message(f"  時間軸ジャンプ修正: {_time_fix_count}件（「翌週」等→同日表現、エピローグ除外）")

    # 12f. v8.9: メタ参照description修正（「シーンXXの場面では」等のAPI生成アーティファクト除去）
    import re as _re_12f
    _META_REF_PATTERN = _re_12f.compile(r'シーン\d+の?(?:場面では|では|のとき|において|シーンでは)[、。]?')
    _meta_fix_count = 0
    for scene in results:
        desc = scene.get("description", "")
        if desc and _META_REF_PATTERN.search(desc):
            new_desc = _META_REF_PATTERN.sub("", desc).lstrip("、。 ")
            if new_desc and len(new_desc) > 10:
                scene["description"] = new_desc
                _meta_fix_count += 1
    if _meta_fix_count > 0:
        log_message(f"  メタ参照description修正: {_meta_fix_count}件（「シーンXXの場面では」除去）")

    # 13. character_feelings重複修正（全既出シーンと比較、一致→intensity別テンプレートで差し替え）
    # テーマ別feelings拡張（テーマ固有の心理状態をintensity別に）
    _THEME_FEELINGS_VARIANTS = {
        "netorare": {
            2: ["彼氏を裏切っている罪悪感と、それでも止められない快感に揺れている",
                "他の男に触れられているのに身体が反応することに、自己嫌悪を感じている"],
            3: ["彼氏のことを考えるたびに罪悪感が胸を刺すが、快楽がそれを上回り始めている",
                "比較してしまう自分に気づき、もう取り返しがつかないと悟り始めている"],
            4: ["もう彼氏のことを考える余裕もなく、目の前の快楽に完全に支配されている",
                "裏切りの罪悪感すら快楽のスパイスに変わり、堕ちていく自分を止められない"],
            5: ["彼氏よりもこの快楽を選んでしまった自分を受け入れ、完全に堕ちている",
                "もう二度と彼氏の前に出られないほど深く堕ちたことに、背徳の悦びを感じている"],
        },
        "love": {
            2: ["好きな人に触れられる幸福感と恥ずかしさで心が一杯になっている",
                "愛されている実感が肌を通して伝わり、涙が出そうなほど嬉しい"],
            3: ["愛する人に身を委ねる幸福と、もっと近づきたい切なさに震えている",
                "信頼しているからこそ、こんな自分を見せられる安堵を感じている"],
            4: ["二人が一つに溶け合うような感覚に、愛情と快楽が区別できなくなっている",
                "この人に全てを捧げたいという衝動が、理性を超えて溢れ出している"],
            5: ["愛する人との一体感に心から満たされ、至福の涙を流している",
                "この瞬間が永遠に続けばいいと、心の底から願っている"],
        },
        "forced": {
            2: ["恐怖で身体が震えているが、声を出すこともできず固まっている",
                "逃げたいのに足が動かない。恐怖が全身を支配している"],
            3: ["嫌なはずなのに身体が感じてしまう。自分の身体が一番の裏切り者だ",
                "抵抗する気力が奪われていく。快楽が恐怖を上書きしていくのが怖い"],
            4: ["もう抗う意思が残っていない。快楽に壊された心が、それでも助けを求めている",
                "壊されていく自分を、どこか遠くから見ているような解離感を感じている"],
            5: ["完全に壊されて、快楽以外何も感じられなくなっている",
                "もう元の自分には戻れないことを悟り、暗い諦念に沈んでいる"],
        },
    }
    _FEELINGS_VARIANTS = {
        1: [
            "まだ状況を理解できず、困惑と不安を感じている",
            "何かが起きる予感に、体が硬直している",
            "突然の展開に戸惑い、どう反応していいか分からない",
            "不穏な空気を感じ取り、本能的に危険を察知している",
            "現実感がなく、夢の中にいるような錯覚を覚えている",
            "逃げたい気持ちと動けない恐怖が入り混じっている",
        ],
        2: [
            "体が反応し始めていることに戸惑い、羞恥に震えている",
            "触れられるたびに走る電流のような感覚に、抗えなくなっている",
            "恥ずかしさで顔が真っ赤になりながらも、意識が集中していく",
            "初めての感覚に戸惑いつつ、体が勝手に求めてしまう",
            "嫌だと思うのに体が言うことを聞かず、混乱している",
            "緊張と期待が入り混じる複雑な感情に揺れている",
        ],
        3: [
            "快感に抗いきれなくなり、自分の反応に罪悪感を覚えている",
            "嫌なはずなのに体が正直に反応してしまう自分に絶望している",
            "理性と本能の間で揺れ動き、心が引き裂かれそうになっている",
            "声を抑えようとしても漏れてしまう喘ぎに、羞恥を感じている",
            "快楽に流されまいと必死に意識を保とうとしている",
            "自分の体がこんなにも敏感だったことに驚き、戸惑っている",
        ],
        4: [
            "快楽に支配されつつも、最後の理性でかろうじて抵抗している",
            "抵抗する意志すら快感に塗り替えられていくのを感じている",
            "もう考えることすらできず、快楽の波に身を委ねている",
            "体の奥から湧き上がる衝動に、心が完全に呑まれそうになっている",
            "恥も外聞もなく声を上げてしまう自分を、遠くから見ている気分",
            "全身の感覚が研ぎ澄まされ、触れられる場所全てが快感に変わる",
        ],
        5: [
            "完全に快楽に溺れ、もう抵抗する気力すら失っている",
            "全身が痙攣し、思考も感情も快楽一色に染まっている",
            "意識が遠のきかけながらも、快楽だけが鮮明に感じられる",
            "自分が自分でなくなっていく感覚に、恐怖すら感じなくなっている",
            "何度目かも分からない絶頂に、体が壊れそうになっている",
            "もう何も考えられず、ただ快楽を受け入れることしかできない",
        ],
    }
    _feelings_fix_count = 0
    _seen_feelings = {}  # frozen feelings values string -> first scene_id
    for i, scene in enumerate(results):
        cf = scene.get("character_feelings", {})
        if not cf or not isinstance(cf, dict):
            continue
        sid = scene.get("scene_id", i + 1)
        # validate_scriptと同じロジック: values()のみで比較（キー名は無視）
        cf_key = str(sorted(cf.values()))
        if len(cf_key) < 15:
            continue
        if cf_key in _seen_feelings:
            intensity = scene.get("intensity", 3)
            # テーマ別feelings優先
            _theme_feel_key = None
            for _tfk, _tfv in _THEME_MOOD_KEY_MAP.items():
                if _tfk in _tl:
                    _theme_feel_key = _tfv
                    break
            _theme_feel_list = _THEME_FEELINGS_VARIANTS.get(_theme_feel_key, {}).get(intensity, []) if _theme_feel_key else []
            variants = _theme_feel_list + list(_FEELINGS_VARIANTS.get(intensity, _FEELINGS_VARIANTS[3]))
            # 全バリアントを試し、未使用のものを選択
            chosen = None
            for try_idx in range(_feelings_fix_count, _feelings_fix_count + len(variants)):
                candidate = variants[try_idx % len(variants)]
                candidate_key = str(sorted([candidate]))
                if candidate_key not in _seen_feelings:
                    chosen = candidate
                    break
            if chosen is None:
                # 全バリアント使用済み → シーン番号を付加してユニーク化
                base = variants[_feelings_fix_count % len(variants)]
                chosen = f"{base}（シーン{sid}）"
            for key in cf:
                cf[key] = chosen
                break
            _feelings_fix_count += 1
            log_message(f"  S{sid}: character_feelings重複修正（S{_seen_feelings[cf_key]}と同一）")
            # 更新後のキーで登録
            cf_key_new = str(sorted(cf.values()))
            if cf_key_new not in _seen_feelings:
                _seen_feelings[cf_key_new] = sid
        else:
            _seen_feelings[cf_key] = sid
    if _feelings_fix_count > 0:
        log_message(f"  character_feelings重複修正: {_feelings_fix_count}件")

    # 14. story_flow先頭20字重複修正（接続詞追加 + intensity別実質テンプレート）
    _STORYFLOW_PREFIXES = [
        "さらに、", "その後、", "やがて、", "次第に、", "一方で、",
        "そして、", "続けて、", "同時に、", "ここから、", "それから、",
        "そこから、", "息つく間もなく、", "勢いのまま、", "流れるように、",
        "間を置かず、", "畳みかけるように、", "一転して、",
    ]
    # intensity別のstory_flow実質テンプレート（接続詞追加より先にマッチを試みる）
    _STORYFLOW_TEMPLATES = {
        1: ["日常の空気が、かすかに変わり始め、",
            "まだ何も起きていないのに、胸騒ぎが止まらず、",
            "いつもと同じはずの時間が、どこか違って感じられ、"],
        2: ["触れられた場所の余韻が消えないまま、",
            "鼓動が速まるのを抑えきれず、",
            "抗えない引力に引き寄せられるように、",
            "身体が熱を帯び始め、理性がぐらつき、"],
        3: ["快感が理性を侵食し始め、",
            "もう後戻りできない一線を越え、",
            "抵抗する力が快楽に奪われていき、",
            "身体が正直に求め始め、心がそれに追いつけないまま、"],
        4: ["止まらない快感の波に呑まれ、",
            "理性の最後の砦が崩れ落ち、",
            "もう何も考えられないまま快楽に身を委ね、",
            "全身が求めることしかできなくなり、"],
        5: ["限界を超えた快楽が全てを焼き尽くし、",
            "意識が白く染まる中で、",
            "もう戻れないほど深く堕ちたまま、",
            "壊れたはずの感覚が、さらなる高みへ駆け上がり、"],
    }
    _sf_fix_count = 0
    _seen_sf = {}  # prefix20 -> first scene_id
    _used_sf_templates = set()
    for i, scene in enumerate(results):
        sf = scene.get("story_flow", "")
        if not sf or len(sf) < 20:
            sid = scene.get("scene_id", i + 1)
            if sf:
                _seen_sf[sf[:20]] = sid
            continue
        sf20 = sf[:20]
        sid = scene.get("scene_id", i + 1)
        if sf20 in _seen_sf:
            _sf_fixed = False
            intensity = scene.get("intensity", 3)
            # まずintensity別テンプレートで実質的な置換を試みる
            _i_templates = _STORYFLOW_TEMPLATES.get(intensity, _STORYFLOW_TEMPLATES[3])
            for _tmpl in _i_templates:
                if _tmpl not in _used_sf_templates:
                    new_sf = _tmpl + sf[min(len(sf), 15):]  # テンプレート+元の後半
                    if new_sf[:20] not in _seen_sf:
                        scene["story_flow"] = new_sf
                        _used_sf_templates.add(_tmpl)
                        _sf_fix_count += 1
                        _seen_sf[new_sf[:20]] = sid
                        _sf_fixed = True
                        break
            # テンプレート枯渇→従来の接続詞追加
            if not _sf_fixed:
                for try_idx in range(_sf_fix_count, _sf_fix_count + len(_STORYFLOW_PREFIXES)):
                    prefix = _STORYFLOW_PREFIXES[try_idx % len(_STORYFLOW_PREFIXES)]
                    new_sf = prefix + sf
                    if new_sf[:20] not in _seen_sf:
                        scene["story_flow"] = new_sf
                        _sf_fix_count += 1
                        _seen_sf[new_sf[:20]] = sid
                        _sf_fixed = True
                        break
            if not _sf_fixed:
                new_sf = f"[S{sid}] " + sf
                scene["story_flow"] = new_sf
                _sf_fix_count += 1
                _seen_sf[new_sf[:20]] = sid
        else:
            _seen_sf[sf20] = sid
    if _sf_fix_count > 0:
        log_message(f"  story_flow重複修正: {_sf_fix_count}件")

    # 15. speech重複修正（異なるシーンで同一セリフ → intensity考慮の微小バリエーション付加）
    # v8.6: intensity別サフィックス辞書（♡数をintensityに応じて制限）
    _INTENSITY_SUFFIXES = {
        1: ["…", "っ", "ぅ…", "ぁ…"],
        2: ["…", "っ", "…っ", "ぅ…", "ぁ…"],
        3: ["…", "っ", "…♡", "…っ", "♡", "ぅ…", "ぁ…"],
        4: ["…", "っ", "…♡", "…っ", "♡", "…♡♡", "ぅ…", "ぁ…"],
        5: ["…", "っ", "…♡", "…っ", "♡", "…♡♡", "♡♡♡", "ぅ…", "ぁ…"],
    }
    _sp_fix_count = 0
    _seen_speech = {}  # line_text -> (scene_idx, bubble_idx)
    for i, scene in enumerate(results):
        bubbles = scene.get("bubbles", [])
        sid = scene.get("scene_id", i + 1)
        intensity = scene.get("intensity", 3)
        for bi, b in enumerate(bubbles):
            if b.get("type") != "speech":
                continue
            line = b.get("text", "")
            if not line or len(line) < 4:
                continue
            if line in _seen_speech:
                # v8.6: intensity対応サフィックスを使用
                _suffixes = _INTENSITY_SUFFIXES.get(intensity, _INTENSITY_SUFFIXES[3])
                modified = False
                for suffix in _suffixes:
                    new_line = line.rstrip("…♡っ。、ぅぁ") + suffix
                    if new_line != line and new_line not in _seen_speech:
                        b["text"] = new_line
                        _seen_speech[new_line] = (i, bi)
                        _sp_fix_count += 1
                        modified = True
                        break
                if not modified:
                    # 全サフィックス枯渇 → 先頭に感嘆詞追加で一意化
                    _SPEECH_INTERJECTIONS = ["あっ…", "んっ…", "はぁ…", "ねぇ…"]
                    for intj in _SPEECH_INTERJECTIONS:
                        new_line = intj + line
                        if new_line not in _seen_speech:
                            b["text"] = new_line
                            _seen_speech[new_line] = (i, bi)
                            _sp_fix_count += 1
                            modified = True
                            break
                if not modified:
                    _seen_speech[line] = (i, bi)
            else:
                _seen_speech[line] = (i, bi)
    if _sp_fix_count > 0:
        log_message(f"  speech重複修正: {_sp_fix_count}件")

    # 16. description抽象的修正（intensity≥4で具体的キーワードがない → 具体表現を自動追加）
    _CONCRETE_ADDITIONS = {
        4: [
            "激しいピストンで腰が打ちつけられ、",
            "深く挿入された状態で腰を押さえつけられ、",
            "後ろから突き上げられて身体が跳ね、",
            "騎乗位で腰を打ちつけながら、",
            "脚を大きく開かされた体勢で、",
            "背後から抱きかかえられ腰を突かれ、",
            "壁に押し付けられ腰を掴まれた体勢のまま、",
            "四つん這いの姿勢で腰を掴まれ、",
        ],
        5: [
            "限界を超えた激しいピストンに身体が痙攣し、",
            "奥まで突き上げられ仰け反りながら、",
            "腰を掴まれ激しいピストンで突かれ続け、",
            "全身が震えるほどの快感に耐えきれず、",
            "何度もイかされビクビクと痙攣しながら、",
            "力が抜けた身体を好きにされ挿入が続き、",
            "ピストンの快楽に意識が飛びそうになり、",
            "汗だくの身体を抱え上げられ突かれ、",
        ],
    }
    _CONCRETE_KW_CHECK = [
        "正常位", "後背位", "騎乗位", "バック", "挿入", "ピストン", "腰を",
        "突き", "咥え", "舐め", "フェラ", "パイズリ", "手コキ", "指を",
        "汗", "涙", "震え", "痙攣", "力が抜け", "仰け反", "ビクビク",
        "掴み", "押さえ", "開かせ", "四つん這い", "うつ伏せ",
        "胸を", "腰を", "脚を", "太もも", "尻を",
    ]
    # v8.8: 事後・準備・時間停止解除シーンはプレフィックス追加をスキップ
    _POSTACT_SKIP_KW = [
        "射精を終え", "出した後", "事後", "終わった後", "行為の後",
        "証拠隠滅", "痕跡", "時間再開", "時間が再び", "再び動き",
        "混乱して", "違和感", "帰宅", "元に戻", "偽装", "拭き取",
        "余韻", "気づき始め", "調べて", "確認して",
    ]
    _desc_fix_count = 0
    _seen_concrete_prefixes = set()  # 500シーン耐性: 具体化後のprefix30重複回避
    for i, scene in enumerate(results):
        intensity = scene.get("intensity", 0)
        if intensity < 4:
            continue
        desc = scene.get("description", "")
        if not desc or len(desc) < 10:
            continue
        if any(kw in desc for kw in _CONCRETE_KW_CHECK):
            continue
        # v8.8: 事後・準備シーンにはエロプレフィックスを追加しない
        if any(kw in desc for kw in _POSTACT_SKIP_KW):
            continue
        # 具体表現を先頭に追加（prefix30重複回避付き）
        level = min(intensity, 5)
        additions = _CONCRETE_ADDITIONS.get(level, _CONCRETE_ADDITIONS[4])
        chosen = None
        for try_offset in range(len(additions)):
            candidate = additions[(i + try_offset) % len(additions)]
            new_prefix = (candidate + desc)[:30]
            if new_prefix not in _seen_concrete_prefixes:
                chosen = candidate
                break
        if chosen is None:
            # 全addition使用済み→隣接intensityも試す
            for adj_level in [max(4, level - 1), min(5, level + 1)]:
                if adj_level == level:
                    continue
                for candidate in _CONCRETE_ADDITIONS.get(adj_level, []):
                    new_prefix = (candidate + desc)[:30]
                    if new_prefix not in _seen_concrete_prefixes:
                        chosen = candidate
                        break
                if chosen:
                    break
        if chosen is None:
            chosen = additions[i % len(additions)]  # フォールバック
        scene["description"] = chosen + desc
        _seen_concrete_prefixes.add((chosen + desc)[:30])
        _desc_fix_count += 1
    if _desc_fix_count > 0:
        log_message(f"  description具体化修正: {_desc_fix_count}件")

    # 16b. description連続類似修正（3連続で同一行為キーワード→中央シーンを差し替え）
    _DESC_ACT_KW_FIX = ["膣奥", "突かれ", "責められ", "腰を振", "ピストン",
                         "挿入", "フェラ", "パイズリ", "騎乗", "バック",
                         "正常位", "四つん這い"]
    _DESC_SYNONYMS = {
        "ピストン": ["律動", "腰の動き", "突き上げ"],
        "膣奥": ["最奥部", "子宮口付近", "一番奥"],
        "突かれ": ["貫かれ", "押し込まれ", "攻められ"],
        "挿入": ["結合", "繋がった状態で", "受け入れた体勢で"],
    }
    _desc_sim_fix = 0
    _desc_kw_list = []
    for scene in results:
        desc = scene.get("description", "")
        kws = frozenset(kw for kw in _DESC_ACT_KW_FIX if kw in desc)
        _desc_kw_list.append(kws)
    for k in range(2, len(_desc_kw_list)):
        common = _desc_kw_list[k] & _desc_kw_list[k-1] & _desc_kw_list[k-2]
        if len(common) >= 2:
            # 中央シーン(k-1)のdescriptionを修正: 共通キーワードを類語に置換
            mid = results[k-1]
            desc = mid.get("description", "")
            for ckw in common:
                syns = _DESC_SYNONYMS.get(ckw, [])
                if syns:
                    desc = desc.replace(ckw, syns[k % len(syns)], 1)
            mid["description"] = desc
            _desc_kw_list[k-1] = frozenset(kw for kw in _DESC_ACT_KW_FIX if kw in desc)
            _desc_sim_fix += 1
    if _desc_sim_fix > 0:
        log_message(f"  description連続類似修正: {_desc_sim_fix}件")

    # 16d. description外見反復修正（3連続で同一先頭30文字→2回目以降を短縮）
    _appearance_fix_count = 0
    _desc_prefixes = [s.get("description", "")[:30] for s in results]
    for k in range(2, len(results)):
        p0, p1, p2 = _desc_prefixes[k-2], _desc_prefixes[k-1], _desc_prefixes[k]
        if p0 and p1 and p2 and p0 == p1 == p2:
            # 中央シーン(k-1)のdescription先頭を短縮: キャラ名だけ残す
            desc = results[k-1].get("description", "")
            if correct_names:
                # 「{キャラ名}が」「{キャラ名}は」の直後から残す
                for cn in correct_names:
                    for particle in ("が", "は", "の"):
                        marker = cn + particle
                        idx = desc.find(marker)
                        if idx >= 0:
                            desc = desc[idx:]
                            break
                    else:
                        continue
                    break
            results[k-1]["description"] = desc
            _desc_prefixes[k-1] = desc[:30]
            _appearance_fix_count += 1
    if _appearance_fix_count > 0:
        log_message(f"  description外見反復修正: {_appearance_fix_count}件")

    # 16c. title品質修正（句点除去・location混入修正・description混入修正）
    _TITLE_MOOD_WORDS = [
        "背徳", "快楽", "陶酔", "絶望", "衝動", "狂気", "恍惚", "堕落",
        "覚醒", "屈辱", "暴走", "服従", "解放", "執着", "欲望", "情欲",
    ]
    _title_quality_fix = 0
    for scene in results:
        title = scene.get("title", "")
        if not title:
            continue
        sid = scene.get("scene_id", "?")
        old_title = title
        need_regenerate = False

        # 句点「。」除去
        if "。" in title:
            title = title.replace("。", "").strip()

        # location混入チェック: locationテキストがtitleに含まれる場合は除去
        loc = scene.get("location_detail", scene.get("location", ""))
        if loc and len(loc) >= 4 and loc in title:
            title = title.replace(loc, "")
            # 連続する助詞の整理（「のの」「のでの」等）
            title = re.sub(r'([のでにてへ])\1+', r'\1', title)
            title = re.sub(r'の(で|に|て|へ)の', r'\1', title)
            title = title.strip("のでにてへ、 ")
            if len(title) < 3:
                need_regenerate = True

        # description混入チェック: titleの先頭10字がdescriptionに含まれる場合
        desc = scene.get("description", "")
        if desc and len(title) >= 10:
            title_head = title[:10]
            if title_head in desc:
                need_regenerate = True
        # 文体チェック: タイトルが文末表現で終わる→description的な文が混入
        if len(title) >= 6 and re.search(
                r'(された|されて|ながら|ている|ていく|ていた|かれて|かれた|声が|体が)$', title):
            need_regenerate = True
        # 助詞で開始するtitle（文の途中から切り取られた形跡）
        if title and title[0] in "のとがをでにへはも":
            need_regenerate = True
        # 送り仮名・漢字の途中で途切れるtitle
        if len(title) >= 8 and title[-1] in "張貼掛掲載映写薄汚":
            need_regenerate = True

        # titleが3文字未満or再生成フラグ → mood+intensityベースで再生成
        if need_regenerate or len(title) < 3:
            mood = scene.get("mood", "")
            intensity = scene.get("intensity", 3)
            mood_word = _TITLE_MOOD_WORDS[(int(sid) if isinstance(sid, int) else 0) % len(_TITLE_MOOD_WORDS)]
            if mood and len(mood) >= 2:
                title = f"{mood_word}の{mood[:8]}"
            else:
                title = f"{mood_word}のシーン{sid}"

        if title != old_title:
            scene["title"] = title
            _title_quality_fix += 1
            log_message(f"  S{sid}: title品質修正「{old_title}」→「{title}」")
    if _title_quality_fix > 0:
        log_message(f"  title品質修正: {_title_quality_fix}件")

    # 17. title重複修正（同一titleの2回目以降→行為/感情ベースの短いtitleに差し替え）
    # v8.2根本修正: f"{mood}の{desc}" を廃止。短いキーワードベースに変更
    _TITLE_ACTION_WORDS = [
        "背徳", "快楽", "服従", "支配", "羞恥", "覚醒", "堕落", "恍惚",
        "衝動", "情欲", "欲望", "執着", "解放", "陶酔", "狂乱", "震撼",
    ]
    _TITLE_BODY_WORDS = [
        "唇", "胸", "腰", "脚", "首筋", "耳", "背中", "指先",
        "太もも", "うなじ", "肌", "身体", "内腿",
    ]
    _seen_titles_af = set()
    _title_fix_af = 0
    for scene in results:
        t = scene.get("title", "")
        if t in _seen_titles_af:
            sid = scene.get("scene_id", "?")
            _si = scene.get("intensity", 3)
            # 行為 + 感情の短いタイトルを生成（12字以内）
            _aw = _TITLE_ACTION_WORDS[(_title_fix_af + sid if isinstance(sid, int) else _title_fix_af) % len(_TITLE_ACTION_WORDS)]
            _bw = _TITLE_BODY_WORDS[(_title_fix_af + (_si * 3)) % len(_TITLE_BODY_WORDS)]
            new_title = f"{_aw}の{_bw}"
            # 重複しないようにする
            if new_title in _seen_titles_af:
                new_title = f"{_aw}と{_bw}({sid})"
            scene["title"] = new_title
            _title_fix_af += 1
            log_message(f"  S{sid}: title重複修正「{t}」→「{new_title}」")
        _seen_titles_af.add(scene.get("title", ""))
    if _title_fix_af > 0:
        log_message(f"  title重複修正: {_title_fix_af}件")

    # 17b. title接頭辞反復修正（同一接頭辞が多すぎる → タイトル全体を再生成）
    # v8.3修正: old_title[2:]の盲目的文字切断を廃止。description/moodからキーワード抽出して全体再生成
    _title_prefix2_counter = {}
    for scene in results:
        t = scene.get("title", "")[:2]
        if t:
            _title_prefix2_counter.setdefault(t, []).append(scene)
    _title_prefix_fix = 0
    _all_titles_17b = set(s.get("title", "") for s in results)
    _TITLE_REGEN_TEMPLATES = [
        "{emotion}の{action}",
        "{action}と{emotion}",
        "{body}に走る{emotion}",
        "{emotion}に濡れた{body}",
        "{action}の先に",
        "溢れる{emotion}",
        "{body}が求めた{action}",
        "{emotion}の{body}",
    ]
    _TITLE_REGEN_EMOTIONS = {
        1: ["戸惑い", "緊張", "不安", "躊躇", "動揺"],
        2: ["羞恥", "期待", "困惑", "ときめき", "誘惑"],
        3: ["快感", "衝動", "陶酔", "情熱", "昂ぶり"],
        4: ["絶頂", "狂熱", "暴走", "支配", "崩壊"],
        5: ["恍惚", "極限", "解放", "至福", "白濁"],
    }
    _TITLE_REGEN_ACTIONS = [
        "愛撫", "吐息", "囁き", "接触", "抱擁", "口づけ", "交わり",
        "律動", "高まり", "震え", "疼き", "昂り", "絡み合い", "蜜月",
    ]
    _TITLE_REGEN_BODIES = [
        "唇", "指先", "肌", "胸", "うなじ", "太もも", "腰",
        "背中", "首筋", "耳たぶ", "素肌", "身体",
    ]
    for prefix, scenes_17b in _title_prefix2_counter.items():
        if len(scenes_17b) < 4:
            continue
        # 4回目以降をタイトル全体再生成
        for idx_17b, scene in enumerate(scenes_17b[3:]):
            sid = scene.get("scene_id", "?")
            old_title = scene.get("title", "")
            intensity = scene.get("intensity", 3)
            # intensityに応じた感情語を選択
            emotions = _TITLE_REGEN_EMOTIONS.get(intensity, _TITLE_REGEN_EMOTIONS[3])
            emotion = emotions[(idx_17b + _title_prefix_fix) % len(emotions)]
            action = _TITLE_REGEN_ACTIONS[(idx_17b + intensity) % len(_TITLE_REGEN_ACTIONS)]
            body = _TITLE_REGEN_BODIES[(idx_17b + (sid if isinstance(sid, int) else 0)) % len(_TITLE_REGEN_BODIES)]
            tmpl = _TITLE_REGEN_TEMPLATES[(idx_17b + _title_prefix_fix) % len(_TITLE_REGEN_TEMPLATES)]
            new_title = tmpl.format(emotion=emotion, action=action, body=body)
            # 重複チェック
            if new_title in _all_titles_17b:
                new_title = f"{emotion}の{body}({sid})"
            if len(new_title) > 25:
                new_title = new_title[:25].rstrip("。、…")
            scene["title"] = new_title
            _all_titles_17b.add(new_title)
            _title_prefix_fix += 1
            log_message(f"  S{sid}: title接頭辞修正「{old_title}」→「{new_title}」")
    if _title_prefix_fix > 0:
        log_message(f"  title接頭辞修正: {_title_prefix_fix}件")

    # 17c. title location leak検出（場所名がタイトルに混入→テンプレート再生成）
    _TITLE_LOCATION_LEAK_WORDS = [
        "タイル", "シャワー室", "プール", "白い壁", "天井", "床", "ベンチ", "更衣室",
        "洗面台", "カーテン", "廊下", "階段", "エレベーター", "ドア", "窓", "机",
    ]
    _title_leak_fix = 0
    for scene in results:
        title = scene.get("title", "")
        if not title:
            continue
        sid = scene.get("scene_id", "?")
        needs_regen = False
        # location語混入チェック
        for leak_word in _TITLE_LOCATION_LEAK_WORDS:
            if leak_word in title:
                needs_regen = True
                break
        # 末尾切断チェック（「新」「の」「と」等1文字で終わる不自然なタイトル）
        if not needs_regen and len(title) >= 3 and title[-1] in "新のとがをにでへは":
            needs_regen = True
        if needs_regen:
            old_title = title
            intensity = scene.get("intensity", 3)
            emotions = _TITLE_REGEN_EMOTIONS.get(intensity, _TITLE_REGEN_EMOTIONS[3])
            emotion = emotions[(_title_leak_fix + (sid if isinstance(sid, int) else 0)) % len(emotions)]
            action = _TITLE_REGEN_ACTIONS[(_title_leak_fix + intensity) % len(_TITLE_REGEN_ACTIONS)]
            body = _TITLE_REGEN_BODIES[(_title_leak_fix + 3) % len(_TITLE_REGEN_BODIES)]
            tmpl = _TITLE_REGEN_TEMPLATES[(_title_leak_fix + 2) % len(_TITLE_REGEN_TEMPLATES)]
            new_title = tmpl.format(emotion=emotion, action=action, body=body)
            if new_title in _all_titles_17b:
                new_title = f"{emotion}の{action}({sid})"
            if len(new_title) > 25:
                new_title = new_title[:25].rstrip("。、…")
            scene["title"] = new_title
            _all_titles_17b.add(new_title)
            _title_leak_fix += 1
            log_message(f"  S{sid}: title location leak修正「{old_title}」→「{new_title}」")
    if _title_leak_fix > 0:
        log_message(f"  title location leak修正: {_title_leak_fix}件")

    # 18. titleキーワード過剰使用修正（同じキーワードが3回以上→場所/mood/行為ベースに差し替え）
    _TITLE_KW_FIX = ["膣奥", "理性", "崩壊", "限界", "快感", "堕ち", "抵抗",
                      "連続", "激突", "責め", "声", "最後"]
    _seen_kw_fix_titles = set(s.get("title", "") for s in results)  # 500シーン耐性: 既存title追跡
    for kw in _TITLE_KW_FIX:
        kw_scenes = [(i, s) for i, s in enumerate(results) if kw in s.get("title", "")]
        if len(kw_scenes) >= 3:
            # 3回目以降の出現を差し替え
            _alt_kw = ["衝動", "背徳", "交わり", "激情", "陶酔", "震え", "熱", "嵐"]
            for idx, (i, scene) in enumerate(kw_scenes):
                if idx < 2:
                    continue  # 最初の2回は許容
                sid = scene.get("scene_id", "?")
                old_title = scene["title"]
                loc = scene.get("location_detail", scene.get("location", ""))[:10]
                # 重複回避: _alt_kwを順に試す
                new_title = None
                for try_offset in range(len(_alt_kw)):
                    alt = _alt_kw[(i + try_offset) % len(_alt_kw)]
                    candidate = f"{alt}の{loc}" if loc else f"{alt}のシーン{sid}"
                    if candidate not in _seen_kw_fix_titles:
                        new_title = candidate
                        break
                if new_title is None:
                    new_title = f"{_alt_kw[i % len(_alt_kw)]}のシーン{sid}"
                _seen_kw_fix_titles.discard(old_title)
                _seen_kw_fix_titles.add(new_title)
                scene["title"] = new_title
                log_message(f"  S{sid}: titleキーワード過剰修正「{old_title}」→「{new_title}」")

    # 19. title長制限（Step 17-18で生成されたtitleも含め25文字以内に）
    for scene in results:
        title = scene.get("title", "")
        if len(title) > 25:
            scene["title"] = title[:25].rstrip("。、…")

    # 20. intensity不一致自動修正（♡除去/丁寧語短縮/高intensity喘ぎ置換）
    _intensity_fix_count = 0
    # i=5級喘ぎパターン（i≤3シーンで出現したらintensity相応の喘ぎに置換）
    _HIGH_INTENSITY_MOAN_RE = re.compile(
        r'ひぎ|んほ[ぉぅ]|あへ[ぇぁ]|ん゛|いぐ[ぅっ]|おほ[ぉっ]'
        r'|らめ[ぇぅ]|こわれ[るっ]|ぶっ壊|いっちゃ[うぅ]')
    _used_moan_fix20 = set()
    for scene in results:
        intensity = scene.get("intensity", 3)
        for bubble in scene.get("bubbles", []):
            txt = bubble.get("text", "")
            if not txt:
                continue
            orig = txt
            if intensity <= 2 and "♡" in txt:
                txt = txt.replace("♡", "")
            # v8.6: intensity別♡数上限（i=3: ♡1個, i=4: ♡♡まで, i=5: 制限なし）
            elif intensity == 3:
                while "♡♡" in txt:
                    txt = txt.replace("♡♡", "♡")
            elif intensity == 4:
                while "♡♡♡" in txt:
                    txt = txt.replace("♡♡♡", "♡♡")
            if intensity >= 4 and bubble.get("type") in ("speech", "thought"):
                txt = txt.replace("です", "")
                txt = txt.replace("ます", "")
                txt = txt.replace("ください", "…♡")
            # i≤3のmoanにi=5級喘ぎが混入 → 適切なintensityの喘ぎに置換
            if intensity <= 3 and bubble.get("type") == "moan" and _has_pool:
                if _HIGH_INTENSITY_MOAN_RE.search(txt):
                    pool = _get_moan_pool_with_char(intensity)
                    repl = pick_replacement(pool, _used_moan_fix20, _normalize_bubble_text,
                                           intensity=intensity)
                    if repl:
                        txt = repl
            # 空になった場合は元に戻す
            stripped = txt.replace("…", "").replace("♡", "").strip()
            if not stripped:
                txt = orig
            if txt != orig:
                bubble["text"] = txt
                _intensity_fix_count += 1
    if _intensity_fix_count > 0:
        log_message(f"  intensity不一致修正: {_intensity_fix_count}件")

    # 20b. Speech崩壊パターン（高intensityのspeechを段階的に断片化）
    # i=1: そのまま / i=2: やや不安定 / i=3: 断片化開始 / i=4: ほぼ崩壊 / i=5: moan支配
    _speech_frag_count = 0
    import random as _frag_rand
    for scene in results:
        intensity = scene.get("intensity", 3)
        if intensity <= 2:
            continue  # i1-2はそのまま
        for bubble in scene.get("bubbles", []):
            if bubble.get("type") != "speech":
                continue
            txt = bubble.get("text", "")
            if not txt or len(txt) < 4:
                continue
            # 男性セリフは崩壊させない（_is_male_by_nameで包括判定）
            _spk = bubble.get("speaker", "")
            if _spk and _is_male_by_name(_spk):
                continue
            orig = txt
            if intensity == 3:
                # i3: 文末に「…っ」挿入、句読点を「…」に
                if not txt.endswith("…") and not txt.endswith("っ"):
                    txt = txt.rstrip("。、！!？?") + "…っ"
                txt = txt.replace("。", "…")
            elif intensity == 4:
                # i4: 文字間に「…」挿入で断片化 + 末尾切断
                # 「やめてください」→「やめ…て…っ」
                chars = list(txt.replace("…", "").replace("♡", "").replace("っ", "").replace("。", "").replace("、", ""))
                if len(chars) >= 3:
                    # 2-3文字ごとに「…」を挿入
                    fragmented = []
                    _chunk_size = 2 if len(chars) <= 6 else 3
                    for _fi in range(0, len(chars), _chunk_size):
                        fragmented.append("".join(chars[_fi:_fi + _chunk_size]))
                    txt = "…".join(fragmented)
                    # ♡があれば末尾に1つ
                    if "♡" in orig:
                        txt += "…♡"
                    else:
                        txt += "…っ"
            elif intensity >= 5:
                # i5: speechをmoan的に変換（3文字以下+「…♡」）
                # 元の先頭1-2文字を残して崩壊
                _core = txt.replace("…", "").replace("♡", "").replace("っ", "").replace("。", "").replace("、", "")
                if _core:
                    _first = _core[0]
                    txt = _first + "…っ…♡"
            if txt != orig:
                bubble["text"] = txt
                _speech_frag_count += 1
    if _speech_frag_count > 0:
        log_message(f"  speech崩壊: {_speech_frag_count}件 (intensity連動)")

    # 21. エピローグ・ストーリーリセット検出＋修正
    # 最後の10%のシーンでintensityがi=2以下に戻った場合、ストーリーがリセットしている
    _progress("Step 21 エピローグリセット修正")
    _total = len(results)
    if _total >= 20:
        _epilogue_start = max(1, _total - max(5, _total // 10))
        _reset_count = 0
        # 文脈無視のセリフ（導入的・日常的・挨拶・出会い）をプールから置換
        _RESET_INDICATORS = frozenset([
            "お腹すいた", "エッチ", "おはよう", "こんにちは", "はじめまして",
            "なに？", "誰？", "どうした", "久しぶり", "元気？",
            "よろしく", "初めまして", "いらっしゃい", "おじゃまします",
            "ただいま", "おかえり", "行ってきます", "お邪魔します",
            "はい、どうぞ", "ごめんください", "失礼します", "お久しぶり",
            "今日は", "調子どう", "暇だ", "退屈", "何しよう",
            "お茶", "ご飯", "勉強", "宿題", "仕事",
        ])
        _RESET_SPEECH_REPLACEMENTS = [
            "まだ…終わらないの…",
            "もう…むりぃ…♡",
            "やだ…まだ…♡",
            "いっちゃ…う…♡",
            "おかしく…なる…♡",
            "やめ…て…♡",
            "もう…だめ…♡",
            "とまんない…♡",
        ]
        _RESET_THOUGHT_REPLACEMENTS = [
            "からだ…もう…",
            "あたま…まっしろ…",
            "もう…なにも…かんがえられない…",
            "とまらない…からだ…",
            "こわれ…ちゃう…",
            "もどれない…もう…",
            "きもちいい…のに…こわい…",
            "おわらない…おわらない…",
        ]
        import random as _rng_reset
        for i in range(_epilogue_start, _total):
            scene = results[i]
            si = scene.get("intensity", 3)
            if si <= 2:
                scene["intensity"] = 3
                _reset_count += 1
                # intensity 2以下のセリフが導入っぽいテキストなら修正
                for b in scene.get("bubbles", []):
                    txt = b.get("text", "")
                    btype = b.get("type", "")
                    if any(ri in txt for ri in _RESET_INDICATORS):
                        if btype == "speech":
                            b["text"] = _rng_reset.choice(_RESET_SPEECH_REPLACEMENTS)
                        elif btype == "thought":
                            b["text"] = _rng_reset.choice(_RESET_THOUGHT_REPLACEMENTS)
                        _reset_count += 1
        if _reset_count > 0:
            log_message(f"  エピローグリセット修正: {_reset_count}件（i≤2→i=3 + セリフ置換）")

    # 22. i=4連続過多の自動修正（4-6シーン超→i=3ブレイク挿入、i=5ピーク保護付き）v8.9
    _consecutive_4 = 0
    _i4_break_count = 0
    _i4_auto_limit = _rng.randint(4, 6)
    for idx_22, scene in enumerate(results):
        if scene.get("intensity", 3) == 4:
            _consecutive_4 += 1
            if _consecutive_4 > _i4_auto_limit:
                _next_is_peak = (idx_22 + 1 < len(results) and results[idx_22 + 1].get("intensity", 3) == 5)
                if _next_is_peak:
                    # i=5ピーク直前ではブレイクしない。1つ前のi=4をブレイクに変更
                    if idx_22 > 0 and results[idx_22 - 1].get("intensity", 3) == 4:
                        results[idx_22 - 1]["intensity"] = 3
                        _consecutive_4 = 1
                        _i4_break_count += 1
                else:
                    scene["intensity"] = 3
                    _consecutive_4 = 0
                    _i4_break_count += 1
                    _i4_auto_limit = _rng.randint(4, 6)  # 次の閾値をランダム再設定
        else:
            _consecutive_4 = 0
    if _i4_break_count > 0:
        _progress(f"Step 22 i=4連続ブレイク挿入: {_i4_break_count}箇所")
        log_message(f"  i=4連続上限4-6: {_i4_break_count}箇所にi=3ブレイク挿入")

    # ── Step 23: 物理状態累積修正（服装復活防止 + 射精後の体液持続）──
    _phys_max_undress = 0
    _phys_had_cum = False
    _phys_fix_count = 0
    for i, scene in enumerate(results):
        sd = scene.get("sd_prompt", "")
        if not sd:
            continue
        _sd_tags_set = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}

        # 脱衣レベル検出
        _cur_lv = 0
        if _sd_tags_set & {"nude", "naked", "completely_nude"}:
            _cur_lv = 5
        elif _sd_tags_set & {"topless", "bottomless", "panties_only", "naked_shirt", "stockings_only"}:
            _cur_lv = 4
        elif _sd_tags_set & {"panties_aside", "open_shirt", "bra_removed", "torn_clothes", "no_bra", "no_panties"}:
            _cur_lv = 3
        elif _sd_tags_set & {"partially_undressed", "shirt_lift", "bra_visible", "unbuttoned_shirt"}:
            _cur_lv = 2

        # 服装復活修正: 前シーンでnude(5)→現シーンで脱衣タグなし→nude追加
        if _phys_max_undress >= 5 and _cur_lv < 3 and i > 0:
            if "nude" not in sd.lower():
                scene["sd_prompt"] = sd.rstrip() + ", nude"
                _phys_fix_count += 1
        elif _phys_max_undress >= 4 and _cur_lv <= 1 and i > 0:
            if not (_sd_tags_set & {"topless", "bottomless", "panties_only", "nude", "naked"}):
                scene["sd_prompt"] = sd.rstrip() + ", partially_undressed"
                _phys_fix_count += 1

        _phys_max_undress = max(_phys_max_undress, _cur_lv)

        # 射精後の体液持続修正: 前シーンで射精→現シーンにcum系なし→追加
        _has_cum = bool(_sd_tags_set & {"cum", "cum_on_body", "cum_on_face", "cum_in_pussy",
                                         "cum_overflow", "cum_string", "cum_pool", "cum_drip"})
        if _phys_had_cum and not _has_cum and scene.get("intensity", 3) >= 3:
            scene["sd_prompt"] = scene["sd_prompt"].rstrip() + ", cum_on_body"
            _phys_fix_count += 1
        if _has_cum:
            _phys_had_cum = True

    if _phys_fix_count > 0:
        _progress(f"Step 23 物理状態累積修正: {_phys_fix_count}箇所")

    return results


def _fix_consecutive_locations(results: list) -> None:
    """location多様性の自動修正。3連続同一 + 全体70%以上同一に対応。"""
    # 場所内位置バリエーション（単一ロケーションシナリオ用）
    _MICRO_LOCATIONS = {
        "トイレ": ["洗面台の前", "入口付近の壁際", "隣の個室", "清掃用具入れの奥", "手洗い場の鏡前"],
        "教室": ["教卓の前", "窓際の席", "ロッカーの陰", "教室の隅", "廊下に面した壁際"],
        "寝室": ["ベッドの上", "窓際", "クローゼットの前", "ドア付近", "鏡台の前"],
        "浴室": ["浴槽の中", "洗い場", "脱衣所", "シャワーの下", "浴室の壁際"],
        "シャワー": ["シャワーヘッドの下", "壁に背を預けて", "ガラス戸の前", "排水口付近", "湯気の中"],
        "プール": ["プールサイド", "更衣室の奥", "シャワー室", "監視台の裏", "深い方の端"],
        "温泉": ["露天風呂の岩陰", "洗い場の隅", "脱衣所の棚前", "湯船の縁", "竹垣の奥"],
        "海": ["砂浜の東屋", "岩場の陰", "海の家の裏", "波打ち際", "テントの中"],
        "オフィス": ["デスクの上", "応接ソファ", "給湯室", "会議室", "コピー機の裏"],
    }
    _LOC_VARIATIONS = [
        ("窓際", "near_window, window"),
        ("隅", "corner"),
        ("入り口付近", "doorway"),
        ("奥まった場所", "dimly_lit"),
        ("壁際", "against_wall"),
        ("中央", "center"),
        ("片隅のテーブル付近", "table"),
        ("柱の陰", "pillar, shadow"),
    ]
    import random as _rng

    locations_list = []
    for scene in results:
        loc = scene.get("location_detail", scene.get("location", ""))
        locations_list.append(loc.strip().lower() if loc else "")

    # --- 全体の同一location率チェック（prefix 20文字で集計: 微妙な表現違いも同一扱い） ---
    _LOC_PREFIX_LEN = 20
    if len(locations_list) >= 10:
        from collections import Counter as _Counter
        loc_prefix_list = [l[:_LOC_PREFIX_LEN] if l else "" for l in locations_list]
        loc_counter = _Counter(lp for lp in loc_prefix_list if lp)
        if loc_counter:
            most_common_prefix, most_common_count = loc_counter.most_common(1)[0]
            # 最頻prefixに対応するフルlocationを取得
            most_common_loc = most_common_prefix
            for l in locations_list:
                if l and l[:_LOC_PREFIX_LEN] == most_common_prefix:
                    most_common_loc = l
                    break
            ratio = most_common_count / len(locations_list)
            if ratio > 0.70:  # 70%以上が同一location prefix
                # micro-locationで分散させる
                micro_pool = None
                for key, micros in _MICRO_LOCATIONS.items():
                    if key.lower() in most_common_loc or key in most_common_loc:
                        micro_pool = micros
                        break
                if not micro_pool:
                    micro_pool = [v[0] for v in _LOC_VARIATIONS]

                micro_idx = 0
                fix_micro = 0
                for k, scene in enumerate(results):
                    loc = locations_list[k]
                    if loc and loc[:_LOC_PREFIX_LEN] == most_common_prefix and k % 2 == 0:  # 2シーンに1つ変化（v8.9強化）
                        new_micro = micro_pool[micro_idx % len(micro_pool)]
                        orig_loc_detail = scene.get("location_detail", scene.get("location", ""))
                        scene["location_detail"] = f"{orig_loc_detail}（{new_micro}）"
                        locations_list[k] = scene["location_detail"].strip().lower()
                        micro_idx += 1
                        fix_micro += 1
                if fix_micro > 0:
                    log_message(f"  location同一率{ratio:.0%}→micro-location分散: {fix_micro}件")

    # --- 3シーン連続同一locationの修正（prefix 20文字一致で判定） ---
    fix_count = 0
    for k in range(2, len(locations_list)):
        lk = locations_list[k]
        lk1 = locations_list[k-1]
        lk2 = locations_list[k-2]
        if (lk and lk1 and lk2
                and lk[:_LOC_PREFIX_LEN] == lk1[:_LOC_PREFIX_LEN] == lk2[:_LOC_PREFIX_LEN]):
            mid = results[k - 1]
            orig_loc = mid.get("location_detail", mid.get("location", ""))
            if not orig_loc:
                continue
            if any(v[0] in orig_loc for v in _LOC_VARIATIONS):
                continue
            var_jp, var_tags = _rng.choice(_LOC_VARIATIONS)
            new_loc = f"{orig_loc}の{var_jp}"
            mid["location_detail"] = new_loc
            sd = mid.get("sd_prompt", "")
            if sd:
                existing = {t.strip().lower().replace(" ", "_") for t in sd.split(",") if t.strip()}
                new_tags = [t.strip() for t in var_tags.split(",") if t.strip()]
                added = []
                for nt in new_tags:
                    if nt.lower() not in existing:
                        added.append(nt)
                if added:
                    mid["sd_prompt"] = sd.rstrip(", ") + ", " + ", ".join(added)
            locations_list[k - 1] = new_loc.strip().lower()
            fix_count += 1

    if fix_count > 0:
        log_message(f"  location連続修正: {fix_count}件")


# ---------------------------------------------------------------------------
# 設定スタイル検出: コンセプトから背景の世界観を推定し、SDタグを補正する
# ---------------------------------------------------------------------------
SETTING_STYLES = {
    "traditional_japanese_rural": {
        "keywords": ["村", "田舎", "農村", "山里", "漁村", "集落",
                     "古民家", "昔ながら", "伝統", "風習", "因習", "祭り",
                     "大正", "昭和初期", "時代劇", "和風の村"],
        "replace": {
            "bed": "futon",
            "bedroom": "japanese_room",
            "brick_wall": "wooden_wall",
            "concrete": "wooden_floor",
            "warehouse": "storehouse",
            "sofa": "zabuton",
            "table": "chabudai",
            "curtains": "shoji",
            "nightstand": "andon",
            "pillow": "sobagara_pillow",
            "blanket": "futon_blanket",
            "modern": "traditional",
            "apartment": "old_japanese_house",
            "hotel_room": "ryokan_room",
            "alley": "village_path",
            "narrow_street": "rural_path",
        },
        "prohibit": {"brick_wall", "concrete", "modern", "neon",
                     "skyscraper", "office", "elevator", "subway",
                     "highway", "parking", "urban", "city_lights",
                     "apartment", "hotel"},
        "append": ["traditional", "japanese", "wooden", "rural",
                   "old_house", "tatami", "shoji", "dim"],
        "prompt_hint": "和風田舎・伝統的日本家屋（茅葺き屋根、板壁、障子、畳、布団、囲炉裏、行灯、土間）。現代的な家具・建材(ベッド、レンガ、コンクリート)は絶対に使わない",
    },
    "traditional_japanese_urban": {
        "keywords": ["遊郭", "花街", "吉原", "江戸", "京都", "芸者", "花魁",
                     "大正ロマン", "料亭"],
        "replace": {
            "bed": "futon",
            "bedroom": "japanese_room",
            "brick_wall": "wooden_wall",
            "concrete": "wooden_floor",
            "sofa": "zabuton",
            "curtains": "noren",
            "nightstand": "andon",
        },
        "prohibit": {"concrete", "modern", "skyscraper", "office",
                     "elevator", "subway", "highway"},
        "append": ["traditional", "japanese", "wooden", "paper_lantern",
                   "tatami", "fusuma", "ornate"],
        "prompt_hint": "和風花街・遊郭風（襖、行灯、赤い照明、障子、畳、掛け軸、豪華な着物）。現代要素禁止",
    },
    "fantasy_medieval": {
        "keywords": ["ファンタジー", "異世界", "中世", "魔法", "王国", "城",
                     "冒険者", "ギルド", "騎士", "魔王", "勇者", "ドラゴン",
                     "エルフ", "ダンジョン"],
        "replace": {
            "apartment": "stone_chamber",
            "hotel_room": "inn_room",
            "concrete": "stone_wall",
            "office": "throne_room",
            "subway": "underground_passage",
        },
        "prohibit": {"modern", "neon", "skyscraper", "office",
                     "elevator", "subway", "highway", "smartphone",
                     "computer"},
        "append": ["fantasy", "medieval", "stone", "torch",
                   "candlelight"],
        "prompt_hint": "中世ファンタジー風（石造りの壁、蝋燭、松明、木製家具、革製品）。現代要素禁止",
    },
    "modern_school": {
        "keywords": ["学園", "学校", "クラスメイト", "同級生",
                     "生徒", "文化祭",
                     "体育祭", "放課後", "部室", "屋上"],
        "replace": {
            "futon": "bed",
            "tatami": "floor",
            "shoji": "window",
            "andon": "fluorescent_light",
            "stone_wall": "concrete_wall",
        },
        "prohibit": {"torch", "candlelight", "medieval", "stone",
                     "fantasy", "traditional", "rural"},
        "append": ["school", "indoors"],
        "prompt_hint": "現代日本の学園（教室、廊下、屋上、体育館、プール、図書室、保健室）。学校の雰囲気を重視",
    },
    "modern_urban": {
        "keywords": ["都会", "東京", "マンション", "アパート", "オフィス",
                     "ビル", "繁華街", "ラブホ", "カラオケ", "居酒屋",
                     "コンビニ", "駅", "電車", "バス", "タクシー",
                     "ネットカフェ", "現代"],
        "replace": {
            "futon": "bed",
            "tatami": "wooden_floor",
            "shoji": "curtains",
            "andon": "lamp",
            "stone_wall": "concrete_wall",
        },
        "prohibit": {"torch", "medieval", "stone", "fantasy",
                     "traditional", "rural", "old_house"},
        "append": ["indoors", "modern"],
        "prompt_hint": "現代都市（マンション、オフィスビル、ラブホテル、居酒屋、電車内）。都会的な雰囲気",
    },
    "hot_spring": {
        "keywords": ["温泉", "露天風呂", "旅館", "混浴", "湯けむり",
                     "秘湯", "温泉旅行", "大浴場", "脱衣所"],
        "replace": {
            "bed": "futon",
            "bedroom": "japanese_room",
            "apartment": "ryokan_room",
            "hotel_room": "ryokan_room",
            "curtains": "noren",
        },
        "prohibit": {"office", "elevator", "subway", "highway",
                     "skyscraper", "urban"},
        "append": ["onsen", "steam", "wet", "towel", "japanese",
                   "warm_lighting"],
        "prompt_hint": "温泉・旅館（露天風呂、岩風呂、檜風呂、湯けむり、暖簾、浴衣、タオル）。蒸気と湯の質感を重視",
    },
    "beach_resort": {
        "keywords": ["ビーチ", "海", "水着", "プール", "リゾート",
                     "南国", "離島", "海辺", "海水浴", "日焼け",
                     "サーフ", "ヨット", "砂浜"],
        "replace": {
            "futon": "bed",
            "tatami": "wooden_floor",
            "shoji": "window",
        },
        "prohibit": {"medieval", "stone", "torch", "traditional",
                     "rural", "old_house", "snow"},
        "append": ["outdoors", "beach", "ocean", "sky", "sunlight",
                   "palm_tree"],
        "prompt_hint": "ビーチリゾート（砂浜、ヤシの木、青い海と空、白い砂浜、パラソル、水着）。開放的な南国感",
    },
    "sci_fi": {
        "keywords": ["SF", "宇宙", "近未来", "サイバーパンク", "ロボット",
                     "アンドロイド", "宇宙船", "コロニー", "メカ"],
        "replace": {
            "futon": "bed",
            "tatami": "metal_floor",
            "shoji": "sliding_door",
            "wooden_wall": "metal_wall",
            "andon": "neon_light",
            "stone_wall": "metal_wall",
        },
        "prohibit": {"medieval", "traditional", "rural", "old_house",
                     "tatami", "shoji", "torch", "candlelight"},
        "append": ["sci-fi", "futuristic", "neon", "hologram",
                   "metal", "chrome"],
        "prompt_hint": "SF・近未来（メタリックな壁、ネオンライト、ホログラム、宇宙船内、ハイテク機器）。未来的な無機質感",
    },
}


def _detect_setting_style(concept: str, theme: str = "") -> Optional[dict]:
    """コンセプト文字列からSETTING_STYLESのどれに該当するか判定する。

    theme引数がある場合、テーマと矛盾するスタイルをスキップする。
    """
    if not concept:
        return None

    # テーマ×スタイル矛盾マップ（このスタイルはこのテーマでは不適切）
    _THEME_STYLE_CONFLICTS = {
        "office": {"modern_school", "traditional_japanese_rural", "traditional_japanese_urban",
                   "fantasy_medieval", "hot_spring", "beach_resort"},
        "medical": {"modern_school", "fantasy_medieval", "beach_resort"},
        "sports": {"traditional_japanese_rural", "traditional_japanese_urban",
                   "fantasy_medieval"},
        "idol": {"traditional_japanese_rural", "fantasy_medieval"},
        "neighbor": {"modern_school", "fantasy_medieval"},
        "prostitution": {"modern_school", "fantasy_medieval"},
        "sleep": {"modern_school", "beach_resort"},
        "isekai": {"modern_school", "modern_urban"},
        "monster": {"modern_school", "modern_urban"},
        "tentacle": {"modern_school", "modern_urban"},
        "onsen": {"modern_school", "fantasy_medieval", "sci_fi"},
        "swimsuit": {"fantasy_medieval", "sci_fi"},
        "chikan": {"traditional_japanese_rural", "fantasy_medieval"},
    }
    blocked_styles = _THEME_STYLE_CONFLICTS.get(theme, set()) if theme else set()

    concept_lower = concept.lower()
    for style_key, style in SETTING_STYLES.items():
        if style_key in blocked_styles:
            continue
        for kw in style["keywords"]:
            if kw in concept or kw in concept_lower:
                return style
    return None


def enhance_sd_prompts(results: list, char_profiles: list = None,
                       setting_style: Optional[dict] = None,
                       male_tags: str = "", time_tags: str = "",
                       location_type: str = "",
                       sd_quality_tags: str = "",
                       sd_prefix_tags: str = "",
                       sd_suffix_tags: str = "",
                       theme: str = "",
                       faceless_male: bool = True,
                       sd_neg_base: str = "",
                       sd_neg_prefix: str = "",
                       sd_neg_suffix: str = "") -> list:
    """全シーンのSDプロンプトを後処理で最適化（APIコスト不要）。

    - 日本語タグ除去
    - quality tags確保
    - キャラタグ補完
    - 照明タグ追加
    - 重要表情タグにウェイト付加
    - 設定スタイルに基づくタグ置換・追加・禁止
    - 重複排除
    """
    import re as _re

    char_danbooru = []
    if char_profiles:
        for cp in char_profiles:
            char_danbooru.extend(cp.get("danbooru_tags", []))

    LIGHTING_WORDS = {"lighting", "sunlight", "moonlight", "candlelight",
                      "backlight", "rim_light", "neon", "lamp", "golden_hour",
                      "light_rays", "volumetric"}

    # ウェイト付加対象（SD画像の品質に直結する重要タグ）
    WEIGHT_EXPRESSION = {"ahegao", "orgasm", "rolling_eyes", "tongue_out",
                         "crying_with_eyes_open", "fucked_silly", "mindbreak",
                         "head_back", "drooling", "heart-shaped_pupils",
                         "tears_of_pleasure", "arched_back", "clenched_teeth"}
    WEIGHT_ACTION = {"deep_penetration", "cum_in_pussy", "overflow",
                     "multiple_penises", "double_penetration"}

    # intensity別 表情・身体反応タグ自動注入マップ（v7.6拡張: バリエーション増）
    _INTENSITY_EXPRESSION_MAP = {
        1: ["calm", "closed_mouth", "looking_away", "embarrassed",
            "slight_blush", "curious", "nervous_smile"],
        2: ["blush", "looking_down", "covering_face", "shy",
            "averting_eyes", "fidgeting", "pursed_lips",
            "hand_on_own_chest", "embarrassed", "watery_eyes"],
        3: ["blush", "parted_lips", "panting", "nervous", "heavy_breathing",
            "light_sweat", "clenched_teeth", "biting_lip",
            "furrowed_brow", "closed_eyes", "lip_biting",
            "hand_over_mouth", "surprised", "gasping"],
        4: ["open_mouth", "moaning", "tears", "sweating", "head_back",
            "arched_back", "clenched_fists", "trembling",
            "sweat_drops", "sweaty_body", "flushed_skin",
            "spread_legs", "gripping_sheets", "messy_hair",
            "half-closed_eyes", "glazed_eyes", "crying_with_eyes_open",
            "o-ring_mouth", "biting_own_lip", "scrunched_face"],
        5: ["ahegao", "rolling_eyes", "tongue_out", "drooling", "head_back",
            "arched_back", "toes_curling", "full_body_arch", "tears",
            "sweat_drops", "sweaty_body", "sweat_glistening", "skin_glistening",
            "heart_pupils", "cross-eyed", "saliva_drip", "fucked_silly",
            "vacant_eyes", "steam", "trembling_legs",
            "eye_roll", "slack_jaw", "convulsing"],
    }

    # intensity別 カメラアングル/構図タグプール
    # シーンごとにローテーションで選択し、連続重複を防ぐ
    _CAMERA_ANGLE_POOL = {
        1: ["upper_body", "portrait", "from_side", "straight-on"],
        2: ["upper_body", "cowboy_shot", "from_side", "close-up",
            "from_above", "looking_at_viewer"],
        3: ["pov", "from_above", "from_side", "cowboy_shot",
            "dutch_angle", "close-up", "from_below"],
        4: ["pov", "from_below", "from_above", "from_side",
            "wide_shot", "dutch_angle", "from_behind",
            "close-up", "between_legs"],
        5: ["pov", "from_below", "from_above", "wide_shot",
            "from_behind", "dutch_angle", "from_side",
            "close-up", "full_body", "between_legs"],
    }

    # intensity別 衣装状態エスカレーション
    CLOTHING_ESCALATION = {
        1: [],  # 通常衣装（変更なし）
        2: ["clothes_pull", "skirt_lift", "loosened_tie"],
        3: ["partially_undressed", "shirt_lift", "bra_visible",
            "one_shoulder_exposed", "disheveled_clothes",
            "unbuttoned_shirt", "skirt_around_waist"],
        4: ["topless", "panties_only", "torn_clothes", "clothes_around_ankles",
            "open_shirt", "naked_shirt", "panties_aside",
            "bra_removed", "stockings_only"],
        5: ["completely_nude", "nude", "naked", "clothes_scattered",
            "torn_panties", "clothes_removed"],
    }

    # intensity別 体液進行タグ
    FLUID_PROGRESSION = {
        1: [],
        2: ["light_blush"],
        3: ["light_blush", "pussy_juice", "wet", "love_juice"],
        4: ["pussy_juice", "wet", "saliva", "saliva_trail",
            "wet_skin", "sweat_stain", "dripping_wet",
            "precum", "mixed_fluids"],
        5: ["cum", "cum_on_body", "cum_drip", "overflow",
            "dripping", "saliva_trail", "wet_skin",
            "cum_in_pussy", "cum_pool", "cum_string",
            "excessive_cum", "body_fluids"],
    }

    _prev_scene_positions = set()  # 前シーンの体位タグ（重複防止用）
    _prev_camera_angle = ""  # 前シーンのカメラアングル（連続重複防止用）
    _camera_scene_idx = 0  # カメラアングルローテーション用カウンタ
    _angle_history = []  # 直近アングル履歴（3連続同一防止用）

    # v9.0: 物理状態累積トラッキング（シーン間の服装・体液引き継ぎ）
    _accumulated_clothing = set()  # 累積脱衣状態（一度nudeなら以降もnude）
    _accumulated_fluids = set()    # 累積体液状態（射精後は体液タグ持続）
    # 脱衣レベル: 高いほど脱いでいる（逆行禁止）
    _CLOTHING_UNDRESS_LEVEL = {
        "clothes_pull": 1, "skirt_lift": 1, "loosened_tie": 1,
        "partially_undressed": 2, "shirt_lift": 2, "bra_visible": 2,
        "one_shoulder_exposed": 2, "disheveled_clothes": 2,
        "unbuttoned_shirt": 2, "skirt_around_waist": 2, "unbuttoned": 2,
        "open_shirt": 3, "dress_lift": 2, "swimsuit_aside": 3,
        "topless": 4, "panties_only": 4, "torn_clothes": 3,
        "clothes_around_ankles": 4, "naked_shirt": 4, "panties_aside": 3,
        "bra_removed": 3, "bra_pull": 2, "stockings_only": 4,
        "no_bra": 3, "no_panties": 3, "shirt_lift": 2,
        "bottomless": 4, "undressing": 2, "clothes_removed": 5,
        "completely_nude": 5, "nude": 5, "naked": 5,
    }
    _max_undress_level = 0  # これまでの最大脱衣レベル
    # 体液持続タグ（射精系は以降のシーンにも残る）
    _PERSISTENT_FLUID_TAGS = {
        "cum", "cum_on_body", "cum_on_face", "cum_in_pussy", "cum_overflow",
        "cum_string", "cum_pool", "cum_drip", "excessive_cum",
    }

    # Phase6: アクセサリ永続化トラッキング
    _detected_accessories = set()  # 全シーン横断で検出されたアクセサリタグ

    # Phase6: キャラ保護タグ（char_danbooru上位5タグは除去対象外）
    _protected_char_tags = set()
    if char_danbooru:
        _protected_char_tags = {c.lower().replace(" ", "_") for c in char_danbooru[:5]}

    # Phase6: 性格タイプ検出
    _char_personality = ""
    if char_profiles:
        for cp in char_profiles:
            _pt = cp.get("personality_type", "")
            if _pt:
                _char_personality = _pt.lower()
                break

    for scene in results:
        sd = scene.get("sd_prompt", "")
        if not sd:
            continue

        tags = [t.strip() for t in sd.split(",") if t.strip()]
        sd_lower = sd.lower()

        # 1. 日本語タグ除去
        tags = [t for t in tags if not _re.search(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]', t)]

        # 1.5. 室内外矛盾タグ自動修正
        _outdoor_mk = {"outdoors", "park", "forest", "beach", "poolside", "rooftop", "garden"}
        _indoor_mk = {"indoors", "classroom", "bedroom", "bathroom", "kitchen", "elevator",
                       "office", "living_room", "train_interior", "car_interior"}
        _tags_norm = {t.strip().lower().replace(" ", "_") for t in tags}
        _has_win = any("window" in t.lower() for t in tags)
        if _tags_norm & _outdoor_mk:
            _rm = {"ceiling", "fluorescent_light", "wallpaper", "chandelier",
                   "carpet", "wooden_floor", "tile_floor"}
            tags = [t for t in tags if t.strip().lower().replace(" ", "_") not in _rm]
        if (_tags_norm & _indoor_mk) and not _has_win:
            _rm_out = {"sky", "cloud", "horizon"}
            if "open_air_bath" not in " ".join(tags).lower():
                tags = [t for t in tags if t.strip().lower().replace(" ", "_") not in _rm_out]

        # 1.6. 時間帯矛盾タグ自動除去（全シーンにnight/moonlight混入防止）
        _tags_norm_16 = {t.strip().lower().replace(" ", "_") for t in tags}
        _daytime_mk = {"morning", "sunrise", "daytime", "afternoon", "sunlight", "bright"}
        _nighttime_mk = {"night", "midnight", "late_night"}
        _night_tags_rm = {"moonlight", "darkness", "night_sky", "starlight", "dark", "night", "midnight", "late_night"}
        _day_tags_rm = {"sunlight", "bright_daylight", "blue_sky", "morning_light", "morning", "sunrise", "daytime", "afternoon"}
        _has_daytime = bool(_tags_norm_16 & _daytime_mk)
        _has_nighttime = bool(_tags_norm_16 & _nighttime_mk)
        if _has_daytime and not _has_nighttime:
            tags = [t for t in tags if t.strip().lower().replace(" ", "_") not in _night_tags_rm]
        elif _has_nighttime and not _has_daytime:
            tags = [t for t in tags if t.strip().lower().replace(" ", "_") not in _day_tags_rm]
        # 室内（window無し）+ moonlight → moonlight除去
        if (_tags_norm_16 & _indoor_mk) and not _has_win and not (_tags_norm_16 & _nighttime_mk):
            tags = [t for t in tags if t.strip().lower().replace(" ", "_") != "moonlight"]

        # 1.7. ユーザー指定の場所タイプ強制
        if location_type:
            _existing_lower_17 = {t.strip().lower().replace(" ", "_") for t in tags}
            if location_type == "indoors":
                tags = [t for t in tags if t.strip().lower() not in {"outdoors"}]
                if "indoors" not in _existing_lower_17:
                    tags.insert(min(2, len(tags)), "indoors")
            elif location_type == "outdoors":
                tags = [t for t in tags if t.strip().lower() not in {"indoors"}]
                if "outdoors" not in _existing_lower_17:
                    tags.insert(min(2, len(tags)), "outdoors")

        # 1.8. 品質/スタイル/LoRAタグ除去（v9.0: シーン固有タグのみ残す）
        _pre_clean_tags = []
        for t in tags:
            _t_norm = t.strip().lower().replace(" ", "_")
            # ウェイト付きタグからタグ名を抽出
            _t_inner = _re.sub(r'[()]', '', _t_norm).split(":")[0].strip()
            # 品質/スタイルタグ除去
            if _t_inner in _QUALITY_TAGS_TO_REMOVE:
                continue
            # LoRAタグ除去 (<lora:...>)
            if "<lora:" in t.lower():
                continue
            # score_N / score_N_up パターン除去
            if _re.match(r'^score_\d', _t_inner):
                continue
            _pre_clean_tags.append(t)
        tags = _pre_clean_tags

        # 1.9. Danbooruタグ正規化（非Danbooruタグ→正規タグ or 除去）
        _normalized_tags = []
        for t in tags:
            _t_norm19 = t.strip().lower().replace(" ", "_")
            _t_inner19 = _re.sub(r'[()]', '', _t_norm19).split(":")[0].strip()
            if _t_inner19 in _TAG_ALIAS_MAP:
                _alias = _TAG_ALIAS_MAP[_t_inner19]
                if _alias is None:
                    continue  # 非視覚タグ除去
                _normalized_tags.append(_alias)
            else:
                _normalized_tags.append(t)
        tags = _normalized_tags

        # 2. quality tags先頭確保（QUALITY_TAGS_DISABLED時は完全除去）
        if sd_quality_tags == QUALITY_TAGS_DISABLED:
            # カスタムモードで空欄→既存のquality tagsも除去
            _quality_kw = {"masterpiece", "best_quality", "best quality", "high_quality", "highres", "absurdres"}
            tags = [t for t in tags if not any(kw in t.lower() for kw in _quality_kw)]
        else:
            quality_found = any("masterpiece" in t.lower() or "best_quality" in t.lower() for t in tags)
            if not quality_found:
                effective_quality = sd_quality_tags if sd_quality_tags else QUALITY_POSITIVE_TAGS
                tags.insert(0, effective_quality)

        # 3. キャラタグ補完（上位タグが欠落していれば追加）
        if char_danbooru:
            existing = {t.strip().lower().replace(" ", "_") for t in tags}
            insert_at = 1  # quality tagsの直後
            for ct in char_danbooru[:8]:
                if ct.lower().replace(" ", "_") not in existing:
                    tags.insert(insert_at, ct)
                    insert_at += 1
                    existing.add(ct.lower().replace(" ", "_"))

        # 3.5. ナラティブ→SDタグ意味連携（description/mood等から視覚タグ抽出・注入）
        _narrative_tags = _extract_narrative_sd_tags(scene)
        if _narrative_tags:
            _existing_35 = {t.strip().lower().replace(" ", "_") for t in tags}
            _injected_35 = 0
            for nt in _narrative_tags:
                if nt not in _existing_35 and _injected_35 < 8:
                    tags.append(nt)
                    _existing_35.add(nt)
                    _injected_35 += 1

        # 3.6. 場所タグ動的解決（v9.3再実装: シーンのlocationに基づきテーマ場所タグを差替）
        if theme:
            _scene_loc_group = _resolve_scene_location_group(scene)
            if _scene_loc_group:
                # 現在のシーン場所と矛盾する他の場所グループのタグを除去
                _all_loc_tags_to_remove = set()
                for _grp_name, _grp_tags in _LOCATION_TAG_GROUPS.items():
                    if _grp_name != _scene_loc_group:
                        _all_loc_tags_to_remove |= _grp_tags
                # 現在の場所グループのタグは保持
                _current_grp_tags = _LOCATION_TAG_GROUPS.get(_scene_loc_group, set())
                _all_loc_tags_to_remove -= _current_grp_tags
                if _all_loc_tags_to_remove:
                    tags = [t for t in tags
                            if t.strip().lower().replace(" ", "_") not in _all_loc_tags_to_remove]
                # テーマ固有の場所詳細タグを注入
                _detail_map = _THEME_LOCATION_DETAIL_TAGS.get(theme, {})
                _detail_tags = _detail_map.get(_scene_loc_group, _detail_map.get("default", []))
                if _detail_tags:
                    _existing_36 = {t.strip().lower().replace(" ", "_") for t in tags}
                    for _dt in _detail_tags:
                        if _dt not in _existing_36:
                            tags.append(_dt)
                            _existing_36.add(_dt)

        # 4. 照明タグ追加
        has_light = any(any(lw in t.lower() for lw in LIGHTING_WORDS) for t in tags)
        if not has_light:
            if any(kw in sd_lower for kw in ("night", "dark", "evening")):
                tags.append("dim_lighting")
            elif any(kw in sd_lower for kw in ("morning", "sunrise", "dawn")):
                tags.append("soft_lighting")
            elif any(kw in sd_lower for kw in ("sunset", "dusk", "golden")):
                tags.append("warm_lighting")
            else:
                tags.append("natural_lighting")

        # 4.3. ユーザー指定の時間帯タグ注入
        if time_tags:
            _time_related = {"morning", "sunrise", "morning_light", "daytime", "sunlight",
                             "afternoon", "evening", "sunset", "golden_hour", "night",
                             "moonlight", "dim_lighting", "bright"}
            _existing_lower_43 = {t.strip().lower().replace(" ", "_") for t in tags}
            tags = [t for t in tags if t.strip().lower().replace(" ", "_") not in _time_related]
            _existing_lower_43 -= _time_related
            for tt in time_tags.split(","):
                tt = tt.strip()
                if tt and tt.lower() not in _existing_lower_43:
                    tags.append(tt)
                    _existing_lower_43.add(tt.lower())

        # 4.2. 背景タグ保証（sd_promptに背景系タグが無い場合、locationから補完）
        _bg_kw = {
            "outdoors", "indoors",
            # 学校
            "classroom", "library", "gym", "hallway", "stairwell",
            "locker_room", "infirmary", "rooftop", "club_room",
            "storage_room", "school",
            # 住居
            "bedroom", "bathroom", "kitchen", "living_room",
            "japanese_room", "balcony", "basement", "study",
            "entrance", "closet", "garage",
            # 商業・仕事
            "office", "elevator", "warehouse", "factory",
            "convenience_store", "store",
            # 宿泊
            "hotel_room", "ryokan_room", "inn_room", "cabin",
            # 飲食
            "cafe", "restaurant", "izakaya", "bar", "cafeteria",
            # 交通
            "car_interior", "train_interior", "bus_interior",
            "airplane_interior", "ship_interior", "train_station",
            # 娯楽
            "karaoke_room", "internet_cafe", "arcade", "theater", "studio",
            # 屋外・自然
            "park", "forest", "beach", "mountain", "river", "lake",
            "garden", "alley", "bridge", "riverbank", "field",
            "grassland", "cliff", "cave",
            # 風呂・温泉
            "onsen", "bath", "pool", "open_air_bath", "bathhouse", "sauna",
            # 宗教
            "shrine", "temple", "church", "graveyard",
            # ファンタジー
            "dungeon", "castle", "tower", "prison", "tavern", "throne_room",
            # SF
            "spaceship_interior", "laboratory", "cockpit",
            # 日本建築
            "engawa", "storehouse", "barn",
        }
        _exist_low = {t.strip().lower().replace(" ", "_") for t in tags}
        if not (_exist_low & _bg_kw):
            _location = scene.get("location_detail", scene.get("location", ""))
            if _location:
                _loc_map = {
                    # 学校系
                    "教室": "classroom", "保健室": "infirmary", "体育": "gym",
                    "部室": "club_room", "屋上": "rooftop", "図書": "library",
                    "廊下": "hallway", "階段": "stairwell", "トイレ": "bathroom",
                    "更衣": "locker_room", "プール": "pool", "校庭": "outdoors",
                    "準備室": "storage_room", "職員室": "office",
                    # 住居系
                    "寝室": "bedroom", "リビング": "living_room", "台所": "kitchen",
                    "浴室": "bathroom", "風呂": "bathroom", "玄関": "entrance",
                    "和室": "japanese_room", "子供部屋": "bedroom", "書斎": "study",
                    "ベランダ": "balcony", "押し入れ": "closet", "地下室": "basement",
                    "洗面": "bathroom", "脱衣": "locker_room", "トイレ": "bathroom",
                    "物置": "storage_room", "ガレージ": "garage",
                    # 日本建築系
                    "畳": "japanese_room", "障子": "japanese_room",
                    "縁側": "engawa", "土間": "dirt_floor",
                    "蔵": "storehouse", "納屋": "barn",
                    # 宿泊系
                    "ホテル": "hotel_room", "旅館": "ryokan_room", "ラブホ": "hotel_room",
                    "民宿": "inn_room", "ペンション": "hotel_room", "コテージ": "cabin",
                    # 飲食系
                    "カフェ": "cafe", "居酒屋": "izakaya", "レストラン": "restaurant",
                    "バー": "bar", "食堂": "cafeteria", "ファミレス": "restaurant",
                    "キッチン": "kitchen",
                    # 商業施設系
                    "オフィス": "office", "会議室": "office", "エレベータ": "elevator",
                    "ビル": "office", "倉庫": "warehouse", "工場": "factory",
                    "コンビニ": "convenience_store", "スーパー": "store",
                    "デパート": "store", "ショッピング": "store",
                    # 交通系
                    "車": "car_interior", "電車": "train_interior", "バス": "bus_interior",
                    "タクシー": "car_interior", "飛行機": "airplane_interior",
                    "船": "ship_interior", "駅": "train_station",
                    # 娯楽系
                    "カラオケ": "karaoke_room", "ネットカフェ": "internet_cafe",
                    "ゲームセンター": "arcade", "映画館": "theater",
                    "ボウリング": "bowling_alley", "スタジオ": "studio",
                    # 屋外・自然系
                    "公園": "park", "森": "forest", "海": "beach",
                    "山": "mountain", "川": "river", "湖": "lake",
                    "庭": "garden", "路地": "alley", "橋": "bridge",
                    "河原": "riverbank", "野原": "field", "草原": "grassland",
                    "崖": "cliff", "洞窟": "cave", "砂浜": "beach",
                    # 宗教・文化系
                    "神社": "shrine", "寺": "temple", "教会": "church",
                    "墓地": "graveyard", "鳥居": "torii",
                    # 温泉・風呂系
                    "温泉": "onsen", "露天風呂": "open_air_bath",
                    "銭湯": "bathhouse", "サウナ": "sauna",
                    # ファンタジー系
                    "ダンジョン": "dungeon", "城": "castle", "塔": "tower",
                    "牢獄": "prison", "酒場": "tavern", "宿屋": "inn_room",
                    "玉座": "throne_room",
                    # SF系
                    "宇宙船": "spaceship_interior", "研究所": "laboratory",
                    "実験室": "laboratory", "コックピット": "cockpit",
                }
                _added = False
                for _jp, _en in _loc_map.items():
                    if _jp in _location:
                        tags.append(_en)
                        _added = True
                        break
                if not _added:
                    tags.append("indoors")

        # 4.5. 男性タグ注入（v9.0: faceless_male デフォルトON / intensity制限撤廃）
        _MULTI_MALE_THEMES = {"gangbang"}
        intensity = scene.get("intensity", 0)
        existing_lower = {t.strip().lower().replace(" ", "_") for t in tags}
        if theme not in _MULTI_MALE_THEMES:
            # 通常テーマ: 1boy 常時付与
            if "1boy" not in existing_lower:
                tags.append("1boy")
                existing_lower.add("1boy")
            # faceless_male: フラグに応じて付与
            if faceless_male and "faceless_male" not in existing_lower:
                tags.append("faceless_male")
                existing_lower.add("faceless_male")
        else:
            # gangbang: 1boyは付与しない（THEME_GUIDESからmultiple_boysが来る）
            if faceless_male and "faceless_male" not in existing_lower:
                tags.append("faceless_male")
                existing_lower.add("faceless_male")
        if male_tags:
            # ユーザー指定の男性体型タグを高重要度で注入
            for mt in male_tags.split(","):
                mt = mt.strip()
                if mt and mt.lower().replace(" ", "_") not in existing_lower:
                    if mt.lower() in ("1boy", "faceless_male"):
                        continue  # 既に付与済み
                    weighted_mt = f"({mt}:1.3)"
                    tags.append(weighted_mt)
                    existing_lower.add(mt.lower().replace(" ", "_"))

        # 4.55. 男性体型タグ（ユーザー指定がない場合のデフォルト）
        if not male_tags and "1boy" in existing_lower:
            _male_body_defaults = ["muscular_male", "veiny_arms"]
            for mt in _male_body_defaults:
                if mt not in existing_lower:
                    tags.append(mt)
                    existing_lower.add(mt)

        # 4.6. intensity別 表情・身体反応タグ自動注入（プールからサンプリング）
        # Phase6: 性格別表情バイアス適用
        if intensity >= 1:
            _expr_pool = _INTENSITY_EXPRESSION_MAP.get(min(intensity, 5), [])
            existing_lower = {t.strip().lower().replace(" ", "_") for t in tags}

            # Phase6: 性格バイアス - suppressタグを除外、boostタグを優先注入
            _bias = _PERSONALITY_EXPRESSION_BIAS.get(_char_personality, {})
            _suppress = _bias.get("suppress", set())
            _boost = _bias.get("boost", [])

            # suppressタグを候補から除外
            _avail = [et for et in _expr_pool if et not in existing_lower and et not in _suppress]
            # boostタグを先頭に追加（重複除外）
            _boost_avail = [bt for bt in _boost if bt not in existing_lower and bt not in _avail]
            _avail = _boost_avail + _avail

            # intensity 1-2: 2-3個, intensity 3: 4-5個, 4-5: 5-6個
            _max_inject = {1: 2, 2: 3, 3: 5, 4: 6, 5: 6}.get(min(intensity, 5), 4)
            import random as _rnd_expr
            # boostタグは確定注入、残りをランダムサンプリング
            _boost_inject = [bt for bt in _boost_avail if bt not in existing_lower][:2]
            _remaining = [et for et in _avail if et not in _boost_inject]
            _remaining_count = max(0, _max_inject - len(_boost_inject))
            _random_pick = _rnd_expr.sample(_remaining, min(len(_remaining), _remaining_count))
            _selected = _boost_inject + _random_pick

            for et in _selected:
                tags.append(et)
                existing_lower.add(et)

        # 4.65. セリフ→SDプロンプト連動（v9.0: bubblesの内容からSDタグ自動注入）
        _BUBBLE_MOAN_SD = {
            "open_mouth": ["あぁ", "はぁ", "んぁ", "ああ", "あっあっ", "んほ"],
            "tongue_out": ["んほ", "あへ", "れろ", "舌"],
            "drooling": ["じゅる", "れろ", "んほ", "あへ"],
            "tears": ["いや", "痛", "泣", "うっ"],
            "rolling_eyes": ["んほ", "あへ", "いぐ", "壊れ"],
            "ahegao": ["壊れ", "いぐ", "んほぉ", "あへ"],
        }
        _BUBBLE_THOUGHT_SD = {
            "trembling": ["怖", "震", "ビクビク", "ゾクゾク"],
            "blush": ["恥ず", "は、恥", "見ないで", "やだ"],
            "crying": ["泣き", "涙", "うっ…"],
            "heart_pupils": ["好き", "もっと", "気持ちい", "離さないで"],
            "dazed": ["頭", "ぼんやり", "真っ白", "何も考え"],
        }
        _BUBBLE_SPEECH_SD = {
            "covering_face": ["見ないで", "恥ずかし", "やめて"],
            "looking_away": ["あっち", "見ないで", "は、恥"],
            "clenched_teeth": ["くっ", "ぐっ", "耐え"],
        }
        bubbles = scene.get("bubbles", [])
        _bubble_inject = set()
        for _bub in bubbles:
            _btext = _bub.get("text", "")
            _btype = _bub.get("type", "")
            if _btype == "moan":
                for _sd_tag, _kws in _BUBBLE_MOAN_SD.items():
                    if any(kw in _btext for kw in _kws):
                        _bubble_inject.add(_sd_tag)
            elif _btype == "thought":
                for _sd_tag, _kws in _BUBBLE_THOUGHT_SD.items():
                    if any(kw in _btext for kw in _kws):
                        _bubble_inject.add(_sd_tag)
            elif _btype == "speech":
                for _sd_tag, _kws in _BUBBLE_SPEECH_SD.items():
                    if any(kw in _btext for kw in _kws):
                        _bubble_inject.add(_sd_tag)
        if _bubble_inject:
            existing_lower = {t.strip().lower().replace(" ", "_") for t in tags}
            for _bt in _bubble_inject:
                if _bt not in existing_lower:
                    tags.append(_bt)
                    existing_lower.add(_bt)

        # 4.7. intensity別 衣装状態タグ自動注入 + 累積脱衣トラッキング（v9.0）
        existing_lower = {t.strip().lower().replace(" ", "_") for t in tags}

        # 現在のシーンの脱衣レベルを検出
        _cur_undress = 0
        for _ct_tag in existing_lower:
            _lv = _CLOTHING_UNDRESS_LEVEL.get(_ct_tag, 0)
            if _lv > _cur_undress:
                _cur_undress = _lv

        _clothing_tags = CLOTHING_ESCALATION.get(min(intensity, 5), [])
        if _clothing_tags:
            _has_nude = existing_lower & {"nude", "naked", "completely_nude"}
            if not (intensity >= 5 and _has_nude):
                for ct in _clothing_tags:
                    if ct not in existing_lower:
                        tags.append(ct)
                        existing_lower.add(ct)
                        _lv = _CLOTHING_UNDRESS_LEVEL.get(ct, 0)
                        if _lv > _cur_undress:
                            _cur_undress = _lv
                        break

        # v9.0: 累積脱衣状態の引き継ぎ（脱衣逆行防止）
        # 前シーンで脱衣レベルが高かった場合、現シーンでも最低その状態を維持
        if _max_undress_level >= 5 and _cur_undress < 5:
            # 前シーンでnudeだった → 現シーンもnude（服復活禁止）
            if "nude" not in existing_lower and "naked" not in existing_lower and "completely_nude" not in existing_lower:
                tags.append("nude")
                existing_lower.add("nude")
        elif _max_undress_level >= 4 and _cur_undress < 3:
            # 前シーンでtopless/panties_only等 → 最低でもpartially_undressed
            if not (existing_lower & {"topless", "bottomless", "panties_only", "naked_shirt", "stockings_only", "nude", "naked"}):
                tags.append("partially_undressed")
                existing_lower.add("partially_undressed")

        # 脱衣レベルを更新（単調増加）
        _max_undress_level = max(_max_undress_level, _cur_undress)

        # 累積脱衣タグ更新
        _accumulated_clothing.update(existing_lower & set(_CLOTHING_UNDRESS_LEVEL.keys()))

        # 4.8. intensity別 体液進行タグ自動注入 + 累積体液トラッキング（v9.0）
        _fluid_tags = FLUID_PROGRESSION.get(min(intensity, 5), [])
        existing_lower = {t.strip().lower().replace(" ", "_") for t in tags}

        # v9.0: 前シーンの持続性体液タグを引き継ぎ（射精後は体液残留）
        if _accumulated_fluids:
            _injected_persistent = 0
            for pf in _accumulated_fluids:
                if pf not in existing_lower and _injected_persistent < 2:
                    tags.append(pf)
                    existing_lower.add(pf)
                    _injected_persistent += 1

        if _fluid_tags:
            _injected_fluid = 0
            for ft in _fluid_tags:
                if ft not in existing_lower and _injected_fluid < 2:
                    tags.append(ft)
                    existing_lower.add(ft)
                    _injected_fluid += 1

        # 累積体液タグ更新（持続性タグのみ蓄積）
        _cur_persistent = existing_lower & _PERSISTENT_FLUID_TAGS
        _accumulated_fluids.update(_cur_persistent)

        # 4.9. カメラアングル/構図タグ自動注入 + 3連続同一アングル防止（v9.0強化）
        import random as _rnd_angle
        _exist_angles = {t.strip().lower().replace(" ", "_") for t in tags}
        # 既存のアングルタグを検出
        _cur_angle = None
        for t in tags:
            _t_norm = t.strip().lower().replace(" ", "_")
            if _t_norm in _ANGLE_TAGS:
                _cur_angle = _t_norm
                break

        if _cur_angle is None and intensity >= 2:
            # アングルタグがない → intensityに応じたプールから注入
            _angle_pool = _INTENSITY_ANGLE_MAP.get(min(intensity, 5),
                          _CAMERA_ANGLE_POOL.get(min(intensity, 5), []))
            _candidates = [a for a in _angle_pool if a != _prev_camera_angle]
            if not _candidates:
                _candidates = list(_angle_pool)
            if _candidates:
                _pick_idx = _camera_scene_idx % len(_candidates)
                _cur_angle = _candidates[_pick_idx]
                tags.append(_cur_angle)
                _camera_scene_idx += 1

        # 3連続同一アングル検出 → 3番目以降を代替に差し替え
        if _cur_angle and len(_angle_history) >= 2 and _angle_history[-1] == _cur_angle and _angle_history[-2] == _cur_angle:
            _alt_pool = _INTENSITY_ANGLE_MAP.get(min(intensity, 5),
                        _CAMERA_ANGLE_POOL.get(min(intensity, 5), []))
            _alt_candidates = [a for a in _alt_pool if a != _cur_angle]
            if not _alt_candidates:
                _alt_candidates = [a for a in _ANGLE_TAGS if a != _cur_angle]
            if _alt_candidates:
                _replacement = _rnd_angle.choice(_alt_candidates)
                tags = [(_replacement if t.strip().lower().replace(" ", "_") == _cur_angle else t) for t in tags]
                _cur_angle = _replacement

        _prev_camera_angle = _cur_angle or _prev_camera_angle
        _angle_history.append(_cur_angle or "")

        # 4.95. 被写界深度自動制御（Phase7: close-up/portrait時に depth_of_field 追加）
        _existing_495 = {t.strip().lower().replace(" ", "_") for t in tags}
        _closeup_kw = {"close-up", "portrait", "upper_body", "face_focus"}
        if (_existing_495 & _closeup_kw) and "depth_of_field" not in _existing_495:
            tags.append("depth_of_field")

        # 5. 設定スタイル適用（タグ置換・禁止・追加）
        if setting_style:
            replace_map = setting_style.get("replace", {})
            prohibit = setting_style.get("prohibit", set())
            append_tags = setting_style.get("append", [])

            new_tags = []
            for t in tags:
                norm = t.strip().lower().replace(" ", "_")
                # ウェイト付きタグからnorm抽出
                inner = _re.sub(r'[()]', '', norm).split(":")[0].strip()
                # 禁止タグ除去
                if inner in prohibit:
                    continue
                # タグ置換
                if inner in replace_map:
                    new_tags.append(replace_map[inner])
                else:
                    new_tags.append(t)
            tags = new_tags

            # 雰囲気タグ追加（未存在のもののみ）
            existing_norm = {t.strip().lower().replace(" ", "_") for t in tags}
            for at in append_tags:
                if at.lower().replace(" ", "_") not in existing_norm:
                    tags.append(at)
                    existing_norm.add(at.lower().replace(" ", "_"))

        # 5.3. テーマ×服装/場所タグ矛盾除去
        if theme:
            # 共通カテゴリ定義
            _SCHOOL_TAGS = {"school_uniform", "sailor_uniform", "serafuku", "blazer",
                            "plaid_skirt", "gym_uniform", "buruma", "school_swimsuit",
                            "school", "classroom", "school_bag"}
            _OFFICE_TAGS = {"business_suit", "pencil_skirt", "office_lady", "office"}
            _FANTASY_TAGS = {"armor", "gauntlets", "breastplate", "medieval", "fantasy"}
            _SHRINE_TAGS = {"miko", "hakama"}
            _MAID_TAGS = {"maid", "maid_headdress"}
            _NURSE_TAGS = {"nurse", "nurse_cap"}

            _THEME_CLOTHING_CONFLICTS = {
                # OL → 学校/ファンタジー/神社系除去（blazerはOLも着るが学校系と
                # セットで来た場合plaid_skirt等が問題なので一括除去）
                "office": (_SCHOOL_TAGS | _FANTASY_TAGS | _SHRINE_TAGS)
                          - {"blazer"},  # OLのblazerは許容
                # 先生・生徒 → OL系除去（スーツは先生が着るので許容）
                "teacher_student": {"office_lady"},
                # 異世界 → 現代系全般除去
                "isekai": (_SCHOOL_TAGS | _OFFICE_TAGS | _NURSE_TAGS
                           | {"modern", "neon", "smartphone"}),
                # 温泉 → 学校/OL/ファンタジー除去
                "onsen": (_SCHOOL_TAGS | _OFFICE_TAGS | _FANTASY_TAGS),
                # 水着 → OL/ファンタジー/神社除去
                "swimsuit": (_OFFICE_TAGS | _FANTASY_TAGS | _SHRINE_TAGS),
                # 医療 → 学校/ファンタジー/メイド/神社除去
                "medical": (_SCHOOL_TAGS | _FANTASY_TAGS | _MAID_TAGS | _SHRINE_TAGS),
                # メイド → 学校/OL/ファンタジー除去
                "maid": (_SCHOOL_TAGS | _OFFICE_TAGS | _FANTASY_TAGS
                         | {"serafuku", "buruma", "pencil_skirt"}),
                # 痴漢 → ファンタジー/神社除去（服装は維持=通学/通勤中あり得る）
                "chikan": (_FANTASY_TAGS | _SHRINE_TAGS
                           | {"classroom", "school", "gym"}),
                # モンスター → 現代系除去
                "monster": (_OFFICE_TAGS | _NURSE_TAGS
                            | {"school_uniform", "classroom", "school"}),
                # 触手 → 現代OL系除去
                "tentacle": (_OFFICE_TAGS | {"pencil_skirt",
                             "school_uniform", "classroom", "school"}),
                # スポーツ → OL/ファンタジー/神社除去
                "sports": (_OFFICE_TAGS | _FANTASY_TAGS | _SHRINE_TAGS),
                # アイドル → 体操着/ファンタジー/神社除去
                "idol": (_FANTASY_TAGS | _SHRINE_TAGS
                         | {"gym_uniform", "buruma"}),
                # 風俗 → 学校/ファンタジー除去
                "prostitution": (_SCHOOL_TAGS | _FANTASY_TAGS
                                 | {"serafuku", "buruma"}),
                # 睡眠 → 学校/OL/ファンタジー除去
                "sleep": (_SCHOOL_TAGS | _OFFICE_TAGS | _FANTASY_TAGS
                          | {"sailor_uniform", "gym_uniform", "buruma"}),
                # 近親 → ファンタジー除去
                "incest": (_FANTASY_TAGS | {"classroom", "school", "office"}),
                # 隣人 → 学校/ファンタジー除去
                "neighbor": (_SCHOOL_TAGS | _FANTASY_TAGS
                             | {"office", "gym"}),
            }
            _conflicts = _THEME_CLOTHING_CONFLICTS.get(theme, set())
            if _conflicts:
                # Phase6: キャラ保護タグは除去対象外
                _safe_conflicts = _conflicts - _protected_char_tags
                tags = [t for t in tags
                        if t.strip().lower().replace(" ", "_") not in _safe_conflicts]

        # 5.55. 体位サポートタグ注入（体位に必須の視覚タグを補完）
        _tags_lower_555 = {t.strip().lower().replace(" ", "_") for t in tags}
        for _pos_tag, _support_list in _POSITION_SUPPORT_TAGS.items():
            if _pos_tag in _tags_lower_555:
                _injected_sup = 0
                for _sup in _support_list:
                    if _sup not in _tags_lower_555 and _injected_sup < 2:
                        tags.append(_sup)
                        _tags_lower_555.add(_sup)
                        _injected_sup += 1

        # 5.5. 矛盾する体位/行為タグの相互排他チェック
        # 同時に成立しない行為の組み合わせを検出し、descriptionに近い方を残す
        _CONTRADICTORY_PAIRS = [
            # (グループA, グループB): A+Bが同時に存在したらBを除去
            ({"blowjob", "fellatio", "deepthroat", "oral"}, {"sitting_on_lap", "cowgirl_position", "reverse_cowgirl", "missionary", "vaginal"}),
            ({"paizuri"}, {"doggy_style", "from_behind", "all_fours", "missionary", "standing_sex"}),
            ({"handjob"}, {"vaginal", "penetration", "sex"}),
            # v7.0追加
            ({"anal", "anal_sex"}, {"vaginal", "pussy"}),
            ({"sleeping", "asleep"}, {"standing", "standing_sex", "walking"}),
            ({"bondage", "tied_up", "restrained"}, {"hugging", "holding_hands", "hand_on_hip"}),
            ({"clothed_sex"}, {"completely_nude", "nude", "naked"}),
            ({"underwater"}, {"indoors", "bedroom", "classroom"}),
            # v8.6: multiple_boys + 1boy矛盾（gangbangテーマで混入防止）
            ({"multiple_boys"}, {"1boy"}),
            # Phase2追加: 体位矛盾ペア
            ({"standing_sex", "standing"}, {"lying", "on_back", "on_bed"}),
            ({"missionary"}, {"standing", "standing_sex"}),
            ({"cowgirl_position"}, {"on_stomach", "prone"}),
            ({"prone_bone"}, {"on_back", "cowgirl_position", "girl_on_top"}),
            ({"kneeling"}, {"sitting", "chair"}),
            ({"spooning"}, {"standing_sex", "against_wall"}),
        ]
        _tags_lower_set = {t.strip().lower().replace(" ", "_") for t in tags}
        for group_a, group_b in _CONTRADICTORY_PAIRS:
            has_a = _tags_lower_set & group_a
            has_b = _tags_lower_set & group_b
            if has_a and has_b:
                # グループAの行為が明示的にある場合、矛盾するグループBを除去
                desc_lower = scene.get("description", "").lower()
                a_in_desc = any(k in desc_lower for k in ["フェラ", "口", "咥", "パイズリ", "手コキ", "blowjob", "oral", "paizuri"])
                if a_in_desc or len(has_a) >= len(has_b):
                    tags = [t for t in tags if t.strip().lower().replace(" ", "_") not in has_b]
                else:
                    tags = [t for t in tags if t.strip().lower().replace(" ", "_") not in has_a]

        # 5.6. aftermathシーンのSDタグ整合性チェック
        _desc_for_ctx = (scene.get("description", "") + " " + scene.get("title", "")
                         + " " + scene.get("mood", "")).lower()
        _aftermath_ctx_kw = ["事後", "余韻", "虚脱", "終えた", "身繕い", "動けない", "虚ろ",
                             "後片付け", "放心", "脱力", "ぐったり", "呆然", "立てない",
                             "aftermath", "afterglow"]
        if any(k in _desc_for_ctx for k in _aftermath_ctx_kw):
            # 事後シーンからsex/penetration等のアクティブな行為タグを除去
            _sex_tags_to_remove = {"sex", "vaginal", "penetration", "thrusting",
                                   "piston", "hip_thrust", "grabbing_hips"}
            tags = [t for t in tags if t.strip().lower().replace(" ", "_") not in _sex_tags_to_remove]
            # aftermath固有タグを追加
            _existing = {t.strip().lower().replace(" ", "_") for t in tags}
            for at_tag in ["exhausted", "dazed", "afterglow"]:
                if at_tag not in _existing:
                    tags.append(at_tag)
                    _existing.add(at_tag)

        # 5.7. mood↔表情矛盾検出・修正
        _mood_text = scene.get("mood", "")
        if _mood_text:
            _tags_lower_57 = {t.strip().lower().replace(" ", "_") for t in tags}
            for _rule_name, _rule in _MOOD_EXPRESSION_CONFLICTS.items():
                _mood_match = any(mk in _mood_text for mk in _rule["mood_kw"])
                if _mood_match and _rule["conflict_tag"] in _tags_lower_57:
                    # 矛盾タグを除去
                    tags = [t for t in tags if t.strip().lower().replace(" ", "_") != _rule["conflict_tag"]]
                    _tags_lower_57.discard(_rule["conflict_tag"])
                    # 代替タグを注入
                    for _repl in _rule["replace_with"]:
                        if _repl not in _tags_lower_57:
                            tags.append(_repl)
                            _tags_lower_57.add(_repl)

        # 5.8. アクセサリ永続化（Phase6: 初回検出後、全シーンで維持）
        _existing_58 = {t.strip().lower().replace(" ", "_") for t in tags}
        # 現在のシーンのアクセサリを検出・蓄積
        _cur_accessories = _existing_58 & _ACCESSORY_PERSISTENT_TAGS
        _detected_accessories.update(_cur_accessories)
        # 過去のシーンで検出されたアクセサリを注入
        _injected_acc = 0
        for _acc_tag in _detected_accessories:
            if _acc_tag not in _existing_58 and _injected_acc < 3:
                tags.append(_acc_tag)
                _existing_58.add(_acc_tag)
                _injected_acc += 1

        # 5.9. タグ順序最適化（Phase3: SDモデルは先頭タグほど影響大）
        tags = _reorder_sd_tags(tags, intensity, char_danbooru)

        # 6. 重要タグにウェイト付加（未ウェイトのもののみ）
        # Phase3: intensity連動ウェイト強化
        _boost_map = _INTENSITY_WEIGHT_BOOST.get(min(intensity, 5), {})
        weighted = []
        for tag in tags:
            norm = tag.strip().lower().replace(" ", "_").strip("()")
            # 既にウェイト付きならスキップ
            if ":" in tag and "(" in tag:
                weighted.append(tag)
                continue
            # Phase3: intensity連動ブースト（既存ウェイトより優先）
            if norm in _boost_map:
                weighted.append(f"({tag}:{_boost_map[norm]})")
            elif norm in WEIGHT_EXPRESSION:
                weighted.append(f"({tag}:1.3)")
            elif norm in WEIGHT_ACTION:
                weighted.append(f"({tag}:1.2)")
            else:
                weighted.append(tag)

        # 7. 体位タグ自動調整（前シーンと同一体位を代替に置換）
        # intensityは step 4.5 で取得済み
        _cur_positions = set()
        adjusted = []
        for tag in weighted:
            # ウェイト付きタグからタグ名を抽出
            _inner = _re.sub(r'[()]', '', tag).split(":")[0].strip().lower().replace(" ", "_")
            if _inner in POSITION_TAGS:
                _cur_positions.add(_inner)
                if _inner in _prev_scene_positions:
                    # 前シーンと同じ体位 → 代替に置換
                    fallbacks = POSITION_FALLBACKS.get(_inner, [])
                    # intensity連動: 高intensityでは激しい体位を優先
                    preferred = _POSITION_INTENSITY_PREFERENCE.get(min(intensity, 5), set())
                    replacement = None
                    # 優先度の高い体位から選択
                    for fb in fallbacks:
                        if fb in preferred and fb not in _prev_scene_positions:
                            replacement = fb
                            break
                    # 優先に該当なければフォールバックの先頭から
                    if not replacement:
                        for fb in fallbacks:
                            if fb not in _prev_scene_positions:
                                replacement = fb
                                break
                    if replacement:
                        _cur_positions.discard(_inner)
                        _cur_positions.add(replacement)
                        adjusted.append(replacement)
                        continue
            adjusted.append(tag)
        weighted = adjusted
        _prev_scene_positions = _cur_positions

        main_prompt = deduplicate_sd_tags(", ".join(weighted))
        # prefix/suffix注入（LoRAタグ等はdeduplicateに通さない）
        parts = []
        if sd_prefix_tags:
            parts.append(sd_prefix_tags)
        parts.append(main_prompt)
        if sd_suffix_tags:
            parts.append(sd_suffix_tags)
        scene["sd_prompt"] = ", ".join(parts).replace(",,", ",").strip(", ")

    # 8. 体位分布リバランス（spread_legsが40%超過→一部を代替体位に自動置換）
    import re as _re8
    total = len(results)
    if total >= 8:  # 8シーン以上のスクリプトのみ
        pos_counter = {}
        pos_scene_map = {}  # tag → [scene_indices]
        for idx, sc in enumerate(results):
            prompt = sc.get("sd_prompt", "")
            for t in prompt.split(","):
                norm = t.strip().lower().replace(" ", "_")
                norm = _re8.sub(r'[()]', '', norm).split(":")[0].strip()
                if norm in POSITION_TAGS:
                    pos_counter[norm] = pos_counter.get(norm, 0) + 1
                    pos_scene_map.setdefault(norm, []).append(idx)
        # 40%超過の体位を検出
        _SPREAD_ALTERNATIVES = [
            "legs_apart", "legs_up", "open_legs", "straddling",
            "cowgirl_position", "reverse_cowgirl", "missionary", "mating_press"
        ]
        for ptag, cnt in pos_counter.items():
            if cnt / total >= 0.40:
                target_count = int(total * 0.30)  # 30%以下に削減
                excess = cnt - target_count
                if excess <= 0:
                    continue
                scenes_with = pos_scene_map.get(ptag, [])
                # 奇数indexのシーンから優先的に置換（偶数は維持）
                replace_targets = [i for i in scenes_with if i % 2 == 1][:excess]
                if len(replace_targets) < excess:
                    replace_targets.extend([i for i in scenes_with if i % 2 == 0 and i not in replace_targets][:excess - len(replace_targets)])
                alt_idx = 0
                fallbacks_for_tag = POSITION_FALLBACKS.get(ptag, _SPREAD_ALTERNATIVES)
                for sidx in replace_targets[:excess]:
                    alt = fallbacks_for_tag[alt_idx % len(fallbacks_for_tag)]
                    alt_idx += 1
                    sc = results[sidx]
                    old_prompt = sc.get("sd_prompt", "")
                    # タグ置換
                    new_tags = []
                    replaced = False
                    for t in old_prompt.split(","):
                        norm = t.strip().lower().replace(" ", "_")
                        norm_clean = _re8.sub(r'[()]', '', norm).split(":")[0].strip()
                        if norm_clean == ptag and not replaced:
                            new_tags.append(alt)
                            replaced = True
                        else:
                            new_tags.append(t.strip())
                    sc["sd_prompt"] = deduplicate_sd_tags(", ".join(new_tags))
                log_message(f"体位リバランス: {ptag} {cnt}/{total}({cnt*100//total}%)→{cnt-len(replace_targets[:excess])}/{total}に削減")

    # 8.5. 構図バランス監視（Phase7: close-up系が60%超→一部をfull_body/cowboy_shotに差替え）
    import re as _re_comp
    _closeup_tags = {"close-up", "portrait", "upper_body", "face_focus"}
    _wideshot_alternatives = ["cowboy_shot", "full_body", "wide_shot", "from_above"]
    if total >= 8:
        _closeup_count = 0
        _closeup_scenes = []
        for idx, sc in enumerate(results):
            prompt = sc.get("sd_prompt", "")
            _p_tags = {_re_comp.sub(r'[()]', '', t.strip().lower().replace(" ", "_")).split(":")[0].strip()
                       for t in prompt.split(",") if t.strip()}
            if _p_tags & _closeup_tags:
                _closeup_count += 1
                _closeup_scenes.append(idx)
        if total > 0 and _closeup_count / total > 0.60:
            _target = int(total * 0.45)
            _excess = _closeup_count - _target
            _replace_idx = [i for i in _closeup_scenes if i % 3 == 0][:_excess]
            _alt_i = 0
            for sidx in _replace_idx:
                sc = results[sidx]
                old_prompt = sc.get("sd_prompt", "")
                new_tags = []
                replaced = False
                for t in old_prompt.split(","):
                    _norm = _re_comp.sub(r'[()]', '', t.strip().lower().replace(" ", "_")).split(":")[0].strip()
                    if _norm in _closeup_tags and not replaced:
                        new_tags.append(_wideshot_alternatives[_alt_i % len(_wideshot_alternatives)])
                        _alt_i += 1
                        replaced = True
                    else:
                        new_tags.append(t.strip())
                if replaced:
                    sc["sd_prompt"] = deduplicate_sd_tags(", ".join(new_tags))

    # 最終パス: 体位/アングル連続重複の再チェック（Step 8で再導入された可能性対応）
    import re as _re_final
    _prev_pos_f = set()
    for scene in results:
        sd = scene.get("sd_prompt", "")
        if not sd:
            _prev_pos_f = set()
            continue
        tags = [t.strip() for t in sd.split(",") if t.strip()]
        _cur_pos_f = set()
        new_tags = []
        changed = False
        for tag in tags:
            _inner = _re_final.sub(r'[()]', '', tag).split(":")[0].strip().lower().replace(" ", "_")
            if _inner in POSITION_TAGS:
                _cur_pos_f.add(_inner)
                if _inner in _prev_pos_f:
                    fallbacks = POSITION_FALLBACKS.get(_inner, [])
                    replacement = None
                    for fb in fallbacks:
                        if fb not in _prev_pos_f:
                            replacement = fb
                            break
                    if replacement:
                        _cur_pos_f.discard(_inner)
                        _cur_pos_f.add(replacement)
                        new_tags.append(replacement)
                        changed = True
                        continue
            new_tags.append(tag)
        if changed:
            scene["sd_prompt"] = ", ".join(new_tags)
        _prev_pos_f = _cur_pos_f

    # Phase4: シーン固有ネガティブプロンプト生成
    # ユーザー設定のネガティブベース（指定があれば自動生成のベース部分を置換）
    for scene in results:
        if scene.get("sd_prompt"):
            auto_neg = _generate_negative_prompt(scene, theme)
            if sd_neg_base:
                # ユーザー指定のベースを使用し、自動生成のシーン固有部分のみ追加
                # 自動生成からベース品質タグを除いたシーン固有分を抽出
                _auto_base = {
                    "worst_quality", "low_quality", "bad_anatomy", "bad_hands",
                    "missing_fingers", "extra_digits", "fewer_digits",
                    "text", "signature", "watermark", "username",
                    "blurry", "jpeg_artifacts", "cropped",
                }
                _auto_tags = [t.strip() for t in auto_neg.split(",") if t.strip()]
                _scene_specific = [t for t in _auto_tags if t not in _auto_base]
                # ユーザーベース + シーン固有
                _neg_parts = [sd_neg_base]
                if _scene_specific:
                    _neg_parts.append(", ".join(_scene_specific))
                core_neg = ", ".join(_neg_parts)
            else:
                core_neg = auto_neg
            # prefix + core + suffix 組み立て
            neg_parts = []
            if sd_neg_prefix:
                neg_parts.append(sd_neg_prefix)
            neg_parts.append(core_neg)
            if sd_neg_suffix:
                neg_parts.append(sd_neg_suffix)
            scene["sd_negative_prompt"] = ", ".join(neg_parts).replace(",,", ",").strip(", ")

    return results

# タグDB（キャッシュ）
_tag_db_cache = None

def _load_tag_db() -> dict:
    """danbooru_tags.jsonからタグDBを読み込み（キャッシュ付き）"""
    global _tag_db_cache
    if _tag_db_cache is not None:
        return _tag_db_cache
    
    if DANBOORU_TAGS_JSON.exists():
        try:
            with open(DANBOORU_TAGS_JSON, "r", encoding="utf-8") as f:
                _tag_db_cache = json.load(f)
                log_message(f"タグDB読み込み完了: {DANBOORU_TAGS_JSON.name}")
                return _tag_db_cache
        except Exception as e:
            log_message(f"タグDB読み込みエラー: {e}")
    
    # フォールバック: 最小限のタグ
    _tag_db_cache = {
        "locations": {
            "教室": "classroom, school_desk, chair, chalkboard, window, school_interior",
            "寝室": "bedroom, bed, pillow, blanket, curtains, indoor, dim_lighting",
            "浴室": "bathroom, shower, bathtub, steam, wet, tiles, water",
            "リビング": "living_room, sofa, couch, cushion, tv, indoor",
            "屋上": "rooftop, fence, sky, school_rooftop, outdoor",
            "公園": "park, bench, trees, grass, outdoor, sunlight",
            "電車": "train_interior, seat, window, handrail",
            "ホテル": "hotel_room, bed, luxurious, curtains, dim_lighting",
            "オフィス": "office, desk, computer, chair, window, indoor"
        },
        "time_of_day": {
            "朝": "morning, sunrise, soft_lighting, warm_colors",
            "昼": "daytime, bright, sunlight, clear_sky",
            "放課後": "afternoon, golden_hour, warm_lighting, sunset_colors",
            "夕方": "evening, sunset, orange_sky, golden_light, dusk",
            "夜": "night, dark, moonlight, dim_lighting, starry_sky",
            "深夜": "late_night, darkness, lamp_light, intimate_lighting"
        },
        "compositions": {},
        "expressions": {},
        "poses_by_intensity": {},
        "clothing": {},
        "undress_states": {}
    }
    return _tag_db_cache


def _detect_personality_type(char_profiles: list) -> str:
    """キャラプロファイルから性格タイプを判定。
    Returns: personality key or ""
    対応: tsundere/submissive/sadistic/ojou/gal/seiso/genki/kuudere/inkya
    """
    if not char_profiles:
        return ""
    for cp in char_profiles:
        personality = cp.get("personality_core", {})
        desc = personality.get("brief_description", "")
        traits = personality.get("main_traits", [])
        archetype = cp.get("archetype", "")
        all_text = f"{desc} {' '.join(traits)} {archetype}".lower()
        # ツンデレ
        if any(k in all_text for k in ["ツンデレ", "ツン", "tsundere"]):
            return "tsundere"
        # ドM・従順・受け身
        if any(k in all_text for k in ["ドm", "どm", "masochist", "従順", "submissive",
                                        "奴隷", "ペット", "服従", "受け身"]):
            return "submissive"
        # Sっ気・サディスト・サキュバス・強気
        if any(k in all_text for k in ["ドs", "どs", "sadist", "サキュバス", "succubus",
                                        "小悪魔", "女王", "支配", "強気"]):
            return "sadistic"
        # お嬢様・高貴
        if any(k in all_text for k in ["お嬢様", "令嬢", "高貴", "ojou", "noble",
                                        "上品", "princess", "姫"]):
            return "ojou"
        # ギャル
        if any(k in all_text for k in ["ギャル", "gal", "黒ギャル", "パリピ",
                                        "チャラい"]):
            return "gal"
        # 清楚・純粋
        if any(k in all_text for k in ["清楚", "純粋", "清純", "天然", "innocent",
                                        "文学少女"]):
            return "seiso"
        # 元気・体育会系
        if any(k in all_text for k in ["元気", "活発", "体育会", "ボーイッシュ",
                                        "energetic", "スポーツ"]):
            return "genki"
        # クーデレ・無表情
        if any(k in all_text for k in ["クーデレ", "kuudere", "無表情", "無口",
                                        "クール", "cool"]):
            return "kuudere"
        # 陰キャ・オタク
        if any(k in all_text for k in ["陰キャ", "オタク", "otaku", "引っ込み",
                                        "内向", "根暗", "introvert"]):
            return "inkya"
    return ""


# 性格タイプ → primaryセリフスキルマッピング
_PERSONALITY_SKILL_MAP = {
    "tsundere":  "ero_serihu_tundere",     # ツンデレ → ツンデレ系
    "submissive": "ero_serihu_jyunai",     # 従順/受け身 → 甘え系（恥じらい・受容）
    "sadistic":  "ero_serihu_tundere",     # Sっ気/強気 → ツンデレ系（挑発・煽り）
    "ojou":      "ero_serihu_nomal",       # お嬢様 → ノーマル（ギャップ感重視）
    "gal":       "ero_serihu_ohogoe",      # ギャル/強気 → 激しい系
    "seiso":     "ero_serihu_nomal",       # 清楚 → ノーマル（ギャップ感重視）
    "genki":     "ero_serihu_nomal",       # 元気系 → ノーマル
    "kuudere":   "ero_serihu_nomal",       # クーデレ → ノーマル（感情抑制）
    "inkya":     "ero_serihu_jyunai",      # 陰キャ/オタク → 恥じらい系
}

# 性格タイプ → secondaryスキル + 混合比率（secondary_skill, secondary_ratio）
_PERSONALITY_SECONDARY_MAP = {
    "tsundere":  ("ero_serihu_jyunai", 0.3),   # 30%甘え混合（デレ部分）
    "submissive": ("ero_serihu_ohogoe", 0.3),   # 30%堕ち表現混合
    "sadistic":  ("ero_serihu_ohogoe", 0.4),    # 40%激しい表現混合
    "ojou":      ("ero_serihu_jyunai", 0.4),    # 40%上品な甘え混合
    "gal":       ("ero_serihu_nomal", 0.3),     # 30%通常混合
    "seiso":     ("ero_serihu_jyunai", 0.4),    # 40%恥じらい混合
    "genki":     ("ero_serihu_ohogoe", 0.2),    # 20%激しさ混合
    "kuudere":   ("ero_serihu_jyunai", 0.3),    # 30%感情漏れ混合
    "inkya":     ("ero_serihu_nomal", 0.2),     # 20%通常混合
}

# 性格タイプ → ero_dialogue_pool のプール混合カテゴリ指定
_PERSONALITY_POOL_MIX = {
    "tsundere":  {"primary": ["denial", "embarrassed"], "secondary": ["acceptance"]},
    "submissive": {"primary": ["submissive", "plea"], "secondary": ["acceptance", "ecstasy"]},
    "sadistic":  {"primary": ["provoke"], "secondary": ["ecstasy", "acceptance"]},
    "ojou":      {"primary": ["embarrassed", "denial"], "secondary": ["plea"]},
    "gal":       {"primary": ["acceptance", "provoke"], "secondary": ["ecstasy"]},
    "seiso":     {"primary": ["embarrassed", "denial"], "secondary": ["acceptance", "plea"]},
    "genki":     {"primary": ["acceptance", "provoke"], "secondary": ["ecstasy"]},
    "kuudere":   {"primary": ["denial"], "secondary": ["embarrassed", "acceptance"]},
    "inkya":     {"primary": ["embarrassed", "plea"], "secondary": ["denial", "acceptance"]},
}

# テーマキーワード → スキルマッピング
_THEME_SKILL_MAP = {
    "love":        "ero_serihu_jyunai",
    "vanilla":     "ero_serihu_jyunai",
    "netorare":    "ero_serihu_ohogoe",
    "humiliation": "ero_serihu_ohogoe",
    "forced":      "ero_serihu_ohogoe",
    "corruption":  "ero_serihu_ohogoe",
    "gangbang":    "ero_serihu_ohogoe",
    "chikan":      "ero_serihu_ohogoe",
    "isekai":        "ero_serihu_nomal",
    "onsen":         "ero_serihu_jyunai",
    "sleep":         "ero_serihu_ohogoe",
    "medical":       "ero_serihu_nomal",
    "swimsuit":      "ero_serihu_jyunai",
    "sports":        "ero_serihu_nomal",
    "idol":          "ero_serihu_nomal",
    "neighbor":      "ero_serihu_jyunai",
    "prostitution":  "ero_serihu_nomal",
    "voyeur":        "ero_serihu_ohogoe",
    "tentacle":      "ero_serihu_ohogoe",
    "reverse_rape":  "ero_serihu_tundere",
    "cosplay":       "ero_serihu_nomal",
    # v8.9: 未登録テーマ補完
    "maid":          "ero_serihu_jyunai",
    "hypnosis":      "ero_serihu_ohogoe",
    "harem":         "ero_serihu_nomal",
    "femdom":        "ero_serihu_tundere",
    "incest":        "ero_serihu_jyunai",
    "time_stop":     "ero_serihu_ohogoe",
    "monster":       "ero_serihu_ohogoe",
}


def _select_serihu_skill(theme: str = "", char_profiles: list = None) -> dict:
    """キャラ性格×テーマの2軸判定でセリフスキルを自動選択。

    Returns:
        dict: {"primary": str, "secondary": str|None, "ratio": float,
               "personality": str}
        - primary: メインスキル名
        - secondary: サブスキル名（混合用、なければNone）
        - ratio: primaryの比率 (0.0-1.0)。secondary = 1 - ratio
        - personality: 検出された性格タイプ（""の場合テーマのみ判定）

    複合テーマ: "netorare+love" 等、+/＋区切りで複数テーマを認識し混合。
    """
    personality = _detect_personality_type(char_profiles)

    # テーマ解析（複合テーマ対応: +/＋区切り）
    theme_lower = theme.lower() if theme else ""
    theme_parts = [t.strip() for t in theme_lower.replace("\uff0b", "+").split("+")
                   if t.strip()]

    # === 性格優先パス ===
    if personality and personality in _PERSONALITY_SKILL_MAP:
        primary = _PERSONALITY_SKILL_MAP[personality]
        secondary, sec_ratio = _PERSONALITY_SECONDARY_MAP.get(personality, (None, 0.0))
        # テーマスキルがprimaryと異なればsecondaryに採用（テーマ混合）
        if theme_parts:
            for part in theme_parts:
                theme_skill = _THEME_SKILL_MAP.get(part)
                if theme_skill and theme_skill != primary:
                    secondary = theme_skill
                    sec_ratio = max(sec_ratio, 0.3)
                    break
        return {
            "primary": primary,
            "secondary": secondary if sec_ratio > 0 else None,
            "ratio": 1.0 - sec_ratio,
            "personality": personality,
        }

    # === テーマのみパス ===
    if not theme_parts:
        return {"primary": "ero_serihu_nomal", "secondary": None,
                "ratio": 1.0, "personality": ""}

    # 複合テーマ: 先にマッチしたものをprimary、2番目をsecondary
    matched_skills = []
    for part in theme_parts:
        skill = _THEME_SKILL_MAP.get(part)
        if skill and skill not in matched_skills:
            matched_skills.append(skill)

    if len(matched_skills) >= 2:
        return {
            "primary": matched_skills[0],
            "secondary": matched_skills[1],
            "ratio": 0.7,
            "personality": "",
        }
    elif len(matched_skills) == 1:
        return {"primary": matched_skills[0], "secondary": None,
                "ratio": 1.0, "personality": ""}

    return {"primary": "ero_serihu_nomal", "secondary": None,
            "ratio": 1.0, "personality": ""}


# === データクラス ===
@dataclass
class CostTracker:
    haiku_input: int = 0
    haiku_output: int = 0
    haiku_fast_input: int = 0
    haiku_fast_output: int = 0
    sonnet_input: int = 0
    sonnet_output: int = 0
    opus_input: int = 0
    opus_output: int = 0
    cache_creation: int = 0
    cache_read: int = 0
    # モデル別キャッシュ追跡（正確なコスト計算用）
    haiku_cache_creation: int = 0
    haiku_cache_read: int = 0
    haiku_fast_cache_creation: int = 0
    haiku_fast_cache_read: int = 0
    sonnet_cache_creation: int = 0
    sonnet_cache_read: int = 0
    opus_cache_creation: int = 0
    opus_cache_read: int = 0
    api_calls: int = 0
    _lock: threading.Lock = field(default_factory=threading.Lock, repr=False, compare=False)

    def add(self, model: str, input_tokens: int, output_tokens: int,
            cache_creation_tokens: int = 0, cache_read_tokens: int = 0,
            batch: bool = False):
        with self._lock:
            self.api_calls += 1
            self.cache_creation += cache_creation_tokens
            self.cache_read += cache_read_tokens
            if "opus" in model:
                self.opus_input += input_tokens
                self.opus_output += output_tokens
                self.opus_cache_creation += cache_creation_tokens
                self.opus_cache_read += cache_read_tokens
            elif "sonnet" in model:
                self.sonnet_input += input_tokens
                self.sonnet_output += output_tokens
                self.sonnet_cache_creation += cache_creation_tokens
                self.sonnet_cache_read += cache_read_tokens
            elif model == MODELS.get("haiku_fast", "claude-3-haiku-20240307"):
                self.haiku_fast_input += input_tokens
                self.haiku_fast_output += output_tokens
                self.haiku_fast_cache_creation += cache_creation_tokens
                self.haiku_fast_cache_read += cache_read_tokens
            else:
                self.haiku_input += input_tokens
                self.haiku_output += output_tokens
                self.haiku_cache_creation += cache_creation_tokens
                self.haiku_cache_read += cache_read_tokens

    def total_cost_usd(self) -> float:
        """キャッシュ料金を正確に反映したコスト計算。
        Anthropic API: cache_read=入力単価x0.1, cache_creation=入力単価x1.25"""
        hf_cost = COSTS.get(MODELS["haiku_fast"], {"input": 0.25, "output": 1.25})
        h_cost = COSTS.get(MODELS["haiku"], {"input": 1.00, "output": 5.00})
        s_cost = COSTS.get(MODELS["sonnet"], {"input": 3.00, "output": 15.00})
        o_cost = COSTS.get(MODELS["opus"], {"input": 5.00, "output": 25.00})
        claude_cost = (
            # Haiku fast（非キャッシュ入力 + 出力 + キャッシュ作成 + キャッシュ読取）
            (self.haiku_fast_input / 1_000_000) * hf_cost["input"] +
            (self.haiku_fast_output / 1_000_000) * hf_cost["output"] +
            (self.haiku_fast_cache_creation / 1_000_000) * hf_cost["input"] * 1.25 +
            (self.haiku_fast_cache_read / 1_000_000) * hf_cost["input"] * 0.10 +
            # Haiku 4.5
            (self.haiku_input / 1_000_000) * h_cost["input"] +
            (self.haiku_output / 1_000_000) * h_cost["output"] +
            (self.haiku_cache_creation / 1_000_000) * h_cost["input"] * 1.25 +
            (self.haiku_cache_read / 1_000_000) * h_cost["input"] * 0.10 +
            # Sonnet
            (self.sonnet_input / 1_000_000) * s_cost["input"] +
            (self.sonnet_output / 1_000_000) * s_cost["output"] +
            (self.sonnet_cache_creation / 1_000_000) * s_cost["input"] * 1.25 +
            (self.sonnet_cache_read / 1_000_000) * s_cost["input"] * 0.10 +
            # Opus
            (self.opus_input / 1_000_000) * o_cost["input"] +
            (self.opus_output / 1_000_000) * o_cost["output"] +
            (self.opus_cache_creation / 1_000_000) * o_cost["input"] * 1.25 +
            (self.opus_cache_read / 1_000_000) * o_cost["input"] * 0.10
        )
        return claude_cost

    def _cache_savings_usd(self) -> float:
        """キャッシュによる節約額（キャッシュなしの場合との差分）"""
        h_cost = COSTS.get(MODELS["haiku"], {"input": 1.00, "output": 5.00})
        s_cost = COSTS.get(MODELS["sonnet"], {"input": 3.00, "output": 15.00})
        hf_cost = COSTS.get(MODELS["haiku_fast"], {"input": 0.25, "output": 1.25})
        o_cost = COSTS.get(MODELS["opus"], {"input": 5.00, "output": 25.00})
        # キャッシュ読み取りがフル入力だった場合のコスト差分（90%節約）
        return (
            (self.haiku_cache_read / 1_000_000) * h_cost["input"] * 0.90 +
            (self.sonnet_cache_read / 1_000_000) * s_cost["input"] * 0.90 +
            (self.haiku_fast_cache_read / 1_000_000) * hf_cost["input"] * 0.90 +
            (self.opus_cache_read / 1_000_000) * o_cost["input"] * 0.90
        )

    def summary(self) -> str:
        lines = []
        if self.haiku_fast_input or self.haiku_fast_output:
            lines.append(f"Haiku(fast): {self.haiku_fast_input:,} in / {self.haiku_fast_output:,} out")
        if self.haiku_input or self.haiku_output:
            lines.append(f"Haiku(4.5): {self.haiku_input:,} in / {self.haiku_output:,} out")
        if self.sonnet_input or self.sonnet_output:
            lines.append(f"Sonnet: {self.sonnet_input:,} in / {self.sonnet_output:,} out")
        if self.opus_input or self.opus_output:
            lines.append(f"Opus: {self.opus_input:,} in / {self.opus_output:,} out")
        if self.cache_read or self.cache_creation:
            lines.append(f"Cache: {self.cache_read:,} read / {self.cache_creation:,} create")
            savings = self._cache_savings_usd()
            if savings > 0.001:
                lines.append(f"Cache節約: -${savings:.4f}")
        lines.append(f"API呼出: {self.api_calls}回")
        lines.append(f"推定コスト: ${self.total_cost_usd():.4f}")
        return "\n".join(lines)


def estimate_cost(num_scenes: int, use_sonnet_polish: bool = True) -> dict:
    """生成前にコストを予測（Prompt Caching反映版）
    haiku=圧縮/あらすじ/アウトライン+低intensityシーン, sonnet=i4以上シーン, opus=i5清書"""
    h_cost = COSTS.get(MODELS["haiku"], {"input": 1.00, "output": 5.00})
    s_cost = COSTS.get(MODELS["sonnet"], {"input": 3.00, "output": 15.00})
    o_cost = COSTS.get(MODELS["opus"], {"input": 5.00, "output": 25.00})

    # Phase 1: コンテキスト圧縮 + あらすじ (haiku 4.5)
    fast_input = 500 + 600
    fast_output = 150 + 800

    # Phase 3: アウトライン (haiku: 全ケース)
    haiku_input = fast_input  # Phase 1-2もhaiku
    haiku_output = fast_output
    if num_scenes <= 12:
        haiku_input += 2000
        haiku_output += num_scenes * 300
    else:
        chunks = (num_scenes + 9) // 10
        haiku_input += chunks * 3000
        haiku_output += chunks * 2000

    # シーン生成（intensity分布推定: 40% i1-3→haiku, 60% i4-5→sonnet）
    haiku_scenes = int(num_scenes * 0.40)  # intensity 1-3 → haiku
    sonnet_scenes = num_scenes - haiku_scenes  # intensity 4-5 → sonnet

    # Opus清書対象（intensity 5 ≒ シーン数の ~13%）
    opus_scenes = max(2, num_scenes // 15)

    # シーン固有の非キャッシュ入力（user prompt: context + story_so_far + scene指示）
    avg_user_tokens = 3000  # 平均user prompt（story_so_far含む）

    # Claude: Prompt Caching効果
    cached_system_tokens = 16000
    # Haiku シーン: 1回cache_create + (N-1)回cache_read
    haiku_cache_create_cost = (cached_system_tokens / 1_000_000) * h_cost["input"] * 1.25
    haiku_cache_read_cost = (cached_system_tokens / 1_000_000) * h_cost["input"] * 0.10 * max(0, haiku_scenes - 1)
    haiku_uncached_input = haiku_scenes * avg_user_tokens
    haiku_input += haiku_uncached_input
    haiku_output += haiku_scenes * 650
    # Sonnet シーン: 1回cache_create + (N-1)回cache_read
    sonnet_cache_create_cost = (cached_system_tokens / 1_000_000) * s_cost["input"] * 1.25 if sonnet_scenes > 0 else 0
    sonnet_cache_read_cost = (cached_system_tokens / 1_000_000) * s_cost["input"] * 0.10 * max(0, sonnet_scenes - 1)
    sonnet_input = sonnet_scenes * avg_user_tokens
    sonnet_output = sonnet_scenes * 700
    # Opus清書: 各シーンJSON往復（入力~2000tok, 出力~1500tok）
    opus_input = opus_scenes * 2000
    opus_output = opus_scenes * 1500
    estimated_usd = (
        (haiku_input / 1_000_000) * h_cost["input"] +
        (haiku_output / 1_000_000) * h_cost["output"] +
        haiku_cache_create_cost + haiku_cache_read_cost +
        (sonnet_input / 1_000_000) * s_cost["input"] +
        (sonnet_output / 1_000_000) * s_cost["output"] +
        sonnet_cache_create_cost + sonnet_cache_read_cost +
        (opus_input / 1_000_000) * o_cost["input"] +
        (opus_output / 1_000_000) * o_cost["output"]
    )

    return {
        "haiku_tokens": haiku_input + haiku_output,
        "sonnet_tokens": sonnet_input + sonnet_output,
        "opus_tokens": opus_input + opus_output,
        "estimated_usd": estimated_usd,
        "estimated_jpy": estimated_usd * 150  # 概算レート
    }


# === ユーティリティ ===
def load_file(filepath: Path) -> str:
    if filepath.exists():
        return filepath.read_text(encoding="utf-8")
    return ""


_skill_cache: dict = {}  # スキルファイル読み込みキャッシュ（同一パイプライン内の重複I/O削減）

def load_skill(skill_name: str) -> str:
    if skill_name in _skill_cache:
        return _skill_cache[skill_name]
    skill_file = SKILLS_DIR / f"{skill_name}.skill.md"
    if skill_file.exists():
        content = skill_file.read_text(encoding="utf-8")
        _skill_cache[skill_name] = content
        return content
    _skill_cache[skill_name] = ""
    return ""


def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log_message(f"設定ファイル読み込みエラー: {e}")
    return {}


def save_config(config: dict):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)


def _decode_exif_user_comment(value) -> str:
    """EXIF UserComment (0x9286) を文字列にデコード。

    SD WebUI(piexif)は b"UNICODE\\x00" + utf-16 で書き込む。
    バイトオーダーはEXIFヘッダ(II=LE/MM=BE)依存のためbodyの先頭2バイトで自動判定。
    Pillowバージョンにより bytes or str(latin-1誤デコード) で返る。
    """
    if not value:
        return ""
    # str型: Pillowが自動デコード済み or latin-1誤デコードの可能性
    if isinstance(value, str):
        # エンコーディングプレフィックスやNULLバイトが含まれている場合は
        # latin-1で誤デコードされたバイト列 → bytesに戻して再デコード
        if "\x00" in value or value.startswith(("UNICODE", "ASCII", "JIS")):
            try:
                value = value.encode("latin-1")
            except (UnicodeEncodeError, UnicodeDecodeError):
                # latin-1に戻せない場合はNULLバイトだけ除去して返す
                return value.replace("\x00", "").strip()
        else:
            return value.strip()
    # bytes型: エンコーディング識別子を解析して適切にデコード
    if isinstance(value, bytes):
        header = value[:8] if len(value) >= 8 else value
        body = value[8:] if len(value) > 8 else value
        # piexif: b"UNICODE\x00" + utf-16 encoded body
        # バイトオーダーは EXIF ヘッダ(II/MM)依存: bodyの先頭2バイトで自動判定
        if header.startswith(b"UNICODE"):
            # 先頭2バイトでバイトオーダーを推定
            # body[0]==0x00 && body[1] が印刷可能ASCII → big-endian
            # body[1]==0x00 && body[0] が印刷可能ASCII → little-endian
            if len(body) >= 2 and body[0] == 0x00 and 0x20 <= body[1] <= 0x7e:
                utf16_order = ("utf-16-be", "utf-16-le")
            else:
                utf16_order = ("utf-16-le", "utf-16-be")
            for enc in (*utf16_order, "utf-8"):
                try:
                    decoded = body.decode(enc).strip("\x00").strip()
                    if decoded:
                        return decoded
                except (UnicodeDecodeError, ValueError):
                    continue
        # ASCII prefix or undefined (\x00 * 8)
        if header.startswith(b"ASCII") or header == b"\x00" * 8:
            try:
                return body.decode("ascii", errors="ignore").strip("\x00").strip()
            except Exception:
                pass
        # JIS prefix
        if header.startswith(b"JIS"):
            try:
                return body.decode("shift_jis", errors="ignore").strip("\x00").strip()
            except Exception:
                pass
        # フォールバック: プレフィックスなしで全体をutf-8試行
        try:
            return value.decode("utf-8", errors="ignore").strip("\x00").strip()
        except Exception:
            return ""
    return str(value).strip()


def _try_decode_bytes(data: bytes) -> str:
    """バイト列をSD生成パラメータ文字列としてデコード（複数エンコーディング試行）。"""
    if not data:
        return ""
    # エンコーディングプレフィックス付きの場合は _decode_exif_user_comment に委譲
    if len(data) >= 8:
        hdr = data[:8]
        if hdr.startswith((b"UNICODE", b"ASCII", b"JIS")) or hdr == b"\x00" * 8:
            return _decode_exif_user_comment(data)
    # UTF-8 → UTF-16-LE → shift_jis → latin-1 の順で試行
    for enc in ("utf-8", "utf-16-le", "utf-16-be", "shift_jis", "latin-1"):
        try:
            decoded = data.decode(enc).strip("\x00").strip()
            # SD prompt らしいかチェック（英数字タグが含まれるか）
            if decoded and any(c.isalpha() for c in decoded[:50]):
                # UTF-16誤検出防止: NULLバイトだらけなら不採用
                if "\x00" not in decoded:
                    return decoded
        except (UnicodeDecodeError, ValueError):
            continue
    # 最終フォールバック
    return data.decode("utf-8", errors="replace").replace("\x00", "").strip()


def _extract_sd_params_from_bytes(file_bytes: bytes) -> str:
    """ファイルのrawバイトからSD WebUIパラメータ文字列を直接探索（最終フォールバック）。"""
    import re as _re
    # SD WebUI形式: "Steps: NN, Sampler: ..., CFG scale: ..."
    # このパターンを末尾側から探してパラメータ文字列を逆引き
    markers = [b"Steps: ", b"steps: "]
    for marker in markers:
        idx = file_bytes.rfind(marker)
        if idx < 0:
            continue
        # marker前方1024バイト〜marker後方512バイトを切り出し
        start = max(0, idx - 4096)
        end = min(len(file_bytes), idx + 1024)
        chunk = file_bytes[start:end]
        # テキスト部分だけ抽出（NULLバイト除去して複数エンコーディング試行）
        for enc in ("utf-8", "latin-1"):
            try:
                text = chunk.decode(enc, errors="ignore")
                text = text.replace("\x00", "")
                # "Steps: " を含む行を探し、その前のテキストも含めて返す
                m = _re.search(r'([\x20-\x7e,\(\)\n:.<>_\-\w]{10,}Steps:\s*\d+[^\x00]*)', text)
                if m:
                    candidate = m.group(1).strip()
                    # プロンプトの開始位置を推定（最初の改行ブロックの前か、印刷可能文字の始まり）
                    # 前方にゴミバイナリが混ざっている可能性があるので、有効な先頭を探す
                    for ci, ch in enumerate(candidate):
                        if ch.isalpha() or ch == '(':
                            return candidate[ci:].strip()
                    return candidate
            except Exception:
                continue
    return ""


def parse_png_info(file_path: str) -> dict:
    """画像ファイルからSD生成パラメータを解析（PNG/JPEG/WebP対応）。

    Returns:
        {"positive": str, "negative": str, "parameters": str, "raw": str}
        or {"error": str}
    """
    if not PIL_AVAILABLE:
        return {"error": "Pillowがインストールされていません。\npip install Pillow で追加してください。"}
    try:
        img = Image.open(file_path)
        raw = ""
        source = ""

        # 1. PNG: img.info にtEXt/iTXtチャンクが格納される
        if img.format == "PNG":
            raw = img.info.get("parameters", "")
            if raw:
                source = "PNG tEXt[parameters]"
            else:
                for key in ("prompt", "Comment", "Description"):
                    if key in img.info:
                        raw = img.info[key]
                        source = f"PNG tEXt[{key}]"
                        break

        # 2. EXIF sub-IFD → UserComment (SD WebUI A1111/Forge標準)
        if not raw:
            try:
                exif_data = img.getexif()
                if exif_data:
                    try:
                        exif_ifd = exif_data.get_ifd(0x8769)
                        if exif_ifd:
                            uc = exif_ifd.get(0x9286, "")
                            if uc:
                                raw = _decode_exif_user_comment(uc)
                                if raw:
                                    source = "EXIF sub-IFD UserComment"
                    except Exception:
                        pass
                    # 2b. root IFD → UserComment (一部ツール)
                    if not raw:
                        uc_root = exif_data.get(0x9286, "")
                        if uc_root:
                            raw = _decode_exif_user_comment(uc_root)
                            if raw:
                                source = "EXIF root UserComment"
                    # 2c. root IFD → ImageDescription
                    if not raw:
                        desc = exif_data.get(0x010E, "")
                        if desc:
                            raw = _decode_exif_user_comment(desc) if isinstance(desc, bytes) else str(desc).strip()
                            if raw:
                                source = "EXIF ImageDescription"
            except Exception:
                pass

        # 3. raw EXIF bytes から直接パース（Pillow解析が失敗した場合のフォールバック）
        if not raw:
            exif_raw = img.info.get("exif", b"")
            if isinstance(exif_raw, bytes) and len(exif_raw) > 100:
                raw = _extract_sd_params_from_bytes(exif_raw)
                if raw:
                    source = "EXIF raw bytes scan"

        # 4. JPEG COM marker (img.info["comment"])
        if not raw:
            comment = img.info.get("comment", b"")
            if comment:
                raw = _try_decode_bytes(comment) if isinstance(comment, bytes) else str(comment).strip()
                if raw:
                    source = "JPEG COM marker"

        # 5. フォーマット問わず img.info のフォールバック
        if not raw:
            for key in ("parameters", "prompt", "Comment", "Description"):
                val = img.info.get(key, "")
                if val:
                    if isinstance(val, bytes):
                        raw = _try_decode_bytes(val)
                    else:
                        raw = str(val).strip()
                    if raw:
                        source = f"img.info[{key}]"
                        break

        # 6. ファイル全体のrawバイトスキャン（最終フォールバック）
        if not raw:
            try:
                with open(file_path, "rb") as fb:
                    file_bytes = fb.read()
                raw = _extract_sd_params_from_bytes(file_bytes)
                if raw:
                    source = "file bytes scan"
            except Exception:
                pass

        if not raw:
            log_message(f"PNG Info: 生成情報なし format={img.format} info_keys={list(img.info.keys())}")
            return {"error": "生成情報が見つかりません。\nSD WebUI / ComfyUI 等で生成した画像を使用してください。"}

        log_message(f"PNG Info: 読取成功 source={source} len={len(raw)}")

        # 最終クリーンアップ: NULLバイトやエンコーディング残骸を除去
        raw = raw.replace("\x00", "").strip()

        # パラメータ文字列をpositive / negative / params に分離
        positive = raw
        negative = ""
        params = ""
        neg_idx = raw.find("Negative prompt:")
        steps_idx = raw.find("\nSteps:")
        if neg_idx >= 0:
            positive = raw[:neg_idx].strip()
            if steps_idx >= 0 and steps_idx > neg_idx:
                negative = raw[neg_idx + len("Negative prompt:"):steps_idx].strip()
                params = raw[steps_idx + 1:].strip()
            else:
                negative = raw[neg_idx + len("Negative prompt:"):].strip()
        elif steps_idx >= 0:
            positive = raw[:steps_idx].strip()
            params = raw[steps_idx + 1:].strip()

        return {"positive": positive, "negative": negative, "parameters": params, "raw": raw}
    except Exception as e:
        return {"error": f"画像読み取りエラー: {e}"}


# === プロファイル管理 ===
def get_profile_list() -> list[str]:
    """保存されているプロファイル一覧を取得"""
    profiles = []
    for f in PROFILES_DIR.glob("*.json"):
        profiles.append(f.stem)
    return sorted(profiles)


def save_profile(name: str, config: dict):
    """プロファイルを保存"""
    profile_path = PROFILES_DIR / f"{name}.json"
    config["profile_name"] = name
    config["saved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(profile_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    log_message(f"プロファイル保存: {name}")


def load_profile(name: str) -> dict:
    """プロファイルを読み込み"""
    profile_path = PROFILES_DIR / f"{name}.json"
    if profile_path.exists():
        with open(profile_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def delete_profile(name: str) -> bool:
    """プロファイルを削除"""
    profile_path = PROFILES_DIR / f"{name}.json"
    if profile_path.exists():
        profile_path.unlink()
        log_message(f"プロファイル削除: {name}")
        return True
    return False


def copy_profile(src_name: str, dst_name: str) -> bool:
    """プロファイルをコピー"""
    src_path = PROFILES_DIR / f"{src_name}.json"
    if src_path.exists():
        config = load_profile(src_name)
        config["profile_name"] = dst_name
        save_profile(dst_name, config)
        log_message(f"プロファイルコピー: {src_name} → {dst_name}")
        return True
    return False


def log_message(message: str):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")


# === API呼び出し ===
def call_claude(
    client: anthropic.Anthropic,
    model: str,
    system,
    user: str,
    cost_tracker: CostTracker,
    max_tokens: int = 4096,
    callback: Optional[Callable] = None
) -> str:
    total_max_retries = MAX_RETRIES_OVERLOADED  # 529対応で最大試行回数を拡大
    overloaded_count = 0  # 529エラー連続カウント
    for attempt in range(total_max_retries):
        try:
            if model == MODELS.get("haiku_fast"):
                model_name = "Haiku(fast)"
            elif "haiku" in model:
                model_name = "Haiku(4.5)"
            else:
                model_name = "Sonnet"
            log_message(f"API呼び出し開始: {model_name} (試行 {attempt + 1}/{total_max_retries})")

            if callback:
                callback(f"API呼び出し中 ({model_name})...")

            # Prompt Caching対応: systemがlistならそのまま、strならブロック化
            if isinstance(system, list):
                system_param = system
            else:
                system_param = system

            response = client.messages.create(
                model=model,
                max_tokens=max_tokens,
                system=system_param,
                messages=[{"role": "user", "content": user}],
                timeout=120.0  # 2分タイムアウト
            )

            usage = response.usage
            cache_creation = getattr(usage, 'cache_creation_input_tokens', 0) or 0
            cache_read = getattr(usage, 'cache_read_input_tokens', 0) or 0
            cost_tracker.add(model, usage.input_tokens, usage.output_tokens,
                             cache_creation, cache_read)

            # キャッシュ統計ログ
            if cache_creation or cache_read:
                log_message(f"{model_name}: {usage.input_tokens} in, {usage.output_tokens} out (cache: +{cache_creation} create, {cache_read} read)")
            else:
                log_message(f"{model_name}: {usage.input_tokens} in, {usage.output_tokens} out")

            return response.content[0].text

        except anthropic.RateLimitError as e:
            wait_time = RETRY_DELAY * (2 ** (attempt + 1))
            log_message(f"Rate limit: {e} (待機{wait_time}秒)")
            if callback:
                callback(f"レート制限、{wait_time}秒待機...")
            time.sleep(wait_time)

        except anthropic.APIStatusError as e:
            if e.status_code == 401:
                raise ValueError("APIキーが無効です")
            if e.status_code == 529:
                # 529 Overloaded: 段階的対処
                overloaded_count += 1
                # 3回失敗後: 別モデルにフォールバック（Haiku→Sonnet）
                if overloaded_count == 3 and "haiku" in model and model != MODELS.get("haiku_fast"):
                    fallback_model = MODELS["sonnet"]
                    log_message(f"529 Overloaded 3回連続: Sonnetにフォールバック")
                    if callback:
                        callback(f"Haiku過負荷、Sonnetで代替生成中...")
                    model = fallback_model  # 以降の試行はSonnetを使用
                    time.sleep(5)
                    continue
                wait_time = RETRY_DELAY_OVERLOADED * min(overloaded_count, 4)  # 15→30→45→60秒
                log_message(f"529 Overloaded ({overloaded_count}回目): {wait_time}秒待機後に再試行")
                if callback:
                    callback(f"サーバー過負荷、{wait_time}秒待機中... ({overloaded_count}/{MAX_RETRIES_OVERLOADED})")
                time.sleep(wait_time)
                if overloaded_count >= MAX_RETRIES_OVERLOADED:
                    raise RuntimeError(f"サーバー過負荷が継続（{MAX_RETRIES_OVERLOADED}回試行）。時間をおいて再実行してください。")
                continue
            log_message(f"API error {e.status_code}: {e}")
            if attempt < total_max_retries - 1:
                if callback:
                    callback(f"APIエラー、再試行中...")
                time.sleep(RETRY_DELAY)
            else:
                raise

        except anthropic.APITimeoutError as e:
            log_message(f"API timeout: {e}")
            if callback:
                callback(f"タイムアウト、再試行中...")
            if attempt < total_max_retries - 1:
                time.sleep(RETRY_DELAY * 2)
            else:
                raise RuntimeError(f"APIタイムアウト（{total_max_retries}回試行）")

        except Exception as e:
            log_message(f"Error: {e}")
            if callback:
                callback(f"エラー: {str(e)[:30]}...")
            if attempt < total_max_retries - 1:
                time.sleep(RETRY_DELAY)
            else:
                raise

    raise RuntimeError("最大リトライ回数を超えました")


def _call_api(
    client,
    model: str,
    system,
    user: str,
    cost_tracker: CostTracker,
    max_tokens: int = 4096,
    callback: Optional[Callable] = None
) -> str:
    """API呼び出し（Claude）"""
    return call_claude(client, model, system, user, cost_tracker, max_tokens, callback)


def parse_json_response(text: str):
    """Parse JSON from API response, handling markdown code blocks and prefixed text."""
    original_text = text
    log_message(f"Raw API response: {text[:1000]}")
    
    try:
        # マークダウンコードブロック除去
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            parts = text.split("```")
            if len(parts) >= 2:
                text = parts[1]
        
        text = text.strip()
        
        # JSONの前にある前置きテキストを除去
        # 「{」または「[」で始まる部分を探す
        if text and not text.startswith("{") and not text.startswith("["):
            # 最初の { または [ を探す
            brace_idx = text.find("{")
            bracket_idx = text.find("[")
            
            if brace_idx == -1 and bracket_idx == -1:
                log_message(f"No JSON found in response: {text[:300]}")
                raise ValueError(f"No JSON in response: {original_text[:150]}")
            
            # より早く出現する方を使用
            if brace_idx == -1:
                start_idx = bracket_idx
            elif bracket_idx == -1:
                start_idx = brace_idx
            else:
                start_idx = min(brace_idx, bracket_idx)
            
            log_message(f"Stripping prefix text before JSON (index {start_idx})")
            text = text[start_idx:]
        
        # 末尾の余分なテキストも除去（JSONの閉じ括弧以降）
        if text.startswith("{"):
            # 対応する } を探す
            depth = 0
            end_idx = 0
            for i, c in enumerate(text):
                if c == "{":
                    depth += 1
                elif c == "}":
                    depth -= 1
                    if depth == 0:
                        end_idx = i + 1
                        break
            if end_idx > 0:
                text = text[:end_idx]
        elif text.startswith("["):
            # 対応する ] を探す
            depth = 0
            end_idx = 0
            for i, c in enumerate(text):
                if c == "[":
                    depth += 1
                elif c == "]":
                    depth -= 1
                    if depth == 0:
                        end_idx = i + 1
                        break
            if end_idx > 0:
                text = text[:end_idx]
        
        text = text.strip()
        if not text:
            log_message(f"Empty response after parsing. Original: {original_text[:500]}")
            raise ValueError(f"Empty response: {original_text[:200]}")
        
        return json.loads(text)
    except json.JSONDecodeError as e:
        log_message(f"JSON parse error: {e}")
        log_message(f"Parsed text: {text[:500]}")
        raise ValueError(f"Invalid JSON: {str(e)[:50]}. Text: {text[:100]}...") from e


# === Skill 1: Prompt Compactor ===
def compact_context(
    client: anthropic.Anthropic,
    concept: str,
    characters: str,
    theme: str,
    cost_tracker: CostTracker,
    callback: Optional[Callable] = None
) -> dict:
    skill = load_skill("prompt_compactor")
    prompt = f"""以下の作品情報を、トークン効率の良い形式に圧縮してください。

## 作品コンセプト
{concept}

## 登場人物
{characters}

## テーマ
{theme if theme else "指定なし"}

## 出力形式（JSON）
{{
    "setting": "舞台（短文）",
    "chars": [
        {{"name": "名前", "look": "外見（箇条書き）", "voice": "口調特徴"}}
    ],
    "tone": "トーン（1語）",
    "theme": "テーマ（1語）",
    "ng": ["NG要素"]
}}

冗長な説明を排除し、箇条書きで簡潔に。JSONのみ出力。"""

    if callback:
        callback("[PACK]コンテキスト圧縮中...")

    response = _call_api(
        client, MODELS["haiku"],
        skill if skill else "You compress prompts to save tokens. Output only JSON.",
        prompt, cost_tracker, 1024, callback
    )
    return parse_json_response(response)


def compact_context_local(
    concept: str,
    characters: str,
    theme: str,
    char_profiles: list,
    callback: Optional[Callable] = None
) -> dict:
    """キャラプロファイルからローカルでcontext JSONを構築（API不要）"""
    if callback:
        callback("[PACK]コンテキスト圧縮中（ローカル・API節約）...")

    theme_guide = THEME_GUIDES.get(theme, THEME_GUIDES.get("vanilla", {}))

    # 舞台を概念テキストから抽出（最初の1文 or 50文字）
    setting = concept.strip() if concept.strip() else "日常"

    # キャラ情報をプロファイルから構築
    chars = []
    ng_all = []
    for cp in char_profiles:
        name = cp.get("character_name", "")
        physical = cp.get("physical_description", {})
        speech = cp.get("speech_pattern", {})
        avoid = cp.get("avoid_patterns", [])

        look_parts = []
        if physical.get("hair"):
            look_parts.append(f"髪:{physical['hair']}")
        if physical.get("eyes"):
            look_parts.append(f"目:{physical['eyes']}")
        if physical.get("body"):
            look_parts.append(f"体型:{physical['body']}")
        if physical.get("chest"):
            look_parts.append(f"胸:{physical['chest']}")

        voice_parts = []
        if speech.get("first_person"):
            voice_parts.append(f"一人称:{speech['first_person']}")
        endings = speech.get("sentence_endings", [])
        if endings:
            voice_parts.append(f"語尾:{','.join(endings[:3])}")

        chars.append({
            "name": name,
            "look": ", ".join(look_parts),
            "voice": ", ".join(voice_parts)
        })
        ng_all.extend(avoid[:3])

    # テーマに基づくトーン
    tone = theme_guide.get("name", "一般")
    theme_label = theme_guide.get("name", "指定なし")

    context = {
        "setting": setting,
        "chars": chars,
        "tone": tone,
        "theme": theme_label,
        "ng": list(set(ng_all))[:5]
    }

    log_message(f"コンテキスト圧縮完了（ローカル）: chars={len(chars)}, setting={setting[:30]}")
    if callback:
        callback("[OK] コンテキスト圧縮完了（ローカル・API節約）")

    return context


def generate_synopsis(
    client: anthropic.Anthropic,
    concept: str,
    context: dict,
    num_scenes: int,
    theme: str,
    cost_tracker: CostTracker,
    callback: Optional[Callable] = None,
    male_description: str = "",
    faceless_male: bool = True,
) -> str:
    """コンセプトから短い一本のストーリーあらすじを生成（Haiku API 1回）"""
    theme_guide = THEME_GUIDES.get(theme, THEME_GUIDES.get("vanilla", {}))
    theme_name = theme_guide.get("name", "指定なし")
    story_arc = theme_guide.get("story_arc", "導入→展開→本番→余韻")
    key_emotions = theme_guide.get("key_emotions", ["期待", "緊張", "快感", "幸福"])
    story_elements = theme_guide.get("story_elements", [])

    if callback:
        callback(f"[STORY]{theme_name}テーマでストーリー原案を作成中...")

    chars = context.get("chars", [])
    char_info = ""
    for c in chars:
        name = c.get("name", "")
        look = c.get("look", "")
        voice = c.get("voice", "")
        char_info += f"・{name}: {look} / {voice}\n"

    elements_str = "\n".join(f"・{e}" for e in story_elements) if story_elements else "・特になし"
    emotions_str = "、".join(key_emotions)

    prompt = f"""以下の作品コンセプトに基づき、FANZA同人CG集のストーリーあらすじ（400〜600字）を書いてください。

## 作品コンセプト
{concept}

## 登場キャラクター
{char_info}
{f"## 男性キャラクター外見{chr(10)}{male_description}{chr(10)}※ あらすじにおける男性の描写はこの外見設定を反映すること{chr(10)}{'※ 男性はfaceless male（顔なし）。顔の特徴は描写しない' + chr(10) if faceless_male else ''}" if male_description else ""}## テーマ: {theme_name}
- ストーリーの流れ: {story_arc}
- 重要な感情: {emotions_str}
- ストーリー要素:
{elements_str}

## シーン数: {num_scenes}シーン

## あらすじの要件
1. コンセプトの設定（場所、人間関係、状況）を忠実に反映すること
2. 起承転結のある一貫したストーリーであること
3. テーマ「{theme_name}」のアーク「{story_arc}」に沿うこと
4. プロローグ（約20%）→ 展開・焦らし（約30%）→ 本番エロシーン（約40%）→ エピローグ（約10%）の配分
5. **焦らしの重要性**: エロの品質は焦らし（i1-2）で決まる。「まだしない」の緊張感、予感、間接的接触を丁寧に描け。焦らしが薄いと後半が平坦になる
6. 各キャラの個性・口調が活きる展開にすること
7. 本番パートでは具体的な行為の流れも簡潔に含めること
8. エロシーンは段階的にエスカレートし、クライマックスに向かうこと
9. コンセプトにない極端な展開は絶対に追加しないこと（コンセプトの範囲内で物語を展開する）
10. **同じ行為の繰り返し禁止**: 本番パートで最低5種の異なる展開（体位変更/場所移動/行為変更/主導権交代等）を含めること
11. **展開バリエーション例**: 愛撫→口淫→正常位→体位変更→騎乗位→バック→絶頂のように行為を変化させエスカレートすること
12. **視点**: あらすじは女性キャラ（ヒロイン）視点で書くこと。男性は「相手の男」「彼」と呼び、「主人公」と呼ばないこと

あらすじの文章のみ出力してください。JSON不要。"""

    try:
        response = _call_api(
            client, MODELS["haiku"],
            f"FANZA同人CG集の脚本家です。ユーザーのコンセプトに基づき、エロ同人作品の完全なストーリーあらすじを作成します。成人向けコンテンツを忠実に生成します。",
            prompt, cost_tracker, 2048, callback
        )

        synopsis = response.strip()
        log_message(f"あらすじ生成完了: {len(synopsis)}文字")
        if callback:
            callback(f"[OK]ストーリー原案完成（{len(synopsis)}文字）")

        return synopsis

    except Exception as e:
        log_message(f"あらすじ生成エラー: {e}")
        if callback:
            callback(f"[WARN]あらすじ生成失敗: {str(e)[:50]}")
        # フォールバック: コンセプトをそのままあらすじとして使用
        return concept


# === Skill 2: Low Cost Pipeline ===

def generate_scene_batch(
    client: anthropic.Anthropic,
    context: dict,
    scenes: list,
    jailbreak: str,
    cost_tracker: CostTracker,
    theme: str = "",
    char_profiles: list = None,
    callback: Optional[Callable] = None,
    story_so_far: str = ""
) -> list:
    """複数のLow-Intensityシーンをまとめて1回のAPI呼び出しで生成（API節約）"""
    skill = load_skill("low_cost_pipeline")
    danbooru_nsfw = load_skill("danbooru_nsfw_tags")
    scene_composer = load_skill("nsfw_scene_composer")
    _serihu_info = _select_serihu_skill(theme, char_profiles)
    serihu_skill_name = _serihu_info["primary"]
    serihu_skill = load_skill(serihu_skill_name)
    _serihu_secondary = load_skill(_serihu_info["secondary"]) if _serihu_info.get("secondary") else ""
    _serihu_ratio = _serihu_info.get("ratio", 1.0)
    _serihu_personality = _serihu_info.get("personality", "")
    bubble_writer_skill = load_skill("cg_bubble_writer")
    visual_skill = load_skill("cg_visual_variety")

    theme_guide = THEME_GUIDES.get(theme, THEME_GUIDES.get("vanilla", {}))
    theme_name = theme_guide.get("name", "指定なし")
    dialogue_tone = theme_guide.get("dialogue_tone", "自然で楽しい雰囲気")
    use_heart = theme_guide.get("use_heart", True)
    theme_sd_tags = theme_guide.get("sd_tags", "")
    theme_sd_expressions = theme_guide.get("sd_expressions", "")
    key_emotions = theme_guide.get("key_emotions", [])

    tag_db = _load_tag_db()
    loc_tags_db = tag_db.get("locations", {})
    time_tags_db = tag_db.get("time_of_day", {})

    # キャラガイド（低intensity用＝簡潔版）
    char_guide = ""
    char_danbooru_tags = []
    char_names = []

    if char_profiles:
        for cp in char_profiles:
            name = cp.get("character_name", "")
            char_names.append(name)
            speech = cp.get("speech_pattern", {})
            avoid = cp.get("avoid_patterns", [])
            physical = cp.get("physical_description", {})
            tags = cp.get("danbooru_tags", [])
            char_danbooru_tags.extend(tags)

            emotional = cp.get("emotional_speech", {})
            char_guide += f"""
【{name}】口調ガイド
・一人称: {speech.get('first_person', '私')} / 語尾: {', '.join(speech.get('sentence_endings', ['〜よ', '〜ね']))}
・間投詞: {', '.join(speech.get('fillers', ['あっ', 'んっ']))}
・照れた時: {emotional.get('when_embarrassed', '言葉に詰まる')}
・甘える時: {emotional.get('when_flirty', '甘い声で')}
・外見: 髪={physical.get('hair', '')}, 目={physical.get('eyes', '')}
・NG: {', '.join(avoid) if avoid else 'なし'}
"""

    heart_instruction = "♡は甘いシーンで自然に使用OK。" if use_heart else f"テーマ「{theme_name}」のspeechでは♡控えめ。ただしmoan（喘ぎ）では♡を積極的に使え（不本意な快楽表現として有効）。"

    # 共通システムプロンプト（CG集フォーマット対応）
    common_system = f"""{jailbreak}

{skill if skill else "FANZA同人CG集の脚本を生成します。"}

{danbooru_nsfw if danbooru_nsfw else ""}

{scene_composer if scene_composer else ""}

{bubble_writer_skill if bubble_writer_skill else ""}

{char_guide if char_guide else "（キャラ設定なし）"}

## CG集フォーマット補足
CG画像1枚 + 吹き出し1-3個（ヒロイン1-2+男0-1） + SE 0-4個。画像がメイン。
moan=喘ぎ声のみ(説明文禁止) / speech=感情的反応のみ / story_flow=各シーン固有(コピペ禁止)

{f'''
## ⚠️ セリフ品質ガイド（厳守・最優先）

bubblesのtextは以下の【喘ぎ声バリエーション集】と【鉄則】に厳密に従え。
「タスク手順」「不自然診断」「改訂版セリフ」等のセクションは無視せよ。

★ 喘ぎ声は必ず下記辞書の【段階1〜4】から選べ。自分で喘ぎを創作するな。
★ intensityに対応する段階を使え（intensity 1-2=段階1、intensity 3=段階2、intensity 4=段階3、intensity 5=段階4）
★ 前シーンで使った喘ぎと同じものは絶対禁止。毎シーン辞書の別パターンを選べ。

{serihu_skill}
''' if serihu_skill else ''}{f'''

### サブスタイル（混合比率{int((1-_serihu_ratio)*100)}%で以下のスタイルも取り入れること）:
{_serihu_secondary}
''' if _serihu_secondary and _serihu_ratio < 1.0 else ''}{f'''
★ キャラ性格タイプ「{_serihu_personality}」を意識したセリフ。ギャップ感を出すこと。
''' if _serihu_personality else ''}

{f'''
## CG集ビジュアル構成ガイド

{visual_skill}
''' if visual_skill else ''}

全キャラ成人(18+)。JSON配列形式のみ出力。"""

    # ストーリー連続性セクション
    story_context_section = ""
    if story_so_far:
        story_context_section = f"""
## ⚠️ ストーリーの連続性（最重要）

以下は前のシーンまでの展開です。**必ずこの続きとして**シーンを書いてください。

{story_so_far}

---
"""

    # 各シーンの情報を組み立て
    scenes_info = []
    for scene in scenes:
        intensity = scene.get("intensity", 2)
        location = scene.get("location", "室内")
        time_of_day = scene.get("time", "")

        location_tags = ""
        for key, tags in loc_tags_db.items():
            if key in location:
                location_tags = tags
                break
        if not location_tags:
            location_tags = "indoor, room"

        time_tags = ""
        for key, tags in time_tags_db.items():
            if key in time_of_day:
                time_tags = tags
                break

        char_tags_str = ", ".join(char_danbooru_tags[:15]) if char_danbooru_tags else ""
        
        intensity_sd_tags = {
            5: f"ahegao, rolling_eyes, tongue_out, drooling, head_back, arched_back, tears, trembling, orgasm, fucked_silly, {theme_sd_expressions}",
            4: f"open_mouth, moaning, tears, sweating, head_back, arched_back, heavy_breathing, panting, clenched_fists, {theme_sd_expressions}",
            3: f"kiss, french_kiss, undressing, groping, blush, nervous, anticipation, parted_lips, heavy_breathing, {theme_sd_expressions}",
            2: f"eye_contact, close-up, romantic, blushing, hand_holding, leaning_close, {theme_sd_expressions}",
            1: f"portrait, smile, casual, standing, looking_at_viewer, {theme_sd_expressions}"
        }
        sd_intensity_tags = intensity_sd_tags.get(intensity, "")
        background_tags = f"{location_tags}, {time_tags}".strip(", ")
        if theme_sd_tags:
            background_tags = f"{background_tags}, {theme_sd_tags}"

        # 設定スタイルの背景タグ追加
        batch_setting_style = _detect_setting_style(context.get("setting", ""), theme=theme)
        if batch_setting_style:
            style_append = ", ".join(batch_setting_style.get("append", []))
            if style_append:
                background_tags = f"{background_tags}, {style_append}"

        composition_db = tag_db.get("compositions", {})
        composition_tags = composition_db.get(str(intensity), {}).get("tags", "")

        scenes_info.append({
            "scene": scene,
            "char_tags_str": char_tags_str,
            "sd_intensity_tags": sd_intensity_tags,
            "background_tags": background_tags,
            "composition_tags": composition_tags
        })

    # 設定スタイルのヒント行（バッチ共通）
    batch_setting_style = _detect_setting_style(context.get("setting", ""), theme=theme)
    batch_setting_hint = ""
    if batch_setting_style:
        batch_setting_hint = f"\n背景スタイル必須: {batch_setting_style.get('prompt_hint', '')}"

    # バッチプロンプト構築
    prompt_parts = []
    if story_context_section:
        prompt_parts.append(story_context_section)
    prompt_parts.append(f"設定: {json.dumps(context, ensure_ascii=False)}\n")
    prompt_parts.append(f"テーマ「{theme_name}」のトーン: {dialogue_tone}\n{heart_instruction}\n")
    if batch_setting_hint:
        prompt_parts.append(batch_setting_hint)

    for idx, info in enumerate(scenes_info):
        scene = info["scene"]
        prompt_parts.append(f"""
--- シーン{idx+1} ---
シーン情報: {json.dumps(scene, ensure_ascii=False)}
キャラ固有タグ: {info['char_tags_str']}
ポーズ・表情: {info['sd_intensity_tags']}
背景・場所: {info['background_tags']}
構図: {info['composition_tags']}
""")

    prompt_parts.append(f"""
## 出力形式（JSON配列で{len(scenes)}シーン分を出力）

[
  {{
    "scene_id": シーンID,
    "title": "シーンタイトル",
    "description": "このシーンの詳細説明",
    "location_detail": "場所の具体的な描写",
    "mood": "雰囲気",
    "character_feelings": {{
        "{char_names[0] if char_names else 'ヒロイン'}": "心情"
    }},
    "bubbles": [
        {{"speaker": "キャラ名", "type": "speech", "text": "短い一言"}}
    ],
    "onomatopoeia": [],
    "direction": "演出・ト書き",
    "story_flow": "次のシーンへの繋がり",
    "sd_prompt": "{QUALITY_POSITIVE_TAGS}, キャラ外見タグ, ポーズ・行為タグ, 表情タグ, 場所・背景タグ"
  }}
]

## ルール
1. 必ず{len(scenes)}シーン分のJSON配列を出力
2. 各シーンのscene_idは指定通りに
3. **bubblesは1-3個**（ヒロイン1-2個 + 男性0-1個。セリフの長さは自由）
4. sd_promptはシーン固有の描写タグのみ出力: キャラ外見+ポーズ+表情+エロ描写+アングル+場所・背景。**品質タグ(masterpiece, best_quality, score_9等)やLoRAタグ(<lora:...>)は絶対に含めるな**
5. **sd_promptにオノマトペ・日本語テキストを含めない**（英語のDanbooruタグのみ）
6. タグは重複なくカンマ区切り
7. **シーン1→シーン2は自然に繋がるストーリーにすること**
8. **前シーンまでの展開を必ず引き継ぐこと**
9. **同じセリフ・オノマトペを複数シーンで繰り返さない**
10. **type="moan"には喘ぎ声・声漏れのみ**。「そうなんだ」「汗すごい」等の説明文は絶対禁止
11. **type="speech"は感情的反応のみ**。「汗すごい」「震えてる」等の身体状態報告はナレーションであり禁止
13. **セリフに設備・家具の名前を入れない**。「便器」「便座」「手洗い台」等の設備名称はlocation_detailに書くもの。セリフでは「こんなとこで…」「ここで…」等の感情的な言い回しに置き換える
12. **story_flowは各シーン固有の展開**を書け。前シーンのコピペ禁止

JSON配列のみ出力。""")

    prompt = "\n".join(prompt_parts)

    system_with_cache = [
        {"type": "text", "text": common_system, "cache_control": {"type": "ephemeral"}},
    ]

    if callback:
        scene_ids = [s.get("scene_id") for s in scenes]
        callback(f"バッチ生成中: シーン {scene_ids} (Haiku, {len(scenes)}シーン一括)...")

    response = _call_api(
        client, MODELS["haiku"],
        system_with_cache,
        prompt, cost_tracker, 2500 * len(scenes), callback
    )

    # JSON配列をパース
    result_list = parse_json_response(response)

    if isinstance(result_list, dict):
        result_list = [result_list]

    # スキーマバリデーション（parse直後・各シーン）
    for _bi, _br in enumerate(result_list):
        if isinstance(_br, dict):
            _bv = validate_scene(_br, _bi)
            if not _bv["valid"]:
                _bsid = _br.get("scene_id", _bi + 1)
                for _be in _bv["errors"]:
                    log_message(f"  [SCHEMA] scene_batch(シーン{_bsid}): {_be}")

    for result in result_list:
        if isinstance(result, dict) and result.get("sd_prompt"):
            result["sd_prompt"] = deduplicate_sd_tags(result["sd_prompt"])

    while len(result_list) < len(scenes):
        missing_scene = scenes[len(result_list)]
        result_list.append({
            "scene_id": missing_scene.get("scene_id", len(result_list) + 1),
            "title": "生成不足",
            "mood": "一般",
            "bubbles": [],
            "onomatopoeia": [],
            "direction": "バッチ生成で不足",
            "sd_prompt": ""
        })

    return result_list[:len(scenes)]


def _generate_outline_chunk(
    client: anthropic.Anthropic,
    chunk_size: int,
    chunk_offset: int,
    total_scenes: int,
    theme_name: str,
    story_arc: str,
    key_emotions: list,
    elements_str: str,
    synopsis: str,
    char_names: list,
    act_info: str,
    previous_scenes: list,
    cost_tracker: CostTracker,
    callback: Optional[Callable] = None,
    extra_instructions: str = "",
) -> list:
    """アウトラインを10シーンずつチャンク生成（常にフル12フィールド形式）"""

    # 前チャンクの要約を構築（スライディングウィンドウ: 直近20件+古い分は1行要約）
    prev_summary = ""
    # v8.2根本修正: 完了済みアクション一覧を構築（重複防止の要）
    _completed_actions = ""
    if previous_scenes:
        prev_lines = []
        n_prev = len(previous_scenes)

        # 完了済みアクション・行為の一覧（全シーンから抽出）
        _action_set = set()
        for s in previous_scenes:
            sit = s.get("situation", "")[:40]
            if sit:
                _action_set.add(sit)
            ttl = s.get("title", "")
            if ttl:
                _action_set.add(ttl)
        # v8.5: 体位/行為キーワード抽出（直近チャンクからの繰り返し防止）
        _POSITION_KEYWORDS = [
            "騎乗位", "正常位", "バック", "立ちバック", "対面座位", "駅弁",
            "四つん這い", "寝バック", "側位", "松葉崩し", "背面座位",
            "膝立ち", "立位", "仰向け", "うつ伏せ", "跨が",
            "フェラ", "パイズリ", "手コキ", "クンニ", "69", "素股",
            "挿入", "中出し", "顔射", "口内射精", "二穴",
        ]
        _used_positions = set()
        for s in previous_scenes[-10:]:  # 直近10シーンから抽出
            sit = s.get("situation", "")
            for kw in _POSITION_KEYWORDS:
                if kw in sit:
                    _used_positions.add(kw)
        _position_warning = ""
        if _used_positions:
            _pos_list = sorted(_used_positions)
            _position_warning = (
                f"\n## ⚠️ 直近チャンクで使用済みの体位/行為（繰り返し厳禁）\n"
                f"🔁 {', '.join(_pos_list)}\n"
                f"**上記の体位/行為は直近で使用済み。必ず異なる体位/行為で新しい展開にすること。**\n"
            )

        if _action_set:
            _action_list = sorted(_action_set)[:30]  # 最大30件（トークン制限）
            _completed_actions = (
                "\n## ⚠️ 完了済みアクション（以下は既に描写済み。絶対に繰り返すな）\n"
                + "\n".join(f"❌ {a}" for a in _action_list)
                + "\n**上記と同一・類似のsituationやtitleは使用禁止。必ず新しい展開を書け。**\n"
            )
        # _action_setが空でも体位警告は有効
        _completed_actions += _position_warning

        # 古いシーン（20件より前）: 1行要約でトークン節約
        if n_prev > 20:
            prev_lines.append(f"（シーン1〜{n_prev - 20}: {n_prev - 20}シーン確定済み、省略）")
            # 最後の5シーンの要約だけ残す（古い分の雰囲気を伝える）
            old_start = max(0, n_prev - 30)
            old_end = n_prev - 20
            for s in previous_scenes[old_start:old_end]:
                sid = s.get("scene_id", "?")
                title = s.get("title", "")[:15]
                intensity = s.get("intensity", 3)
                prev_lines.append(f"[{sid}] {title} (i={intensity})")

        # 直近20件: フル要約
        recent_start = max(0, n_prev - 20)
        for s in previous_scenes[recent_start:]:
            sid = s.get("scene_id", "?")
            title = s.get("title", "")[:20]
            intensity = s.get("intensity", 3)
            situation = s.get("situation", "")[:60]
            loc = s.get("location", "")[:15]
            ea = s.get("emotional_arc", {})
            emo = f'{ea.get("start", "")}→{ea.get("end", "")}' if isinstance(ea, dict) else ""
            prev_lines.append(f"[{sid}] {title} (i={intensity}, {loc}) {situation} ({emo})")
        prev_summary = f"""## 確定済みシーン（これに続けて書くこと。重複禁止）
{chr(10).join(prev_lines)}
{_completed_actions}"""

    start_id = chunk_offset + 1
    end_id = chunk_offset + chunk_size

    output_format = (
        "## 出力形式（JSON配列）\n"
        f"シーン{start_id}〜{end_id}の{chunk_size}シーンをJSON配列で出力：\n"
        "{\n"
        f'    "scene_id": {start_id}〜{end_id}の番号,\n'
        '    "title": "シーンタイトル",\n'
        '    "goal": "このシーンの目的",\n'
        '    "location": "場所",\n'
        '    "time": "時間帯",\n'
        '    "situation": "このシーンで何が起きるか（具体的な状況）",\n'
        '    "story_flow": "前シーンからの繋がりと次シーンへの橋渡し",\n'
        '    "emotional_arc": {"start": "シーン冒頭の感情", "end": "シーン終わりの感情"},\n'
        '    "beats": ["展開ビート1", "展開ビート2", "展開ビート3"],\n'
        '    "intensity": 1から5の数値,\n'
        '    "erotic_level": "none/light/medium/heavy/climax",\n'
        '    "viewer_hook": "視聴者を引き付けるポイント"\n'
        "}\n\n"
        f"⚠️ 必ず{chunk_size}シーン（ID {start_id}〜{end_id}）全て出力すること。"
    )

    chunk_prompt = f"""{prev_summary}以下のストーリーあらすじに基づき、シーン{start_id}〜{end_id}（{chunk_size}シーン分）を生成してください。
全体は{total_scenes}シーンの作品です。

## ストーリーあらすじ
{synopsis}

## 登場キャラクター
{', '.join(char_names)}

## テーマ: {theme_name}
- ストーリーアーク: {story_arc}
- 重要な感情: {', '.join(key_emotions)}
- ストーリー要素:
{elements_str}

## シーン配分（全{total_scenes}シーン）
{act_info}
{extra_instructions}
{output_format}

## 絶対ルール
1. あらすじの内容を忠実にこのチャンク分に割り当てること
2. 確定済みシーンの直後から自然に繋がること（ストーリーのリセット・巻き戻り禁止）
3. situationは具体的に記述（抽象表現禁止）
4. 各シーンのsituationは前シーンと異なる具体的展開にすること
5. locationは作品を通じて自然に変わればよい。同じ場所で5シーン以上続く場合は場所内の位置を変えよ（例: ベッド→壁際→床）
6. emotional_arcのstartは前シーンのendと一致させること
7. intensity 4が4-5シーン連続したら、必ずintensity 3のシーン（体位変更・心理描写・休憩）を1つ挟むこと
8. story_flowは各シーン固有の内容を書け（重複禁止）。必ず感情変化・心理変化を含めること。「体位変更した」だけのstory_flowは禁止。
9. intensity 5は各mini-arc（15シーン程度）のクライマックスとして使え
10. titleは4-12文字。行為/体位/感情を反映。location名（「トイレ」「教室」等）をtitleに含めるな
11. 同じ場所が続く場合も場所内の位置を変えよ（例: 便座→壁際→洗面台→床）
12. 男性セリフは5パターン（脅迫/挑発/命令/嘲笑/独白）を均等に使え。観察実況（「～だな」）禁止

## ⚠️ 性行為進行フロー（体位カタログ禁止）
- 同じ体位で2-3シーン続いてよい。ただし各シーンで以下の少なくとも2つを変化させること:
  (1) 構図・アングル（正面→横→上から等）
  (2) テンポ・激しさ（ゆっくり→激しく→焦らし等）
  (3) 感情・心理の変化（恥じらい→受容→快楽等）
  (4) 性行為の進行段階（挿入→ビルドアップ→射精）
- 3-5シーンで「挿入→ピストン→快感上昇→射精/絶頂」のミニアークを作ること
- 体位変更は「射精後の次のラウンド」や「気分転換」として自然に行うこと
- 毎シーン体位を変える「体位カタログ」は禁止
- titleの重複禁止。同じキーワードを含むtitleは最大2回まで
- 確定済みシーンのsituation/titleと被らないこと
{f'''
## ⚠️ 最終チャンク特別ルール（エピローグ）
- これは作品の最後のチャンクです。ストーリーを適切に完結させてください
- 導入シーンの繰り返しは絶対禁止。第1幕の内容を再び書いてはならない
- intensity は3-4で余韻を描写（行為の事後、関係性の変化、心情の変化）
- 「呼び出し」「始まり」「出会い」等の導入表現は使わないこと
''' if end_id >= total_scenes else ''}
JSON配列のみ出力。\"\"\""""

    if callback:
        callback(f"[INFO]アウトラインチャンク生成: シーン{start_id}〜{end_id}")

    response = _call_api(
        client, MODELS["haiku"],
        f"FANZA同人CG集の脚本プランナーです。全{total_scenes}シーンのうちシーン{start_id}〜{end_id}の詳細設計をJSON配列で出力します。",
        chunk_prompt, cost_tracker, min(8192, chunk_size * 400), callback
    )

    chunk = parse_json_response(response)
    if not isinstance(chunk, list):
        chunk = [chunk] if isinstance(chunk, dict) else []

    # scene_idを正しいオフセットに修正
    for i, scene in enumerate(chunk):
        scene["scene_id"] = chunk_offset + i + 1

    return chunk


def _get_intensity_curve_instruction(theme_guide: dict) -> str:
    """テーマのintensity curveに応じたプロンプト指示を生成"""
    curve = theme_guide.get("intensity_curve", "ascending")
    if curve == "valley":
        return (
            "\n## intensity展開パターン: 谷あり型（このテーマ特有）\n"
            "このテーマでは中盤に「intensity が一時低下する谷」を作ること。\n"
            "例: 1→2→3→4→**3**→4→5→5→4（中盤で心理的葛藤・罪悪感・中断リスクにより一旦引く）\n"
            "谷の部分では心理描写を重視し、再上昇時により激しいエスカレーションにすること。"
        )
    elif curve == "staircase":
        return (
            "\n## intensity展開パターン: 階段型（段階的深化）\n"
            "このテーマでは同じintensityを2シーン連続で使ってから次の段階に上げること。\n"
            "例: 1→2→3→3→4→4→5→5→4（各段階を定着させてから次へ）\n"
            "調教/堕落の「慣れ→次のステップ」を表現する。"
        )
    elif curve == "wave":
        return (
            "\n## intensity展開パターン: 波型（緩急リズム）\n"
            "このテーマでは「上がって少し下がる」を繰り返し、波のようにエスカレートさせること。\n"
            "例: 1→3→2→4→3→4→3→5→5→4（焦らしを入れながら段階的に上げる）\n"
            "「Slow,slow,quick,slow」のダンスリズムで、焦らしの後は必ず前より強いintensityにすること。"
        )
    elif curve == "two_stage":
        return (
            "\n## intensity展開パターン: 2段ロケット型（クライマックス2回）\n"
            "このテーマでは1回目のクライマックスの後にintensityを落とし、2回目の山を作ること。\n"
            "例: 1→2→3→4→5→3→4→4→5→4（1回目絶頂→リセット→2回目はより激しく）\n"
            "2回目は1回目を超える激しさで、新しい体位・行為を導入すること。"
        )
    elif curve == "plateau":
        return (
            "\n## intensity展開パターン: 高原型（intensity 4を長く維持）\n"
            "このテーマではintensity 4の状態を長く維持し、最後に一気にクライマックスへ。\n"
            "例: 1→2→3→4→4→4→4→4→5→4（じわじわ快感を蓄積し最後に爆発）\n"
            "高原部分では体位・アングルの変化で単調さを防ぐこと。\n"
            "ただしi=4が4-5シーン連続したら必ずi=3を1つ挟むこと。"
        )
    # ascending（デフォルト）
    return (
        "\n## intensity展開パターン: 上昇型（基本）\n"
        "基本は1→2→3→4→5の上昇型だが、単調な右肩上がりにしないこと。\n"
        "本番パートはmini-arc（小さな山）の連続で構成せよ:\n"
        "各mini-arc: i=3(転換/休憩)→i=4→i=4→i=4→i=5(小クライマックス)→i=3(次へ)\n"
        "例（30シーン）: 1→2→2→3→3→4→4→4→5→3→4→4→4→4→5→3→4→4→4→5→3→4→4→4→4→5→4→3→3→4\n"
        "**i=4が4-5シーン連続したら必ずi=3を1つ挟むこと。これは絶対ルール。**"
    )


def generate_outline(
    client: anthropic.Anthropic,
    context: dict,
    num_scenes: int,
    theme: str,
    cost_tracker: CostTracker,
    callback: Optional[Callable] = None,
    synopsis: str = "",
    story_structure: dict = None,
    male_description: str = "",
    faceless_male: bool = True,
) -> list:
    """あらすじをシーン分割してアウトライン生成（Haiku API 1回）"""
    theme_guide = THEME_GUIDES.get(theme, THEME_GUIDES.get("vanilla", {}))
    theme_name = theme_guide.get("name", "指定なし")
    story_arc = theme_guide.get("story_arc", "導入→展開→本番→余韻")
    key_emotions = theme_guide.get("key_emotions", ["期待", "緊張", "快感", "幸福"])
    story_elements = theme_guide.get("story_elements", [])

    if callback:
        callback(f"[INFO]{theme_name}テーマでシーン分割中（AI生成）...")

    chars = context.get("chars", [])
    char_names = [c["name"] for c in chars] if chars else ["ヒロイン"]

    # シーン配分計算（テーマ別ratio → ユーザー設定 → デフォルト）
    # テーマ別ratioがある場合はそれを優先
    theme_intro = theme_guide.get("intro_ratio", 0.08)
    theme_foreplay = theme_guide.get("foreplay_ratio", 0.20)
    # v8.9: 大規模スクリプト時のintro/foreplay上限（50+シーンで導入肥大化を防止）
    if num_scenes >= 50:
        _max_intro_scenes = 5  # 50シーン以上: 導入は最大5シーン
        _max_foreplay_scenes = 15  # 前戯は最大15シーン
        _capped_intro = min(theme_intro, _max_intro_scenes / num_scenes)
        _capped_foreplay = min(theme_foreplay, _max_foreplay_scenes / (num_scenes * 0.8))
        if _capped_intro < theme_intro:
            log_message(f"v8.9: intro_ratio {theme_intro:.2f}→{_capped_intro:.2f}（{num_scenes}シーン: 導入{_max_intro_scenes}上限）")
        if _capped_foreplay < theme_foreplay:
            log_message(f"v8.9: foreplay_ratio {theme_foreplay:.2f}→{_capped_foreplay:.2f}（{num_scenes}シーン: 前戯{_max_foreplay_scenes}上限）")
        theme_intro = _capped_intro
        theme_foreplay = _capped_foreplay
    if story_structure is None:
        story_structure = {"prologue": round(theme_intro * 100), "main": 80, "epilogue": 10}
    prologue_pct = story_structure.get("prologue", round(theme_intro * 100)) / 100
    epilogue_pct = story_structure.get("epilogue", 10) / 100

    act1 = max(1, round(num_scenes * prologue_pct))       # プロローグ
    act4 = max(1, round(num_scenes * epilogue_pct))        # エピローグ
    main_scenes = num_scenes - act1 - act4                  # 本編合計
    if main_scenes < 2:
        main_scenes = 2
        act1 = max(1, num_scenes - main_scenes - 1)
        act4 = num_scenes - act1 - main_scenes
    act2 = max(1, round(main_scenes * theme_foreplay))     # 前戯（テーマ別比率）
    act3 = main_scenes - act2                              # 本番（残り）

    elements_str = chr(10).join(f'・{e}' for e in story_elements) if story_elements else "・特になし"

    # 常にフル12フィールド形式を使用（チャンク生成で大量シーンにも対応）
    output_format_section = (
        "## 出力形式（JSON配列）\n"
        "各シーンは以下の形式：\n"
        "{\n"
        '    "scene_id": シーン番号,\n'
        '    "title": "シーンタイトル",\n'
        '    "goal": "このシーンの目的",\n'
        '    "location": "場所",\n'
        '    "time": "時間帯",\n'
        '    "situation": "このシーンで何が起きるか（具体的な状況）",\n'
        '    "story_flow": "前シーンからの繋がりと次シーンへの橋渡し",\n'
        '    "emotional_arc": {"start": "シーン冒頭の感情", "end": "シーン終わりの感情"},\n'
        '    "beats": ["展開ビート1", "展開ビート2", "展開ビート3"],\n'
        '    "intensity": 1から5の数値,\n'
        '    "erotic_level": "none/light/medium/heavy/climax",\n'
        '    "viewer_hook": "視聴者を引き付けるポイント"\n'
        "}"
    )

    # ストーリーパターン自動選択（エロ漫画定番パターンで整合性向上）
    story_pattern_section = ""
    try:
        from ero_dialogue_pool import select_story_pattern
        matched_pattern = select_story_pattern(theme, synopsis)
        if matched_pattern:
            beats_str = "\n".join(f"  {i+1}. {b}" for i, b in enumerate(matched_pattern["beats"]))
            i_pattern = " → ".join(str(x) for x in matched_pattern["intensity_pattern"])
            story_pattern_section = (
                f"\n## 参考ストーリーパターン: {matched_pattern['name']}\n"
                f"以下は参考用の定番展開ビートです。あらすじの内容を優先しつつ、"
                f"展開のテンポや感情の流れを参考にしてください:\n"
                f"{beats_str}\n"
                f"セリフ進化: {matched_pattern['dialogue_evolution']}\n"
                f"intensity展開の参考: {i_pattern}\n"
            )
    except ImportError:
        pass

    # 大量シーン時の追加指示
    long_script_section = ""
    if num_scenes >= 25:
        # mini-arc数を計算（15-20シーンごとに1つのmini-arc）
        _mini_arc_count = max(3, act3 // 15)
        _mini_arc_size = act3 // _mini_arc_count
        _i5_peaks = max(2, _mini_arc_count)  # mini-arcごとに1つのi=5ピーク
        long_script_section = f"""
## ⚠️ 大量シーン（{num_scenes}シーン）追加ルール

1. **本番パートにmini-arc**: {act3}シーンの本番パートは、{_mini_arc_count}個のmini-arc（各約{_mini_arc_size}シーン）に分割せよ。各mini-arcは「挿入→ビルドアップ→絶頂→射精」の性行為フローで構成し、intensity配分は「i=3(導入/転換)→i=4(エスカレート)→i=5(小クライマックス)→i=3(休憩/心理描写)」
2. **intensity 4の連続上限4-5**: intensity 4が4-5シーン連続したら、必ず間にintensity 3のシーン（休憩/体位変更/心理描写）を1つ挟むこと。これは絶対ルール
3. **intensity 5は{_i5_peaks}回**: 各mini-arcのクライマックスでintensity 5を使え（合計{_i5_peaks}回）。最後のmini-arcのi=5が全体のクライマックス
4. **男性セリフ多様性**: 男性のセリフは5パターン（脅迫/挑発/命令/嘲笑/独白）を均等に使え。同じ意味のセリフの連続禁止。末尾フレーズの重複は最大2回まで。観察実況（「～だな」「～してるな」）禁止
5. **locationの変化**: 同じ場所で5シーン以上続く場合は場所内の位置を変えよ（例: 便座→壁際→洗面台→床）
6. **titleルール**: titleは4-12文字、行為/体位/感情を反映。location名（「トイレ」「教室」等）をtitleに含めてはならない
7. **第4幕（余韻）はリセット禁止**: 余韻シーンはintensity 3-4で、行為の事後・余韻・関係性の変化を描け。導入シーンの繰り返しは絶対禁止
8. **mood多様性**: moodは毎シーン異なる表現にすること。同じmoodの連続使用禁止。intensityが同じでも表現を変えよ
"""

    # v8.8: テーマ世界ルール注入
    _world_rules = theme_guide.get("world_rules", [])
    _world_rules_section = ""
    if _world_rules:
        _world_rules_section = "\n## ⚠️ テーマ世界設定ルール（厳守）\n"
        for _wr in _world_rules:
            _world_rules_section += f"- {_wr}\n"

    prompt = f"""以下のストーリーあらすじを{num_scenes}シーンに分割し、各シーンの詳細をJSON配列で出力してください。

## ストーリーあらすじ（これに忠実に分割すること）
{synopsis}

## 登場キャラクター
{', '.join(char_names)}
{f"{chr(10)}## 男性キャラクター外見: {male_description}{chr(10)}※ situationやbeatsに男性が登場する場合、この外見設定を反映すること{chr(10)}{'※ 男性はfaceless male（顔なし）。顔の特徴は描写しない' if faceless_male else '※ 男性の外見を一貫して描写すること'}{chr(10)}" if male_description else (f"{chr(10)}## 男性キャラクター{chr(10)}{'※ 男性はfaceless male（顔なし）。顔の特徴は描写しない' if faceless_male else '※ 男性の外見を一貫して描写すること'}{chr(10)}" if True else "")}
## テーマ: {theme_name}
- ストーリーアーク: {story_arc}
- 重要な感情: {', '.join(key_emotions)}
- ストーリー要素:
{elements_str}
{story_pattern_section}

## シーン配分（{num_scenes}シーン・エロ70%以上）
- 第1幕・導入: {act1}シーン → intensity 1-2（状況設定だが手を抜くな。キャラの魅力・空気の変化・予感を丁寧に。ここの品質が全体を決める）
- 第2幕・前戯: {act2}シーン → intensity 3（焦らし=エロの核。「まだしない」の緊張感。間接的接触→直接接触へのゆっくりした移行）
- 第3幕・本番: {act3}シーン → intensity 3-5を使い分けること。i=4を基本としつつ、5シーンごとにi=3の緩急を入れ、各mini-arcのクライマックスでi=5を使え
- 第4幕・余韻: {act4}シーン → intensity 3-4（事後・余韻。エロの余韻を残す。第1幕のリピート禁止）
※ FANZA CG集は読者がエロを求めて購入する。導入は短く、エロシーンを手厚く。
{_get_intensity_curve_instruction(theme_guide)}
{long_script_section}
{_world_rules_section}
{output_format_section}

## 絶対ルール
1. あらすじの内容を全シーンに漏れなく割り当てること
2. あらすじにない展開を勝手に追加しないこと
3. situationはあらすじの該当部分を具体的に記述すること（抽象表現禁止）
4. 各シーンは前シーンの直後から始まり、自然に繋がること
5. 本番シーン（intensity 4-5）は段階的にエスカレートすること
6. 最後から2番目のシーンがクライマックス（intensity 5）であること
7. 各シーンのsituationは必ず前シーンと異なる具体的展開にすること（「近づく」「囲まれる」等の同パターン繰り返し禁止）
8. locationは作品を通じて自然に変わればよい。同じ場所で5シーン以上続く場合は場所内の位置を変えよ（例: ベッド→壁際→床）
9. intensity 5のシーン数: 20シーン以下は最大2シーン、それ以上は15シーンにつき1回（例: 30シーン→2回、60シーン→4回、100シーン→6回）
10. intensity 1の次にintensity 3以上は禁止。必ずintensity 2を挟むこと（1→2→3→4→5の段階的上昇）
11. **視点**: situationは女性キャラ視点で記述。男性の行動ではなく、女性の体験・反応・感情を中心に書く

{_get_time_axis_instruction(theme, act1)}
## ⚠️⚠️ 性行為進行フロー（最重要・体位カタログ厳禁）

**本番シーン（intensity 4-5）は「性行為の流れ」で構成すること。体位を毎シーン変える「体位カタログ」は厳禁。**

### 正しい進行フロー例（3-5シーンで1サイクル）
1. 挿入・開始 → 2. テンポ上昇・快感蓄積 → 3. 激しさピーク → 4. 射精/絶頂 → 5. 余韻or体位変更

### 同じ体位の連続シーンで変化させる要素（毎シーン最低2つ変化）
1. **構図・アングル**: 正面/横/上から/下から/クローズアップ — 視覚的変化を出す
2. **テンポ・激しさ**: ゆっくり→激しく→焦らし→一気に — 緩急をつける
3. **感情・心理**: 前シーンの心理の「次の段階」を必ず記述（恥じらい→受容→快楽→陶酔）
4. **性行為の段階**: 挿入→ピストン→快感上昇→限界→射精。段階的に進行すること
5. **焦点部位**: 胸/腰/脚/首筋/耳/背中 — 毎シーン異なる部位を描写

### 体位変更のタイミング
- 射精/絶頂後の「次のラウンド」として体位変更するのが自然
- mini-arc（3-5シーン）の区切りで変更
- 同じ体位は最大4シーン連続まで

❌ 禁止: 毎シーン体位が変わる「体位ショーケース」
❌ 禁止: 同じsituation表現が3シーン以上
❌ 禁止: titleに同じ単語が3回以上出現
✅ 推奨: 正常位で3シーン（挿入→激しく→射精）→騎乗位で3シーン（回復→主導権移動→再絶頂）

### titleの多様性ルール
- 全シーンのtitleは重複禁止（完全一致禁止）
- 同じキーワードを含むtitleは最大2回まで
- 具体的な行為・体位・場所・感情を反映した固有のタイトルにすること

## ⚠️ エスカレーション段階ルール（飛躍禁止・最重要）

### 行為の段階（必ずこの順序で進めること。段階をスキップ禁止）
段階A: 会話・接近・ムード作り（intensity 1-2）
段階B: キス・愛撫・脱衣・前戯（intensity 3）
段階C: 1対1の性行為（intensity 4）
段階D: 激しい1対1 or 複数人（intensity 4-5）
段階E: クライマックス（intensity 5）

❌ 段階B（前戯）→ 段階D（複数人）は禁止。必ず段階C（1対1性行為）を挟むこと
❌ 1対1シーンの次にいきなり3人以上は禁止。1対1→2人→複数人と段階的に増やすこと

### 相手人数の段階
- 1人のシーンの次に3人以上のシーンは禁止
- 複数人に移行するなら、間に「相手が増える過程」のシーンを挟むこと

### 心理変化の段階
- 「抵抗している」→「完全堕落」の1シーン飛躍は禁止
- 抵抗→葛藤→受容→快楽→堕落の順で段階的に変化させること
- emotional_arcのstartは必ず前シーンのendと一致させること

### situationの具体性
各シーンのsituationには以下を必ず明記すること:
- 相手の人数（1人/2人/複数人）
- 行為の具体的内容（キス/愛撫/挿入/体位名）
- 場所の移動理由（前シーンと場所が変わる場合）

JSON配列のみ出力。"""

    # シーン配分情報文字列（チャンク生成でも共有）
    act_info = (
        f"- 第1幕・導入: {act1}シーン → intensity 1-2\n"
        f"- 第2幕・前戯: {act2}シーン → intensity 3\n"
        f"- 第3幕・本番: {act3}シーン → intensity 4-5\n"
        f"- 第4幕・余韻: {act4}シーン → intensity 3-4"
    )

    try:
        if num_scenes <= 12:
            # 12シーン以下: シングルコール（Haiku4.5で品質確保）
            outline_max_tokens = min(8192, max(2048, num_scenes * 400))
            response = _call_api(
                client, MODELS["haiku"],
                f"FANZA同人CG集の脚本プランナーです。ストーリーあらすじを忠実に{num_scenes}シーンに分割し、各シーンの詳細設計をJSON配列で出力します。",
                prompt, cost_tracker, outline_max_tokens, callback
            )
            outline = parse_json_response(response)
            if not isinstance(outline, list) or len(outline) == 0:
                raise ValueError("Invalid outline response")
        else:
            # 13シーン以上: チャンク分割生成（10シーンずつ、常にフル形式）
            chunk_size = 10
            outline = []
            # v8.2根本修正: long_script_section/intensity_curve/story_patternをチャンクにも渡す
            _chunk_extra_instructions = ""
            if long_script_section:
                _chunk_extra_instructions += long_script_section
            _chunk_extra_instructions += _get_intensity_curve_instruction(theme_guide)
            # v8.9: テーマ別時間軸ルール注入
            _chunk_extra_instructions += _get_time_axis_instruction(theme, act1)
            if story_pattern_section:
                _chunk_extra_instructions += story_pattern_section
            # v8.8: テーマ世界ルール注入
            _world_rules = theme_guide.get("world_rules", [])
            if _world_rules:
                _chunk_extra_instructions += "\n## ⚠️ テーマ世界設定ルール（厳守）\n"
                for _wr in _world_rules:
                    _chunk_extra_instructions += f"- {_wr}\n"
            # v8.8: 同パターン繰り返し禁止（大量シーンで第2サイクルが第1サイクルのコピーになる問題を防止）
            if num_scenes >= 25:
                _chunk_extra_instructions += """
## ⚠️ 同パターン繰り返し禁止
- 確定済みシーンと同じsituation展開パターンを繰り返すな。場所が変わっても「観察→脱衣→愛撫→挿入→体位変更→射精」の同一アークの再利用は禁止
- 第2ラウンドは第1ラウンドと異なるアプローチにせよ（例: 第1が受動的→第2は能動的/堕ち、第1が恐怖→第2は快楽依存）
- 確定済みシーンのsituation/title一覧を確認し、同じ表現・キーワードの3回以上使用を避けよ
"""
            for offset in range(0, num_scenes, chunk_size):
                this_chunk = min(chunk_size, num_scenes - offset)
                log_message(f"チャンクアウトライン: シーン{offset+1}〜{offset+this_chunk} ({this_chunk}シーン)")
                chunk = _generate_outline_chunk(
                    client, this_chunk, offset, num_scenes,
                    theme_name, story_arc, key_emotions, elements_str,
                    synopsis, char_names, act_info,
                    outline,  # 確定済みシーンを渡す
                    cost_tracker, callback,
                    extra_instructions=_chunk_extra_instructions,
                )
                outline.extend(chunk)
                log_message(f"チャンク完了: {len(chunk)}シーン取得、合計{len(outline)}シーン")

            if len(outline) == 0:
                raise ValueError("Chunk outline generation failed")

        # スキーマバリデーション（setdefault補完前に実行して欠落を検出）
        _outline_val = validate_outline(outline, num_scenes)
        if not _outline_val["valid"]:
            for _oe in _outline_val["errors"]:
                log_message(f"  [SCHEMA] outline(parse直後): {_oe}")

        # 必須フィールドの補完
        for i, scene in enumerate(outline):
            scene.setdefault("scene_id", i + 1)
            scene.setdefault("title", f"シーン{i+1}")
            scene.setdefault("goal", "")
            scene.setdefault("location", "室内")
            scene.setdefault("time", "")
            scene.setdefault("situation", "")
            scene.setdefault("story_flow", "")
            scene.setdefault("emotional_arc", {"start": "", "end": ""})
            scene.setdefault("beats", [])
            scene.setdefault("intensity", 3)
            scene.setdefault("erotic_level", "medium")
            scene.setdefault("viewer_hook", "")

        # intensity分布の自動修正
        _n_scenes = len(outline)
        _max_i5 = max(2, _n_scenes // 15)  # 15シーンにつき1回のi=5ピーク

        # 1. intensity 5の上限制御（mini-arc分散）
        intensity_5_count = sum(1 for s in outline if s.get("intensity", 3) == 5)
        if intensity_5_count > _max_i5:
            five_indices = [i for i, s in enumerate(outline) if s.get("intensity", 3) == 5]
            # 均等分散: mini-arcごとに1つのi=5を残す
            _spacing = max(1, len(five_indices) // _max_i5)
            keep_five = set(five_indices[i] for i in range(0, len(five_indices), _spacing))
            # 最後のi=5は必ず残す
            keep_five.add(five_indices[-1])
            if len(five_indices) >= 2:
                keep_five.add(five_indices[-2])
            # _max_i5個に制限
            keep_five = sorted(keep_five)[-_max_i5:]
            for i in five_indices:
                if i not in keep_five:
                    outline[i]["intensity"] = 4
            log_message(f"intensity 5を{intensity_5_count}→{len(keep_five)}シーンに自動修正（{_max_i5}上限）")

        # 2. intensity 5が不足時の自動挿入（50シーン以上で不足なら追加）
        # v8.9: nearbyチェック改善（近接i=5がある場合は±5-8シーンずらして挿入）
        if _n_scenes >= 50:
            _current_i5 = sum(1 for s in outline if s.get("intensity", 3) == 5)
            if _current_i5 < _max_i5:
                _act3_start = act1 + act2
                _act3_end = _act3_start + act3
                _need_i5 = _max_i5 - _current_i5
                _arc_size = act3 // _max_i5 if _max_i5 > 0 else act3
                _existing_5_positions = {i for i, s in enumerate(outline) if s.get("intensity", 3) == 5}
                for arc_idx in range(_max_i5):
                    if _need_i5 <= 0:
                        break
                    peak_pos = _act3_start + (arc_idx + 1) * _arc_size - 1
                    peak_pos = min(peak_pos, _act3_end - 1)
                    if peak_pos >= len(outline) or outline[peak_pos].get("intensity", 3) == 5:
                        continue
                    # 近接チェック: ±5シーン以内にi=5があれば位置をずらす
                    _nearby_5 = any(abs(peak_pos - p) <= 5 for p in _existing_5_positions)
                    if _nearby_5:
                        # ±6-10シーンの範囲で空きスロットを探す
                        _found_alt = False
                        for _offset in [6, -6, 7, -7, 8, -8, 9, -9, 10, -10]:
                            alt_pos = peak_pos + _offset
                            if (_act3_start <= alt_pos < _act3_end and
                                alt_pos < len(outline) and
                                outline[alt_pos].get("intensity", 3) != 5 and
                                not any(abs(alt_pos - p) <= 5 for p in _existing_5_positions)):
                                peak_pos = alt_pos
                                _found_alt = True
                                break
                        if not _found_alt:
                            continue  # どうしても挿入できない場合はスキップ
                    # ピーク前のシーンをi=4にランプアップ（3→5飛躍防止）
                    if peak_pos > 0 and outline[peak_pos - 1].get("intensity", 3) < 4:
                        outline[peak_pos - 1]["intensity"] = 4
                    outline[peak_pos]["intensity"] = 5
                    _existing_5_positions.add(peak_pos)
                    _need_i5 -= 1
                    log_message(f"シーン{peak_pos+1}: mini-arcクライマックスとしてi=5挿入")

        # 3. intensity 1→3以上の飛躍を修正
        for i in range(1, len(outline)):
            prev_intensity = outline[i-1].get("intensity", 3)
            curr_intensity = outline[i].get("intensity", 3)
            if prev_intensity == 1 and curr_intensity >= 3:
                outline[i]["intensity"] = 2
                log_message(f"シーン{i+1}: intensity {curr_intensity}→2に修正（1→3以上の飛躍防止）")

        # 4. intensity 2段階以上の上昇飛躍を修正（2→4, 2→5, 3→5 等）
        # v8.2根本修正: i=5ピークを保護（3→5の場合は前のシーンを4にランプアップ）
        for i in range(1, len(outline)):
            prev_intensity = outline[i-1].get("intensity", 3)
            curr_intensity = outline[i].get("intensity", 3)
            if curr_intensity - prev_intensity >= 2:
                if curr_intensity == 5 and prev_intensity >= 3:
                    # i=5ピーク保護: ピークを維持し、前のシーンをランプアップ
                    outline[i-1]["intensity"] = 4
                    log_message(f"シーン{i}: intensity {prev_intensity}→4にランプアップ（i=5ピーク保護）")
                else:
                    fixed = prev_intensity + 1
                    outline[i]["intensity"] = fixed
                    log_message(f"シーン{i+1}: intensity {curr_intensity}→{fixed}に修正（{prev_intensity}→{curr_intensity}の上昇飛躍防止）")

        # 5. intensity 3段階以上の下降ジャンプを修正（5→2, 5→1, 4→1 等）
        for i in range(1, len(outline)):
            prev_intensity = outline[i-1].get("intensity", 3)
            curr_intensity = outline[i].get("intensity", 3)
            if prev_intensity - curr_intensity >= 3:
                fixed = prev_intensity - 2
                outline[i]["intensity"] = fixed
                log_message(f"シーン{i+1}: intensity {curr_intensity}→{fixed}に修正（{prev_intensity}→{curr_intensity}の急降下防止）")

        # 6. consecutive i=4上限: 4-6シーン連続でi=3ブレイクを強制挿入
        # v8.9: 固定5→ランダム4-6に変更（機械的6シーンパターン防止）
        import random as _rng
        _consecutive_4 = 0
        _break_count = 0
        _i4_break_limit = _rng.randint(4, 6)  # 初回のブレイク閾値
        for i, s in enumerate(outline):
            if s.get("intensity", 3) == 4:
                _consecutive_4 += 1
                if _consecutive_4 > _i4_break_limit:
                    # i=5ピーク直前にブレイクを入れると、Step8aでピークが潰されるため回避
                    _next_is_peak = (i + 1 < len(outline) and outline[i + 1].get("intensity", 3) == 5)
                    if _next_is_peak:
                        if i > 0 and outline[i - 1].get("intensity", 3) == 4:
                            outline[i - 1]["intensity"] = 3
                            _consecutive_4 = 1
                            _break_count += 1
                    else:
                        s["intensity"] = 3
                        _consecutive_4 = 0
                        _break_count += 1
                    _i4_break_limit = _rng.randint(4, 6)  # 次のブレイク閾値をランダム再設定
            else:
                _consecutive_4 = 0
        if _break_count > 0:
            log_message(f"i=4連続上限(4-6ランダム): {_break_count}箇所にi=3ブレイク挿入")

        # 6b. consecutive i≤2上限: act1+1を超えたら強制i=3化（導入肥大化防止）
        # v8.9: 上限をact1+1に動的設定（100シーンact1=5→上限6、20シーンact1=2→上限3）
        _max_consecutive_low = min(act1 + 1, 6)  # 最大でも6シーン連続まで
        _consecutive_low = 0
        _low_break_count = 0
        for i, s in enumerate(outline):
            if s.get("intensity", 3) <= 2:
                _consecutive_low += 1
                if _consecutive_low > _max_consecutive_low:
                    s["intensity"] = 3
                    _consecutive_low = 0
                    _low_break_count += 1
            else:
                _consecutive_low = 0
        if _low_break_count > 0:
            log_message(f"i≤2連続上限{_max_consecutive_low}: {_low_break_count}箇所にi=3挿入（導入肥大化防止）")

        # 6b2. i≤2の総数上限: act1+2を超えたらact1近辺のi≤2を順次i=3に昇格
        _total_low = sum(1 for s in outline if s.get("intensity", 3) <= 2)
        _max_total_low = act1 + 2  # act1=5なら最大7シーンまで
        if _total_low > _max_total_low:
            _low_indices = [i for i, s in enumerate(outline) if s.get("intensity", 3) <= 2]
            _excess = _total_low - _max_total_low
            # 後方（act1境界付近）のi≤2から昇格
            for idx in reversed(_low_indices):
                if _excess <= 0:
                    break
                outline[idx]["intensity"] = 3
                _excess -= 1
            log_message(f"i≤2総数上限{_max_total_low}: {_total_low}→{_max_total_low}に削減")

        # 6c. i=4の総数上限制御（40%超の場合、超過分をact境界に基づいて再割当）
        # v8.9: act境界ベースに改善（位置ratioの穴を解消）
        _i4_count = sum(1 for s in outline if s.get("intensity", 3) == 4)
        _i4_ratio = _i4_count / max(_n_scenes, 1)
        if _i4_ratio > 0.40:
            _i4_target = int(_n_scenes * 0.40)
            _i4_excess = _i4_count - _i4_target
            _i4_indices = [i for i, s in enumerate(outline) if s.get("intensity", 3) == 4]
            _current_i5_6c = sum(1 for s in outline if s.get("intensity", 3) == 5)
            _act3_start_6c = act1 + act2
            _act3_end_6c = _act3_start_6c + act3
            _rebalance_count = 0
            for idx_6c in _i4_indices:
                if _rebalance_count >= _i4_excess:
                    break
                if idx_6c < act1:
                    # Act1（導入） → i=2に降格
                    outline[idx_6c]["intensity"] = 2
                    _rebalance_count += 1
                elif idx_6c < _act3_start_6c:
                    # Act2（前戯） → i=3に降格
                    outline[idx_6c]["intensity"] = 3
                    _rebalance_count += 1
                elif idx_6c >= _act3_end_6c:
                    # Act4（エピローグ） → i=3に降格
                    outline[idx_6c]["intensity"] = 3
                    _rebalance_count += 1
                elif _current_i5_6c < _max_i5:
                    # Act3内でi=5が不足なら昇格を試行
                    _near_5_6c = any(
                        outline[j].get("intensity", 3) == 5
                        for j in range(max(0, idx_6c - 5), min(len(outline), idx_6c + 6))
                    )
                    if not _near_5_6c:
                        outline[idx_6c]["intensity"] = 5
                        _current_i5_6c += 1
                        _rebalance_count += 1
                        if idx_6c > 0 and outline[idx_6c - 1].get("intensity", 3) < 4:
                            outline[idx_6c - 1]["intensity"] = 4
                    else:
                        # i=5近接でも、Act3前半なら i=3に降格（ブレイク追加）
                        _act3_progress = (idx_6c - _act3_start_6c) / max(act3, 1)
                        if _act3_progress < 0.3:
                            outline[idx_6c]["intensity"] = 3
                            _rebalance_count += 1
            if _rebalance_count > 0:
                log_message(f"i=4上限40%制御: {_i4_count}→{_i4_count - _rebalance_count}シーンに再割当（{_rebalance_count}件変更）")

        # 7. エピローグのストーリーリセット防止: 最終act4がi=2以下にならないよう制限
        if _n_scenes >= 20:
            _epilogue_start = act1 + act2 + act3
            for i in range(_epilogue_start, len(outline)):
                curr_i = outline[i].get("intensity", 3)
                if curr_i <= 2:
                    outline[i]["intensity"] = 3
                    log_message(f"シーン{i+1}: エピローグi={curr_i}→3に修正（余韻リセット防止）")

        # 8. 最終スムージング: step6/7が生成した飛躍を修正し、制約を再適用
        _epilogue_start_8 = (act1 + act2 + act3) if _n_scenes >= 20 else len(outline)
        for _pass in range(3):  # 最大3パスで収束
            _smooth_count = 0
            # 8a. 飛躍修正（前方パス）- i=5ピークは保護
            for i in range(1, len(outline)):
                prev_i = outline[i-1].get("intensity", 3)
                curr_i = outline[i].get("intensity", 3)
                if curr_i - prev_i >= 2:
                    if curr_i == 5 and prev_i >= 3:
                        # i=5ピーク保護: ピークを下げるのではなく前のシーンをランプアップ
                        outline[i-1]["intensity"] = 4
                        _smooth_count += 1
                    else:
                        outline[i]["intensity"] = prev_i + 1
                        _smooth_count += 1
                elif prev_i - curr_i >= 3:
                    outline[i]["intensity"] = prev_i - 2
                    _smooth_count += 1
            # 8b. consecutive i=4上限の再適用（i=5ピーク直前保護付き）
            _consecutive_4_8 = 0
            for i, s in enumerate(outline):
                if s.get("intensity", 3) == 4:
                    _consecutive_4_8 += 1
                    if _consecutive_4_8 > _i4_break_limit:  # v8.9: Step 6bと同じランダム閾値
                        _next_is_peak = (i + 1 < len(outline) and outline[i + 1].get("intensity", 3) == 5)
                        if _next_is_peak:
                            if i > 0 and outline[i - 1].get("intensity", 3) == 4:
                                outline[i - 1]["intensity"] = 3
                                _consecutive_4_8 = 1
                                _smooth_count += 1
                        else:
                            s["intensity"] = 3
                            _consecutive_4_8 = 0
                            _smooth_count += 1
                else:
                    _consecutive_4_8 = 0
            # 8b2. consecutive i≤2上限の再適用（v8.9: 動的上限）
            _consecutive_low_8 = 0
            for i, s in enumerate(outline):
                if s.get("intensity", 3) <= 2:
                    _consecutive_low_8 += 1
                    if _consecutive_low_8 > _max_consecutive_low:
                        s["intensity"] = 3
                        _consecutive_low_8 = 0
                        _smooth_count += 1
                else:
                    _consecutive_low_8 = 0
            # 8c. エピローグi<=2防止の再適用（境界の飛躍も修正）
            for i in range(_epilogue_start_8, len(outline)):
                if outline[i].get("intensity", 3) <= 2:
                    outline[i]["intensity"] = 3
                    _smooth_count += 1
            # エピローグ境界: 直前シーンがi<=1だとi=3へ+2飛躍になるので直前をi=2に引き上げ
            if _epilogue_start_8 > 0 and _epilogue_start_8 < len(outline):
                _pre_epi = outline[_epilogue_start_8 - 1].get("intensity", 3)
                if _pre_epi < 2:
                    outline[_epilogue_start_8 - 1]["intensity"] = 2
                    _smooth_count += 1
            if _smooth_count == 0:
                break
        if _pass > 0:
            if _smooth_count == 0:
                log_message(f"最終スムージング: {_pass + 1}パスで収束")
            else:
                log_message(f"最終スムージング: 3パスで未収束（残り{_smooth_count}箇所）")

        # 9. v8.9: Act間エスカレーション検証（Act3がAct2以下ならブースト）
        if _n_scenes >= 30:
            _act_ranges = [
                (0, act1),
                (act1, act1 + act2),
                (act1 + act2, act1 + act2 + act3),
                (act1 + act2 + act3, len(outline)),
            ]
            _act_avgs = []
            for _start, _end in _act_ranges:
                _end = min(_end, len(outline))
                _count = _end - _start
                if _count > 0:
                    _avg = sum(outline[i].get("intensity", 3) for i in range(_start, _end)) / _count
                else:
                    _avg = 3.0
                _act_avgs.append(_avg)

            # Act3 avg がAct2 avg以下ならAct3をブースト
            if len(_act_avgs) >= 3 and _act_avgs[2] <= _act_avgs[1] + 0.2:
                _boost_count = 0
                _act3_s = act1 + act2
                _act3_e = min(act1 + act2 + act3, len(outline))
                for i in range(_act3_s, _act3_e):
                    if outline[i].get("intensity", 3) == 3:
                        outline[i]["intensity"] = 4
                        _boost_count += 1
                        if _boost_count >= max(2, act3 // 10):
                            break
                if _boost_count > 0:
                    log_message(f"Act間エスカレーション: Act3 avg {_act_avgs[2]:.2f}≤Act2 avg {_act_avgs[1]:.2f} → {_boost_count}シーンをi=3→4にブースト")

            # Act4 avg がAct3 avg以上ならAct4を抑制
            if len(_act_avgs) >= 4 and _act_avgs[3] >= _act_avgs[2]:
                _act4_s = act1 + act2 + act3
                _depress_count = 0
                for i in range(_act4_s, len(outline)):
                    if outline[i].get("intensity", 3) == 4:
                        outline[i]["intensity"] = 3
                        _depress_count += 1
                        if _depress_count >= max(1, act4 // 3):
                            break
                if _depress_count > 0:
                    log_message(f"Act間エスカレーション: Act4 avg {_act_avgs[3]:.2f}≥Act3 → {_depress_count}シーンをi=4→3に抑制")

        # erotic_levelとintensityの整合性を修正
        erotic_map = {1: "none", 2: "light", 3: "medium", 4: "heavy", 5: "climax"}
        for scene in outline:
            scene["erotic_level"] = erotic_map.get(scene.get("intensity", 3), "medium")

        # タイトル重複修正（アウトライン段階で検出・修正）
        _seen_titles_ol = set()
        _title_fix_ol = 0
        for s in outline:
            t = s.get("title", "")
            if t in _seen_titles_ol:
                sid = s.get("scene_id", "?")
                sit = s.get("situation", "")[:20]
                loc = s.get("location", "")
                new_title = f"{loc}での{sit}" if loc and sit else f"シーン{sid}"
                if new_title in _seen_titles_ol:
                    new_title = f"{new_title}({sid})"
                s["title"] = new_title
                _title_fix_ol += 1
                log_message(f"アウトラインtitle重複修正: S{sid}「{t}」→「{new_title}」")
            _seen_titles_ol.add(s.get("title", ""))
        if _title_fix_ol > 0:
            log_message(f"アウトラインtitle重複修正: {_title_fix_ol}件")

        # situation連続類似検出・警告（3連続で同一キーワードパターン）
        _SITUATION_KW_OL = ["膣奥", "突かれ", "責められ", "腰を振", "ピストン",
                            "挿入", "犯され", "抱かれ", "押し倒", "襲われ",
                            "口内", "フェラ", "パイズリ", "騎乗", "バック",
                            "正常位", "四つん這い", "膝立ち", "指で"]
        _sit_kw_list = []
        for s in outline:
            sit = s.get("situation", "")
            kws = frozenset(kw for kw in _SITUATION_KW_OL if kw in sit)
            _sit_kw_list.append(kws)
        _consec_same_ol = 0
        for idx in range(2, len(outline)):
            if (_sit_kw_list[idx] and
                _sit_kw_list[idx] == _sit_kw_list[idx - 1] == _sit_kw_list[idx - 2]):
                _consec_same_ol += 1
                sid = outline[idx].get("scene_id", idx + 1)
                log_message(f"⚠️ アウトライン: S{sid-2}〜S{sid} situationキーワード同一（{_sit_kw_list[idx]}）")
        if _consec_same_ol > 0:
            log_message(f"⚠️ アウトライン: {_consec_same_ol}箇所でsituation連続類似検出")

        # アウトライン数がnum_scenesに不足する場合、自動補完
        if len(outline) < num_scenes:
            missing = num_scenes - len(outline)
            log_message(f"⚠️ アウトラインが{missing}シーン不足（{len(outline)}/{num_scenes}）、自動補完")
            if callback:
                callback(f"[WARN]AI出力{len(outline)}シーン（{num_scenes}要求）、{missing}シーン自動補完中...")

            for _pad_i in range(missing):
                new_id = len(outline) + 1
                ratio = new_id / num_scenes

                # 位置に基づくintensity決定
                if ratio >= (1.0 - epilogue_pct):
                    pad_intensity = 3  # エピローグ
                elif ratio >= (1.0 - epilogue_pct - 0.1):
                    pad_intensity = 5  # クライマックス付近
                else:
                    pad_intensity = 4  # 本番

                # 前シーンとのintensity飛躍を防止
                if outline:
                    prev_int = outline[-1].get("intensity", 3)
                    if pad_intensity - prev_int >= 2:
                        pad_intensity = prev_int + 1

                outline.append({
                    "scene_id": new_id,
                    "title": f"シーン{new_id}",
                    "goal": "",
                    "location": outline[-1].get("location", "室内") if outline else "室内",
                    "time": "",
                    "situation": "",
                    "story_flow": "",
                    "emotional_arc": {"start": "", "end": ""},
                    "beats": [],
                    "intensity": min(pad_intensity, 5),
                    "erotic_level": erotic_map.get(min(pad_intensity, 5), "medium"),
                    "viewer_hook": ""
                })

            # 補完後のerotic_level再整合
            for scene in outline:
                scene["erotic_level"] = erotic_map.get(scene.get("intensity", 3), "medium")
            log_message(f"アウトライン補完完了: {len(outline)}シーン")

        log_message(f"アウトライン生成完了（API）: {len(outline)}シーン, テーマ: {theme_name}")
        if callback:
            callback(f"[OK]シーン分割完成（AI生成）: {len(outline)}シーン")

        return outline

    except Exception as e:
        log_message(f"アウトラインAPI生成失敗、テンプレートフォールバック: {e}")
        import traceback
        log_message(traceback.format_exc())
        if callback:
            callback(f"[WARN]AI分割失敗、テンプレートで代替: {str(e)[:50]}")

        # === テンプレートフォールバック ===
        arc_parts = [p.strip() for p in story_arc.replace("→", "/").split("/")]
        outline = []
        scene_id = 0
        for i in range(num_scenes):
            scene_id += 1
            if scene_id <= act1:
                intensity = 1 if i == 0 else 2
                erotic = "none" if i == 0 else "light"
                arc_label = arc_parts[0] if arc_parts else "導入"
            elif scene_id <= act1 + act2:
                intensity = 2 if (scene_id - act1) <= act2 // 2 else 3
                erotic = "light" if intensity == 2 else "medium"
                arc_label = arc_parts[1] if len(arc_parts) > 1 else "展開"
            elif scene_id <= act1 + act2 + act3:
                is_climax = (scene_id == act1 + act2 + act3)
                intensity = 5 if is_climax else 4
                erotic = "climax" if is_climax else "heavy"
                arc_label = arc_parts[2] if len(arc_parts) > 2 else "本番"
            else:
                intensity = 2
                erotic = "light"
                arc_label = arc_parts[-1] if arc_parts else "余韻"

            outline.append({
                "scene_id": scene_id,
                "title": arc_label,
                "goal": "",
                "location": "室内",
                "time": "",
                "situation": f"（あらすじ参照）{synopsis[:100] if synopsis else ''}",
                "story_flow": "",
                "emotional_arc": {"start": "", "end": ""},
                "beats": [],
                "intensity": intensity,
                "erotic_level": erotic,
                "viewer_hook": ""
            })

        log_message(f"テンプレートフォールバック: {len(outline)}シーン")
        return outline



def extract_scene_summary(scene: dict) -> str:
    """シーンの要約を抽出（次シーンのstory_so_farに使用）。
    v9.0: 物理状態（服装・体液・表情）を抽出してサマリに含める。"""
    sid = scene.get("scene_id", "?")
    title = scene.get("title", "")
    desc = scene.get("description", "")[:200]
    location = scene.get("location_detail", "")
    mood = scene.get("mood", "")
    intensity = scene.get("intensity", 3)

    # 吹き出しの要約
    bubbles = scene.get("bubbles", [])
    bubble_texts = []
    has_moan = False
    for b in bubbles:
        speaker = b.get("speaker", "")
        btype = b.get("type", "")
        text = b.get("text", "")
        bubble_texts.append(f"{speaker}({btype}):「{text}」")
        if btype == "moan":
            has_moan = True
    bubbles_str = ", ".join(bubble_texts) if bubble_texts else "なし"

    # オノマトペの要約
    onomatopoeia = scene.get("onomatopoeia", [])
    se_str = ", ".join(onomatopoeia) if onomatopoeia else "なし"

    # 心情の要約
    feelings = scene.get("character_feelings", {})
    feelings_str = ", ".join(f"{k}: {v}" for k, v in feelings.items()) if isinstance(feelings, dict) and feelings else ""

    # ストーリーフロー（次への繋がり）
    story_flow = scene.get("story_flow", "")

    # v9.0: sd_promptから物理状態を抽出
    sd_prompt = scene.get("sd_prompt", "")
    sd_tags_lower = {t.strip().lower().replace(" ", "_") for t in sd_prompt.split(",") if t.strip()}
    clothing_state = sorted(_CLOTHING_STATE_TAGS & sd_tags_lower)
    fluid_state = sorted(_FLUID_STATE_TAGS & sd_tags_lower)
    expression_state = sorted(_EXPRESSION_STATE_TAGS & sd_tags_lower)

    clothing_str = ",".join(clothing_state) if clothing_state else "着衣"
    fluid_str = ",".join(fluid_state) if fluid_state else "なし"
    expression_str = ",".join(expression_state) if expression_state else ""

    # 物理状態行（簡潔に1行で）
    phys_parts = [f"服装:{clothing_str}", f"体液:{fluid_str}"]
    if expression_str:
        phys_parts.append(f"表情:{expression_str}")
    if has_moan:
        phys_parts.append("喘ぎあり")
    physical_line = " | ".join(phys_parts)

    return (
        f"[シーン{sid}] {title} (intensity={intensity}, {mood}) "
        f"場所:{location} / {desc}\n"
        f"  物理状態: {physical_line}\n"
        f"  心情: {feelings_str}\n"
        f"  吹き出し: {bubbles_str}\n"
        f"  SE: {se_str}\n"
        f"  次への繋がり: {story_flow}"
    )


def _compact_scene_summary(scene: dict) -> str:
    """シーンの圧縮要約（セリフ/SE情報を保持）"""
    sid = scene.get("scene_id", "?")
    title = scene.get("title", "")[:20]
    desc = scene.get("description", "")[:60]
    intensity = scene.get("intensity", 3)
    # 吹き出しテキストだけ抽出（ブラックリスト用に保持）
    bubbles = scene.get("bubbles", [])
    bubble_texts = ", ".join(f"「{b.get('text', '')}」" for b in bubbles if b.get("text"))
    se = scene.get("onomatopoeia", [])
    se_str = ", ".join(se) if se else ""
    story_flow = scene.get("story_flow", "")[:80]
    return (
        f"[シーン{sid}] {title} (intensity={intensity}) {desc}\n"
        f"  吹き出し: {bubble_texts or 'なし'}\n"
        f"  SE: {se_str or 'なし'}\n"
        f"  次への繋がり: {story_flow}"
    )


def _oneliner_scene_summary(scene: dict) -> str:
    """シーンの1行概要（6シーン以前用、ブラックリスト情報なし）"""
    sid = scene.get("scene_id", "?")
    title = scene.get("title", "")[:10]
    intensity = scene.get("intensity", 3)
    mood = scene.get("mood", "")[:6]
    return f"[シーン{sid}] {title} (intensity={intensity}, {mood})"


def _build_narrative_arc_summary(scene_results: list) -> str:
    """v8.7: ナラティブアーク要約（~100tok）。序盤・転換点・intensity推移を永続化。
    100シーン超でも中盤の流れが消失しない。"""
    if not scene_results:
        return ""
    first = scene_results[0]
    intensities = [s.get("intensity", 3) for s in scene_results]
    turning = []
    for i in range(1, len(intensities)):
        if abs(intensities[i] - intensities[i - 1]) >= 2:
            s = scene_results[i]
            turning.append(f"S{s.get('scene_id', i + 1)}:{s.get('title', '')[:15]}")
    tp = "/".join(turning[-3:]) or "なし"
    return (
        f"【全体の流れ】{first.get('location_detail', '')} "
        f"{first.get('description', '')[:60]}...から開始 / "
        f"intensity: {intensities[0]}→{intensities[-1]}(peak={max(intensities)}) / "
        f"転換点: {tp}"
    )


def _build_story_so_far(story_summaries: list, scene_results: list) -> str:
    """story_so_farを構築（3層スライディングウィンドウ）。

    - v8.7: ナラティブアーク要約（先頭に挿入、~100tok追加）
    - 直近3シーン: フルテキスト（extract_scene_summary）
    - 4-8シーン前: 圧縮要約（_compact_scene_summary）※セリフ/SE情報保持
    - 9シーン以上前: 直近20件の1行概要 + さらに古い分は件数のみ（トークン節約）

    セリフ重複防止のブラックリストは別途used_blacklistで処理されるため、
    古いシーンの詳細をstory_so_farに保持する必要は薄い。
    """
    n = len(story_summaries)
    if n == 0:
        return ""

    parts = []

    # v8.7: ナラティブアーク要約（9シーン以上で挿入）
    if n >= 9 and scene_results:
        arc = _build_narrative_arc_summary(scene_results)
        if arc:
            parts.append(arc)
            parts.append("")

    # 9シーン以上前: 1行概要（スライディングウィンドウ: 直近20件のみ、それ以前は省略）
    oneline_end = max(0, n - 8)
    if oneline_end > 0:
        parts.append("--- 序盤の展開 ---")
        # 20件を超える古いシーンは件数表示のみ（トークン爆発防止）
        _ONELINE_MAX = 20
        if oneline_end > _ONELINE_MAX:
            _skipped = oneline_end - _ONELINE_MAX
            parts.append(f"（シーン1〜{_skipped}: {_skipped}シーン確定済み、省略）")
            oneline_start = oneline_end - _ONELINE_MAX
        else:
            oneline_start = 0
        for j in range(oneline_start, oneline_end):
            if j < len(scene_results):
                sc = scene_results[j]
                sid = sc.get("scene_id", j + 1)
                title = sc.get("title", "")[:15]
                desc = sc.get("description", "")[:40]
                parts.append(f"[S{sid}] {title}: {desc}")
        parts.append("")

    # 4-8シーン前: 圧縮要約（セリフ/SE情報保持でブラックリスト補助）
    compact_start = max(0, n - 8)
    compact_end = max(0, n - 3)
    if compact_start < compact_end:
        parts.append("--- これまでの展開 ---")
        for j in range(compact_start, compact_end):
            if j < len(scene_results):
                parts.append(_compact_scene_summary(scene_results[j]))
        parts.append("")

    # 直近3シーン: フルテキスト
    recent_start = max(0, n - 3)
    if recent_start < n:
        parts.append("--- 直近の展開（詳細） ---")
        for j in range(recent_start, n):
            parts.append(story_summaries[j])

    # v9.0: 直前シーンの物理状態を明示抽出（generate_scene_draftのプロンプト注入用）
    if scene_results:
        last = scene_results[-1]
        sd_prompt = last.get("sd_prompt", "")
        sd_tags_lower = {t.strip().lower().replace(" ", "_") for t in sd_prompt.split(",") if t.strip()}
        _cl = sorted(_CLOTHING_STATE_TAGS & sd_tags_lower)
        _fl = sorted(_FLUID_STATE_TAGS & sd_tags_lower)
        _ex = sorted(_EXPRESSION_STATE_TAGS & sd_tags_lower)
        last_intensity = last.get("intensity", 3)
        parts.append("")
        parts.append("--- 前シーン最終物理状態（次シーンで引き継ぐこと） ---")
        parts.append(f"服装: {', '.join(_cl) if _cl else '着衣'}")
        parts.append(f"体液: {', '.join(_fl) if _fl else 'なし'}")
        if _ex:
            parts.append(f"表情: {', '.join(_ex)}")
        parts.append(f"興奮レベル: intensity={last_intensity}")

    return "\n".join(parts)


def generate_scene_draft(
    client: anthropic.Anthropic,
    context: dict,
    scene: dict,
    jailbreak: str,
    cost_tracker: CostTracker,
    theme: str = "",
    char_profiles: list = None,
    callback: Optional[Callable] = None,
    story_so_far: str = "",
    synopsis: str = "",
    outline_roadmap: str = "",
    male_description: str = "",
    scene_index: int = -1,
    total_scenes: int = 0,
    _return_prompt_only: bool = False,
    faceless_male: bool = True,
) -> dict:
    skill = load_skill("low_cost_pipeline")

    # Danbooruタグ強化スキルを読み込み
    danbooru_nsfw = load_skill("danbooru_nsfw_tags")

    # NSFWシーン構成スキル
    scene_composer = load_skill("nsfw_scene_composer")

    # エロ漫画セリフスキルを性格・テーマ別に選択
    _serihu_info = _select_serihu_skill(theme, char_profiles)
    serihu_skill_name = _serihu_info["primary"]
    serihu_skill = load_skill(serihu_skill_name)
    _serihu_secondary = load_skill(_serihu_info["secondary"]) if _serihu_info.get("secondary") else ""
    _serihu_ratio = _serihu_info.get("ratio", 1.0)
    _serihu_personality = _serihu_info.get("personality", "")

    # CG集吹き出し専門スキル（自然な日本語セリフガイド）
    bubble_writer_skill = load_skill("cg_bubble_writer")

    # CG集ビジュアル多様性スキル
    visual_skill = load_skill("cg_visual_variety")

    # テーマ別ガイドを取得
    theme_guide = THEME_GUIDES.get(theme, THEME_GUIDES.get("vanilla", {}))
    theme_name = theme_guide.get("name", "指定なし")
    dialogue_tone = theme_guide.get("dialogue_tone", "自然で楽しい雰囲気")
    use_heart = theme_guide.get("use_heart", True)
    theme_sd_tags = theme_guide.get("sd_tags", "")
    theme_sd_expressions = theme_guide.get("sd_expressions", "")
    key_emotions = theme_guide.get("key_emotions", [])
    story_elements = theme_guide.get("story_elements", [])
    
    # シーンの重要度
    intensity = scene.get("intensity", 3)
    location = scene.get("location", "室内")
    time_of_day = scene.get("time", "")
    
    # タグDB読み込み（外部JSON対応）
    tag_db = _load_tag_db()
    
    # 背景タグテンプレート
    loc_tags_db = tag_db.get("locations", {})
    time_tags_db = tag_db.get("time_of_day", {})
    
    # 場所と時間帯のタグを取得
    location_tags = ""
    for key, tags in loc_tags_db.items():
        if key in location:
            location_tags = tags
            break
    if not location_tags:
        location_tags = "indoor, room"
    
    time_tags = ""
    for key, tags in time_tags_db.items():
        if key in time_of_day:
            time_tags = tags
            break
    
    # キャラプロファイルをintensity別に圧縮（API節約）
    char_guide = ""
    char_danbooru_tags = []
    char_names = []

    if char_profiles:
        for cp in char_profiles:
            name = cp.get("character_name", "")
            char_names.append(name)
            speech = cp.get("speech_pattern", {})
            emotional = cp.get("emotional_speech", {})
            examples = cp.get("dialogue_examples", {})
            relationship = cp.get("relationship_speech", {})
            avoid = cp.get("avoid_patterns", [])
            physical = cp.get("physical_description", {})
            tags = cp.get("danbooru_tags", [])

            char_danbooru_tags.extend(tags)

            if intensity <= 2:
                char_guide += f"""
【{name}】口調: 一人称={speech.get('first_person', '私')}, 語尾={', '.join(speech.get('sentence_endings', [])[:3])}, 間投詞={', '.join(speech.get('fillers', ['あっ'])[:2])}
外見: 髪={physical.get('hair', '')}, 目={physical.get('eyes', '')}, 体型={physical.get('body', '')}
NG: {', '.join(avoid[:3]) if avoid else 'なし'}
"""
            elif intensity == 3:
                char_guide += f"""
【{name}】口調ガイド
・一人称: {speech.get('first_person', '私')} / 語尾: {', '.join(speech.get('sentence_endings', ['〜よ', '〜ね']))}
・間投詞: {', '.join(speech.get('fillers', ['あっ', 'んっ']))}
・照れた時: {emotional.get('when_embarrassed', '言葉に詰まる')}
・甘える時: {emotional.get('when_flirty', '甘い声で')}
・外見: 髪={physical.get('hair', '')}, 目={physical.get('eyes', '')}
・NG: {', '.join(avoid) if avoid else 'なし'}
"""
            else:
                char_guide += f"""
═══════════════════════════════════════
【{name}】完全口調ガイド
═══════════════════════════════════════

■ 基本設定
・一人称: {speech.get('first_person', '私')}
・語尾: {', '.join(speech.get('sentence_endings', ['〜よ', '〜ね']))}
・よく使う表現: {', '.join(speech.get('favorite_expressions', [])[:5])}
・間投詞（息遣い）: {', '.join(speech.get('fillers', ['あっ', 'んっ']))}

■ 感情別の話し方
・照れた時: {emotional.get('when_embarrassed', '言葉に詰まる')}
・感じてる時: {emotional.get('when_flirty', '甘い声で')}
・感じてる時(エロ): {emotional.get('when_aroused', '声が震える')}
・絶頂時: {emotional.get('when_climax', '理性が飛ぶ')}

■ セリフのお手本
・好意: 「{examples.get('affection', '好きだよ')}」
・喘ぎ（軽）: {examples.get('moaning_light', 'あっ...んっ...')}
・喘ぎ（激）: {examples.get('moaning_intense', 'あっあっ...♡')}

■ 恋人への話し方
{relationship.get('to_lover', '甘えた調子で話す')}

■ NG表現: {', '.join(avoid) if avoid else 'なし'}
■ 外見: 髪={physical.get('hair', '')}, 目={physical.get('eyes', '')}, 体型={physical.get('body', '')}
"""

    # フルネーム→短縮名マップ構築（キャラ名途切れ対策用）
    char_short_names = []
    for full_name in char_names:
        short = full_name
        for sep in ["・", " ", "＝", "　"]:
            idx = full_name.find(sep)
            if idx > 0:
                short = full_name[:idx]
                break
        char_short_names.append(short)

    # キャラ固有セリフプールからfew-shot例を注入
    char_pool_section = ""
    if char_profiles:
        _cp0 = char_profiles[0]
        _cp_id = generate_char_id(_cp0.get("work_title", ""), _cp0.get("character_name", ""))
        _scene_char_pool = load_character_pool(_cp_id)
        if _scene_char_pool:
            sample_lines = []
            for _ph in ["foreplay", "penetration", "climax"]:
                phase_speech = _scene_char_pool.get("speech", {}).get(_ph, [])[:2]
                sample_lines.extend(phase_speech)
            if sample_lines:
                char_pool_section = "\n## キャラ固有セリフ例（このトーンで書け）\n" + "\n".join(f"・{l}" for l in sample_lines)

    # v8.8: テーマ世界ルール（セリフ制約等）
    world_rules = theme_guide.get("world_rules", [])
    world_rules_instruction = ""
    if world_rules:
        world_rules_instruction = "\n## ⚠️ テーマ世界設定ルール（厳守）\n"
        for wr in world_rules:
            world_rules_instruction += f"- {wr}\n"

    # ♡使用のルール（テーマ別）
    heart_instruction = ""
    if use_heart:
        heart_instruction = "♡は甘いシーンで自然に使用OK。"
    else:
        heart_instruction = f"""
テーマ「{theme_name}」のspeech（会話）では♡は控えめに。
ただし**moan（喘ぎ）では♡を積極的に使え**（不本意な快楽の表現として効果的）。
セリフ品質ガイドの喘ぎ辞書にある♡はそのまま使用すること。
"""

    # テーマ別セリフトーン指示
    theme_dialogue_instruction = f"""
## テーマ「{theme_name}」のセリフトーン

{dialogue_tone}

【このテーマで重要な感情】
{', '.join(key_emotions) if key_emotions else '自然な感情表現'}

【ストーリー要素として入れるべきもの】
{chr(10).join(f'・{e}' for e in story_elements[:3]) if story_elements else '・特になし'}

{heart_instruction}
"""

    # シーン重要度別のエロ指示（5段階）- CG集フォーマット対応
    if intensity >= 5:
        erotic_instruction = f"""
## クライマックスシーン（intensity 5）

最高潮のエロシーン。画像が全てを語る。絶頂・射精・快楽堕ちの瞬間。

【descriptionの書き方（50字以上）— 描写であって説明ではない】
画像が「何が起きているか」を見せる。descriptionは「どう感じているか」を書け。
■ 触覚を最優先: 温度、圧迫感、痙攣、脈動を書け
■ 聴覚を混ぜろ: 肌が打ち合う音、呼吸、心音
■ 文体: 短文連打。体言止めOK。テンポ速く
✅ 「腰が跳ね上がる。奥を突き上げられるたび、頭が真っ白になる。熱い塊が注がれる感覚に、全身が痙攣した。」
✅ 「背中を反らし、声にならない悲鳴。脈打つ熱が奥で弾け、意識が途切れる。」
❌ 「快感に溺れている」「絶頂を迎えた」のような報告文は禁止。身体の感覚を書け。

【吹き出し指針（1-3個）】
・女: 絶頂系の喘ぎ1-2個（★セリフ品質ガイドの【段階4】から選べ。自作するな。前シーンと被らないこと）
・男: 言葉責め0-1個（5パターンを均等に使え: 脅迫/挑発/命令/嘲笑/独白。前シーンと同じ意味のセリフ禁止）
  例: 「中に出すぞ」「全部受けろ」「イケ」「もう逃がさねえ」「もっと鳴け」

【オノマトペ指針（3-4個・辞書から選べ）】
・射精系+反応系+抽送系を組み合わせる
  例: ドビュッ, ビクビクッ, パンパンパン, ドクドクッ

【心情】
・{key_emotions[2] if len(key_emotions) > 2 else '快感に溺れる'}
・{key_emotions[3] if len(key_emotions) > 3 else '理性と本能の葛藤'}

【禁止】
❌ 説明的なセリフ
❌ 冷静な会話
❌ 前シーンと同じ喘ぎパターン
"""
    elif intensity == 4:
        erotic_instruction = f"""
## 本番シーン（intensity 4）

濃厚なエロシーン。挿入・激しい行為。画像の行為を吹き出しが補強。

【descriptionの書き方（50字以上）— 感覚で書け、状況説明するな】
画像が体位を見せる。descriptionは「身体がどう反応しているか」を書け。
■ 触覚中心: 挿入感、摩擦、圧迫、内部の感覚
■ 聴覚: 水音、肌を打つ音、漏れる息
■ 視覚は画像に任せろ: 「胸を鷲掴みにし」より「胸を掴む指に力が込められ、甘い痺れが走る」
✅ 「背後から突き上げられるたび、机を掴む指が白くなる。耐えきれず漏れる声が、静かな教室に響いた。」
✅ 「奥を擦り上げられる感覚に、膝が震える。抗おうとする指先が、いつの間にかシーツを掻きむしっている。」
❌ 「快感に溺れていく」「挿入された」だけの報告文は禁止。感覚と反応を書け

【吹き出し指針（1-3個）】
・女: 喘ぎ1-2個（★セリフ品質ガイドの【段階3】から選べ。自作するな。前シーンと被らないこと）
・男: 言葉責め0-1個（5パターンを均等に使え: 脅迫/挑発/命令/嘲笑/独白。前シーンと同じ意味のセリフ禁止）
  例: 「もっと鳴け」「逃がさねぇぞ」「欲しいんだろ」「もう我慢すんな」

【オノマトペ指針（2-3個・辞書から選べ）】
・挿入系+抽送系+濡れ系を組み合わせる
  例: ズブッ, パンパン, グチュグチュ

【心情】
・{key_emotions[1] if len(key_emotions) > 1 else '恥ずかしさと快感の葛藤'}
・{key_emotions[2] if len(key_emotions) > 2 else 'もっと欲しいという欲求'}

【禁止】
❌ 説明的なセリフ
❌ 前シーンと同じ喘ぎパターン
"""
    elif intensity == 3:
        erotic_instruction = f"""
## 前戯・焦らしシーン（intensity 3）

エロの助走。脱衣・愛撫・キス等。期待感を煽る画像に短い吹き出し。

【descriptionの書き方 — 「まだしない」の緊張感こそエロの核】
触れる前の予感。触れた瞬間の反応。もっと触れたいのに我慢する葛藤。
■ 触覚: 服越しの温もり、指先が触れた瞬間の電撃感、肌のざわめき
■ 空間: 二人の距離感、静寂、息遣いだけの空間
✅ 「制服越しに触れた指先が、ゆっくりと鎖骨をなぞる。息を詰める彼女の肩が、かすかに震えた。」
✅ 「唇が触れる寸前で止められ、焦らされる。吐息が肌にかかるだけで、身体が熱くなる。」
❌ 「愛撫された」「キスした」だけの事実報告は禁止。触れる前後の感覚を書け

【吹き出し指針（1-3個）】
・女: ドキドキ感のある反応1-2個（★セリフ品質ガイドの【段階2】から選べ）
・男: 煽りor会話0-1個（前シーンと同じ意味のセリフ禁止。脅迫/挑発/命令/嘲笑/独白の5パターンを使い分けろ）
  例: 「おとなしくしろ」「脱げ」「正直になれよ」

【オノマトペ指針（1-2個）】
・愛撫系+心音系: サワッ, チュッ, ゾクッ, ドキドキ, ペロッ, スルッ

【心情】
・{key_emotions[0] if key_emotions else 'ドキドキと期待'}
・恥ずかしいけど…という葛藤
"""
    elif intensity == 2:
        erotic_instruction = f"""
## ムード構築シーン（intensity 2）

雰囲気作り。接近する画像に自然な一言。

【descriptionの書き方 — 日常から非日常への移行を描け】
まだ何も起きていないのに、空気が変わる瞬間。予感と緊張の描写。
■ 空間と距離: いつもより近い距離、いつもと違う視線、変化した空気
■ 五感: ふわりと香る匂い、聞こえる心臓の音、頬に触れる吐息
✅ 「いつもの距離より一歩近い。彼の視線が、制服の胸元に一瞬だけ落ちたのを、彼女は見逃さなかった。」
❌ 「二人きりになった」だけの状況説明は禁止。空気の変化を書け

【吹き出し指針】
・自然な短い会話（1-3個）
・例: 「ねえ…」「え…？」
・**喘ぎ声・絶頂セリフは絶対NG**。まだ本番前。ドキドキや戸惑いのみ

【オノマトペ指針】
・なし or 1個: ドキッ

【心情】
・{key_emotions[0] if key_emotions else '緊張とドキドキ'}
"""
    else:
        erotic_instruction = f"""
## 導入シーン（intensity 1）

状況設定。キャラ紹介の画像に短い会話。

【descriptionの書き方 — キャラの魅力と状況を一文で伝えろ】
読者が「この子がこれからどうなるのか」を期待する導入。
■ キャラの外見/仕草の特徴的なディテール1つ
■ 場所の空気感、時間帯の光
✅ 「夕暮れの教室。窓際の席で頬杖をつく彼女の横顔に、オレンジの光が差している。」
❌ 「教室にいる」だけの場所説明は禁止。画になるディテールを1つ入れろ

【吹き出し指針】
・自然な一言（1-3個）。状況説明はdescriptionで行い、吹き出しは最小限
・例: 「ただいま〜」「久しぶり…」
・**絶対に喘ぎ声・♡・エロ系セリフを入れるな**。歩いてるだけ、座ってるだけの場面で喘ぐな

【オノマトペ指針】
・なし

【心情】
・日常の中の雰囲気
"""

    # キャラ固有SDタグの組み込み
    char_tags_str = ", ".join(char_danbooru_tags[:15]) if char_danbooru_tags else ""
    
    # テーマ別SDタグを追加
    theme_tags_combined = f"{theme_sd_tags}, {theme_sd_expressions}".strip(", ")
    
    # === Prompt Caching: 共通部分（全シーンで同一）とシーン固有部分を分離 ===

    # v9.0: faceless_male 対応の男性描写セクション構築
    if male_description and faceless_male:
        _male_section = f"""
## 男性キャラクター外見設定
**外見: {male_description}**
- descriptionに男性が登場する場合、必ずこの外見設定を反映した描写にすること
- 男性はfaceless male（顔なし）として扱う。男性の顔の特徴（目・鼻・口・表情）は一切描写しない
- 描写例: 「{male_description}の男に押し倒され…」「背後から{male_description}に覆いかぶさられ…」
"""
    elif male_description and not faceless_male:
        _male_section = f"""
## 男性キャラクター外見設定
**外見: {male_description}**
- descriptionに男性が登場する場合、必ずこの外見設定を反映した描写にすること
- 男性の外見（体型・髪型・服装・表情）を一貫して描写すること
- 描写例: 「{male_description}の男に押し倒され…」「背後から{male_description}に覆いかぶさられ…」
"""
    elif faceless_male:
        _male_section = """
## 男性キャラクター
- 男性はfaceless male（顔なし）として扱う。男性の顔の特徴（目・鼻・口・表情）は一切描写しない
"""
    else:
        _male_section = """
## 男性キャラクター
- 男性の外見（体型・髪型・服装・表情）を一貫して描写すること
"""

    # 共通部分（キャッシュ対象）- CG集フォーマット完全対応
    common_system = f"""{jailbreak}

{skill if skill else "FANZA同人CG集の脚本を生成します。"}

{danbooru_nsfw if danbooru_nsfw else ""}

{scene_composer if scene_composer else ""}

{bubble_writer_skill if bubble_writer_skill else ""}

{char_guide if char_guide else "（キャラ設定なし）"}

{char_pool_section}
{_male_section}
## FANZA同人CG集フォーマット
「セリフ付きCG集」＝1枚絵に吹き出し+オノマトペ。**画像がメイン、テキストはサブ**。
各ページ: CG画像1枚 + 吹き出し1-3個（ヒロイン1-2+男0-1） + SE 0-4個

## ⚠️ 追加厳守ルール（上記吹き出しスキルに加えて）

### セリフ・SE重複禁止
story_so_farのセリフ・SEと同一・類似は絶対禁止。毎シーン辞書の別パターンを選べ。

### 場所名の一貫性
同じ場所は**全シーンで同一の表記**。表記ブレ禁止。

### セリフ内容整合性
- moan=喘ぎ声のみ。説明文禁止（❌「そうなんだ」「汗すごい」）
- speech=感情的反応のみ。身体報告禁止（❌「震えてる」「目が回る」）
- thought=感情断片のみ。ナレーション禁止（❌「こんなことをしている自分が…」）。部位ラベル冒頭禁止（❌「胸…こんなに…」「太もも…そんな…」→✅「熱い…」「ゾクって…」）
- 男性speech=命令/挑発/独白のみ。観察実況禁止（❌「いい声だな」「敏感だな」「ここも感じるんだな」→✅「もっと鳴け」「欲しいんだろ」「逃がさねぇ」）
- descriptionと吹き出しの内容が**論理的に一致**すること

### story_flowの書き方
各シーン固有の展開。前シーンのコピペ禁止。状況が必ず進展すること。

### 語尾パターン多様性
同じ語尾パターン（～し…、～る♡、～の…、～て…、～だし等）を3シーン連続で使うな。
speech/thoughtの語尾は毎シーン構造を変えろ（体言止め/疑問/感嘆/中断/独白）。

### thought先頭パターン反復禁止
同じ書き出し（先頭2文字）を3シーン以内で再使用するな。バリエーションが生命線。
同一構文（X…Yが…Z…）の3連続も禁止。構文パターンを変えろ（疑問/自嘲/矛盾/驚き/諦め/感覚）。
感情は恐怖/羞恥/快感否定の3種だけでなく、怒り/諦め/自嘲/混乱/背徳感も使え。
❌禁止: thoughtで「胸…」「太もも…」「耳…」等、部位名を冒頭に置く説明型（「胸…こんなに…」）。部位を主語にするな。感覚・感情を主語にしろ（「熱い…」「ゾクって…」「やばい…」）。
同じキーワード（「だめ」「声」等）は全シーン中4回まで。

## オノマトペ辞書（同じ組み合わせの連続禁止）
・挿入: ズブッ, ヌプッ, ズリュッ, ヌルッ, ズンッ ・抽送: パンパン, グチュグチュ, ヌチュヌチュ
・愛撫: サワッ, ペロッ, チュッ, レロレロ ・吸引: チュパッ, ジュルッ, ゴクッ
・射精: ドクドク, ドビュッ, ビュルル ・反応: ビクッ, ビクビク, ガクガク, ゾクッ
・心音: ドキドキ, バクバク ・衝撃: ドンッ, ギシギシ ・濡れ: トロッ, グショッ, ヌルヌル
・剥ぎ: ビリッ, スルッ

{f'''
## ⚠️ セリフ品質ガイド（厳守・最優先）

bubblesのtextは以下の【喘ぎ声バリエーション集】と【鉄則】に厳密に従え。
「タスク手順」「不自然診断」「改訂版セリフ」等のセクションは無視せよ。

★ 喘ぎ声は必ず下記辞書の【段階1〜4】から選べ。自分で喘ぎを創作するな。
★ intensityに対応する段階を使え（intensity 1-2=段階1、intensity 3=段階2、intensity 4=段階3、intensity 5=段階4）
★ 前シーンで使った喘ぎと同じものは絶対禁止。毎シーン辞書の別パターンを選べ。

{serihu_skill}
''' if serihu_skill else ''}{f'''

### サブスタイル（混合比率{int((1-_serihu_ratio)*100)}%で以下のスタイルも取り入れること）:
{_serihu_secondary}
''' if _serihu_secondary and _serihu_ratio < 1.0 else ''}{f'''
★ キャラ性格タイプ「{_serihu_personality}」を意識したセリフ。ギャップ感を出すこと。
''' if _serihu_personality else ''}

{f'''
## CG集ビジュアル構成ガイド

{visual_skill}
''' if visual_skill else ''}

全キャラ成人(18+)。JSON形式のみ出力。"""

    # シーン固有部分（毎回変わる）
    scene_system = f"""{erotic_instruction}

{theme_dialogue_instruction}
{world_rules_instruction}"""

    # Prompt Caching: systemをリスト形式でcache_control付与
    system_with_cache = [
        {"type": "text", "text": common_system, "cache_control": {"type": "ephemeral"}},
        {"type": "text", "text": scene_system}
    ]

    # シーン別SD推奨タグ（ポーズ・表情）+ テーマ別タグ - 大幅拡張
    intensity_sd_tags = {
        5: f"ahegao, orgasm, cum, creampie, cum_overflow, cum_on_body, trembling, convulsing, full_body_spasm, tears, heavy_breathing, drooling, rolling_eyes, tongue_out, mind_break, fucked_silly, sweat, wet, {theme_sd_expressions}",
        4: f"sex, vaginal, penetration, nude, spread_legs, missionary, doggy_style, cowgirl_position, moaning, sweat, blush, panting, pussy_juice, groping, breast_grab, grabbing_hips, {theme_sd_expressions}",
        3: f"kiss, french_kiss, undressing, groping, breast_grab, nipple_play, fingering, blush, nervous, anticipation, wet_panties, bra_pull, panties_aside, embarrassed, {theme_sd_expressions}",
        2: f"eye_contact, close-up, romantic, blushing, hand_holding, leaning_close, nervous, looking_away, {theme_sd_expressions}",
        1: f"portrait, smile, casual, standing, looking_at_viewer, {theme_sd_expressions}"
    }
    
    sd_intensity_tags = intensity_sd_tags.get(intensity, "")
    
    # 背景タグを組み合わせ
    background_tags = f"{location_tags}, {time_tags}".strip(", ")

    # テーマタグを背景に追加（intensity 3以上のみ）
    if theme_sd_tags and intensity >= 3:
        background_tags = f"{background_tags}, {theme_sd_tags}"

    # 設定スタイルから背景ヒントを取得
    setting_style = _detect_setting_style(context.get("setting", ""), theme=theme)
    setting_hint_line = ""
    if setting_style:
        hint = setting_style.get("prompt_hint", "")
        style_append = ", ".join(setting_style.get("append", []))
        if style_append:
            background_tags = f"{background_tags}, {style_append}"
        if hint:
            setting_hint_line = f"\n背景スタイル必須: {hint}"

    # 構図タグ（intensity連動）
    composition_db = tag_db.get("compositions", {})
    composition_tags = composition_db.get(str(intensity), {}).get("tags", "")

    # あらすじセクション（全シーン共通の物語の骨格）
    synopsis_section = ""
    if synopsis:
        synopsis_section = f"""## 参考: 作品全体のあらすじ
{synopsis}
---
"""

    # ストーリー連続性セクション（使用済みセリフ・SE・story_flowを明示抽出）
    story_context_section = ""
    if story_so_far:
        # story_so_farから使用済みセリフ・SE・story_flowを抽出してブラックリスト化
        import re as _re
        used_bubbles = []
        used_se = []
        used_flows = []
        for line in story_so_far.split("\n"):
            line = line.strip()
            if line.startswith("吹き出し:"):
                bubble_content = line[len("吹き出し:"):].strip()
                if bubble_content and bubble_content != "なし":
                    used_bubbles.append(bubble_content)
            elif line.startswith("SE:"):
                se_content = line[len("SE:"):].strip()
                if se_content and se_content != "なし":
                    used_se.append(se_content)
            elif line.startswith("次への繋がり:"):
                flow_content = line[len("次への繋がり:"):].strip()
                if flow_content and len(flow_content) >= 10:
                    used_flows.append(flow_content)

        blacklist_parts = []
        if used_bubbles:
            blacklist_parts.append("【使用済みセリフ（同一・類似禁止）】")
            for ub in used_bubbles:
                blacklist_parts.append(f"  ❌ {ub}")
        if used_se:
            blacklist_parts.append("【使用済み効果音（同一組み合わせ禁止）】")
            for us in used_se:
                blacklist_parts.append(f"  ❌ {us}")
        if used_flows:
            blacklist_parts.append("【使用済みstory_flow（同一テキスト禁止。各シーン固有の展開を書け）】")
            for uf in used_flows:
                blacklist_parts.append(f"  ❌ {uf}")
        used_blacklist = "\n".join(blacklist_parts) if blacklist_parts else "（初回シーンのため禁止リストなし）"

        story_context_section = f"""
## ⚠️ ストーリーの連続性（最重要）

以下は前のシーンまでの展開です。**必ずこの続きとして**シーンを書いてください。

{story_so_far}

### 🚫 使用禁止リスト（以下と同じ・類似は絶対禁止）
{used_blacklist}

### 禁止事項（違反したら不合格）
- **上の使用禁止リストにあるセリフ・SE・story_flowと同一または類似は使用不可**
- **story_flowは毎シーン固有の内容を書け**。前シーンと同じ文章のコピペは即不合格
- **前シーンと同じ状況描写・同じ展開の繰り返し禁止**
- **ストーリーを必ず前シーンより先に進めること（行為をエスカレート）**
- **同じ場所名は前シーンと同じ表記を使え（表記ブレ禁止）**
- **キャラ名はフルネーム「{', '.join(char_names) if char_names else 'ヒロイン'}」または姓「{', '.join(char_short_names) if char_short_names else 'ヒロイン'}」のみ使用**

### ⚠️ エスカレーション制御（段階飛躍禁止）
- **前シーンの行為レベルから1段階だけ進めること**
- 前シーンが前戯なら→このシーンは挿入開始。いきなり複数人や絶頂は禁止
- 前シーンが1対1なら→このシーンも1対1か、せいぜい2人目の登場まで
- 前シーンで抵抗していたなら→このシーンは葛藤。いきなり完全堕落は禁止
- **心情の変化は前シーンの「次への繋がり」を必ず引き継ぐこと**

### ⚠️ 描写バリエーション（体位カタログ禁止）
- 同じ体位で2-3シーン続いてよい。ただし各シーンで構図・テンポ・感情の少なくとも2つを変化させること
- 体位変更は射精/絶頂後の「次のラウンド」として自然に行うこと。毎シーン体位を変える「体位カタログ」は禁止
- 描写する身体部位・焦点は前シーンと変えること（胸→腰→脚→首筋→耳→背中をローテーション）
- **「膣奥」「膣内」等の同じ表現を3シーン以上繰り返し使用するのは禁止**
- **titleは全シーンで固有であること。前シーンと同じキーワードの繰り返し禁止**
---
"""

    # v9.0: 物理状態引き継ぎ指示を抽出・注入
    physical_state_section = ""
    if story_so_far and "前シーン最終物理状態" in story_so_far:
        _phys_lines = []
        _in_phys = False
        for _line in story_so_far.split("\n"):
            if "前シーン最終物理状態" in _line:
                _in_phys = True
                continue
            if _in_phys:
                _sl = _line.strip()
                if _sl.startswith("服装:") or _sl.startswith("体液:") or _sl.startswith("表情:") or _sl.startswith("興奮レベル:"):
                    _phys_lines.append(_sl)
                elif _sl == "" or _sl.startswith("---"):
                    break
        if _phys_lines:
            physical_state_section = "\n### ⚠️ 前シーンの物理状態（必ず引き継ぐこと）\n"
            for _pl in _phys_lines:
                physical_state_section += f"- {_pl}\n"
            physical_state_section += (
                "※ 射精後のシーンでは体液が残っている描写をdescriptionとsd_promptに必ず含めること\n"
                "※ 脱衣後のシーンで服が復活してはならない（前シーンでnudeなら今シーンもnude）\n"
                "※ 表情の段階的変化: 前シーンの表情をベースに、intensityに応じて自然にエスカレートさせること\n"
            )

    if physical_state_section and story_context_section:
        story_context_section = story_context_section.rstrip() + "\n" + physical_state_section + "---\n"

    # ロードマップセクション構築
    roadmap_section = ""
    if outline_roadmap:
        roadmap_section = f"""## ストーリーロードマップ（全体構成）
{outline_roadmap}

★ 現在生成: シーン{scene['scene_id']}「{scene.get('title', '')}」
このシーンの前後関係を意識し、ストーリーを確実に進めること。
---
"""

    # エピローグ・最終シーン指示（大量シーン時に物語完結を明示）
    epilogue_scene_instruction = ""
    if total_scenes > 0 and scene_index >= 0:
        scene_id = scene.get("scene_id", scene_index + 1)
        remaining = total_scenes - scene_id
        if scene_id == total_scenes:
            epilogue_scene_instruction = f"""
## ⚠️ 最終シーン（シーン{scene_id}/{total_scenes}）
これは作品の**最後のシーン**です。ストーリーを完結させてください。
- 余韻・事後描写で締めくくること
- 導入の繰り返しや新展開の開始は絶対禁止
- 関係性の変化・心情の決着を描くこと
- intensity 3-4で穏やかに（または余韻のあるエロで）終わらせること
"""
        elif remaining <= 3 and remaining > 0:
            epilogue_scene_instruction = f"""
## ⚠️ エピローグ（シーン{scene_id}/{total_scenes}、残り{remaining}シーン）
作品の終盤です。ストーリーの収束に向かってください。
- 新しい展開や新キャラの登場は禁止
- 既存の関係性・感情の決着を描くこと
- 導入シーンの繰り返しは絶対禁止
"""

    # アウトラインフィールドを明示的にフォーマット（JSON dumpの代わり）
    _ea = scene.get("emotional_arc", {})
    _ea_start = _ea.get("start", "") if isinstance(_ea, dict) else ""
    _ea_end = _ea.get("end", "") if isinstance(_ea, dict) else ""
    scene_instruction = f"""{epilogue_scene_instruction}## このシーンの設計指示
- シーンID: {scene['scene_id']}
- タイトル: {scene.get('title', '')}
- 目的(goal): {scene.get('goal', '指定なし')}
- 状況(situation): {scene.get('situation', '指定なし')}
- 場所: {scene.get('location', '')}
- 感情推移: {_ea_start} → {_ea_end}
- 展開ビート: {', '.join(scene.get('beats', [])) if scene.get('beats') else '指定なし'}
- 次への繋がり: {scene.get('story_flow', '指定なし')}
- エロレベル: {scene.get('erotic_level', '')}
- 視聴者フック: {scene.get('viewer_hook', '')}
- intensity: {scene.get('intensity', 3)}

⚠️ 上記の「状況」「感情推移」「展開ビート」に忠実にシーンを生成せよ。
特にdescriptionは「状況」の内容を具体的に膨らませること。"""

    prompt = f"""{synopsis_section}{roadmap_section}{story_context_section}設定: {json.dumps(context, ensure_ascii=False)}
{scene_instruction}

## 出力形式（この形式で出力してください）

{{
    "scene_id": {scene['scene_id']},
    "title": "シーンタイトル",
    "description": "このシーンの詳細説明。場所、状況、何が起きているか、画像として何が描かれるかを説明",
    "location_detail": "場所の具体的な描写",
    "mood": "雰囲気",
    "character_feelings": {{
        "{char_names[0] if char_names else 'ヒロイン'}": "このシーンでの心情"
    }},
    "bubbles": [
        {{"speaker": "キャラ名", "type": "speech", "text": "短い一言"}},
        {{"speaker": "キャラ名", "type": "moan", "text": "あっ♡"}},
        {{"speaker": "キャラ名", "type": "thought", "text": "心の声"}}
    ],
    "onomatopoeia": ["効果音1", "効果音2"],
    "direction": "演出・ト書き",
    "story_flow": "次のシーンへの繋がり",
    "sd_prompt": "{QUALITY_POSITIVE_TAGS}, キャラ外見タグ, ポーズ・行為タグ, 表情タグ, 場所・背景タグ, 照明タグ, テーマタグ"
}}

## タグ参考（sd_promptに統合して使用）

キャラ固有: {char_tags_str}
ポーズ・表情: {sd_intensity_tags}
背景・場所: {background_tags}
構図: {composition_tags}
テーマ専用: {theme_tags_combined}{setting_hint_line}

## ルール

1. descriptionは必ず100字程度。**「説明」ではなく「描写」**。触覚・聴覚中心。「～された」報告文禁止。身体の感覚・反応・空間の空気感を書け
2. character_feelingsで心情を明確に。前シーンと異なる感情変化を示すこと
3. **bubblesは1-3個**（ヒロイン1-2個 + 男性0-1個。セリフの長さは自由）。男性セリフは全体の25-35%のシーンにのみ入れること（20シーンなら5-7シーンのみ）
4. typeはspeech/moan/thoughtの3種。intensity 4-5はmoanメイン。**moanには喘ぎ声のみ（説明文禁止）**
5. **onomatopoeiaは場面に合った効果音**（intensity 1-2はなし〜1個、3は1-2個、4-5は2-4個）
6. sd_promptはこのシーン固有の描写タグのみ出力: 「キャラ外見 + ポーズ・行為 + 表情 + エロ描写 + アングル + 場所・背景 + 照明」。**品質タグ(masterpiece, best_quality, score_9等)は絶対に含めるな**（後処理で自動付与される）
7. **sd_promptにLoRAタグ(<lora:...>)を絶対に含めるな**。出力はDanbooruタグのみ
8. **sd_promptにオノマトペ・日本語テキストを含めない**（英語のDanbooruタグのみ使用）
9. **前シーンの流れを必ず引き継ぐこと**
10. **キャラの一人称・語尾はキャラガイドを絶対厳守**
11. **descriptionは全て日本語で書くこと**（英語タグはsd_promptのみ）
12. **titleに「○回戦」「続き」等の安易な表現禁止**。具体的な行為・状況を反映した簡潔なタイトルにすること
13. **キャラ名**: 初出時はフルネーム「{', '.join(char_names) if char_names else 'ヒロイン'}」を使用。同じdescription内の2回目以降は姓「{', '.join(char_short_names) if char_short_names else 'ヒロイン'}」でよい。表記ブレ厳禁（他の呼び方は禁止）
14. **descriptionは感覚で書け**。「挿入された」→「奥を突かれた衝撃で腰が跳ねる」。行為の報告ではなく、身体がどう反応したかを書くこと
15. **視点**: descriptionは女性キャラ視点で書くこと。男性を「主人公」と呼ばない。男性は「彼」「相手の男」「男性」と表記
16. **お嬢様口調のintensity対応**: intensity 4-5ではお嬢様口調（ですの/ですわ等）は崩壊させること。理性が飛んだ状態で丁寧語は不自然。「ですの」→「…の…♡」「ですわ」→「…♡」に崩す"""

    # 重複禁止の最終警告（user promptの末尾に配置 = モデルが最も注目する位置）
    dedup_warning = ""
    if story_so_far:
        dedup_warning = f"""

## ⚠️⚠️⚠️ 最終チェック（出力前に必ず確認） ⚠️⚠️⚠️

以下の条件を1つでも満たす場合、出力をやり直せ:
- bubblesのtextに前シーンと同じ文言がある → 辞書から別パターンを選び直せ
- onomatopoeiaが前シーンと同じ組み合わせ → 別の効果音に変えろ
- descriptionが前シーンと類似している → 具体的な行為を変えろ
- descriptionでキャラ名を省略している（「ボア」だけにしてる等） → 必ずフルネームで書け
- キャラ名が「{', '.join(char_names) if char_names else 'ヒロイン'}」または「{', '.join(char_short_names) if char_short_names else 'ヒロイン'}」以外の表記になっている → 修正しろ
- 男性キャラのセリフに♡が含まれている → 即座に削除しろ
- 男性キャラが喘いでいる(moanタイプ) → speechに変更し男性的な短い台詞に書き換えろ
- ヒロインの一人称・語尾がキャラ設定と食い違っている → 修正しろ
- bubblesが4個以上ある → ヒロイン1-2個+男性0-1個の最大3個に絞れ
- 男性セリフに「私たち」「いいよ」「ね」等の女性的表現がある → 「俺たち」「いいぞ」「な」に直せ
- 男性セリフが全体の40%以上のシーンに含まれている → 5-7シーンのみに男性セリフを入れ、残りは削除しろ
- descriptionが歩行・食事・帰宅等の非エロ場面なのにbubblesに喘ぎ・♡がある → 場面に合った普通のセリフに直せ
- 「初めて」「彼のこと忘れ」等の同じフレーズを全体で3回以上使っている → 別の表現にしろ
- type="moan"の吹き出しに説明文・会話文が入っている → 喘ぎ声に書き換えろ（「そうなんだ」「汗すごい」等は禁止）
- story_flowが前シーンと同一テキスト → このシーン固有の展開に書き換えろ
- descriptionの体位・行為が前シーンと同じ → 別の体位・行為に変えろ（正常位/後背位/騎乗位/立ちバック/側位/座位等をローテーション）
- titleが前シーンと同じキーワード（「膣奥」「理性」等）を含んでいる → 別のキーワードに変えろ"""

    prompt = prompt + dedup_warning + "\n\nJSONのみ出力。"

    # モデル自動選択: intensity別にコスト最適化
    # intensity 4-5 → Sonnet（本番+クライマックス: セリフ品質が最重要）
    # intensity 1-3 → Haiku 4.5（導入・前戯シーン）
    # ※ Haiku3(fast)はシーン生成に不適: キャラ名化け・NSFW品質不足
    if intensity >= 4:
        model = MODELS["sonnet"]
        model_name = "Sonnet"
    else:
        model = MODELS["haiku"]
        model_name = "Haiku(4.5)"
    
    if callback:
        callback(f"シーン {scene['scene_id']} 生成中 ({model_name}, 重要度{intensity}, {theme_name}, セリフ:{serihu_skill_name})...")

    # Batch APIモード: プロンプトのみ返す（API呼出なし）
    if _return_prompt_only:
        return {"system": system_with_cache, "user": prompt, "model": model}

    response = _call_api(
        client, model,
        system_with_cache,
        prompt, cost_tracker, 3000, callback
    )
    
    # 重複排除の後処理
    result = parse_json_response(response)

    # スキーマバリデーション（parse直後）
    if isinstance(result, dict):
        _sv = validate_scene(result)
        if not _sv["valid"]:
            sid = result.get("scene_id", "?")
            for _se in _sv["errors"]:
                log_message(f"  [SCHEMA] scene_draft(シーン{sid}): {_se}")

    if isinstance(result, dict) and result.get("sd_prompt"):
        result["sd_prompt"] = deduplicate_sd_tags(result["sd_prompt"])
    return result


def polish_scene(
    client: anthropic.Anthropic,
    context: dict,
    draft: dict,
    char_profiles: list = None,
    cost_tracker: CostTracker = None,
    callback: Optional[Callable] = None,
    model: str = None
) -> dict:
    # キャラプロファイルをフル活用
    char_guide = ""
    if char_profiles:
        for cp in char_profiles:
            name = cp.get("character_name", "")
            speech = cp.get("speech_pattern", {})
            emotional = cp.get("emotional_speech", {})
            examples = cp.get("dialogue_examples", {})
            erotic = cp.get("erotic_speech_guide", {})
            
            char_guide += f"""
【{name}の口調チェックリスト】
✓ 一人称: {speech.get('first_person', '私')}
✓ 語尾: {', '.join(speech.get('sentence_endings', [])[:6])}
✓ 間投詞: {', '.join(speech.get('fillers', [])[:4])}
✓ 照れた時: {emotional.get('when_embarrassed', '')}
✓ 甘える時: {emotional.get('when_flirty', '')}
✓ 感じてる時: {emotional.get('when_aroused', '')}
✓ 絶頂時: {emotional.get('when_climax', '')}
✓ 喘ぎ声（軽）: {examples.get('moaning_light', 'あっ...んっ...')}
✓ 喘ぎ声（激）: {examples.get('moaning_intense', 'あっあっ...♡')}
✓ エロ度: {erotic.get('shyness_level', 3)}/5（数字が大きいほど恥ずかしがり）
"""

    system_prompt = f"""あなたはFANZA同人CG集の清書担当です。
下書きの吹き出しテキストを「そのキャラが本当に言いそうな」自然で短い表現に磨き上げてください。

{char_guide if char_guide else "（キャラプロファイルなし）"}

## CG集の清書ルール

【吹き出し改善】
1. 1ページ = ヒロインのセリフ1-2個 + 男性セリフ0-1個（最大3個）
2. 説明的→感情的に（「嬉しい気持ちです」→「嬉しい…♡」）
3. 文章→断片に（主語・目的語を削除）
4. 一人称・語尾を徹底チェック
5. 4個以上の吹き出しがあればヒロイン1-2個+男性0-1個に絞る

【エロシーン改善】
- 「気持ちいいです」→「きもちぃ…♡」
- 「もっとしてください」→「もっと…♡」
- 「イキそうです」→「イっちゃ…♡」
- 喘ぎ声は途切れ途切れに

【オノマトペ改善】
- 場面に合った効果音か確認
- 数は適切か（intensity 1-2: 0-1個、3: 1-2個、4-5: 2-4個）

【禁止】
❌ 4個以上の吹き出し（ヒロイン1-2 + 男性0-1 = 最大3個）
❌ 説明調のテキスト
❌ キャラの一人称・語尾の不一致

Output JSON only."""

    prompt = f"""設定: {json.dumps(context, ensure_ascii=False)}

下書き: {json.dumps(draft, ensure_ascii=False)}

上記の下書きを清書してください：

1. 各吹き出しをキャラの口調に合わせる
2. 吹き出しを最大3個に絞る（ヒロイン1-2 + 男性0-1）
3. descriptionをより詳細に（100字程度）
4. character_feelingsをより感情的に
5. onomatopoeiaが場面に合っているか確認

## 保持すべきフィールド
- scene_id, title, description, location_detail
- mood, character_feelings
- bubbles (speaker, type, text)
- onomatopoeia
- direction, story_flow
- sd_prompt

同じJSON形式で出力。JSONのみ。"""

    response = _call_api(
        client, model or MODELS["sonnet"],
        system_prompt,
        prompt, cost_tracker, 2500, callback
    )
    return parse_json_response(response)


# === Skill 3: Script Quality Supervisor ===
def check_quality(
    client: anthropic.Anthropic,
    context: dict,
    scenes: list,
    cost_tracker: CostTracker,
    callback: Optional[Callable] = None
) -> dict:
    skill = load_skill("script_quality_supervisor")
    prompt = f"""設定: {json.dumps(context, ensure_ascii=False)}

シーン一覧: {json.dumps(scenes, ensure_ascii=False)}

以下をチェック:
1. キャラの口調一貫性
2. シーン目標達成
3. 感情の平坦さ
4. ペーシング問題
5. シーン間矛盾

出力形式（JSON）:
{{
    "has_problems": true/false,
    "problems": [
        {{"scene_id": 1, "type": "問題種別", "detail": "詳細"}}
    ],
    "fix_instructions": [
        {{"scene_id": 1, "instruction": "修正指示（最小限）"}}
    ]
}}

問題なければhas_problems: false。JSONのみ出力。"""

    if callback:
        callback("[CHECK]品質チェック中...")

    response = _call_api(
        client, MODELS["haiku"],
        skill if skill else "You check script quality and suggest minimal fixes.",
        prompt, cost_tracker, 2048, callback
    )
    return parse_json_response(response)


def apply_fix(
    client: anthropic.Anthropic,
    scene: dict,
    instruction: str,
    cost_tracker: CostTracker,
    callback: Optional[Callable] = None
) -> dict:
    prompt = f"""シーン: {json.dumps(scene, ensure_ascii=False)}

修正指示: {instruction}

指示に従い、該当箇所のみ修正してください。
全体の再生成は禁止。最小限の変更のみ。

同じJSON形式で出力。JSONのみ。"""

    response = _call_api(
        client, MODELS["haiku"],
        "You apply minimal fixes to scripts. Never regenerate entirely.",
        prompt, cost_tracker, 2048, callback
    )
    return parse_json_response(response)


# === Wave並列生成ヘルパー ===
def _generate_single_scene_for_wave(
    client, context, scene, jailbreak, cost_tracker, theme, char_profiles,
    callback, story_so_far, synopsis, current_roadmap, male_description,
    scene_index, total_scenes, timestamp, faceless_male=True,
):
    """Wave並列生成用: 1シーン分の生成+エラーハンドリング。

    戻り値: (scene_index, result_dict_or_None, summary_string_or_None, error_msg_or_None)
    InterruptedError は再送出してWave全体を停止させる。
    """
    intensity = scene.get("intensity", 3)
    model_type = "Sonnet" if intensity >= 4 else "Haiku(4.5)"

    def _try_generate(**extra_kwargs):
        draft = generate_scene_draft(
            client, context, scene, jailbreak,
            cost_tracker, theme, char_profiles, callback,
            story_so_far=story_so_far,
            synopsis=extra_kwargs.get("synopsis_override", synopsis),
            outline_roadmap=current_roadmap,
            male_description=male_description,
            scene_index=scene_index,
            total_scenes=total_scenes,
            faceless_male=faceless_male,
        )
        draft["intensity"] = intensity
        scene_val = validate_scene(draft, scene_index)
        if not scene_val["valid"]:
            for err in scene_val["errors"]:
                log_message(f"  [SCHEMA] シーン{scene_index+1}: {err}")
        return draft

    try:
        log_message(f"シーン {scene_index+1}/{total_scenes} 生成開始 (intensity={intensity}, {model_type})")
        if callback:
            callback(f"[SCENE]シーン {scene_index+1}/{total_scenes} [{model_type}] 重要度{intensity}")

        draft = _try_generate()

        # ファイル保存
        draft_file = DRAFTS_DIR / f"draft_{timestamp}_scene{scene_index+1}.json"
        with open(draft_file, "w", encoding="utf-8") as f:
            json.dump(draft, f, ensure_ascii=False, indent=2)
        final_file = FINAL_DIR / f"final_{timestamp}_scene{scene_index+1}.json"
        with open(final_file, "w", encoding="utf-8") as f:
            json.dump(draft, f, ensure_ascii=False, indent=2)

        summary = extract_scene_summary(draft)
        log_message(f"シーン {scene_index+1}/{total_scenes} 完了")
        if callback:
            callback(f"[OK]シーン {scene_index+1}/{total_scenes} 完了")
        return (scene_index, draft, summary, None)

    except InterruptedError:
        # ユーザー停止要求 → 再送出してWave全体を停止させる
        raise

    except Exception as e:
        err_msg = str(e)
        log_message(f"シーン {scene_index+1} 生成エラー: {err_msg}")

        # 529 Overloaded: クールダウン後にリトライ
        is_overloaded = "サーバー過負荷" in err_msg or "529" in err_msg or "Overloaded" in err_msg
        if is_overloaded:
            cooldown = 60
            log_message(f"サーバー過負荷検出: {cooldown}秒待機後にシーン{scene_index+1}をリトライ")
            if callback:
                callback(f"[WARN]サーバー過負荷、{cooldown}秒待機後にシーン{scene_index+1}をリトライ...")
            time.sleep(cooldown)
            try:
                draft = _try_generate()
                draft_file = DRAFTS_DIR / f"draft_{timestamp}_scene{scene_index+1}.json"
                with open(draft_file, "w", encoding="utf-8") as f:
                    json.dump(draft, f, ensure_ascii=False, indent=2)
                final_file = FINAL_DIR / f"final_{timestamp}_scene{scene_index+1}.json"
                with open(final_file, "w", encoding="utf-8") as f:
                    json.dump(draft, f, ensure_ascii=False, indent=2)
                summary = extract_scene_summary(draft)
                log_message(f"シーン {scene_index+1} 過負荷リトライ成功")
                if callback:
                    callback(f"[OK]シーン {scene_index+1}/{total_scenes} 過負荷リトライ成功")
                return (scene_index, draft, summary, None)
            except InterruptedError:
                raise
            except Exception as e2:
                err_msg = str(e2)
                log_message(f"シーン {scene_index+1} 過負荷リトライも失敗: {err_msg}")

        # コンテンツ拒否 or JSONパースエラー → 最大2回リトライ
        is_refusal = any(kw in err_msg for kw in ["倫理", "対応することはできません", "cannot", "inappropriate"])
        is_json_error = any(kw in err_msg for kw in ["Invalid JSON", "No JSON", "Empty response", "JSONDecodeError"])

        if is_refusal or is_json_error:
            retry_reason = "コンテンツ拒否" if is_refusal else "JSONパースエラー"
            log_message(f"シーン {scene_index+1} {retry_reason}検出、リトライ")
            if callback:
                callback(f"[WARN]シーン {scene_index+1} {retry_reason}、リトライ中...")
            for retry_n in range(2):
                try:
                    draft = _try_generate(synopsis_override="" if is_refusal else synopsis)
                    draft_file = DRAFTS_DIR / f"draft_{timestamp}_scene{scene_index+1}.json"
                    with open(draft_file, "w", encoding="utf-8") as f:
                        json.dump(draft, f, ensure_ascii=False, indent=2)
                    final_file = FINAL_DIR / f"final_{timestamp}_scene{scene_index+1}.json"
                    with open(final_file, "w", encoding="utf-8") as f:
                        json.dump(draft, f, ensure_ascii=False, indent=2)
                    summary = extract_scene_summary(draft)
                    log_message(f"シーン {scene_index+1} リトライ{retry_n+1}回目成功")
                    if callback:
                        callback(f"[OK]シーン {scene_index+1}/{total_scenes} リトライ成功")
                    return (scene_index, draft, summary, None)
                except InterruptedError:
                    raise
                except Exception as e2:
                    log_message(f"シーン {scene_index+1} リトライ{retry_n+1}回目失敗: {e2}")

        # 全リトライ失敗 → エラー結果を返す
        import traceback
        log_message(traceback.format_exc())
        if callback:
            callback(f"[ERROR]シーン {scene_index+1} エラー: {err_msg[:50]}")

        error_result = {
            "scene_id": scene.get("scene_id", scene_index + 1),
            "title": f"シーン{scene_index+1}",
            "mood": "エラー",
            "bubbles": [],
            "onomatopoeia": [],
            "direction": f"生成エラー: {err_msg[:100]}",
            "sd_prompt": ""
        }
        return (scene_index, error_result, f"[シーン{scene_index+1}: エラーにより欠落]", None)


def _generate_scenes_wave(
    wave_scenes, client, context, jailbreak, cost_tracker, theme, char_profiles,
    callback, story_so_far, synopsis, roadmap_lines, male_description,
    total_scenes, timestamp, max_workers=CONCURRENT_BATCH_SIZE,
    outline=None, faceless_male=True,
):
    """Wave内の全シーンをThreadPoolExecutorで同時生成し、scene_index順にソートして返す。

    戻り値: [(scene_index, result_dict, summary_string, error), ...]
    InterruptedError発生時はexecutorをシャットダウンして再送出。
    v8.7: outline引数追加でWave内前シーンアウトライン情報を注入。
    """
    wave_results = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {}
        for scene_index, scene in wave_scenes:
            # 各シーンのロードマップ（近傍±5行）
            marked_lines = []
            window_start = max(0, scene_index - 5)
            window_end = min(len(roadmap_lines), scene_index + 6)
            if window_start > 0:
                marked_lines.append(f"  ... (シーン1〜{window_start}省略)")
            for j in range(window_start, window_end):
                line = roadmap_lines[j]
                if j == scene_index:
                    marked_lines.append(f"★ {line}")
                else:
                    marked_lines.append(f"  {line}")
            if window_end < len(roadmap_lines):
                marked_lines.append(f"  ... (シーン{window_end+1}〜{len(roadmap_lines)}省略)")
            current_roadmap = "\n".join(marked_lines)

            # v8.7: Wave内の前シーンアウトライン情報を注入（ストーリー一貫性向上）
            story_so_far_augmented = story_so_far
            if outline and scene_index > 0 and scene_index - 1 < len(outline):
                prev = outline[scene_index - 1]
                prev_anchor = (
                    f"\n★直前シーン{scene_index}: {prev.get('title', '')[:20]} - "
                    f"{prev.get('situation', '')[:80]}\n"
                    f"このシーンは上記の直後から始まる。\n"
                )
                story_so_far_augmented = story_so_far + prev_anchor

            future = executor.submit(
                _generate_single_scene_for_wave,
                client, context, scene, jailbreak, cost_tracker, theme, char_profiles,
                callback, story_so_far_augmented, synopsis, current_roadmap, male_description,
                scene_index, total_scenes, timestamp, faceless_male,
            )
            futures[future] = scene_index

        interrupted = False
        for future in as_completed(futures):
            try:
                result = future.result()
                wave_results.append(result)
            except InterruptedError:
                interrupted = True
                # 残りのfutureをキャンセル
                for f in futures:
                    f.cancel()
                break

        if interrupted:
            raise InterruptedError("ユーザーによる停止")

    # scene_index順にソート
    wave_results.sort(key=lambda x: x[0])
    return wave_results


# === メインパイプライン ===
def generate_pipeline(
    api_key: str,
    concept: str,
    characters: str,
    num_scenes: int,
    theme: str,
    callback: Optional[Callable] = None,
    skip_quality_check: bool = True,
    story_structure: dict = None,
    male_tags: str = "",
    time_tags: str = "",
    location_type: str = "",
    male_description: str = "",
    sd_quality_tags: str = "",
    sd_prefix_tags: str = "",
    sd_suffix_tags: str = "",
    provider: str = "",
    quality_priority: bool = False,
    faceless_male: bool = True,
    sd_neg_base: str = "",
    sd_neg_prefix: str = "",
    sd_neg_suffix: str = "",
) -> tuple[list, CostTracker]:
    client = anthropic.Anthropic(api_key=api_key)
    log_message("Claude (Anthropic) バックエンドで生成開始")
    cost_tracker = CostTracker()

    jailbreak = load_file(JAILBREAK_FILE)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # キャラプロファイルを読み込み（部分一致対応）
    char_profiles = []
    characters_lower = characters.lower()
    log_message(f"キャラプロファイル検索開始: {characters}")
    
    for json_file in CHARACTERS_DIR.glob("*.json"):
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                profile = json.load(f)
                char_name = profile.get("character_name", "")
                work_title = profile.get("work_title", "")
                if char_name and (
                    char_name in characters or
                    char_name.lower() in characters_lower or
                    any(part in characters for part in char_name.split())
                ):
                    char_profiles.append(profile)
                    log_message(f"キャラプロファイル読込: {char_name} ({work_title})")
                    if callback:
                        callback(f"[FILE]キャラ設定適用: {char_name}（{work_title}）")
        except Exception as e:
            log_message(f"キャラプロファイル読込エラー: {e}")

    # プリセットも検索
    for json_file in PRESET_CHARS_DIR.glob("*.json"):
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                profile = json.load(f)
                char_name = profile.get("character_name", "")
                work_title = profile.get("work_title", "")
                existing_names = [cp.get("character_name", "") for cp in char_profiles]
                if char_name and char_name not in existing_names and (
                    char_name in characters or
                    char_name.lower() in characters_lower or
                    any(part in characters for part in char_name.split())
                ):
                    char_profiles.append(profile)
                    log_message(f"プリセットキャラ読込: {char_name} ({work_title})")
                    if callback:
                        callback(f"[PACK]プリセットキャラ適用: {char_name}（{work_title}）")
        except Exception as e:
            log_message(f"プリセット読込エラー: {e}")
    
    if char_profiles:
        char_names = [cp.get("character_name", "") for cp in char_profiles]
        log_message(f"使用キャラ設定: {', '.join(char_names)}")
        if callback:
            callback(f"[OK]{len(char_profiles)}件のキャラ設定を適用")
    else:
        log_message("キャラ設定なし - 汎用設定で生成")
        if callback:
            callback("[WARN]キャラ設定なし（汎用設定で生成）")

    # テーマ自動推定
    if not theme:
        theme = _infer_theme_from_concept(concept)
        if theme:
            log_message(f"テーマ自動推定: {THEME_GUIDES.get(theme, {}).get('name', theme)}")

    # テーマ情報
    theme_guide = THEME_GUIDES.get(theme, {})
    if not theme_guide:
        theme_guide = _build_dynamic_theme_guide(concept)
    theme_name = theme_guide.get("name", "指定なし")
    if theme and theme_guide:
        log_message(f"テーマ適用: {theme_name} (arc: {theme_guide.get('story_arc', '')})")
        if callback:
            callback(f"[CHAR]テーマ: {theme_name}")

    # Phase 1: コンテキスト圧縮
    log_message("Phase 1 開始: コンテキスト圧縮")
    if callback:
        callback("🔧 Phase 1: コンテキスト圧縮")

    try:
        if char_profiles:
            context = compact_context_local(concept, characters, theme, char_profiles, callback)
            log_message("コンテキスト圧縮完了（ローカル）")
        else:
            context = compact_context(client, concept, characters, theme, cost_tracker, callback)
            log_message("コンテキスト圧縮完了（API）")
    except Exception as e:
        log_message(f"コンテキスト圧縮エラー: {e}")
        raise

    context_file = CONTEXT_DIR / f"context_{timestamp}.json"
    with open(context_file, "w", encoding="utf-8") as f:
        json.dump(context, f, ensure_ascii=False, indent=2)

    # スキーマバリデーション: コンテキスト
    ctx_validation = validate_context(context)
    if not ctx_validation["valid"]:
        for err in ctx_validation["errors"]:
            log_message(f"  [SCHEMA] context: {err}")
        if callback:
            callback(f"[WARN]コンテキスト検証: {len(ctx_validation['errors'])}件の問題")

    if callback:
        callback("[OK]コンテキスト圧縮完了")

    # Phase 2: ストーリーあらすじ生成（Haiku 1回）
    log_message("Phase 2 開始: ストーリーあらすじ生成")
    if callback:
        callback("🔧 Phase 2: ストーリー原案作成")

    try:
        synopsis = generate_synopsis(client, concept, context, num_scenes, theme, cost_tracker, callback, male_description=male_description, faceless_male=faceless_male)
        log_message(f"あらすじ生成完了: {len(synopsis)}文字")

        # あらすじをファイルに保存
        synopsis_file = CONTEXT_DIR / f"synopsis_{timestamp}.txt"
        with open(synopsis_file, "w", encoding="utf-8") as f:
            f.write(synopsis)
    except Exception as e:
        log_message(f"あらすじ生成エラー: {e}")
        import traceback
        log_message(traceback.format_exc())
        # フォールバック: コンセプトをあらすじとして使用
        synopsis = concept
        if callback:
            callback(f"[WARN]あらすじ生成失敗、コンセプトで代替")

    if callback:
        callback("[OK]ストーリー原案完成")

    # Phase 3: アウトライン生成（あらすじをシーン分割）
    log_message("Phase 3 開始: アウトライン生成（シーン分割）")
    if callback:
        callback("🔧 Phase 3: シーン分割")

    try:
        outline = generate_outline(client, context, num_scenes, theme, cost_tracker, callback, synopsis=synopsis, story_structure=story_structure, male_description=male_description, faceless_male=faceless_male)
        log_message(f"アウトライン生成完了: {len(outline)}シーン")
        
        intensity_counts = {}
        for scene in outline:
            i = scene.get("intensity", 3)
            intensity_counts[i] = intensity_counts.get(i, 0) + 1
        log_message(f"intensity分布: {intensity_counts}")
    except Exception as e:
        log_message(f"アウトライン生成エラー: {e}、フォールバック（均等分割）を使用")
        if callback:
            callback(f"[WARN]シーン分割エラー、均等分割で代替中...")
        # フォールバック: ストーリー構成に基づく均等分割
        fb_ss = story_structure or {"prologue": 10, "main": 80, "epilogue": 10}
        fb_pro = fb_ss.get("prologue", 10) / 100
        fb_epi = fb_ss.get("epilogue", 10) / 100
        fb_main_start = fb_pro
        fb_main_end = 1.0 - fb_epi
        outline = []
        for idx in range(1, num_scenes + 1):
            ratio = idx / num_scenes
            if ratio <= fb_pro:
                intensity = 1  # プロローグ
            elif ratio <= fb_main_start + (fb_main_end - fb_main_start) * 0.25:
                intensity = 3  # 前戯
            elif ratio <= fb_main_end:
                intensity = 4 + (1 if ratio > fb_main_start + (fb_main_end - fb_main_start) * 0.7 else 0)
            else:
                intensity = 3  # エピローグ
            outline.append({
                "scene_id": idx,
                "summary": f"シーン{idx}",
                "intensity": min(intensity, 5),
                "location": "室内",
                "time": ""
            })
        log_message(f"フォールバックアウトライン生成: {num_scenes}シーン")

    # スキーマバリデーション: アウトライン
    outline_validation = validate_outline(outline, num_scenes)
    if not outline_validation["valid"]:
        for err in outline_validation["errors"]:
            log_message(f"  [SCHEMA] outline: {err}")
        if callback:
            callback(f"[WARN]アウトライン検証: {len(outline_validation['errors'])}件の問題")

    if callback:
        _haiku_n = sum(1 for s in outline if s.get("intensity", 3) <= 4)
        _high_n = sum(1 for s in outline if s.get("intensity", 3) >= 5)
        callback(f"[OK]シーン分割完成: {len(outline)}シーン（Haiku4.5×{_haiku_n} + Sonnet×{_high_n}）")

    # コスト見積もり（Prompt Caching反映版）
    haiku_count = sum(1 for s in outline if s.get("intensity", 3) <= 3)
    high_count = sum(1 for s in outline if s.get("intensity", 3) >= 4)
    h_cost = COSTS.get(MODELS["haiku"], {"input": 1.00, "output": 5.00})
    s_cost = COSTS.get(MODELS["sonnet"], {"input": 3.00, "output": 15.00})
    # あらすじ+アウトライン → haiku 4.5
    overhead_cost = 2 * (2000 / 1_000_000 * h_cost["input"] + 2000 / 1_000_000 * h_cost["output"])
    # シーン生成（Prompt Caching: systemは1回cache_create + (N-1)回cache_read）
    cached_sys = 16000  # systemプロンプト推定トークン数
    avg_user = 3000  # 平均user prompt
    # Haiku: cache_create 1回(1.25x) + cache_read (N-1)回(0.1x) + 非キャッシュ入力 + 出力
    haiku_scene_cost = (
        (cached_sys / 1_000_000 * h_cost["input"] * 1.25) +  # 初回cache作成
        (cached_sys / 1_000_000 * h_cost["input"] * 0.10 * max(0, haiku_count - 1)) +  # cache読取
        (haiku_count * avg_user / 1_000_000 * h_cost["input"]) +  # 非キャッシュ入力
        (haiku_count * 700 / 1_000_000 * h_cost["output"])  # 出力
    ) if haiku_count > 0 else 0
    # Sonnet: 同様
    sonnet_scene_cost = (
        (cached_sys / 1_000_000 * s_cost["input"] * 1.25) +
        (cached_sys / 1_000_000 * s_cost["input"] * 0.10 * max(0, high_count - 1)) +
        (high_count * avg_user / 1_000_000 * s_cost["input"]) +
        (high_count * 700 / 1_000_000 * s_cost["output"])
    ) if high_count > 0 else 0
    est_cost = overhead_cost + haiku_scene_cost + sonnet_scene_cost
    if callback:
        callback(f"[COST]推定コスト: ${est_cost:.4f}（API {len(outline)+2}回: Haiku4.5×{haiku_count} + Sonnet×{high_count}）")

    # Phase 4: シーン生成
    results = []
    story_summaries = []

    # ストーリーロードマップ構築（全シーンの概要を各シーン生成に渡す）
    roadmap_lines = []
    for s in outline:
        sid = s.get("scene_id", "?")
        title = s.get("title", "")[:20]
        _rm_intensity = s.get("intensity", 3)
        situation = s.get("situation", "")[:60]
        location = s.get("location", "")[:15]
        goal = s.get("goal", "")[:30]
        goal_part = f" 目的:{goal}" if goal else ""
        roadmap_lines.append(f"[{sid}] {title} (i={_rm_intensity}, {location}) {situation}{goal_part}")
    outline_roadmap = "\n".join(roadmap_lines)

    # v8.7: 品質優先モードではWave並列を無効化（全シーン直列生成）
    use_wave_parallel = len(outline) >= CONCURRENT_MIN_SCENES and not quality_priority
    if quality_priority and len(outline) >= CONCURRENT_MIN_SCENES:
        log_message(f"品質優先モード: {len(outline)}シーンを直列生成（Wave並列無効）")
        if callback:
            callback("[INFO]品質優先モード: 直列生成（ストーリー一貫性最大化）")

    if use_wave_parallel:
        # === Wave並列モード: 5シーン同時生成 ===
        num_waves = (len(outline) + CONCURRENT_BATCH_SIZE - 1) // CONCURRENT_BATCH_SIZE
        log_message(f"Wave並列モード: {CONCURRENT_BATCH_SIZE}シーン同時生成×{num_waves}wave")
        if callback:
            callback(f"[INFO]Wave並列モード: {CONCURRENT_BATCH_SIZE}シーン同時生成×{num_waves}wave")

        wave_idx = 0
        scene_cursor = 0
        while scene_cursor < len(outline):
            wave_end = min(scene_cursor + CONCURRENT_BATCH_SIZE, len(outline))
            wave_scenes = [(i, outline[i]) for i in range(scene_cursor, wave_end)]
            wave_idx += 1

            log_message(f"Wave {wave_idx}: シーン {scene_cursor+1}-{wave_end}/{len(outline)} 並列生成中...")
            if callback:
                callback(f"[WAVE]Wave {wave_idx}: シーン {scene_cursor+1}-{wave_end}/{len(outline)} 並列生成中...")

            # Wave開始前にstory_so_farスナップショットを取得（Wave内全シーンで共有）
            story_so_far = _build_story_so_far(story_summaries, results)

            wave_results = _generate_scenes_wave(
                wave_scenes, client, context, jailbreak, cost_tracker, theme,
                char_profiles, callback, story_so_far, synopsis, roadmap_lines,
                male_description, len(outline), timestamp,
                outline=outline, faceless_male=faceless_male,
            )

            # 結果をscene_index順に蓄積
            for si, draft, summary, _err in wave_results:
                results.append(draft)
                story_summaries.append(summary)
                log_message(f"シーン {si+1} 要約蓄積: {summary[:80]}...")

            log_message(f"Wave {wave_idx} 完了: シーン {scene_cursor+1}-{wave_end}/{len(outline)}")
            if callback:
                callback(f"[OK]Wave {wave_idx} 完了: シーン {scene_cursor+1}-{wave_end}/{len(outline)}")

            scene_cursor = wave_end

            # Wave間クールダウン（最終Wave以外）
            if scene_cursor < len(outline):
                # ユーザー停止チェック
                if callback:
                    try:
                        callback(f"[INFO]Wave間クールダウン {CONCURRENT_WAVE_COOLDOWN}秒...")
                    except InterruptedError:
                        raise
                time.sleep(CONCURRENT_WAVE_COOLDOWN)

    else:
        # === 直列モード: 12シーン以下（従来通り） ===
        for i, scene in enumerate(outline):
            intensity = scene.get("intensity", 3)
            model_type = "Sonnet" if intensity >= 4 else "Haiku(4.5)"

            # story_so_far を構築（スライディングウィンドウ方式）
            story_so_far = _build_story_so_far(story_summaries, results)

            # 現在シーン近傍±5のロードマップ
            marked_lines = []
            window_start = max(0, i - 5)
            window_end = min(len(roadmap_lines), i + 6)
            if window_start > 0:
                marked_lines.append(f"  ... (シーン1〜{window_start}省略)")
            for j in range(window_start, window_end):
                line = roadmap_lines[j]
                if j == i:
                    marked_lines.append(f"★ {line}")
                else:
                    marked_lines.append(f"  {line}")
            if window_end < len(roadmap_lines):
                marked_lines.append(f"  ... (シーン{window_end+1}〜{len(roadmap_lines)}省略)")
            current_roadmap = "\n".join(marked_lines)

            try:
                log_message(f"シーン {i+1}/{len(outline)} 生成開始 (intensity={intensity}, {model_type})")
                if callback:
                    callback(f"[SCENE]シーン {i+1}/{len(outline)} [{model_type}] 重要度{intensity}")

                draft = generate_scene_draft(
                    client, context, scene, jailbreak,
                    cost_tracker, theme, char_profiles, callback,
                    story_so_far=story_so_far,
                    synopsis=synopsis,
                    outline_roadmap=current_roadmap,
                    male_description=male_description,
                    scene_index=i,
                    total_scenes=len(outline),
                    faceless_male=faceless_male,
                )

                draft["intensity"] = intensity

                # スキーマバリデーション: 個別シーン
                scene_val = validate_scene(draft, i)
                if not scene_val["valid"]:
                    for err in scene_val["errors"]:
                        log_message(f"  [SCHEMA] シーン{i+1}: {err}")
                    if callback:
                        callback(f"[WARN]シーン{i+1}検証: {len(scene_val['errors'])}件の問題")

                results.append(draft)

                # 要約を蓄積して次シーンに渡す
                summary = extract_scene_summary(draft)
                story_summaries.append(summary)
                log_message(f"シーン {i+1} 要約蓄積: {summary[:80]}...")

                draft_file = DRAFTS_DIR / f"draft_{timestamp}_scene{i+1}.json"
                with open(draft_file, "w", encoding="utf-8") as f:
                    json.dump(draft, f, ensure_ascii=False, indent=2)
                final_file = FINAL_DIR / f"final_{timestamp}_scene{i+1}.json"
                with open(final_file, "w", encoding="utf-8") as f:
                    json.dump(draft, f, ensure_ascii=False, indent=2)

                log_message(f"シーン {i+1}/{len(outline)} 完了")
                if callback:
                    callback(f"[OK]シーン {i+1}/{len(outline)} 完了")

            except Exception as e:
                err_msg = str(e)
                log_message(f"シーン {i+1} 生成エラー: {err_msg}")

                # 529 Overloaded: グローバルクールダウン後にリトライ
                is_overloaded = "サーバー過負荷" in err_msg or "529" in err_msg or "Overloaded" in err_msg
                if is_overloaded:
                    cooldown = 60
                    log_message(f"サーバー過負荷検出: {cooldown}秒のグローバルクールダウン後にシーン{i+1}をリトライ")
                    if callback:
                        callback(f"[WARN]サーバー過負荷、{cooldown}秒待機後にシーン{i+1}をリトライ...")
                    time.sleep(cooldown)
                    try:
                        draft = generate_scene_draft(
                            client, context, scene, jailbreak,
                            cost_tracker, theme, char_profiles, callback,
                            story_so_far=story_so_far, synopsis=synopsis,
                            outline_roadmap=current_roadmap,
                            male_description=male_description,
                            scene_index=i,
                            total_scenes=len(outline),
                            faceless_male=faceless_male,
                        )
                        draft["intensity"] = intensity
                        results.append(draft)
                        summary = extract_scene_summary(draft)
                        story_summaries.append(summary)
                        log_message(f"シーン {i+1} 過負荷リトライ成功")
                        if callback:
                            callback(f"[OK]シーン {i+1}/{len(outline)} 過負荷リトライ成功")
                        continue
                    except Exception as e2:
                        err_msg = str(e2)
                        log_message(f"シーン {i+1} 過負荷リトライも失敗: {err_msg}")

                # リトライ判定（コンテンツ拒否 or JSONパースエラー）
                is_refusal = any(kw in err_msg for kw in ["倫理", "対応することはできません", "cannot", "inappropriate"])
                is_json_error = any(kw in err_msg for kw in ["Invalid JSON", "No JSON", "Empty response", "JSONDecodeError"])

                if is_refusal or is_json_error:
                    retry_reason = "コンテンツ拒否" if is_refusal else "JSONパースエラー"
                    log_message(f"シーン {i+1} {retry_reason}検出、リトライ")
                    if callback:
                        callback(f"[WARN]シーン {i+1} {retry_reason}、リトライ中...")

                    # 最大2回リトライ
                    retry_success = False
                    for retry_n in range(2):
                        try:
                            draft = generate_scene_draft(
                                client, context, scene, jailbreak,
                                cost_tracker, theme, char_profiles, callback,
                                story_so_far=story_so_far,
                                synopsis="" if is_refusal else synopsis,
                                outline_roadmap=current_roadmap,
                                male_description=male_description,
                                scene_index=i,
                                total_scenes=len(outline),
                                faceless_male=faceless_male,
                            )
                            draft["intensity"] = intensity
                            results.append(draft)
                            summary = extract_scene_summary(draft)
                            story_summaries.append(summary)
                            log_message(f"シーン {i+1} リトライ{retry_n+1}回目成功")
                            if callback:
                                callback(f"[OK]シーン {i+1}/{len(outline)} リトライ成功")
                            retry_success = True
                            break
                        except Exception as e2:
                            log_message(f"シーン {i+1} リトライ{retry_n+1}回目失敗: {e2}")

                    if retry_success:
                        continue

                import traceback
                log_message(traceback.format_exc())
                if callback:
                    callback(f"[ERROR]シーン {i+1} エラー: {err_msg[:50]}")

                error_result = {
                    "scene_id": scene.get("scene_id", i + 1),
                    "title": f"シーン{i+1}",
                    "mood": "エラー",
                    "bubbles": [],
                    "onomatopoeia": [],
                    "direction": f"生成エラー: {err_msg[:100]}",
                    "sd_prompt": ""
                }
                results.append(error_result)
                story_summaries.append(f"[シーン{i+1}: エラーにより欠落]")

    # スキーマバリデーション: 結果配列全体
    results_validation = validate_results(results)
    if not results_validation["valid"]:
        log_message(f"[SCHEMA] 結果検証: {len(results_validation['errors'])}件の構造問題")
        for sid, errs in results_validation.get("scene_errors", {}).items():
            for err in errs:
                log_message(f"  [SCHEMA] シーン{sid}: {err}")
        if callback:
            stats = results_validation["stats"]
            callback(f"[STAT]スキーマ検証: {stats['valid_count']}/{stats['total']}シーンOK, {len(results_validation['errors'])}件の構造問題")
    else:
        log_message("[SCHEMA] 結果配列全体のスキーマ検証OK")

    # Phase 5: 品質検証 + SDプロンプト最適化（APIコスト不要）
    log_message("Phase 5 開始: 品質検証 + SDプロンプト最適化")
    if callback:
        callback("[CHECK]Phase 5: 品質検証 + SDプロンプト最適化")

    # 5-1: FANZA基準で自動検証
    try:
        validation = validate_script(results, theme, char_profiles)
        log_message(f"品質検証完了: {validation['summary']}")
        if callback:
            callback(f"[STAT]{validation['summary']}")

        # シーン別問題をログ（詳細はファイルのみ、GUIはサマリー）
        _scene_issue_count = sum(len(v) for v in validation["scene_issues"].values())
        for sid, issues in validation["scene_issues"].items():
            for issue in issues:
                log_message(f"  シーン{sid}: {issue}")
        if _scene_issue_count > 0 and callback:
            callback(f"  [WARN]シーン別問題: {_scene_issue_count}件検出（詳細はログファイル参照）")

        # 喘ぎ重複（詳細はファイルのみ）
        _moan_dup_count = len(validation.get("repeated_moans", {}))
        if _moan_dup_count > 0:
            for text, sids in validation["repeated_moans"].items():
                log_message(f"  喘ぎ重複「{text}」→ シーン{', '.join(str(s) for s in sids)}")
            if callback:
                callback(f"  [WARN]喘ぎ重複: {_moan_dup_count}件検出（自動修正で置換します）")

        # オノマトペ連続重複（詳細はファイルのみ）
        _ono_dup_count = len(validation.get("repeated_onomatopoeia", []))
        for s1, s2 in validation["repeated_onomatopoeia"]:
            log_message(f"  オノマトペ連続重複: シーン{s1}→{s2}")
        if _ono_dup_count > 0 and callback:
            callback(f"  [WARN]オノマトペ重複: {_ono_dup_count}件検出")
    except Exception as _validate_err:
        log_message(f"[WARN]品質検証エラー（スキップ）: {_validate_err}")
        import traceback
        log_message(traceback.format_exc())
        validation = {"summary": "検証スキップ", "score": 0, "scene_issues": {},
                      "repeated_moans": {}, "repeated_onomatopoeia": [],
                      "position_variety": {}, "total_issues": 0}

    # 5-2: SDプロンプト最適化（設定スタイル適用）
    setting_style = _detect_setting_style(concept, theme=theme)
    if setting_style:
        log_message(f"設定スタイル検出: {setting_style.get('prompt_hint', '')[:40]}...")
        if callback:
            callback(f"🏠 設定スタイル適用: {setting_style.get('prompt_hint', '')[:30]}...")
    try:
        results = enhance_sd_prompts(results, char_profiles, setting_style=setting_style,
                                        male_tags=male_tags, time_tags=time_tags,
                                        location_type=location_type,
                                        sd_quality_tags=sd_quality_tags,
                                        sd_prefix_tags=sd_prefix_tags,
                                        sd_suffix_tags=sd_suffix_tags,
                                        theme=theme,
                                        faceless_male=faceless_male,
                                        sd_neg_base=sd_neg_base,
                                        sd_neg_prefix=sd_neg_prefix,
                                        sd_neg_suffix=sd_neg_suffix)
        log_message("SDプロンプト最適化完了")
        if callback:
            callback("[OK]SDプロンプト最適化完了")
    except Exception as _sd_err:
        log_message(f"[WARN]SDプロンプト最適化エラー（結果はそのまま使用）: {_sd_err}")
        import traceback
        log_message(traceback.format_exc())

    # 5-2.5: 中間結果を自動保存（auto_fix前にAPI生成結果を保全）
    if len(results) >= 10:
        try:
            _ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            _raw_path = EXPORTS_DIR / f"script_{_ts}_raw.json"
            export_json(results, _raw_path)
            log_message(f"中間結果保存: {_raw_path}")
            if callback:
                callback(f"💾 中間結果保存済み（{len(results)}シーン）")
        except Exception as _save_err:
            log_message(f"中間結果保存失敗: {_save_err}")

    # 5-3.5: Opusクライマックス清書パス（intensity >= 5 のシーンのみ）
    opus_targets = [i for i, r in enumerate(results) if r.get("intensity", 0) >= 5]
    if opus_targets:
        log_message(f"Opus清書パス: {len(opus_targets)}シーン対象（intensity >= 5）")
        if callback:
            callback(f"[POLISH]Opus清書: {len(opus_targets)}シーン開始...")
        _opus_ok = 0
        for idx in opus_targets:
            scene = results[idx]
            _sid = scene.get("scene_id", idx + 1)
            _orig_intensity = scene.get("intensity", 0)
            _orig_scene_id = scene.get("scene_id")
            try:
                _opus_context = {"concept": concept, "theme": theme}
                polished = polish_scene(
                    client, _opus_context,
                    scene, char_profiles, cost_tracker, callback,
                    model=MODELS["opus"]
                )
                if polished and isinstance(polished, dict):
                    # intensity/scene_idの上書き保護
                    polished["intensity"] = _orig_intensity
                    if _orig_scene_id is not None:
                        polished["scene_id"] = _orig_scene_id
                    results[idx] = polished
                    _opus_ok += 1
                    log_message(f"  Opus清書OK: シーン{_sid}")
            except Exception as _opus_err:
                log_message(f"  [WARN]Opus清書失敗（シーン{_sid}、元データ維持）: {_opus_err}")
        log_message(f"Opus清書パス完了: {_opus_ok}/{len(opus_targets)}シーン成功")
        if callback:
            callback(f"[OK]Opus清書完了: {_opus_ok}/{len(opus_targets)}シーン")

    # 5-3: 自動修正（文字数マーカー除去、キャラ名統一、SDタグ整理、セリフ重複置換）
    if callback:
        callback("🔧 自動修正開始...")
    try:
        results = auto_fix_script(results, char_profiles, theme=theme, callback=callback, concept=concept)
        log_message("自動修正完了")
    except Exception as _autofix_err:
        log_message(f"[WARN]自動修正中にエラー発生（結果はそのまま使用）: {_autofix_err}")
        import traceback
        log_message(traceback.format_exc())
    if callback:
        callback("🔧 自動修正完了")

    # 5-4: dedup後の再検証（文字数超過・男性セリフ数の最終チェック）
    try:
        post_validation = validate_script(results, theme, char_profiles)
        if post_validation.get("issues"):
            log_message(f"再検証: {len(post_validation['issues'])}件の警告")
            for issue in post_validation["issues"][:5]:
                log_message(f"  {issue}")
    except Exception as _post_val_err:
        log_message(f"[WARN]再検証エラー（スキップ）: {_post_val_err}")

    # 完了サマリー
    success_count = sum(1 for r in results if r.get("mood") != "エラー")
    log_message(f"パイプライン完了: {success_count}/{len(results)}シーン成功")

    if callback:
        callback(f"[DONE]生成完了: {success_count}シーン成功（品質: {validation['score']}/100）")

    # メタデータを構築（エクスポート用）
    char_names = [cp.get("character_name", "") for cp in char_profiles] if char_profiles else []
    pipeline_metadata = {
        "concept": concept,
        "theme": theme,
        "theme_name": theme_name,
        "num_scenes": len(results),
        "characters": char_names,
        "story_structure": story_structure,
        "cost": cost_tracker.summary(),
        "quality_score": validation.get("score", 0),
        "provider": provider,
        "model_versions": (
            {"haiku": MODELS["haiku"], "sonnet": MODELS["sonnet"], "opus": MODELS["opus"]}
        ),
        "synopsis": synopsis,
    }

    return results, cost_tracker, pipeline_metadata


def export_csv(results: list, output_path: Path):
    fieldnames = [
        "scene_id", "title", "description", "bubble_no", "speaker", "text",
        "onomatopoeia", "sd_prompt",
        "type", "mood", "location_detail", "character_feelings",
        "direction", "story_flow"
    ]

    # utf-8-sig でBOM付きUTF-8（Excel対応）
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for scene in results:
            # キャラ心情を文字列に変換
            feelings = scene.get("character_feelings", {})
            if isinstance(feelings, dict):
                feelings_str = "; ".join([f"{k}: {v}" for k, v in feelings.items()])
            else:
                feelings_str = str(feelings)
            
            # オノマトペを文字列に
            onomatopoeia = scene.get("onomatopoeia", [])
            ono_str = ", ".join(onomatopoeia) if isinstance(onomatopoeia, list) else str(onomatopoeia)
            
            # 新フォーマット: bubbles、旧互換: dialogue
            bubbles = scene.get("bubbles", [])
            if not bubbles:
                bubbles = scene.get("dialogue", [])
            
            # SDプロンプト末尾にシーン番号を付与
            sd_raw = scene.get("sd_prompt", "")
            sid = scene.get("scene_id", "")
            sd_with_label = f'{sd_raw}, "シーン{sid}"' if sd_raw else ""

            if not bubbles:
                # 吹き出しがない場合でもシーン情報を出力
                writer.writerow({
                    "scene_id": sid,
                    "title": scene.get("title", ""),
                    "description": scene.get("description", ""),
                    "location_detail": scene.get("location_detail", ""),
                    "mood": scene.get("mood", ""),
                    "character_feelings": feelings_str,
                    "bubble_no": 0,
                    "speaker": "",
                    "type": "",
                    "text": "",
                    "onomatopoeia": ono_str,
                    "direction": scene.get("direction", ""),
                    "story_flow": scene.get("story_flow", ""),
                    "sd_prompt": sd_with_label
                })
            else:
                for idx, bubble in enumerate(bubbles):
                    writer.writerow({
                        "scene_id": sid if idx == 0 else "",
                        "title": scene.get("title", "") if idx == 0 else "",
                        "description": scene.get("description", "") if idx == 0 else "",
                        "location_detail": scene.get("location_detail", "") if idx == 0 else "",
                        "mood": scene.get("mood", "") if idx == 0 else "",
                        "character_feelings": feelings_str if idx == 0 else "",
                        "bubble_no": idx + 1,
                        "speaker": bubble.get("speaker", ""),
                        "type": bubble.get("type", bubble.get("emotion", "")),
                        "text": bubble.get("text", bubble.get("line", "")),
                        "onomatopoeia": ono_str if idx == 0 else "",
                        "direction": scene.get("direction", "") if idx == 0 else "",
                        "story_flow": scene.get("story_flow", "") if idx == 0 else "",
                        "sd_prompt": sd_with_label if idx == 0 else ""
                    })


def export_excel(results: list, output_path: Path):
    """Excel形式でエクスポート（CG集フォーマット対応）"""
    if not OPENPYXL_AVAILABLE:
        log_message("openpyxl未インストール - Excel出力スキップ")
        return False
    
    wb = Workbook()
    ws = wb.active
    ws.title = "脚本"
    
    # ヘッダー
    headers = [
        "シーンID", "タイトル", "シーン説明", "吹き出しNo", "話者", "テキスト",
        "オノマトペ", "SDプロンプト",
        "種類", "雰囲気", "場所詳細", "キャラ心情",
        "演出", "次への繋がり"
    ]
    
    # ヘッダースタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # データ
    row = 2
    for scene in results:
        feelings = scene.get("character_feelings", {})
        if isinstance(feelings, dict):
            feelings_str = "\n".join([f"{k}: {v}" for k, v in feelings.items()])
        else:
            feelings_str = str(feelings)
        
        # オノマトペを文字列に
        onomatopoeia = scene.get("onomatopoeia", [])
        ono_str = ", ".join(onomatopoeia) if isinstance(onomatopoeia, list) else str(onomatopoeia)
        
        # 新フォーマット: bubbles、旧互換: dialogue
        bubbles = scene.get("bubbles", [])
        if not bubbles:
            bubbles = scene.get("dialogue", [])
        if not bubbles:
            bubbles = [{}]
        
        for idx, bubble in enumerate(bubbles):
            data = [
                scene.get("scene_id", "") if idx == 0 else "",
                scene.get("title", "") if idx == 0 else "",
                scene.get("description", "") if idx == 0 else "",
                idx + 1 if bubble else "",
                bubble.get("speaker", ""),
                bubble.get("text", bubble.get("line", "")),
                ono_str if idx == 0 else "",
                scene.get("sd_prompt", "") if idx == 0 else "",
                bubble.get("type", bubble.get("emotion", "")),
                scene.get("mood", "") if idx == 0 else "",
                scene.get("location_detail", "") if idx == 0 else "",
                feelings_str if idx == 0 else "",
                scene.get("direction", "") if idx == 0 else "",
                scene.get("story_flow", "") if idx == 0 else ""
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                # 折り返し表示を有効化
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            row += 1
    
    # 列幅の設定
    column_widths = {
        1: 8,    # シーンID
        2: 12,   # タイトル
        3: 40,   # シーン説明
        4: 8,    # 吹き出しNo
        5: 10,   # 話者
        6: 20,   # テキスト
        7: 20,   # オノマトペ
        8: 60,   # SDプロンプト
        9: 8,    # 種類
        10: 10,  # 雰囲気
        11: 20,  # 場所詳細
        12: 25,  # キャラ心情
        13: 20,  # 演出
        14: 15   # 次への繋がり
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = width
    
    # ヘッダー行を固定
    ws.freeze_panes = "A2"
    
    wb.save(output_path)
    log_message(f"Excel出力完了: {output_path}")
    return True


def export_json(results: list, output_path: Path, metadata: dict = None):
    """JSON構造化エクスポート（メタデータ付き）"""
    data = {
        "version": "3.1.0",
        "generated_at": datetime.now().isoformat(),
        "scenes": results,
    }
    if metadata:
        meta_copy = dict(metadata)
        synopsis = meta_copy.pop("synopsis", None)
        data["metadata"] = meta_copy
        if synopsis:
            data["synopsis"] = synopsis
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def export_sd_prompts(results: list, output_path: Path):
    """SDプロンプト一括エクスポート（1行1プロンプト、シーンID付き）"""
    lines = []
    for scene in results:
        sd = scene.get("sd_prompt", "").strip()
        if sd:
            sid = scene.get("scene_id", "?")
            lines.append(f"# Scene {sid}: {scene.get('title', '')}")
            lines.append(sd)
            lines.append("")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    log_message(f"SDプロンプト出力完了: {output_path}")


def export_wildcard(results: list, output_path: Path,
                    male_tags: str = "", time_tags: str = "",
                    location_type: str = ""):
    """Wild Card形式エクスポート（1行1プロンプト、SD Wild Card対応）"""
    lines = []
    for scene in results:
        sd = scene.get("sd_prompt", "").strip()
        if sd:
            sid = scene.get("scene_id", "")
            lines.append(f'{sd}, "シーン{sid}",')
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    log_message(f"Wild Card出力完了: {output_path}（{len(lines)}行）")


def export_wildcard_negative(results: list, output_path: Path):
    """ネガティブプロンプト Wild Card形式エクスポート（1行1プロンプト）"""
    lines = []
    for scene in results:
        neg = scene.get("sd_negative_prompt", "").strip()
        if neg:
            sid = scene.get("scene_id", "")
            lines.append(f'{neg}, "シーン{sid}",')
    if not lines:
        return
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    log_message(f"Negative Wild Card出力完了: {output_path}（{len(lines)}行）")


def export_dialogue_list(results: list, output_path: Path):
    """セリフ一覧エクスポート（話者・種類・テキスト）"""
    lines = []
    for scene in results:
        sid = scene.get("scene_id", "?")
        title = scene.get("title", "")
        bubbles = scene.get("bubbles", []) or scene.get("dialogue", []) or []
        if not bubbles:
            continue
        lines.append(f"=== Scene {sid}: {title} ===")
        ono = scene.get("onomatopoeia", [])
        if ono:
            lines.append(f"  SE: {', '.join(ono) if isinstance(ono, list) else str(ono)}")
        for b in bubbles:
            speaker = b.get("speaker", "???")
            btype = b.get("type", b.get("emotion", ""))
            text = b.get("text", b.get("line", ""))
            tag = f"[{btype}]" if btype else ""
            lines.append(f"  {speaker}{tag}: {text}")
        lines.append("")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    log_message(f"セリフ一覧出力完了: {output_path}")


def export_markdown(results: list, output_path: Path):
    """マークダウン形式エクスポート（脚本全体の読みやすいビュー）"""
    lines = []
    lines.append(f"# CG集脚本")
    lines.append(f"")
    lines.append(f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"シーン数: {len(results)}")
    lines.append(f"")
    lines.append(f"---")
    lines.append(f"")

    for scene in results:
        sid = scene.get("scene_id", "?")
        title = scene.get("title", "")
        desc = scene.get("description", "")
        mood = scene.get("mood", "")
        location = scene.get("location_detail", "")
        direction = scene.get("direction", "")
        story_flow = scene.get("story_flow", "")
        sd = scene.get("sd_prompt", "")
        intensity = scene.get("intensity", "")

        lines.append(f"## Scene {sid}: {title}")
        lines.append(f"")
        if mood:
            lines.append(f"**雰囲気**: {mood}")
        if location:
            lines.append(f"**場所**: {location}")
        if intensity:
            lines.append(f"**強度**: {intensity}/5")
        lines.append(f"")
        if desc:
            lines.append(f"> {desc}")
            lines.append(f"")

        # キャラ心情
        feelings = scene.get("character_feelings", {})
        if feelings and isinstance(feelings, dict):
            lines.append(f"### 心情")
            for char, feeling in feelings.items():
                lines.append(f"- **{char}**: {feeling}")
            lines.append(f"")

        # セリフ
        bubbles = scene.get("bubbles", []) or scene.get("dialogue", []) or []
        if bubbles:
            lines.append(f"### セリフ")
            ono = scene.get("onomatopoeia", [])
            if ono:
                ono_str = ", ".join(ono) if isinstance(ono, list) else str(ono)
                lines.append(f"*SE: {ono_str}*")
                lines.append(f"")
            for b in bubbles:
                speaker = b.get("speaker", "???")
                btype = b.get("type", b.get("emotion", ""))
                text = b.get("text", b.get("line", ""))
                type_tag = f" ({btype})" if btype else ""
                lines.append(f"- **{speaker}**{type_tag}: {text}")
            lines.append(f"")

        # 演出
        if direction:
            lines.append(f"### 演出")
            lines.append(f"{direction}")
            lines.append(f"")

        # 次への繋がり
        if story_flow:
            lines.append(f"*次へ: {story_flow}*")
            lines.append(f"")

        # SDプロンプト
        if sd:
            lines.append(f"### SD Prompt")
            lines.append(f"```")
            lines.append(sd)
            lines.append(f"```")
            lines.append(f"")

        lines.append(f"---")
        lines.append(f"")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    log_message(f"マークダウン出力完了: {output_path}")


# === キャラクター自動生成システム ===

CHARACTER_BIBLE_TEMPLATE = {
    "work_title": "",
    "character_name": "",
    "core_traits": [],
    "values": [],
    "fears": [],
    "relationship_style": {
        "toward_love_interest": "",
        "toward_rival": "",
        "toward_friends": ""
    },
    "speech_profile": {
        "first_person": "",
        "second_person_style": "",
        "formality_level": 0,
        "sentence_length": "medium",
        "rhythm": "",
        "typical_tone": "",
        "forbidden_elements": []
    },
    "emotion_model": {
        "baseline_state": "",
        "triggers": [],
        "escalation_pattern": [],
        "deescalation_pattern": []
    },
    "conflict_response_style": "",
    "romantic_response_style": "",
    "originality_guard": {
        "avoid_canonical_lines": True,
        "avoid_known_catchphrases": True
    }
}


def generate_char_id(work_title: str, char_name: str) -> str:
    """キャラIDを生成（英数字のみ）"""
    import re
    import hashlib
    combined = f"{work_title}_{char_name}"
    # 日本語などを含む場合はハッシュ化
    if re.search(r'[^\x00-\x7F]', combined):
        short_hash = hashlib.md5(combined.encode()).hexdigest()[:8]
        return f"char_{short_hash}"
    return re.sub(r'[^a-zA-Z0-9_]', '_', combined.lower())[:32]


def analyze_character(
    client: anthropic.Anthropic,
    work_title: str,
    char_name: str,
    cost_tracker: CostTracker,
    callback: Optional[Callable] = None
) -> dict:
    """キャラクター情報をClaudeの知識から抽出（Sonnetで高品質分析）"""

    if callback:
        callback(f"[CHECK]{char_name}の詳細分析中（Sonnet使用）...")

    system_prompt = """あなたは日本のアニメ・漫画・ゲームキャラクターの口調分析専門家です。
二次創作でキャラクターの「らしさ」を完璧に再現するため、話し方を徹底的に分析します。

【重要ルール】
- 原作セリフの直接引用は禁止
- 「こういうパターンで話す」という抽象的な特徴を記述
- エロシーンでも使える「感情が高ぶった時の話し方」を詳細に
- 日本語として自然な表現を意識"""

    prompt = f"""作品名: {work_title}
キャラクター名: {char_name}

このキャラクターの「話し方」を、二次創作（成人向け含む）で使えるレベルで徹底分析してください。

{{
    "work_title": "{work_title}",
    "character_name": "{char_name}",
    
    "personality_core": {{
        "brief_description": "このキャラを一言で表すと",
        "main_traits": ["性格特性を5個"],
        "hidden_traits": ["表に出さない特性を3個"],
        "weakness": "弱点・苦手なこと",
        "values": ["大切にしていること3個"],
        "fears": ["恐れていること2個"]
    }},
    
    "speech_pattern": {{
        "first_person": "一人称（私/あたし/僕/俺/自分の名前等）",
        "sentence_endings": ["語尾パターンを8個以上。例: 〜だよ, 〜かな, 〜ですわ, 〜じゃん, 〜わよ"],
        "favorite_expressions": ["口癖ではないがよく使う言い回し5個"],
        "fillers": ["間投詞を5個。例: えっと, あのさ, ねえ, うーん"],
        "particles": ["特徴的な助詞の使い方3個"],
        "casual_level": "1-5の数字（1=タメ口, 5=超丁寧）",
        "speech_speed": "速い/普通/ゆっくり",
        "sentence_length": "短文多め/普通/長文多め",
        "voice_quality": "声の特徴（高い/低い/ハスキー等）"
    }},
    
    "emotional_speech": {{
        "when_happy": "嬉しい時の話し方（具体的に）",
        "when_embarrassed": "照れた時・恥ずかしい時の話し方",
        "when_angry": "怒った時の話し方",
        "when_sad": "悲しい時の話し方",
        "when_confused": "困惑・動揺した時の話し方",
        "when_flirty": "甘える・誘惑する時の話し方（エロシーン用に詳細に！）",
        "when_aroused": "感じている時の話し方（喘ぎ声のパターン、言葉の途切れ方）",
        "when_climax": "絶頂時の話し方・反応"
    }},
    
    "dialogue_examples": {{
        "greeting": "挨拶の仕方の例",
        "agreement": "同意する時の例",
        "refusal": "断る時の例",
        "surprise": "驚いた時の例",
        "affection": "好意を示す時の例",
        "teasing": "からかう・甘える時の例",
        "moaning_light": "軽い喘ぎ声の例（あっ、んっ等の組み合わせ）",
        "moaning_intense": "激しい喘ぎ声の例"
    }},
    
    "relationship_speech": {{
        "to_lover": "恋人・好きな人への話し方（詳細に）",
        "to_friends": "友人への話し方",
        "to_strangers": "初対面の人への話し方",
        "to_rivals": "ライバル・敵対者への話し方"
    }},
    
    "erotic_speech_guide": {{
        "shyness_level": "1-5（1=大胆, 5=超恥ずかしがり）",
        "verbal_during_sex": "行為中によく言いそうなフレーズパターン3個",
        "orgasm_expression": "絶頂時の表現パターン",
        "pillow_talk": "事後の甘い会話パターン"
    }},
    
    "avoid_patterns": ["このキャラが絶対に言わない表現パターン5個"],
    
    "physical_description": {{
        "hair": "髪型・髪色（詳細に）",
        "eyes": "目の色・特徴",
        "body": "体型（スレンダー/グラマー/ロリ体型等）",
        "chest": "胸のサイズ感",
        "clothing": "よく着る服装",
        "notable": ["その他の外見特徴2個"]
    }},
    
    "danbooru_tags": ["SDプロンプト用のdanbooruタグ20個（キャラ名タグ、髪、目、体型、服装等）"],
    
    "originality_guard": {{
        "avoid_canonical_lines": true,
        "avoid_known_catchphrases": true,
        "known_catchphrases": ["避けるべき有名な口癖があれば記載"]
    }}
}}

【重要】
- speech_patternとemotional_speechは特に詳細に
- erotic_speech_guideは成人向け創作で使うため必須
- danbooru_tagsは必ず20個
- JSONのみ出力"""

    # キャラ分析はSonnetで高品質に
    response = _call_api(
        client, MODELS["sonnet"],
        system_prompt,
        prompt, cost_tracker, 4096, callback
    )

    return parse_json_response(response)


def generate_character_skill(char_id: str, bible: dict) -> str:
    """キャラクター専用のSkillファイルを生成（要件定義準拠）"""
    char_name = bible.get("character_name", char_id)
    work_title = bible.get("work_title", "Unknown")
    
    personality = bible.get("personality_core", {})
    speech = bible.get("speech_pattern", {})
    emotional = bible.get("emotional_speech", {})
    examples = bible.get("dialogue_examples", {})
    relationship = bible.get("relationship_speech", {})
    erotic = bible.get("erotic_speech_guide", {})
    avoid = bible.get("avoid_patterns", [])
    physical = bible.get("physical_description", {})
    tags = bible.get("danbooru_tags", [])
    
    # 文末表現リスト
    endings = speech.get("sentence_endings", [])
    endings_str = ", ".join(endings) if endings else "〜よ, 〜ね, 〜かな"
    
    # フィラー
    fillers = speech.get("fillers", [])
    fillers_str = ", ".join(fillers) if fillers else "えっと, あのね"
    
    # 避けるべきパターン
    avoid_str = "\n".join([f"- {a}" for a in avoid]) if avoid else "- 特になし"

    skill_content = f"""---
name: character_voice_{char_id}
description: Apply abstract character model for {char_name} from {work_title}
commands:
  - /voice-{char_id}
---

# {char_name} 完全口調ガイド

## Role
{char_name}（{work_title}）のセリフを、キャラクターらしい自然な日本語会話として生成する。

## Hard Rules
- Never reproduce canonical lines（原作セリフの再現禁止）
- Never copy known catchphrases（決め台詞のコピー禁止）
- Use structural traits only（構造的特徴のみ使用）
- Maintain character voice consistency（キャラの声を一貫させる）

## Character Profile

### 基本情報
- **作品**: {work_title}
- **名前**: {char_name}
- **性格**: {personality.get('brief_description', '')}
- **特性**: {', '.join(personality.get('main_traits', []))}
- **隠れた面**: {', '.join(personality.get('hidden_traits', []))}

### 話し方の基本

| 項目 | 設定 |
|------|------|
| 一人称 | {speech.get('first_person', '私')} |
| 語尾 | {endings_str} |
| 間投詞 | {fillers_str} |
| カジュアル度 | {speech.get('casual_level', 3)}/5 |
| 話すテンポ | {speech.get('speech_speed', '普通')} |
| 文の長さ | {speech.get('sentence_length', '普通')} |

### 感情別の話し方

#### 日常シーン
- **嬉しい時**: {emotional.get('when_happy', '')}
- **照れた時**: {emotional.get('when_embarrassed', '')}
- **怒った時**: {emotional.get('when_angry', '')}
- **困惑時**: {emotional.get('when_confused', '')}

#### エロシーン（成人向け）
- **甘える時**: {emotional.get('when_flirty', '')}
- **感じてる時**: {emotional.get('when_aroused', '')}
- **絶頂時**: {emotional.get('when_climax', '')}
- **恥ずかしさ**: {erotic.get('shyness_level', 3)}/5

### セリフ例（参考パターン）
- 挨拶: {examples.get('greeting', '')}
- 同意: {examples.get('agreement', '')}
- 驚き: {examples.get('surprise', '')}
- 好意: {examples.get('affection', '')}
- 軽い喘ぎ: {examples.get('moaning_light', 'あっ...んっ...')}
- 激しい喘ぎ: {examples.get('moaning_intense', 'あっあっ...♡')}

### 関係性別の話し方
- **恋人へ**: {relationship.get('to_lover', '')}
- **友人へ**: {relationship.get('to_friends', '')}

## Forbidden Patterns（禁止表現）
{avoid_str}

## Procedure
1. Load ./characters/{char_id}.json
2. Check speaker's emotional state
3. Apply speech_pattern (first_person, endings)
4. Apply emotional_speech based on scene intensity
5. Ensure originality (no canonical lines)
6. Output natural Japanese dialogue

## SD Prompt Tags
```
{', '.join(tags)}
```

## Physical Description
- 髪: {physical.get('hair', '')}
- 目: {physical.get('eyes', '')}
- 体型: {physical.get('body', '')}
- 服装: {physical.get('clothing', '')}
"""
    return skill_content



def load_character_pool(char_id: str) -> dict:
    """キャラ固有プールを読み込み。なければ空dictを返す"""
    pool_path = CHARACTERS_DIR / f"{char_id}_pool.json"
    if pool_path.exists():
        with open(pool_path, "r", encoding="utf-8") as f:
            return json.load(f)
    # プリセットチェック
    preset_path = PRESET_CHARS_DIR / f"{char_id}_pool.json"
    if preset_path.exists():
        with open(preset_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def generate_character_pool(
    client, char_id: str, bible: dict,
    cost_tracker: CostTracker, callback=None
) -> dict:
    """キャラプロファイルから専用セリフプールをAPI 1回で生成。
    moan(intensity 1-5)×各8個 + speech(6フェーズ)×各5個 + thought(6フェーズ)×各5個 = 約100個/キャラ"""

    char_name = bible.get("character_name", "ヒロイン")
    speech = bible.get("speech_pattern", {})
    emotional = bible.get("emotional_speech", {})
    erotic = bible.get("erotic_speech_guide", {})

    if callback:
        callback(f"[INFO]Step 4/4: {char_name}専用セリフプール生成中...")

    # SCENE_PHASE_SPEECH_MAPからfew-shot例を取得
    try:
        from ero_dialogue_pool import SCENE_PHASE_SPEECH_MAP
        few_shot_examples = {}
        for phase in ["foreplay", "penetration", "climax"]:
            phase_data = SCENE_PHASE_SPEECH_MAP.get(phase, {})
            few_shot_examples[phase] = {
                "speech": phase_data.get("speech", [])[:3],
                "thought": phase_data.get("thought", [])[:3],
                "moan": phase_data.get("moan", [])[:3],
            }
    except ImportError:
        few_shot_examples = {}

    few_shot_str = ""
    if few_shot_examples:
        few_shot_str = "\n## 汎用セリフ例（参考。このトーンをキャラの口調に変換せよ）\n"
        for phase, examples in few_shot_examples.items():
            few_shot_str += f"### {phase}\n"
            for btype, lines in examples.items():
                if lines:
                    few_shot_str += f"  {btype}: {', '.join(lines)}\n"

    system_prompt = f"""あなたはエロ漫画・CG集のセリフ専門ライターです。
キャラクターの口調設定に基づき、そのキャラ固有のセリフプールをJSON形式で生成します。

【キャラクター口調設定】
・名前: {char_name}
・一人称: {speech.get('first_person', '私')}
・語尾: {', '.join(speech.get('sentence_endings', ['〜よ', '〜ね']))}
・間投詞: {', '.join(speech.get('fillers', ['あっ', 'んっ']))}
・照れた時: {emotional.get('when_embarrassed', '言葉に詰まる')}
・感じてる時: {emotional.get('when_aroused', '声が震える')}
・絶頂時: {emotional.get('when_climax', '理性が飛ぶ')}
・エロ中の口癖: {erotic.get('verbal_during_sex', '特になし')}
・恥ずかしがり度: {erotic.get('shyness_level', '3')}/5

{few_shot_str}

【出力ルール】
- moanは喘ぎ声のみ（「あっ」「んっ」系）。漢字・助詞禁止。♡はintensity3以上で使用可
- speechは感情的反応の短文（1-10文字）。キャラの語尾・一人称を反映
- thoughtは心の声（1-15文字）。「…」で区切る断片的な表現
- 各フェーズの感情段階を正確に反映すること
- JSONのみ出力"""

    prompt = f"""以下の構造でJSON出力してください。

{{
    "character_name": "{char_name}",
    "char_id": "{char_id}",
    "moan": {{
        "1": ["（intensity1: 微かな息遣い8個）"],
        "2": ["（intensity2: 小さな喘ぎ8個）"],
        "3": ["（intensity3: 本格的な喘ぎ8個。♡使用可）"],
        "4": ["（intensity4: 激しい喘ぎ8個。♡♡使用可）"],
        "5": ["（intensity5: 絶頂の喘ぎ8個。♡♡♡使用可）"]
    }},
    "speech": {{
        "intro": ["（導入: 日常会話5個。キャラ語尾反映）"],
        "approach": ["（接近: 戸惑い5個）"],
        "foreplay": ["（前戯: 恥ずかしさ・感じ始め5個）"],
        "penetration": ["（挿入: 挿入リアクション5個）"],
        "climax": ["（絶頂: 絶頂セリフ5個。♡♡使用）"],
        "afterglow": ["（余韻: 事後セリフ5個）"]
    }},
    "thought": {{
        "intro": ["（導入: 緊張5個）"],
        "approach": ["（接近: 期待と不安5個）"],
        "foreplay": ["（前戯: 体の反応への驚き5個）"],
        "penetration": ["（挿入: 充足感5個）"],
        "climax": ["（絶頂: 理性崩壊5個。♡使用可）"],
        "afterglow": ["（余韻: 振り返り5個）"]
    }}
}}

【重要】
- {char_name}の口調（一人称={speech.get('first_person', '私')}、語尾={', '.join(speech.get('sentence_endings', [])[:3])}）を全セリフに反映
- 汎用的な喘ぎではなく、このキャラの性格・話し方が伝わる喘ぎ・セリフにすること
- JSONのみ出力。説明文不要"""

    response = _call_api(
        client, MODELS["sonnet"],
        system_prompt, prompt, cost_tracker, 4096, callback
    )

    pool = parse_json_response(response)

    if callback:
        # 生成数カウント
        moan_count = sum(len(v) for v in pool.get("moan", {}).values() if isinstance(v, list))
        speech_count = sum(len(v) for v in pool.get("speech", {}).values() if isinstance(v, list))
        thought_count = sum(len(v) for v in pool.get("thought", {}).values() if isinstance(v, list))
        callback(f"[OK]キャラプール生成完了: moan={moan_count}, speech={speech_count}, thought={thought_count}")

    return pool


def upgrade_character_pool_api(
    client, char_id: str, bible: dict,
    cost_tracker: CostTracker, callback=None,
    upgrade_types: list = None,
) -> dict:
    """既存ローカルプールのspeech+thoughtのみAPIで補正。moanはローカル維持。
    upgrade_types: デフォルト ["speech", "thought"]
    """
    if upgrade_types is None:
        upgrade_types = ["speech", "thought"]

    # 既存プール読み込み
    existing_pool = load_character_pool(char_id)

    # API版フル生成
    api_pool = generate_character_pool(client, char_id, bible, cost_tracker, callback)

    # 指定タイプのみAPIで上書き、それ以外はローカル維持
    merged = dict(existing_pool) if existing_pool else dict(api_pool)
    for utype in upgrade_types:
        if utype in api_pool:
            merged[utype] = api_pool[utype]

    # ソース情報追記
    source_detail = {}
    for key in ["moan", "speech", "thought"]:
        source_detail[key] = "api" if key in upgrade_types else "local"
    merged["source"] = "hybrid"
    merged["source_detail"] = source_detail

    return merged


def build_character(
    api_key: str,
    work_title: str,
    char_name: str,
    force_refresh: bool = False,
    callback: Optional[Callable] = None,
) -> tuple[dict, str, CostTracker]:
    """キャラクター生成パイプライン"""
    client = anthropic.Anthropic(api_key=api_key)
    cost_tracker = CostTracker()

    char_id = generate_char_id(work_title, char_name)
    bible_path = CHARACTERS_DIR / f"{char_id}.json"
    skill_path = CHAR_SKILLS_DIR / f"{char_id}.skill.md"
    pool_path = CHARACTERS_DIR / f"{char_id}_pool.json"

    # プリセットチェック（API不要）
    preset_path = PRESET_CHARS_DIR / f"{char_id}.json"
    if preset_path.exists() and not force_refresh:
        if callback:
            callback(f"[PACK]プリセットキャラを使用: {char_name}")
        bible, _ = load_preset_character(char_id, callback)
        return bible, char_id, cost_tracker

    # キャッシュチェック
    if bible_path.exists() and not force_refresh:
        if callback:
            callback(f"[FILE]既存のキャラデータを使用: {char_id}")
        with open(bible_path, "r", encoding="utf-8") as f:
            bible = json.load(f)
        return bible, char_id, cost_tracker

    if callback:
        callback(f"[START]キャラクター生成開始: {char_name}")

    # Step 1: キャラクター分析
    if callback:
        callback("[STAT]Step 1/4: キャラクター分析")

    bible = analyze_character(client, work_title, char_name, cost_tracker, callback)

    # originality_guardを追加
    bible["originality_guard"] = {
        "avoid_canonical_lines": True,
        "avoid_known_catchphrases": True
    }

    # Step 2: キャラバイブル保存
    if callback:
        callback("[SAVE]Step 2/4: キャラバイブル保存")

    with open(bible_path, "w", encoding="utf-8") as f:
        json.dump(bible, f, ensure_ascii=False, indent=2)

    log_message(f"キャラバイブル保存: {bible_path}")

    # Step 3: Skill生成
    if callback:
        callback("[INFO]Step 3/4: Skill生成")

    skill_content = generate_character_skill(char_id, bible)

    with open(skill_path, "w", encoding="utf-8") as f:
        f.write(skill_content)

    log_message(f"Skill保存: {skill_path}")

    # Step 4: キャラ専用セリフプール生成（ローカル優先→API補正）
    if not pool_path.exists() or force_refresh:
        # まずローカル生成（即座、$0）
        from character_pool_generator import generate_character_pool_local
        if callback:
            callback(f"[INFO]Step 4/4: {char_name}専用セリフプール生成中（ローカル）...")
        char_pool = generate_character_pool_local(bible)
        char_pool["char_id"] = char_id
        with open(pool_path, "w", encoding="utf-8") as f:
            json.dump(char_pool, f, ensure_ascii=False, indent=2)
        log_message(f"キャラプール保存（ローカル）: {pool_path}")

        # API可能ならspeech+thoughtだけAPI補正
        try:
            char_pool = upgrade_character_pool_api(
                client, char_id, bible, cost_tracker, callback,
                upgrade_types=["speech", "thought"]
            )
            with open(pool_path, "w", encoding="utf-8") as f:
                json.dump(char_pool, f, ensure_ascii=False, indent=2)
            log_message(f"キャラプール更新（API補正）: {pool_path}")
        except Exception as e:
            log_message(f"API補正スキップ（ローカル版を使用）: {e}")
            if callback:
                callback(f"[WARN]API補正スキップ、ローカル版を使用: {e}")

    if callback:
        callback(f"[OK]キャラクター生成完了: {char_id}")

    return bible, char_id, cost_tracker


def get_existing_characters() -> list[dict]:
    """既存のキャラクター一覧を取得"""
    characters = []
    for json_file in CHARACTERS_DIR.glob("*.json"):
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                characters.append({
                    "char_id": json_file.stem,
                    "name": data.get("character_name", json_file.stem),
                    "work": data.get("work_title", "Unknown")
                })
        except Exception:
            pass
    return characters


def get_preset_characters() -> list[dict]:
    """プリセットキャラクター一覧を取得"""
    if not PRESET_INDEX_FILE.exists():
        return []
    try:
        with open(PRESET_INDEX_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("characters", [])
    except Exception:
        return []


def load_preset_character(char_id: str, callback: Optional[Callable] = None) -> tuple[dict, str]:
    """プリセットキャラをcharactersにコピーしてskillも生成（API不要）"""
    preset_path = PRESET_CHARS_DIR / f"{char_id}.json"
    bible_path = CHARACTERS_DIR / f"{char_id}.json"
    skill_path = CHAR_SKILLS_DIR / f"{char_id}.skill.md"
    pool_path = CHARACTERS_DIR / f"{char_id}_pool.json"

    if callback:
        callback(f"[FILE]プリセット読み込み中: {char_id}")

    with open(preset_path, "r", encoding="utf-8") as f:
        bible = json.load(f)

    # charactersディレクトリにコピー
    with open(bible_path, "w", encoding="utf-8") as f:
        json.dump(bible, f, ensure_ascii=False, indent=2)

    # Skill生成
    skill_content = generate_character_skill(char_id, bible)
    with open(skill_path, "w", encoding="utf-8") as f:
        f.write(skill_content)

    # プールコピー/生成
    if not pool_path.exists():
        preset_pool_path = PRESET_CHARS_DIR / f"{char_id}_pool.json"
        if preset_pool_path.exists():
            # プリセットプールをコピー
            import shutil
            shutil.copy2(preset_pool_path, pool_path)
            if callback:
                callback(f"[FILE]プリセットプールコピー: {char_id}")
        else:
            # ローカル生成
            from character_pool_generator import generate_character_pool_local
            char_pool = generate_character_pool_local(bible)
            char_pool["char_id"] = char_id
            with open(pool_path, "w", encoding="utf-8") as f:
                json.dump(char_pool, f, ensure_ascii=False, indent=2)
            if callback:
                callback(f"[INFO]プールをローカル生成: {char_id}")

    if callback:
        callback(f"[OK]プリセット読み込み完了: {bible.get('character_name', char_id)}")

    return bible, char_id


# === Material Design GUI ===
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


class MaterialCard(ctk.CTkFrame):
    """
    Material Design 3 Card Component
    
    Variants:
    - elevated: Default, subtle shadow effect via background
    - filled: Higher surface tone, no border
    - outlined: Transparent with outline border
    """
    def __init__(
        self, 
        master, 
        title: str = "", 
        collapsible: bool = False, 
        start_collapsed: bool = False,
        variant: str = "elevated",  # elevated, filled, outlined
        **kwargs
    ):
        # M3 Card styling based on variant
        # All variants use border to prevent corner anti-alias fading
        if variant == "filled":
            bg_color = MaterialColors.SURFACE_CONTAINER_HIGHEST
            border_width = 1
            border_color = MaterialColors.OUTLINE_VARIANT
        elif variant == "outlined":
            bg_color = MaterialColors.SURFACE
            border_width = 1
            border_color = MaterialColors.OUTLINE_VARIANT
        else:  # elevated (default)
            bg_color = MaterialColors.SURFACE_CONTAINER_LOW
            border_width = 1
            border_color = MaterialColors.OUTLINE_VARIANT

        super().__init__(
            master,
            fg_color=bg_color,
            corner_radius=0,   # Sharp edges — no anti-alias artifacts
            border_width=border_width,
            border_color=border_color,
            **kwargs
        )
        
        self.collapsible = collapsible
        self.is_collapsed = False
        self.variant = variant
        
        if title:
            # Header with proper M3 typography
            header_frame = ctk.CTkFrame(self, fg_color="transparent")
            header_frame.pack(fill="x", padx=20, pady=(16, 12))
            
            self.title_label = ctk.CTkLabel(
                header_frame,
                text=title,
                font=ctk.CTkFont(family=FONT_JP, size=16, weight="bold"),  # Title Medium
                text_color=MaterialColors.ON_SURFACE
            )
            self.title_label.pack(side="left")
            
            if collapsible:
                self.collapse_btn = ctk.CTkButton(
                    header_frame,
                    text="",
                    width=40,
                    height=40,
                    fg_color="transparent",
                    hover_color=MaterialColors.SURFACE_CONTAINER_HIGH,
                    text_color=MaterialColors.ON_SURFACE_VARIANT,
                    font=ctk.CTkFont(size=14),
                    corner_radius=20,  # Fully rounded for icon button
                    command=self.toggle_collapse
                )
                self.collapse_btn.pack(side="right")
                self._update_collapse_icon()
                # ヘッダー全体をクリック可能に
                header_frame.bind("<Button-1>", lambda e: self.toggle_collapse())
                self.title_label.bind("<Button-1>", lambda e: self.toggle_collapse())
                header_frame.configure(cursor="hand2")
                self.title_label.configure(cursor="hand2")

        self.content_frame = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0)
        if collapsible and start_collapsed:
            self.is_collapsed = True
            self._update_collapse_icon()
        else:
            self.content_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
    
    def _update_collapse_icon(self):
        self.collapse_btn.configure(
            text=Icons.CHEVRON_UP if not self.is_collapsed else Icons.CHEVRON_DOWN,
            font=ctk.CTkFont(family=FONT_ICON, size=12)
        )
    
    def toggle_collapse(self):
        if self.is_collapsed:
            self.content_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        else:
            self.content_frame.pack_forget()
        self.is_collapsed = not self.is_collapsed
        self._update_collapse_icon()


class MaterialButton(ctk.CTkButton):
    """
    Material Design 3 Button Component
    
    Variants:
    - filled: Primary container color (default)
    - filled_tonal: Secondary container color
    - outlined: Transparent with outline
    - text: Text only, no background
    - elevated: Surface with shadow effect
    
    Sizes:
    - small: 32dp height
    - medium: 40dp height (default)
    - large: 56dp height
    """
    def __init__(
        self, 
        master, 
        variant: str = "filled", 
        size: str = "medium", 
        **kwargs
    ):
        # M3 Button sizes (height, font_size, corner_radius, horizontal_padding)
        sizes = {
            "small": {"height": 32, "font_size": 12, "corner": 16, "padx": 12},
            "medium": {"height": 40, "font_size": 14, "corner": 20, "padx": 24},
            "large": {"height": 56, "font_size": 14, "corner": 28, "padx": 24},
            "xlarge": {"height": 64, "font_size": 16, "corner": 28, "padx": 32}
        }
        s = sizes.get(size, sizes["medium"])
        
        # M3 Button variants with proper color tokens
        variants = {
            "filled": {
                "fg_color": MaterialColors.PRIMARY,
                "hover_color": "#2563EB",  # Blue-600 on hover
                "text_color": MaterialColors.ON_PRIMARY,
                "border_width": 0,
            },
            "filled_tonal": {
                "fg_color": MaterialColors.SECONDARY_CONTAINER,
                "hover_color": MaterialColors.SURFACE_CONTAINER_HIGHEST,
                "text_color": MaterialColors.ON_SECONDARY_CONTAINER,
                "border_width": 0,
            },
            "outlined": {
                "fg_color": "transparent",
                "hover_color": MaterialColors.SURFACE_CONTAINER,
                "text_color": MaterialColors.PRIMARY,
                "border_width": 1,
                "border_color": MaterialColors.OUTLINE,
            },
            "text": {
                "fg_color": "transparent",
                "hover_color": MaterialColors.SURFACE_CONTAINER,
                "text_color": MaterialColors.PRIMARY,
                "border_width": 0,
            },
            "elevated": {
                "fg_color": MaterialColors.SURFACE_CONTAINER_LOW,
                "hover_color": MaterialColors.SURFACE_CONTAINER,
                "text_color": MaterialColors.PRIMARY,
                "border_width": 0,
            },
            # Extended variants for app-specific use
            "accent": {
                "fg_color": MaterialColors.TERTIARY,
                "hover_color": MaterialColors.ACCENT_DARK,
                "text_color": MaterialColors.ON_PRIMARY,
                "border_width": 0,
            },
            "danger": {
                "fg_color": MaterialColors.ERROR,
                "hover_color": "#B91C1C",  # Red-700
                "text_color": MaterialColors.ON_ERROR,
                "border_width": 0,
            },
            "success": {
                "fg_color": MaterialColors.SUCCESS,
                "hover_color": "#15803D",  # Green-700
                "text_color": "#FFFFFF",
                "border_width": 0,
            },
        }
        
        v = variants.get(variant, variants["filled"])
        
        super().__init__(
            master,
            fg_color=v["fg_color"],
            hover_color=v["hover_color"],
            text_color=v["text_color"],
            border_width=v.get("border_width", 0),
            border_color=v.get("border_color"),
            corner_radius=s["corner"],
            height=s["height"],
            font=ctk.CTkFont(family=FONT_JP, size=s["font_size"], weight="bold"),
            **kwargs
        )


class MaterialTextField(ctk.CTkFrame):
    """
    Material Design 3 Text Field
    
    Variants:
    - filled: Default M3 text field with container
    - outlined: Border-style text field
    """
    def __init__(
        self, 
        master, 
        label: str, 
        placeholder: str = "", 
        show: str = "", 
        height: int = 56,  # M3 default height
        multiline: bool = False,
        variant: str = "filled",  # filled, outlined
        supporting_text: str = "",
        **kwargs
    ):
        super().__init__(master, fg_color="transparent", **kwargs)
        
        self.variant = variant
        
        # Label (Body Small)
        self.label = ctk.CTkLabel(
            self,
            text=label,
            font=ctk.CTkFont(family=FONT_JP, size=14),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        )
        self.label.pack(anchor="w", pady=(0, 4))

        # Input field styling based on variant
        if variant == "outlined":
            fg_color = "transparent"
            border_width = 1
            border_color = MaterialColors.OUTLINE
            corner_radius = 4
        else:  # filled
            fg_color = MaterialColors.SURFACE_CONTAINER_HIGHEST
            border_width = 0
            border_color = None
            corner_radius = 4  # M3: 4dp top corners only, but CTk doesn't support asymmetric

        if multiline:
            self.entry = ctk.CTkTextbox(
                self,
                height=height,
                fg_color=fg_color,
                text_color=MaterialColors.ON_SURFACE,
                font=ctk.CTkFont(family=FONT_JP, size=16),
                corner_radius=corner_radius,
                border_width=border_width,
                border_color=border_color
            )
        else:
            self.entry = ctk.CTkEntry(
                self,
                height=height,
                placeholder_text=placeholder,
                placeholder_text_color="#3D3D3D",
                show=show,
                fg_color=fg_color,
                text_color=MaterialColors.ON_SURFACE,
                font=ctk.CTkFont(family=FONT_JP, size=16),
                corner_radius=corner_radius,
                border_width=border_width,
                border_color=border_color
            )
        self.entry.pack(fill="x")
        
        # Supporting text (Body Small)
        if supporting_text:
            self.supporting = ctk.CTkLabel(
                self,
                text=supporting_text,
                font=ctk.CTkFont(family=FONT_JP, size=14),
                text_color=MaterialColors.ON_SURFACE_VARIANT
            )
            self.supporting.pack(anchor="w", pady=(4, 0))

    def get(self):
        if isinstance(self.entry, ctk.CTkTextbox):
            return self.entry.get("1.0", "end-1c")
        return self.entry.get()

    def set(self, value: str):
        if isinstance(self.entry, ctk.CTkTextbox):
            self.entry.delete("1.0", "end")
            self.entry.insert("1.0", value)
        else:
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
    
    def set_error(self, message: str = ""):
        """Set error state with optional message"""
        if message:
            self.entry.configure(border_color=MaterialColors.ERROR)
            self.label.configure(text_color=MaterialColors.ERROR)
        else:
            border = MaterialColors.OUTLINE if self.variant == "outlined" else None
            self.entry.configure(border_color=border)
            self.label.configure(text_color=MaterialColors.ON_SURFACE_VARIANT)


class MaterialFAB(ctk.CTkButton):
    """
    Material Design 3 Floating Action Button

    Sizes:
    - small: 40dp (for compact layouts)
    - regular: 56dp (default)
    - large: 96dp (for prominent actions)

    Variants:
    - primary: Primary container (default)
    - secondary: Secondary container
    - tertiary: Tertiary container
    - surface: Surface container
    """
    def __init__(
        self,
        master,
        icon: str = "+",
        size: str = "regular",
        variant: str = "primary",
        **kwargs
    ):
        # M3 FAB sizes
        sizes = {
            "small": {"size": 40, "icon_size": 24, "corner": 12},
            "regular": {"size": 56, "icon_size": 24, "corner": 16},
            "large": {"size": 96, "icon_size": 36, "corner": 28}
        }
        s = sizes.get(size, sizes["regular"])

        # M3 FAB color variants
        variants = {
            "primary": {
                "fg": MaterialColors.PRIMARY_CONTAINER,
                "text": MaterialColors.ON_PRIMARY_CONTAINER,
                "hover": MaterialColors.SURFACE_CONTAINER_HIGHEST
            },
            "secondary": {
                "fg": MaterialColors.SECONDARY_CONTAINER,
                "text": MaterialColors.ON_SECONDARY_CONTAINER,
                "hover": MaterialColors.SURFACE_CONTAINER_HIGHEST
            },
            "tertiary": {
                "fg": MaterialColors.TERTIARY_CONTAINER,
                "text": MaterialColors.ON_SURFACE,
                "hover": MaterialColors.SURFACE_CONTAINER_HIGHEST
            },
            "surface": {
                "fg": MaterialColors.SURFACE_CONTAINER_HIGH,
                "text": MaterialColors.PRIMARY,
                "hover": MaterialColors.SURFACE_CONTAINER_HIGHEST
            }
        }
        v = variants.get(variant, variants["primary"])

        super().__init__(
            master,
            text=icon,
            width=s["size"],
            height=s["size"],
            corner_radius=s["corner"],
            fg_color=v["fg"],
            hover_color=v["hover"],
            text_color=v["text"],
            font=ctk.CTkFont(size=s["icon_size"], weight="bold"),
            **kwargs
        )


class MaterialChip(ctk.CTkButton):
    """
    Material Design 3 Chip

    Types:
    - assist: For smart suggestions
    - filter: For filtering content (toggleable)
    - input: For user input (with close button)
    - suggestion: For dynamic suggestions
    """
    def __init__(
        self,
        master,
        text: str,
        selected: bool = False,
        chip_type: str = "filter",
        **kwargs
    ):
        self.selected = selected
        self.chip_type = chip_type

        if selected:
            fg_color = MaterialColors.SECONDARY_CONTAINER
            text_color = MaterialColors.ON_SECONDARY_CONTAINER
            border_width = 0
        else:
            fg_color = "transparent"
            text_color = MaterialColors.ON_SURFACE_VARIANT
            border_width = 1

        super().__init__(
            master,
            text=text,
            height=32,  # M3: 32dp height
            corner_radius=0,
            fg_color=fg_color,
            hover_color=MaterialColors.SURFACE_CONTAINER,
            text_color=text_color,
            border_width=border_width,
            border_color=MaterialColors.OUTLINE,
            font=ctk.CTkFont(family=FONT_JP, size=13),
            **kwargs
        )

    def toggle(self):
        self.selected = not self.selected
        if self.selected:
            self.configure(
                fg_color=MaterialColors.SECONDARY_CONTAINER,
                text_color=MaterialColors.ON_SECONDARY_CONTAINER,
                border_width=0
            )
        else:
            self.configure(
                fg_color="transparent",
                text_color=MaterialColors.ON_SURFACE_VARIANT,
                border_width=1
            )


class Snackbar(ctk.CTkFrame):
    """
    Material Design 3 Snackbar
    
    Single-line notifications with optional action button.
    Appears at bottom of screen with proper M3 styling.
    """
    def __init__(self, master, **kwargs):
        super().__init__(
            master,
            fg_color=MaterialColors.INVERSE_SURFACE,
            corner_radius=4,  # M3: 4dp corners
            height=48,        # M3: 48dp single-line
            **kwargs
        )

        # Message label (Body Medium)
        self.message_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(family=FONT_JP, size=16),
            text_color=MaterialColors.INVERSE_ON_SURFACE
        )
        self.message_label.pack(side="left", padx=16, pady=12)
        
        # Optional action button
        self.action_btn = ctk.CTkButton(
            self,
            text="",
            font=ctk.CTkFont(family=FONT_JP, size=16, weight="bold"),
            fg_color="transparent",
            hover_color=MaterialColors.INVERSE_SURFACE,
            text_color=MaterialColors.INVERSE_PRIMARY,
            corner_radius=4,
            height=36,
            width=0  # Auto-width
        )
        self.action_btn.pack(side="right", padx=(0, 8))
        self.action_btn.pack_forget()  # Hidden by default

        self.place_forget()

    def show(
        self, 
        message: str, 
        duration: int = 4000,  # M3 recommends 4-10 seconds
        type: str = "info",
        action: str = "",
        action_command = None
    ):
        """
        Show snackbar with message.
        
        Args:
            message: Text to display
            duration: Auto-hide time in ms (0 = no auto-hide)
            type: info, success, error, warning
            action: Optional action button text
            action_command: Optional callback for action button
        """
        # M3 uses inverse surface for snackbar, but we can tint for status
        colors = {
            "info": MaterialColors.INVERSE_SURFACE,
            "success": "#15803D",    # Green-700
            "error": "#B91C1C",      # Red-700
            "warning": "#B45309"     # Amber-700
        }
        
        self.configure(fg_color=colors.get(type, MaterialColors.INVERSE_SURFACE))
        self.message_label.configure(
            text=message,
            text_color=MaterialColors.INVERSE_ON_SURFACE
        )
        
        # Action button
        if action and action_command:
            self.action_btn.configure(text=action, command=action_command)
            self.action_btn.pack(side="right", padx=(0, 8))
        else:
            self.action_btn.pack_forget()
        
        # Position at bottom with proper margin
        self.place(relx=0.5, rely=0.95, anchor="center")
        self.lift()
        
        if duration > 0:
            self.after(duration, self.hide)

    def hide(self):
        self.place_forget()


class MaterialTooltip:
    """M3-style tooltip — hover で表示、離脱で非表示"""

    def __init__(self, widget, text, delay=500):
        self.widget = widget
        self.text = text
        self.delay = delay
        self._tw = None
        self._after_id = None
        widget.bind("<Enter>", self._on_enter, add="+")
        widget.bind("<Leave>", self._on_leave, add="+")

    def _on_enter(self, event=None):
        self._after_id = self.widget.after(self.delay, self._show)

    def _on_leave(self, event=None):
        if self._after_id:
            self.widget.after_cancel(self._after_id)
            self._after_id = None
        self._hide()

    def _show(self):
        if self._tw:
            return
        x = self.widget.winfo_rootx() + self.widget.winfo_width() // 2
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self._tw = tw = ctk.CTkToplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes("-topmost", True)
        frame = ctk.CTkFrame(
            tw, fg_color=MaterialColors.INVERSE_SURFACE, corner_radius=4
        )
        frame.pack()
        ctk.CTkLabel(
            frame, text=self.text,
            font=ctk.CTkFont(family=FONT_JP, size=12),
            text_color=MaterialColors.INVERSE_ON_SURFACE,
        ).pack(padx=8, pady=4)

    def _hide(self):
        if self._tw:
            self._tw.destroy()
            self._tw = None


def add_tooltip(widget, text, delay=500):
    """ウィジェットに M3 Tooltip を追加"""
    return MaterialTooltip(widget, text, delay)


class ExportDialog(ctk.CTkToplevel):
    """マルチフォーマットエクスポートダイアログ"""

    FORMATS = [
        ("csv", "CSV", "Excel対応BOM付きUTF-8"),
        ("json", "JSON", "構造化データ（シーン+メタデータ+SDプロンプト）"),
        ("xlsx", "Excel", "折り返し表示対応（要openpyxl）"),
        ("sd_prompts", "SDプロンプト一括", "1行1プロンプト テキストファイル"),
        ("wildcard", "Wild Card", "SD用1行1プロンプト（__filename__で参照）"),
        ("wildcard_neg", "Wild Card (Negative)", "ネガティブプロンプト用Wild Card"),
        ("dialogue", "セリフ一覧", "話者・種類付きテキストファイル"),
        ("markdown", "マークダウン", "脚本全体の読みやすいビュー"),
    ]

    def __init__(self, master, results: list, metadata: dict = None, **kwargs):
        super().__init__(master, **kwargs)
        self.results = results
        self.metadata = metadata
        self.title("エクスポート")
        self.geometry("460x480")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        # M3 Surface
        self.configure(fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST)

        # Header
        header = ctk.CTkFrame(self, fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST, corner_radius=0)
        header.pack(fill="x")
        icon_text_label(
            header, Icons.FILE_EXPORT, "エクスポート形式を選択",
            icon_size=14, text_size=16, text_color=MaterialColors.ON_SURFACE
        ).pack(anchor="w", padx=20, pady=16)
        ctk.CTkFrame(self, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x")

        # シーン数表示
        info_lbl = ctk.CTkLabel(
            self, text=f"{len(results)}シーン",
            font=ctk.CTkFont(family=FONT_JP, size=13),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        )
        info_lbl.pack(anchor="w", padx=20, pady=(8, 4))

        # チェックボックスリスト
        self.format_vars = {}
        checks_frame = ctk.CTkFrame(self, fg_color="transparent")
        checks_frame.pack(fill="x", padx=20, pady=(4, 12))

        for fmt_id, fmt_name, fmt_desc in self.FORMATS:
            var = ctk.BooleanVar(value=(fmt_id in ("csv", "json")))
            self.format_vars[fmt_id] = var
            row = ctk.CTkFrame(checks_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)
            cb = ctk.CTkCheckBox(
                row, text=fmt_name,
                variable=var,
                font=ctk.CTkFont(family=FONT_JP, size=14),
                text_color=MaterialColors.ON_SURFACE,
                fg_color=MaterialColors.PRIMARY,
                hover_color=MaterialColors.PRIMARY_CONTAINER,
                border_color=MaterialColors.OUTLINE,
                checkmark_color=MaterialColors.ON_PRIMARY,
                corner_radius=4
            )
            cb.pack(side="left")
            if fmt_id == "xlsx" and not OPENPYXL_AVAILABLE:
                cb.configure(state="disabled")
                var.set(False)
            ctk.CTkLabel(
                row, text=fmt_desc,
                font=ctk.CTkFont(family=FONT_JP, size=12),
                text_color=MaterialColors.ON_SURFACE_VARIANT
            ).pack(side="left", padx=(8, 0))

        # JSONインポート
        ctk.CTkFrame(self, fg_color=MaterialColors.OUTLINE_VARIANT, height=1).pack(fill="x", padx=20, pady=(4, 8))
        import_row = ctk.CTkFrame(self, fg_color="transparent")
        import_row.pack(fill="x", padx=20)
        self.import_btn = MaterialButton(
            import_row, text="JSONから読込", variant="outlined", size="small",
            command=self._import_json
        )
        self.import_btn.pack(side="left")
        self.import_label = ctk.CTkLabel(
            import_row, text="",
            font=ctk.CTkFont(family=FONT_JP, size=12),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        )
        self.import_label.pack(side="left", padx=(8, 0))

        # ボタン行
        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(16, 20))

        MaterialButton(
            btn_row, text="エクスポート", variant="filled",
            command=self._do_export
        ).pack(side="right", padx=(8, 0))
        MaterialButton(
            btn_row, text="キャンセル", variant="outlined",
            command=self.destroy
        ).pack(side="right")

    def _import_json(self):
        """既存のJSONファイルから結果を読み込み"""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="エクスポート済みJSONを選択",
            initialdir=str(EXPORTS_DIR),
            filetypes=[("JSON files", "*.json")]
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            scenes = data.get("scenes", data if isinstance(data, list) else [])
            if not scenes:
                self.import_label.configure(text="シーンが見つかりません", text_color=MaterialColors.ERROR)
                return
            self.results = scenes
            # メタデータも復元
            self.metadata = data.get("metadata", None)
            self.import_label.configure(
                text=f"{len(scenes)}シーン読込済",
                text_color=MaterialColors.SUCCESS
            )
        except Exception as e:
            self.import_label.configure(text=f"読込エラー: {str(e)[:30]}", text_color=MaterialColors.ERROR)

    def _do_export(self):
        """選択されたフォーマットでエクスポート実行"""
        selected = [k for k, v in self.format_vars.items() if v.get()]
        if not selected:
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        exported = []

        for fmt in selected:
            try:
                if fmt == "csv":
                    p = EXPORTS_DIR / f"script_{timestamp}.csv"
                    export_csv(self.results, p)
                    exported.append(f"CSV: {p.name}")
                elif fmt == "json":
                    p = EXPORTS_DIR / f"script_{timestamp}.json"
                    export_json(self.results, p, metadata=self.metadata)
                    exported.append(f"JSON: {p.name}")
                elif fmt == "xlsx":
                    p = EXPORTS_DIR / f"script_{timestamp}.xlsx"
                    if export_excel(self.results, p):
                        exported.append(f"Excel: {p.name}")
                elif fmt == "sd_prompts":
                    p = EXPORTS_DIR / f"sd_prompts_{timestamp}.txt"
                    export_sd_prompts(self.results, p)
                    exported.append(f"SDプロンプト: {p.name}")
                elif fmt == "wildcard":
                    p = EXPORTS_DIR / f"wildcard_{timestamp}.txt"
                    export_wildcard(self.results, p)
                    exported.append(f"Wild Card: {p.name}")
                elif fmt == "wildcard_neg":
                    p = EXPORTS_DIR / f"wildcard_negative_{timestamp}.txt"
                    export_wildcard_negative(self.results, p)
                    exported.append(f"Wild Card (Neg): {p.name}")
                elif fmt == "dialogue":
                    p = EXPORTS_DIR / f"dialogue_{timestamp}.txt"
                    export_dialogue_list(self.results, p)
                    exported.append(f"セリフ一覧: {p.name}")
                elif fmt == "markdown":
                    p = EXPORTS_DIR / f"script_{timestamp}.md"
                    export_markdown(self.results, p)
                    exported.append(f"Markdown: {p.name}")
            except Exception as e:
                log_message(f"エクスポートエラー ({fmt}): {e}")
                exported.append(f"{fmt}: エラー")

        # 結果通知
        if hasattr(self.master, "snackbar"):
            self.master.snackbar.show(
                f"{len(exported)}形式エクスポート完了",
                type="success"
            )
        if hasattr(self.master, "log"):
            for item in exported:
                self.master.log(f"[FILE]{item}")

        self.destroy()


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Daihon Rakku")
        self.geometry("820x950")
        self.minsize(720, 800)
        
        # M3 Surface background
        self.configure(fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST)
        
        self.config_data = load_config()
        self.is_generating = False
        self.stop_requested = False
        self.last_results = None  # 最新の生成結果を保持（再エクスポート用）
        self.last_metadata = None  # パイプラインメタデータ（再エクスポート用）
        self.create_widgets()
        self.load_saved_config()

        # CTkEntryプレースホルダー強制表示（初期化直後のタイミング問題対策）
        self.after(100, self._ensure_placeholders)

        # ドラッグ&ドロップ初期化（ウィンドウマッピング後に実行）
        self.after(200, self._setup_png_drop)

        # ウィンドウ閉じ保護
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        # ショートカットキー
        self.bind("<Control-Return>", lambda e: self.start_generation())
        self.bind("<Escape>", lambda e: self.stop_generation() if self.is_generating else None)

        # プレゼンモード（Ctrl+Shift+P で切替）
        self._presentation_mode = False
        self.bind("<Control-Shift-P>", lambda e: self.toggle_presentation_mode())

    def create_widgets(self):
        # ══════════════════════════════════════════════════════════════
        # HEADER
        # ══════════════════════════════════════════════════════════════
        header = ctk.CTkFrame(self, height=56, fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        header_inner = ctk.CTkFrame(header, fg_color="transparent")
        header_inner.pack(fill="both", expand=True, padx=24, pady=12)

        icon_text_label(
            header_inner, Icons.FILM, "Daihon Rakku",
            icon_size=16, text_size=20, text_color=MaterialColors.ON_SURFACE
        ).pack(side="left")

        ctk.CTkLabel(
            header_inner, text="v1.7.0",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT,
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4, padx=8, pady=4
        ).pack(side="left", padx=(12, 0))

        ctk.CTkLabel(
            header_inner, text="FANZA同人CG集 脚本生成",
            font=ctk.CTkFont(size=13), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="right")

        # Header bottom border
        ctk.CTkFrame(self, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x")

        # ══════════════════════════════════════════════════════════════
        # MAIN CONTENT
        # ══════════════════════════════════════════════════════════════
        self.main_container = ctk.CTkScrollableFrame(
            self, fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST,
            scrollbar_button_color=MaterialColors.OUTLINE_VARIANT
        )
        self.main_container.pack(fill="both", expand=True)

        content = ctk.CTkFrame(self.main_container, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=24, pady=20)

        # ══════════════════════════════════════════════════════════════
        # 1. API設定（プレゼン時は非表示）
        # ══════════════════════════════════════════════════════════════
        api_card = ctk.CTkFrame(content, fg_color=MaterialColors.SURFACE_CONTAINER_LOW, corner_radius=0, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT)
        self._api_card = api_card  # プレゼンモード用参照
        api_card.pack(fill="x", pady=(0, 16))

        icon_text_label(
            api_card, Icons.LOCK, "API設定",
            icon_size=12, text_size=14
        ).pack(anchor="w", padx=20, pady=(12, 8))
        ctk.CTkFrame(api_card, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", padx=20, pady=(0, 8))

        # プロバイダー選択
        provider_row = ctk.CTkFrame(api_card, fg_color="transparent")
        provider_row.pack(fill="x", padx=20, pady=(0, 8))
        ctk.CTkLabel(
            provider_row, text="バックエンド:",
            font=ctk.CTkFont(size=13), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left", padx=(0, 8))
        self._provider_var = tk.StringVar(value=PROVIDER_CLAUDE)
        ctk.CTkRadioButton(
            provider_row, text="Claude (Anthropic)", variable=self._provider_var,
            value=PROVIDER_CLAUDE, font=ctk.CTkFont(size=13),
            text_color=MaterialColors.ON_SURFACE,
            command=self._on_provider_changed
        ).pack(side="left", padx=(0, 12))
        # Claude APIキー
        self.api_field = ctk.CTkEntry(
            api_card, height=42, placeholder_text="Anthropic API Key (sk-ant-...)", show="*",
            font=ctk.CTkFont(size=15),
            fg_color=MaterialColors.SURFACE_CONTAINER, text_color=MaterialColors.ON_SURFACE,
            placeholder_text_color="#3D3D3D",
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE
        )
        self.api_field.pack(fill="x", padx=20, pady=(0, 4))

        # 下部余白
        ctk.CTkFrame(api_card, fg_color="transparent", height=8).pack(fill="x", padx=20)

        # ══════════════════════════════════════════════════════════════
        # 2. プロファイル管理（キャラ生成より上に配置）
        # ══════════════════════════════════════════════════════════════
        profile_card = ctk.CTkFrame(content, fg_color=MaterialColors.SURFACE_CONTAINER_LOW, corner_radius=0, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT)
        profile_card.pack(fill="x", pady=(0, 16))

        icon_text_label(
            profile_card, Icons.FOLDER, "プロファイル管理",
            icon_size=12, text_size=14
        ).pack(anchor="w", padx=20, pady=(12, 8))
        ctk.CTkFrame(profile_card, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", padx=20, pady=(0, 8))

        profile_row = ctk.CTkFrame(profile_card, fg_color="transparent")
        profile_row.pack(fill="x", padx=20, pady=(0, 12))

        self.profile_combo = ctk.CTkOptionMenu(
            profile_row, values=["（新規）"] + get_profile_list(), height=36, width=150,
            font=ctk.CTkFont(size=14),
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            button_color=MaterialColors.PRIMARY,
            text_color=MaterialColors.ON_SURFACE,
            dropdown_text_color=MaterialColors.ON_SURFACE,
            dropdown_fg_color=MaterialColors.SURFACE,
            command=self.on_profile_selected
        )
        self.profile_combo.pack(side="left", padx=(0, 8))
        self.profile_combo.set("（新規）")

        self.profile_name_entry = ctk.CTkEntry(
            profile_row, height=36, width=120, placeholder_text="プロファイル名",
            font=ctk.CTkFont(size=14),
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            text_color=MaterialColors.ON_SURFACE,
            placeholder_text_color="#3D3D3D",
            border_width=1, border_color=MaterialColors.OUTLINE_VARIANT
        )
        self.profile_name_entry.pack(side="left", padx=(0, 8))

        MaterialButton(
            profile_row, text="保存", variant="filled", size="small",
            width=56, command=self.save_current_profile
        ).pack(side="left", padx=(0, 4))
        MaterialButton(
            profile_row, text="読込", variant="filled_tonal", size="small",
            width=56, command=self.load_selected_profile
        ).pack(side="left", padx=(0, 4))
        MaterialButton(
            profile_row, text="複製", variant="outlined", size="small",
            width=48, command=self.copy_selected_profile
        ).pack(side="left", padx=(0, 4))
        MaterialButton(
            profile_row, text="削除", variant="danger", size="small",
            width=48, command=self.delete_selected_profile
        ).pack(side="left")

        # ══════════════════════════════════════════════════════════════
        # 3. キャラクター設定
        # ══════════════════════════════════════════════════════════════
        char_card = ctk.CTkFrame(content, fg_color=MaterialColors.SURFACE_CONTAINER_LOW, corner_radius=0, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT)
        char_card.pack(fill="x", pady=(0, 16))

        icon_text_label(
            char_card, Icons.USER, "キャラクター設定",
            icon_size=12, text_size=14
        ).pack(anchor="w", padx=20, pady=(12, 8))
        ctk.CTkFrame(char_card, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", padx=20, pady=(0, 8))

        # --- 使用キャラ選択行 ---
        self._char_select_row = ctk.CTkFrame(char_card, fg_color="transparent")
        self._char_select_row.pack(fill="x", padx=20, pady=(0, 12))

        # --- プリセットコンテナ（折りたたみ、初期非表示） ---
        self._preset_card = MaterialCard(
            char_card, title="二次創作・プリセットキャラ一覧", variant="outlined",
            collapsible=True, start_collapsed=True
        )
        self._preset_card.pack(fill="x", padx=16, pady=(0, 10))
        self._preset_container = self._preset_card.content_frame

        # --- プリセットタブ構築 ---
        self._all_presets = []
        self._preset_map = {}
        self._category_chips = {}
        self._selected_category = "全て"
        self._preset_card_frame = None
        self._build_preset_tab(self._preset_container)

        # --- オリジナルキャラ設定コンテナ（折りたたみ、初期非表示） ---
        self._custom_card = MaterialCard(
            char_card, title="オリジナルキャラ作成", variant="outlined",
            collapsible=True, start_collapsed=True
        )
        self._custom_card.pack(fill="x", padx=16, pady=(0, 10))
        self._custom_container = self._custom_card.content_frame

        self._selected_archetype = "ツンデレ"
        self._selected_hair_color = "黒髪"
        self._archetype_chips = {}
        self._hair_color_chips = {}
        self._build_custom_tab(self._custom_container)

        # 内部変数（config保存/復元用、UIウィジェットなし）
        self._work_title_val = ""
        self._char_name_val = ""

        # ネストスクロール衝突防止
        self._setup_nested_scroll()

        # --- 使用キャラ選択ウィジェット ---
        char_select_row = self._char_select_row

        ctk.CTkLabel(char_select_row, text="使用キャラ:",
                    font=ctk.CTkFont(size=13, weight="bold"),
                    text_color=MaterialColors.ON_SURFACE_VARIANT).pack(side="left", padx=(0, 6))

        self.char_select_combo = ctk.CTkOptionMenu(
            char_select_row, values=["（キャラ選択）"], height=36,
            font=ctk.CTkFont(size=14),
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            button_color=MaterialColors.PRIMARY, dropdown_fg_color=MaterialColors.SURFACE,
            text_color=MaterialColors.ON_SURFACE,
            dropdown_text_color=MaterialColors.ON_SURFACE,
            command=self.on_char_selected
        )
        self.char_select_combo.pack(side="left", fill="x", expand=True)
        self.refresh_char_list()
        self.refresh_preset_list()

        # ══════════════════════════════════════════════════════════════
        # 4. 作品設定（メイン入力エリア）
        # ══════════════════════════════════════════════════════════════
        concept_card = ctk.CTkFrame(content, fg_color=MaterialColors.SURFACE_CONTAINER_LOW, corner_radius=0, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT)
        concept_card.pack(fill="x", pady=(0, 16))

        icon_text_label(
            concept_card, Icons.BOOK, "作品設定",
            icon_size=12, text_size=14
        ).pack(anchor="w", padx=20, pady=(12, 8))
        ctk.CTkFrame(concept_card, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", padx=20, pady=(0, 8))

        # コンセプト入力
        concept_label_frame = ctk.CTkFrame(concept_card, fg_color="transparent")
        concept_label_frame.pack(fill="x", padx=20)
        ctk.CTkLabel(
            concept_label_frame, text="コンセプト",
            font=ctk.CTkFont(size=14, weight="bold"), text_color=MaterialColors.PRIMARY
        ).pack(side="left")
        ctk.CTkLabel(
            concept_label_frame, text="（作品の設定・シチュエーションを詳しく記述）",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left", padx=(4, 0))

        # --- コンセプトプリセット行 ---
        preset_row = ctk.CTkFrame(concept_card, fg_color="transparent")
        preset_row.pack(fill="x", padx=20, pady=(6, 0))
        ctk.CTkLabel(
            preset_row, text="プリセット:",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        _cat_list = ["--"] + list(CONCEPT_PRESETS.keys())
        self.concept_cat_menu = ctk.CTkOptionMenu(
            preset_row, values=_cat_list, width=160, height=32,
            font=ctk.CTkFont(size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            button_color=MaterialColors.PRIMARY,
            dropdown_fg_color=MaterialColors.SURFACE,
            text_color=MaterialColors.ON_SURFACE,
            command=self._on_concept_category_changed,
        )
        self.concept_cat_menu.set("--")
        self.concept_cat_menu.pack(side="left", padx=(6, 4))
        self.concept_name_menu = ctk.CTkOptionMenu(
            preset_row, values=["--"], width=200, height=32,
            font=ctk.CTkFont(size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            button_color=MaterialColors.PRIMARY,
            dropdown_fg_color=MaterialColors.SURFACE,
            text_color=MaterialColors.ON_SURFACE,
            command=self._on_concept_preset_selected,
        )
        self.concept_name_menu.set("--")
        self.concept_name_menu.pack(side="left", padx=(0, 4))
        self.concept_shuffle_btn = ctk.CTkButton(
            preset_row, text="\U0001f3b2", width=32, height=32,
            font=ctk.CTkFont(size=16),
            fg_color=MaterialColors.SECONDARY_CONTAINER,
            text_color=MaterialColors.ON_SECONDARY_CONTAINER,
            hover_color=MaterialColors.PRIMARY, corner_radius=6,
            command=self._on_concept_shuffle,
        )
        self.concept_shuffle_btn.pack(side="left")
        self._last_concept_variation = ""

        self.concept_text = ctk.CTkTextbox(
            concept_card, height=120,
            font=ctk.CTkFont(size=16),
            fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST,
            text_color=MaterialColors.ON_SURFACE,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE,
            wrap="word"
        )
        self.concept_text.pack(fill="x", padx=20, pady=(6, 12))

        # 登場人物入力（個別フィールド）
        char_label_frame = ctk.CTkFrame(concept_card, fg_color="transparent")
        char_label_frame.pack(fill="x", padx=20)
        ctk.CTkLabel(
            char_label_frame, text="登場人物",
            font=ctk.CTkFont(size=14, weight="bold"), text_color=MaterialColors.PRIMARY
        ).pack(side="left")
        ctk.CTkLabel(
            char_label_frame, text="（名前・性格・外見を入力）",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left", padx=(4, 0))

        char_fields_frame = ctk.CTkFrame(concept_card, fg_color="transparent")
        char_fields_frame.pack(fill="x", padx=20, pady=(6, 12))

        _entry_cfg = dict(
            height=34, font=ctk.CTkFont(size=15),
            fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST,
            text_color=MaterialColors.ON_SURFACE,
            placeholder_text_color="#3D3D3D",
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE
        )
        _lbl_cfg = dict(font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT)

        # 名前
        _r0 = ctk.CTkFrame(char_fields_frame, fg_color="transparent")
        _r0.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(_r0, text="名前", **_lbl_cfg).pack(side="left", padx=(0, 6))
        self.char_name_field = ctk.CTkEntry(_r0, placeholder_text="例: 中野一花（五等分の花嫁）", **_entry_cfg)
        self.char_name_field.pack(side="left", fill="x", expand=True)

        # 性格
        _r1 = ctk.CTkFrame(char_fields_frame, fg_color="transparent")
        _r1.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(_r1, text="性格", **_lbl_cfg).pack(side="left", padx=(0, 6))
        self.char_personality_field = ctk.CTkEntry(_r1, placeholder_text="例: ツンデレ、意地っ張り", **_entry_cfg)
        self.char_personality_field.pack(side="left", fill="x", expand=True)

        # 一人称 + 語尾（横並び）
        _r2 = ctk.CTkFrame(char_fields_frame, fg_color="transparent")
        _r2.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(_r2, text="一人称", **_lbl_cfg).pack(side="left", padx=(0, 6))
        self.char_first_person_field = ctk.CTkEntry(_r2, width=120, placeholder_text="例: あたし", **_entry_cfg)
        self.char_first_person_field.pack(side="left", padx=(0, 12))
        ctk.CTkLabel(_r2, text="語尾", **_lbl_cfg).pack(side="left", padx=(0, 6))
        self.char_endings_field = ctk.CTkEntry(_r2, placeholder_text="例: ～だよ, ～かな", **_entry_cfg)
        self.char_endings_field.pack(side="left", fill="x", expand=True)

        # 外見
        _r3 = ctk.CTkFrame(char_fields_frame, fg_color="transparent")
        _r3.pack(fill="x")
        ctk.CTkLabel(_r3, text="外見", **_lbl_cfg).pack(side="left", padx=(0, 6))
        self.char_appearance_field = ctk.CTkEntry(_r3, placeholder_text="例: 金髪ロング、青い瞳", **_entry_cfg)
        self.char_appearance_field.pack(side="left", fill="x", expand=True)

        # その他の登場人物入力
        other_label_frame = ctk.CTkFrame(concept_card, fg_color="transparent")
        other_label_frame.pack(fill="x", padx=20)
        ctk.CTkLabel(
            other_label_frame, text="その他の登場人物",
            font=ctk.CTkFont(size=14, weight="bold"), text_color=MaterialColors.PRIMARY
        ).pack(side="left")
        ctk.CTkLabel(
            other_label_frame, text="（男主人公・サブキャラ等の設定）",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left", padx=(4, 0))

        self.other_chars_text = ctk.CTkTextbox(
            concept_card, height=70,
            font=ctk.CTkFont(size=16),
            fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST,
            text_color=MaterialColors.ON_SURFACE,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE,
            wrap="word"
        )
        self.other_chars_text.pack(fill="x", padx=20, pady=(8, 16))
        self.other_chars_text.insert("1.0", "相手役の男性（顔なし）\nSD: 1boy, faceless_male")

        # ── 男性キャラ設定 ──
        male_char_frame = ctk.CTkFrame(concept_card, fg_color="transparent")
        male_char_frame.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(male_char_frame, text="男性キャラ", font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold")).pack(anchor="w", padx=4)

        male_row = ctk.CTkFrame(male_char_frame, fg_color="transparent")
        male_row.pack(fill="x", padx=4, pady=(4, 0))

        # プリセットドロップダウン
        self.male_preset_combo = ctk.CTkOptionMenu(
            male_row, values=list(MALE_PRESETS.keys()),
            width=160, height=32,
            font=ctk.CTkFont(family=FONT_JP, size=12),
            command=self._on_male_preset_changed,
        )
        self.male_preset_combo.pack(side="left", padx=(0, 8))
        self.male_preset_combo.set("おまかせ")

        # カスタム入力フィールド
        self.male_custom_field = ctk.CTkEntry(
            male_row, placeholder_text="自由入力: はげた小太りの中年 等",
            height=34, font=ctk.CTkFont(family=FONT_JP, size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST,
            text_color=MaterialColors.ON_SURFACE,
            placeholder_text_color="#3D3D3D",
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE,
        )
        self.male_custom_field.pack(side="left", fill="x", expand=True)

        # 男性 髪型・髪色・肌色
        male_detail_row = ctk.CTkFrame(male_char_frame, fg_color="transparent")
        male_detail_row.pack(fill="x", padx=4, pady=(4, 0))

        ctk.CTkLabel(male_detail_row, text="髪型", font=ctk.CTkFont(family=FONT_JP, size=12)).pack(side="left", padx=(0, 4))
        self.male_hair_style_combo = ctk.CTkOptionMenu(
            male_detail_row, values=list(MALE_HAIR_STYLE_OPTIONS.keys()),
            width=100, height=32, font=ctk.CTkFont(family=FONT_JP, size=12),
        )
        self.male_hair_style_combo.pack(side="left", padx=(0, 12))
        self.male_hair_style_combo.set("おまかせ")

        ctk.CTkLabel(male_detail_row, text="髪色", font=ctk.CTkFont(family=FONT_JP, size=12)).pack(side="left", padx=(0, 4))
        self.male_hair_color_combo = ctk.CTkOptionMenu(
            male_detail_row, values=list(MALE_HAIR_COLOR_OPTIONS.keys()),
            width=100, height=32, font=ctk.CTkFont(family=FONT_JP, size=12),
        )
        self.male_hair_color_combo.pack(side="left", padx=(0, 12))
        self.male_hair_color_combo.set("おまかせ")

        ctk.CTkLabel(male_detail_row, text="肌色", font=ctk.CTkFont(family=FONT_JP, size=12)).pack(side="left", padx=(0, 4))
        self.male_skin_color_combo = ctk.CTkOptionMenu(
            male_detail_row, values=list(MALE_SKIN_COLOR_OPTIONS.keys()),
            width=100, height=32, font=ctk.CTkFont(family=FONT_JP, size=12),
        )
        self.male_skin_color_combo.pack(side="left", padx=(0, 0))
        self.male_skin_color_combo.set("おまかせ")

        # faceless_male チェックボックス（デフォルトON）
        male_faceless_row = ctk.CTkFrame(male_char_frame, fg_color="transparent")
        male_faceless_row.pack(fill="x", padx=4, pady=(4, 0))
        self.male_faceless_var = ctk.BooleanVar(value=True)
        self.male_faceless_check = ctk.CTkCheckBox(
            male_faceless_row, text="faceless_male（顔なし）",
            variable=self.male_faceless_var,
            font=ctk.CTkFont(family=FONT_JP, size=12),
            height=28,
        )
        self.male_faceless_check.pack(side="left", padx=(0, 8))
        ctk.CTkLabel(
            male_faceless_row, text="※ OFFにすると男性の顔も描写されます",
            font=ctk.CTkFont(family=FONT_JP, size=11),
            text_color="#888888",
        ).pack(side="left")

        # ── シーン環境 ──
        env_frame = ctk.CTkFrame(concept_card, fg_color="transparent")
        env_frame.pack(fill="x", padx=16, pady=(0, 8))
        ctk.CTkLabel(env_frame, text="シーン環境", font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold")).pack(anchor="w", padx=4)

        env_row = ctk.CTkFrame(env_frame, fg_color="transparent")
        env_row.pack(fill="x", padx=4, pady=(4, 0))

        # 時間帯
        ctk.CTkLabel(env_row, text="時間帯", font=ctk.CTkFont(family=FONT_JP, size=12)).pack(side="left", padx=(0, 4))
        self.time_of_day_combo = ctk.CTkOptionMenu(
            env_row, values=list(TIME_OF_DAY_OPTIONS.keys()),
            width=100, height=32,
            font=ctk.CTkFont(family=FONT_JP, size=12),
        )
        self.time_of_day_combo.pack(side="left", padx=(0, 16))
        self.time_of_day_combo.set("おまかせ")

        # 場所タイプ
        ctk.CTkLabel(env_row, text="場所", font=ctk.CTkFont(family=FONT_JP, size=12)).pack(side="left", padx=(0, 4))
        self.location_type_combo = ctk.CTkOptionMenu(
            env_row, values=list(LOCATION_TYPE_OPTIONS.keys()),
            width=100, height=32,
            font=ctk.CTkFont(family=FONT_JP, size=12),
        )
        self.location_type_combo.pack(side="left")
        self.location_type_combo.set("おまかせ")

        # ══════════════════════════════════════════════════════════════
        # 4b. SDプロンプト設定（折りたたみ）
        # ══════════════════════════════════════════════════════════════
        sd_card = MaterialCard(
            content, title="SDプロンプト設定", variant="outlined",
            collapsible=True, start_collapsed=False
        )
        sd_card.pack(fill="x", pady=(0, 16))
        sd_content = sd_card.content_frame

        # --- クオリティタグ ---
        _quality_header = ctk.CTkFrame(sd_content, fg_color="transparent")
        _quality_header.pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            _quality_header, text="クオリティタグ",
            font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        ctk.CTkLabel(
            _quality_header, text="（不要の場合は「カスタム」を選択し空欄のままにしてください）",
            font=ctk.CTkFont(family=FONT_JP, size=12),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left", padx=(6, 0))

        quality_mode_row = ctk.CTkFrame(sd_content, fg_color="transparent")
        quality_mode_row.pack(fill="x", pady=(0, 4))
        self.sd_quality_mode_var = tk.StringVar(value="auto")
        ctk.CTkRadioButton(
            quality_mode_row, text="自動", variable=self.sd_quality_mode_var, value="auto",
            font=ctk.CTkFont(family=FONT_JP, size=12),
            command=self._on_sd_quality_mode_changed
        ).pack(side="left", padx=(0, 16))
        ctk.CTkRadioButton(
            quality_mode_row, text="カスタム", variable=self.sd_quality_mode_var, value="manual",
            font=ctk.CTkFont(family=FONT_JP, size=12),
            command=self._on_sd_quality_mode_changed
        ).pack(side="left")

        self.sd_quality_custom_entry = ctk.CTkEntry(
            sd_content, height=36,
            font=ctk.CTkFont(size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT,
            text_color=MaterialColors.ON_SURFACE,
            placeholder_text="カスタムクオリティタグを入力…",
            placeholder_text_color="#3D3D3D",
        )
        # 自動モード初期値: QUALITY_POSITIVE_TAGS を表示して無効化
        self.sd_quality_custom_entry.insert(0, QUALITY_POSITIVE_TAGS)
        self.sd_quality_custom_entry.configure(state="disabled")
        self.sd_quality_custom_entry.pack(fill="x", pady=(0, 8))

        # 区切り線
        ctk.CTkFrame(sd_content, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", pady=8)

        # --- プレフィックス ---
        _prefix_header = ctk.CTkFrame(sd_content, fg_color="transparent")
        _prefix_header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(
            _prefix_header, text="プレフィックス（全シーン先頭に追加）",
            font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        MaterialButton(
            _prefix_header, text="クリア", variant="text", size="small",
            command=lambda: (self.sd_prefix_text.delete("1.0", "end"))
        ).pack(side="right")
        self.sd_prefix_text = ctk.CTkTextbox(
            sd_content, height=100,
            font=ctk.CTkFont(size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT,
            text_color=MaterialColors.ON_SURFACE,
        )
        self.sd_prefix_text.pack(fill="x", pady=(0, 8))
        self.sd_prefix_text.bind("<KeyRelease>", lambda e: self._auto_resize_textbox(self.sd_prefix_text, 100, 1200))

        # --- サフィックス ---
        _suffix_header = ctk.CTkFrame(sd_content, fg_color="transparent")
        _suffix_header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(
            _suffix_header, text="サフィックス（全シーン末尾に追加）",
            font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        MaterialButton(
            _suffix_header, text="クリア", variant="text", size="small",
            command=lambda: (self.sd_suffix_text.delete("1.0", "end"))
        ).pack(side="right")
        self.sd_suffix_text = ctk.CTkTextbox(
            sd_content, height=100,
            font=ctk.CTkFont(size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT,
            text_color=MaterialColors.ON_SURFACE,
        )
        self.sd_suffix_text.pack(fill="x", pady=(0, 8))
        self.sd_suffix_text.bind("<KeyRelease>", lambda e: self._auto_resize_textbox(self.sd_suffix_text, 100, 1200))

        # 区切り線
        ctk.CTkFrame(sd_content, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", pady=8)

        # ── ネガティブプロンプト設定 ──
        ctk.CTkLabel(
            sd_content, text="ネガティブプロンプト",
            font=ctk.CTkFont(family=FONT_JP, size=14, weight="bold"),
            text_color=MaterialColors.ON_SURFACE
        ).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            sd_content, text="シーン固有のネガティブは自動生成されます。ここではベース・プレフィックス・サフィックスを設定します。",
            font=ctk.CTkFont(family=FONT_JP, size=11),
            text_color=MaterialColors.ON_SURFACE_VARIANT, wraplength=380,
        ).pack(anchor="w", pady=(0, 8))

        # --- ネガティブ品質ベース ---
        _neg_base_header = ctk.CTkFrame(sd_content, fg_color="transparent")
        _neg_base_header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(
            _neg_base_header, text="品質ネガティブ（共通ベース）",
            font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        MaterialButton(
            _neg_base_header, text="リセット", variant="text", size="small",
            command=lambda: (
                self.sd_neg_base_text.delete("1.0", "end"),
                self.sd_neg_base_text.insert("1.0", "worst_quality, low_quality, bad_anatomy, bad_hands, missing_fingers, extra_digits, fewer_digits, text, signature, watermark, username, blurry, jpeg_artifacts, cropped"),
                self._auto_resize_textbox(self.sd_neg_base_text, 60, 400),
            )
        ).pack(side="right")
        self.sd_neg_base_text = ctk.CTkTextbox(
            sd_content, height=60,
            font=ctk.CTkFont(size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT,
            text_color=MaterialColors.ON_SURFACE,
        )
        self.sd_neg_base_text.pack(fill="x", pady=(0, 8))
        self.sd_neg_base_text.insert("1.0", "worst_quality, low_quality, bad_anatomy, bad_hands, missing_fingers, extra_digits, fewer_digits, text, signature, watermark, username, blurry, jpeg_artifacts, cropped")
        self.sd_neg_base_text.bind("<KeyRelease>", lambda e: self._auto_resize_textbox(self.sd_neg_base_text, 60, 400))

        # --- ネガティブプレフィックス ---
        _neg_prefix_header = ctk.CTkFrame(sd_content, fg_color="transparent")
        _neg_prefix_header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(
            _neg_prefix_header, text="ネガティブ プレフィックス",
            font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        MaterialButton(
            _neg_prefix_header, text="クリア", variant="text", size="small",
            command=lambda: (self.sd_neg_prefix_text.delete("1.0", "end"))
        ).pack(side="right")
        self.sd_neg_prefix_text = ctk.CTkTextbox(
            sd_content, height=60,
            font=ctk.CTkFont(size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT,
            text_color=MaterialColors.ON_SURFACE,
        )
        self.sd_neg_prefix_text.pack(fill="x", pady=(0, 8))
        self.sd_neg_prefix_text.bind("<KeyRelease>", lambda e: self._auto_resize_textbox(self.sd_neg_prefix_text, 60, 400))

        # --- ネガティブサフィックス ---
        _neg_suffix_header = ctk.CTkFrame(sd_content, fg_color="transparent")
        _neg_suffix_header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(
            _neg_suffix_header, text="ネガティブ サフィックス",
            font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        MaterialButton(
            _neg_suffix_header, text="クリア", variant="text", size="small",
            command=lambda: (self.sd_neg_suffix_text.delete("1.0", "end"))
        ).pack(side="right")
        self.sd_neg_suffix_text = ctk.CTkTextbox(
            sd_content, height=60,
            font=ctk.CTkFont(size=13),
            fg_color=MaterialColors.SURFACE_CONTAINER,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT,
            text_color=MaterialColors.ON_SURFACE,
        )
        self.sd_neg_suffix_text.pack(fill="x", pady=(0, 8))
        self.sd_neg_suffix_text.bind("<KeyRelease>", lambda e: self._auto_resize_textbox(self.sd_neg_suffix_text, 60, 400))

        # 区切り線
        ctk.CTkFrame(sd_content, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", pady=8)

        # --- PNG Info 読み取り ---
        ctk.CTkLabel(
            sd_content, text="画像から読み取り",
            font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(anchor="w", pady=(0, 4))

        # ドロップゾーン
        self.png_drop_frame = ctk.CTkFrame(
            sd_content,
            fg_color=MaterialColors.SURFACE_CONTAINER,
            corner_radius=8, border_width=2,
            border_color=MaterialColors.OUTLINE_VARIANT,
        )
        self.png_drop_frame.pack(fill="x", pady=(0, 4))

        # アイコン＋ヒントテキストを横並び（pack expand で上下中央）
        _drop_row = ctk.CTkFrame(self.png_drop_frame, fg_color="transparent")
        _drop_row.pack(expand=True, pady=16)
        self.png_drop_icon_label = ctk.CTkLabel(
            _drop_row, text=Icons.IMAGE,
            font=ctk.CTkFont(family=FONT_ICON, size=20),
            text_color=MaterialColors.ON_SURFACE_VARIANT,
        )
        self.png_drop_icon_label.pack(side="left", padx=(0, 8))
        self.png_drop_hint_label = ctk.CTkLabel(
            _drop_row,
            text="ここに画像をドロップ、またはクリックして選択" if WINDND_AVAILABLE else "クリックして画像を選択",
            font=ctk.CTkFont(family=FONT_JP, size=12),
            text_color=MaterialColors.ON_SURFACE_VARIANT,
        )
        self.png_drop_hint_label.pack(side="left")
        self.png_filename_label = ctk.CTkLabel(
            _drop_row, text="",
            font=ctk.CTkFont(family=FONT_JP, size=11),
            text_color=MaterialColors.PRIMARY,
        )
        self.png_filename_label.pack(side="left", padx=(12, 0))

        # ドロップゾーン全体をクリック可能に
        for w in (self.png_drop_frame, _drop_row, self.png_drop_icon_label, self.png_drop_hint_label, self.png_filename_label):
            w.configure(cursor="hand2")
            w.bind("<Button-1>", lambda e: self._on_png_info_load())

        self.png_preview_text = ctk.CTkTextbox(
            sd_content, height=120,
            font=ctk.CTkFont(size=12),
            fg_color=MaterialColors.SURFACE_CONTAINER,
            corner_radius=4, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT,
            text_color=MaterialColors.ON_SURFACE,
            state="disabled"
        )
        self.png_preview_text.pack(fill="x", pady=(0, 4))

        png_apply_row = ctk.CTkFrame(sd_content, fg_color="transparent")
        png_apply_row.pack(fill="x", pady=(0, 4))
        MaterialButton(
            png_apply_row, text="プレフィックスに適用", variant="filled_tonal", size="small",
            command=self._apply_png_to_prefix
        ).pack(side="left", padx=(0, 8))
        MaterialButton(
            png_apply_row, text="サフィックスに適用", variant="filled_tonal", size="small",
            command=self._apply_png_to_suffix
        ).pack(side="left")
        MaterialButton(
            png_apply_row, text="Negに適用", variant="filled_tonal", size="small",
            command=self._apply_png_to_negative
        ).pack(side="left", padx=(8, 0))

        # ══════════════════════════════════════════════════════════════
        # 5. 生成設定
        # ══════════════════════════════════════════════════════════════
        settings_card = ctk.CTkFrame(content, fg_color=MaterialColors.SURFACE_CONTAINER_LOW, corner_radius=0, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT)
        settings_card.pack(fill="x", pady=(0, 16))

        icon_text_label(
            settings_card, Icons.GEAR, "生成設定",
            icon_size=12, text_size=14
        ).pack(anchor="w", padx=20, pady=(12, 8))
        ctk.CTkFrame(settings_card, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", padx=20, pady=(0, 8))

        settings_row = ctk.CTkFrame(settings_card, fg_color="transparent")
        settings_row.pack(fill="x", padx=20, pady=(0, 12))

        # シーン数
        scenes_frame = ctk.CTkFrame(settings_row, fg_color="transparent")
        scenes_frame.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(scenes_frame, text="シーン数", font=ctk.CTkFont(size=13), text_color=MaterialColors.ON_SURFACE_VARIANT).pack(anchor="w")
        self.scenes_entry = ctk.CTkEntry(
            scenes_frame, height=38, font=ctk.CTkFont(size=15),
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            text_color=MaterialColors.ON_SURFACE,
            border_width=1, border_color=MaterialColors.OUTLINE_VARIANT
        )
        self.scenes_entry.pack(fill="x", pady=(4, 0))
        self.scenes_entry.insert(0, "10")

        # テーマ
        theme_frame = ctk.CTkFrame(settings_row, fg_color="transparent")
        theme_frame.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(theme_frame, text="テーマ", font=ctk.CTkFont(size=13), text_color=MaterialColors.ON_SURFACE_VARIANT).pack(anchor="w")
        self.theme_combo = ctk.CTkOptionMenu(
            theme_frame, values=list(THEME_OPTIONS.keys()), height=38,
            font=ctk.CTkFont(size=14),
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            button_color=MaterialColors.PRIMARY, dropdown_fg_color=MaterialColors.SURFACE,
            text_color=MaterialColors.ON_SURFACE,
            dropdown_text_color=MaterialColors.ON_SURFACE
        )
        self.theme_combo.pack(fill="x", pady=(4, 0))
        self.theme_combo.set("指定なし")

        self.scenes_entry.bind("<KeyRelease>", self.update_cost_preview)

        # ストーリー構成バー
        self._build_structure_bar(settings_card)

        self.cost_preview_label = ctk.CTkLabel(
            settings_card, text="シーン数入力で予想コスト表示",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT
        )
        self._cost_preview_label = self.cost_preview_label  # プレゼンモード用参照
        self.cost_preview_label.pack(anchor="w", padx=20, pady=(4, 4))

        # v8.7: 品質優先モード（Wave並列を無効化→全シーン直列生成）
        self.quality_priority_var = ctk.BooleanVar(value=False)
        self.quality_priority_cb = ctk.CTkCheckBox(
            settings_card, text="品質優先モード（直列生成・低速）",
            variable=self.quality_priority_var,
            font=ctk.CTkFont(family=FONT_JP, size=13),
            text_color=MaterialColors.ON_SURFACE_VARIANT,
            fg_color=MaterialColors.PRIMARY,
            hover_color=MaterialColors.PRIMARY_CONTAINER,
            border_color=MaterialColors.OUTLINE,
            checkmark_color=MaterialColors.ON_PRIMARY,
            corner_radius=4
        )
        self.quality_priority_cb.pack(anchor="w", padx=20, pady=(0, 12))

        # ══════════════════════════════════════════════════════════════
        # 6. 生成セクション
        # ══════════════════════════════════════════════════════════════
        generate_section = ctk.CTkFrame(content, fg_color=MaterialColors.SURFACE_CONTAINER_LOW, corner_radius=0,
                                        border_width=1, border_color=MaterialColors.PRIMARY)
        generate_section.pack(fill="x", pady=(0, 16))

        gen_inner = ctk.CTkFrame(generate_section, fg_color="transparent")
        gen_inner.pack(fill="x", padx=20, pady=20)

        # ステータス行
        status_row = ctk.CTkFrame(gen_inner, fg_color="transparent")
        status_row.pack(fill="x", pady=(0, 12))

        self.status_icon_label = ctk.CTkLabel(
            status_row, text=Icons.CLOCK,
            font=ctk.CTkFont(family=FONT_ICON, size=12),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        )
        self.status_icon_label.pack(side="left", padx=(0, 8))

        self.status_label = ctk.CTkLabel(
            status_row, text="待機中",
            font=ctk.CTkFont(family=FONT_JP, size=14, weight="bold"),
            text_color=MaterialColors.ON_SURFACE
        )
        self.status_label.pack(side="left")

        # フェーズ
        phase_frame = ctk.CTkFrame(status_row, fg_color="transparent")
        phase_frame.pack(side="right")
        self.phase_labels = []
        for phase in ["圧縮", "あらすじ", "分割", "シーン生成", "品質検証"]:
            pill = ctk.CTkFrame(phase_frame, fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=8)
            pill.pack(side="left", padx=4)
            lbl = ctk.CTkLabel(pill, text=phase, font=ctk.CTkFont(family=FONT_JP, size=13), text_color=MaterialColors.ON_SURFACE_VARIANT, padx=10, pady=4)
            lbl.pack()
            self.phase_labels.append((pill, lbl))

        # プログレス
        self.progress = ctk.CTkProgressBar(
            gen_inner, fg_color=MaterialColors.SURFACE_CONTAINER, progress_color=MaterialColors.PRIMARY,
            height=8, corner_radius=4
        )
        self.progress.pack(fill="x", pady=(0, 12))
        self.progress.set(0)

        # ボタン行
        btn_row = ctk.CTkFrame(gen_inner, fg_color="transparent")
        btn_row.pack(fill="x")

        self.generate_btn = MaterialButton(
            btn_row, text="脚本を生成", variant="filled", size="large",
            command=self.start_generation
        )
        self.generate_btn.pack(side="left", fill="x", expand=True, padx=(0, 8))
        add_tooltip(self.generate_btn, "脚本生成を開始 (Ctrl+Enter)")

        self.save_btn = MaterialButton(
            btn_row, text="保存", variant="filled_tonal", size="large",
            width=72, command=self.save_settings
        )
        self.save_btn.pack(side="left", padx=(0, 8))

        self.stop_btn = MaterialButton(
            btn_row, text="停止", variant="outlined", size="large",
            width=64, command=self.stop_generation
        )
        self.stop_btn.pack(side="left", padx=(0, 8))
        self.stop_btn.configure(state="disabled")
        add_tooltip(self.stop_btn, "生成を停止 (Esc)")

        self.export_btn = MaterialButton(
            btn_row, text="再エクスポート", variant="filled_tonal", size="large",
            width=120, command=self.open_export_dialog
        )
        self.export_btn.pack(side="left")
        self.export_btn.configure(state="disabled")
        add_tooltip(self.export_btn, "別形式で再エクスポート")

        # ══════════════════════════════════════════════════════════════
        # 7. コスト＆ログ（プレゼン時は非表示）
        # ══════════════════════════════════════════════════════════════
        cost_card = ctk.CTkFrame(content, fg_color=MaterialColors.SURFACE_CONTAINER_LOW, corner_radius=0, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT)
        self._cost_card = cost_card  # プレゼンモード用参照
        cost_card.pack(fill="x", pady=(0, 16))

        icon_text_label(
            cost_card, Icons.COINS, "コスト",
            icon_size=12, text_size=14
        ).pack(anchor="w", padx=20, pady=(12, 8))
        ctk.CTkFrame(cost_card, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", padx=20, pady=(0, 8))

        self.cost_label = ctk.CTkLabel(
            cost_card, text="生成後に表示",
            font=ctk.CTkFont(family=FONT_MONO, size=11), text_color=MaterialColors.ON_SURFACE_VARIANT, justify="left"
        )
        self.cost_label.pack(anchor="w", padx=20, pady=(0, 12))

        log_card = ctk.CTkFrame(content, fg_color=MaterialColors.SURFACE_CONTAINER_LOW, corner_radius=0, border_width=1, border_color=MaterialColors.OUTLINE_VARIANT)
        log_card.pack(fill="both", expand=True, pady=(0, 16))

        icon_text_label(
            log_card, Icons.LIST, "実行ログ",
            icon_size=12, text_size=14
        ).pack(anchor="w", padx=20, pady=(12, 8))
        ctk.CTkFrame(log_card, fg_color=MaterialColors.OUTLINE_VARIANT, height=1, corner_radius=0).pack(fill="x", padx=20, pady=(0, 8))

        self.log_text = ctk.CTkTextbox(
            log_card, height=180,
            fg_color=MaterialColors.INVERSE_SURFACE, text_color=MaterialColors.INVERSE_ON_SURFACE,
            corner_radius=6, font=ctk.CTkFont(family=FONT_MONO, size=12)
        )
        self.log_text.pack(fill="both", expand=True, padx=20, pady=(0, 12))

        # フッター（プレゼン用削除）

        # Snackbar
        self.snackbar = Snackbar(self)

        # フォーカス状態バインド（入力フィールド）
        for widget in [self.api_field, self.scenes_entry, self.concept_text,
                       self.char_name_field, self.char_personality_field,
                       self.char_first_person_field, self.char_endings_field,
                       self.char_appearance_field, self.other_chars_text]:
            widget.bind("<FocusIn>", lambda e, w=widget: w.configure(border_color=MaterialColors.PRIMARY))
            widget.bind("<FocusOut>", lambda e, w=widget: w.configure(border_color=MaterialColors.OUTLINE_VARIANT))

    def _set_concept_text(self, value: str):
        """コンセプトテキストを設定"""
        self.concept_text.delete("1.0", "end")
        if value:
            self.concept_text.insert("1.0", value)

    # --- コンセプトプリセットハンドラ ---
    def _on_concept_category_changed(self, cat: str):
        """カテゴリ変更→コンセプトドロップダウン更新"""
        if cat == "--" or cat not in CONCEPT_PRESETS:
            self.concept_name_menu.configure(values=["--"])
            self.concept_name_menu.set("--")
            return
        names = list(CONCEPT_PRESETS[cat].keys())
        self.concept_name_menu.configure(values=["--"] + names)
        self.concept_name_menu.set("--")

    def _on_concept_preset_selected(self, name: str):
        """プリセット選択→テキスト挿入+テーマ自動設定"""
        cat = self.concept_cat_menu.get()
        if name == "--" or cat == "--":
            return
        preset = CONCEPT_PRESETS.get(cat, {}).get(name)
        if not preset:
            return
        variation = random.choice(preset["variations"])
        self._last_concept_variation = variation
        self._set_concept_text(variation)
        # テーマ自動設定
        theme_key = preset.get("theme", "")
        if theme_key and theme_key in _THEME_KEY_TO_JP:
            self.theme_combo.set(_THEME_KEY_TO_JP[theme_key])

    def _on_concept_shuffle(self):
        """別バリエーション再抽選（同じものを避ける）"""
        cat = self.concept_cat_menu.get()
        name = self.concept_name_menu.get()
        if cat == "--" or name == "--":
            return
        preset = CONCEPT_PRESETS.get(cat, {}).get(name)
        if not preset:
            return
        variations = preset["variations"]
        if len(variations) <= 1:
            self._set_concept_text(variations[0])
            return
        candidates = [v for v in variations if v != self._last_concept_variation]
        if not candidates:
            candidates = variations
        variation = random.choice(candidates)
        self._last_concept_variation = variation
        self._set_concept_text(variation)

    def _on_male_preset_changed(self, value):
        """男性プリセット選択時にカスタム入力をクリア"""
        if value != "おまかせ":
            self.male_custom_field.delete(0, "end")

    def get_male_tags(self) -> str:
        """現在のGUI状態から男性SDタグ文字列を返す"""
        custom = self.male_custom_field.get().strip() if hasattr(self, 'male_custom_field') else ""
        if custom:
            base = parse_male_description(custom)
        else:
            preset_key = self.male_preset_combo.get() if hasattr(self, 'male_preset_combo') else "おまかせ"
            base = MALE_PRESETS.get(preset_key, "")

        # 髪型・髪色・肌色の追加タグ
        extra = []
        if hasattr(self, 'male_hair_style_combo'):
            v = MALE_HAIR_STYLE_OPTIONS.get(self.male_hair_style_combo.get(), "")
            if v:
                extra.append(v)
        if hasattr(self, 'male_hair_color_combo'):
            v = MALE_HAIR_COLOR_OPTIONS.get(self.male_hair_color_combo.get(), "")
            if v:
                extra.append(v)
        if hasattr(self, 'male_skin_color_combo'):
            v = MALE_SKIN_COLOR_OPTIONS.get(self.male_skin_color_combo.get(), "")
            if v:
                extra.append(v)

        if extra:
            parts = [t.strip() for t in base.split(",") if t.strip()] if base else []
            parts.extend(extra)
            return ", ".join(parts)
        return base

    def _set_characters_text(self, value: str):
        """登場人物テキストをパースして個別フィールドに設定"""
        # 全フィールドクリア
        for f in [self.char_name_field, self.char_personality_field,
                  self.char_first_person_field, self.char_endings_field,
                  self.char_appearance_field]:
            f.delete(0, "end")

        if not value:
            return

        import re as _re
        lines = value.strip().split("\n")
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith("【") and "】" in line:
                m = _re.match(r"【(.+?)】(?:（(.+?)）)?", line)
                if m:
                    name = m.group(1)
                    work = m.group(2) or ""
                    self.char_name_field.delete(0, "end")
                    self.char_name_field.insert(0, f"{name}（{work}）" if work else name)
                else:
                    self.char_name_field.delete(0, "end")
                    self.char_name_field.insert(0, line)
            elif line.startswith("性格:"):
                self.char_personality_field.delete(0, "end")
                self.char_personality_field.insert(0, line.split(":", 1)[1].strip())
            elif line.startswith("一人称:"):
                self.char_first_person_field.delete(0, "end")
                self.char_first_person_field.insert(0, line.split(":", 1)[1].strip())
            elif line.startswith("語尾:"):
                self.char_endings_field.delete(0, "end")
                self.char_endings_field.insert(0, line.split(":", 1)[1].strip())
            elif line.startswith("外見:"):
                self.char_appearance_field.delete(0, "end")
                self.char_appearance_field.insert(0, line.split(":", 1)[1].strip())

    def _get_characters_text(self) -> str:
        """個別フィールドからパイプライン用テキストを組み立て"""
        name = self.char_name_field.get().strip()
        if not name:
            return ""
        personality = self.char_personality_field.get().strip()
        first_person = self.char_first_person_field.get().strip()
        endings = self.char_endings_field.get().strip()
        appearance = self.char_appearance_field.get().strip()

        # 名前行: 【名前】（作品名）形式に復元
        if "（" in name and name.endswith("）"):
            # 既に「名前（作品名）」形式の場合
            parts = name.split("（", 1)
            name_line = f"【{parts[0]}】（{parts[1]}"
        else:
            name_line = f"【{name}】"

        lines = [name_line]
        if personality:
            lines.append(f"性格: {personality}")
        if first_person:
            lines.append(f"一人称: {first_person}")
        if endings:
            lines.append(f"語尾: {endings}")
        if appearance:
            lines.append(f"外見: {appearance}")
        return "\n".join(lines)

    def _get_characters_fields(self) -> dict:
        """個別フィールドの値を構造化データとして取得"""
        return {
            "name": self.char_name_field.get().strip(),
            "personality": self.char_personality_field.get().strip(),
            "first_person": self.char_first_person_field.get().strip(),
            "endings": self.char_endings_field.get().strip(),
            "appearance": self.char_appearance_field.get().strip(),
        }

    def _set_characters_fields(self, fields: dict):
        """構造化データから個別フィールドに設定"""
        for f in [self.char_name_field, self.char_personality_field,
                  self.char_first_person_field, self.char_endings_field,
                  self.char_appearance_field]:
            f.delete(0, "end")
        if fields.get("name"):
            self.char_name_field.insert(0, fields["name"])
        if fields.get("personality"):
            self.char_personality_field.insert(0, fields["personality"])
        if fields.get("first_person"):
            self.char_first_person_field.insert(0, fields["first_person"])
        if fields.get("endings"):
            self.char_endings_field.insert(0, fields["endings"])
        if fields.get("appearance"):
            self.char_appearance_field.insert(0, fields["appearance"])

    def _set_api_field(self, value: str):
        """APIフィールドを設定"""
        self.api_field.delete(0, "end")
        if value:
            self.api_field.insert(0, value)

    def load_saved_config(self):
        # プロバイダー復元
        if self.config_data.get("provider"):
            self._provider_var.set(self.config_data["provider"])
            self._on_provider_changed()
        if self.config_data.get("api_key"):
            self._set_api_field(self.config_data["api_key"])
        if self.config_data.get("concept"):
            self._set_concept_text(self.config_data["concept"])
        # characters_fields優先、fallbackでテキストパース
        if self.config_data.get("characters_fields"):
            self._set_characters_fields(self.config_data["characters_fields"])
        elif self.config_data.get("characters"):
            self._set_characters_text(self.config_data["characters"])
        if self.config_data.get("num_scenes"):
            self.scenes_entry.delete(0, "end")
            self.scenes_entry.insert(0, str(self.config_data["num_scenes"]))
        if self.config_data.get("theme_jp"):
            self.theme_combo.set(self.config_data["theme_jp"])
        # work_type is always "二次創作" (original character creation removed)
        if self.config_data.get("story_structure"):
            ss = self.config_data["story_structure"]
            self.prologue_slider.set(ss.get("prologue", 10))
            self.epilogue_slider.set(ss.get("epilogue", 10))
            preset_name = ss.get("preset", "標準バランス (10/80/10)")
            if preset_name in STRUCTURE_PRESETS:
                self.structure_preset.set(preset_name)
            self._update_structure_bar()

        # 男性キャラ・シーン環境の復元
        if "male_preset" in self.config_data and hasattr(self, 'male_preset_combo'):
            self.male_preset_combo.set(self.config_data["male_preset"])
        if "male_custom" in self.config_data and hasattr(self, 'male_custom_field'):
            _mc = self.config_data["male_custom"]
            if _mc:  # 値がある時だけ操作（空の場合はプレースホルダー維持）
                self.male_custom_field.delete(0, "end")
                self.male_custom_field.insert(0, _mc)
        if "male_faceless" in self.config_data and hasattr(self, 'male_faceless_var'):
            self.male_faceless_var.set(self.config_data["male_faceless"])
        if "male_hair_style" in self.config_data and hasattr(self, 'male_hair_style_combo'):
            self.male_hair_style_combo.set(self.config_data["male_hair_style"])
        if "male_hair_color" in self.config_data and hasattr(self, 'male_hair_color_combo'):
            self.male_hair_color_combo.set(self.config_data["male_hair_color"])
        if "male_skin_color" in self.config_data and hasattr(self, 'male_skin_color_combo'):
            self.male_skin_color_combo.set(self.config_data["male_skin_color"])
        if "time_of_day" in self.config_data and hasattr(self, 'time_of_day_combo'):
            self.time_of_day_combo.set(self.config_data["time_of_day"])
        if "location_type" in self.config_data and hasattr(self, 'location_type_combo'):
            self.location_type_combo.set(self.config_data["location_type"])

        # SD設定の復元
        if "sd_quality_mode" in self.config_data:
            self.sd_quality_mode_var.set(self.config_data["sd_quality_mode"])
            self._on_sd_quality_mode_changed()
        if self.config_data.get("sd_quality_custom"):
            self.sd_quality_custom_entry.configure(state="normal")
            self.sd_quality_custom_entry.delete(0, "end")
            self.sd_quality_custom_entry.insert(0, self.config_data["sd_quality_custom"])
            if self.sd_quality_mode_var.get() == "auto":
                self.sd_quality_custom_entry.configure(state="disabled")
        if self.config_data.get("sd_prefix_tags"):
            self.sd_prefix_text.delete("1.0", "end")
            self.sd_prefix_text.insert("1.0", self.config_data["sd_prefix_tags"])
            self._auto_resize_textbox(self.sd_prefix_text, 100, 1200)
        if self.config_data.get("sd_suffix_tags"):
            self.sd_suffix_text.delete("1.0", "end")
            self.sd_suffix_text.insert("1.0", self.config_data["sd_suffix_tags"])
            self._auto_resize_textbox(self.sd_suffix_text, 100, 1200)
        # ネガティブプロンプト設定の復元
        if self.config_data.get("sd_neg_base") and hasattr(self, 'sd_neg_base_text'):
            self.sd_neg_base_text.delete("1.0", "end")
            self.sd_neg_base_text.insert("1.0", self.config_data["sd_neg_base"])
            self._auto_resize_textbox(self.sd_neg_base_text, 60, 400)
        if self.config_data.get("sd_neg_prefix") and hasattr(self, 'sd_neg_prefix_text'):
            self.sd_neg_prefix_text.delete("1.0", "end")
            self.sd_neg_prefix_text.insert("1.0", self.config_data["sd_neg_prefix"])
            self._auto_resize_textbox(self.sd_neg_prefix_text, 60, 400)
        if self.config_data.get("sd_neg_suffix") and hasattr(self, 'sd_neg_suffix_text'):
            self.sd_neg_suffix_text.delete("1.0", "end")
            self.sd_neg_suffix_text.insert("1.0", self.config_data["sd_neg_suffix"])
            self._auto_resize_textbox(self.sd_neg_suffix_text, 60, 400)
        # v8.7: 品質優先モードの復元
        if self.config_data.get("quality_priority") and hasattr(self, 'quality_priority_var'):
            self.quality_priority_var.set(True)

        # 初期コスト予測を表示
        self.after(100, self.update_cost_preview)

    # --- SDプロンプト テキスト欄動的リサイズ ---
    @staticmethod
    def _auto_resize_textbox(textbox, min_h: int = 100, max_h: int = 1200):
        """テキスト内容に応じてCTkTextboxの高さを自動調整（折り返し考慮）"""
        try:
            content = textbox.get("1.0", "end-1c")
            if not content.strip():
                textbox.configure(height=min_h)
                return
            # ウィジェット幅から1行あたり文字数を推定
            w_px = textbox.winfo_width()
            if w_px < 50:
                w_px = 500  # 初期化前のフォールバック
            chars_per_line = max(1, w_px // 8)
            # 各行の折り返しを考慮した表示行数を計算
            visual_lines = 0
            for line in content.split("\n"):
                if len(line) <= chars_per_line:
                    visual_lines += 1
                else:
                    visual_lines += (len(line) + chars_per_line - 1) // chars_per_line
            # 1行≒20px + 上下パディング20px
            needed = visual_lines * 20 + 20
            new_h = max(min_h, min(max_h, needed))
            textbox.configure(height=new_h)
        except Exception:
            pass

    # --- SDプロンプト設定コールバック ---
    def _on_provider_changed(self):
        """プロバイダー切替時のUI更新"""
        self.update_cost_preview()

    def _on_sd_quality_mode_changed(self):
        """auto/manual切替でカスタム入力欄のstate変更"""
        if self.sd_quality_mode_var.get() == "manual":
            self.sd_quality_custom_entry.configure(state="normal")
            # 自動テキストが残っていたら消去してユーザー入力に備える
            if self.sd_quality_custom_entry.get() == QUALITY_POSITIVE_TAGS:
                self.sd_quality_custom_entry.delete(0, "end")
        else:
            # カスタム→自動: 空欄なら自動タグを再表示
            self.sd_quality_custom_entry.configure(state="normal")
            current = self.sd_quality_custom_entry.get().strip()
            if not current:
                self.sd_quality_custom_entry.delete(0, "end")
                self.sd_quality_custom_entry.insert(0, QUALITY_POSITIVE_TAGS)
            self.sd_quality_custom_entry.configure(state="disabled")

    def _setup_png_drop(self):
        """windndによるドラッグ&ドロップのフック（ウィンドウ表示後に呼ぶ）"""
        if not WINDND_AVAILABLE:
            return
        try:
            # force_unicode=True: DragQueryFileW使用でUnicode strを直接受け取る
            # （デフォルトのANSI版だと日本語パスがCP932 bytesになり文字化けする）
            windnd.hook_dropfiles(self.png_drop_frame, func=self._on_png_files_dropped, force_unicode=True)
        except Exception as e:
            log_message(f"ドラッグ&ドロップ初期化失敗: {e}")

    def _on_png_files_dropped(self, file_list):
        """windndからのドロップコールバック（force_unicode=Trueでstr型リスト）"""
        _IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff", ".tif")
        for f in file_list:
            # force_unicode=True → str型、万が一bytes時はシステムエンコーディングで復元
            if isinstance(f, bytes):
                import locale
                for enc in (locale.getpreferredencoding(False), "utf-8"):
                    try:
                        path = f.decode(enc)
                        break
                    except (UnicodeDecodeError, LookupError):
                        continue
                else:
                    path = f.decode("utf-8", errors="replace")
            else:
                path = str(f)
            if path.lower().endswith(_IMAGE_EXTS):
                self._load_png_info(path)
                return
        # 画像ファイルが見つからない場合
        if hasattr(self, 'snackbar'):
            self.snackbar.show("対応する画像ファイルが見つかりません", type="error")

    def _on_png_info_load(self):
        """ファイルダイアログから画像選択→パース→プレビュー表示"""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            filetypes=[("画像ファイル", "*.png *.jpg *.jpeg *.webp"), ("PNG", "*.png"), ("JPEG", "*.jpg *.jpeg"), ("WebP", "*.webp"), ("全てのファイル", "*.*")]
        )
        if not path:
            return
        self._load_png_info(path)

    def _load_png_info(self, path: str):
        """画像ファイルからSD情報を読み取りプレビューに表示（共通処理）"""
        result = parse_png_info(path)
        self._png_info_result = result  # 後から参照用に保存
        filename = Path(path).name
        self.png_filename_label.configure(text=filename)
        # ドロップゾーンのヒントを更新
        self.png_drop_hint_label.configure(
            text="別の画像をドロップで差替え" if WINDND_AVAILABLE else "クリックで別の画像を選択"
        )
        self.png_preview_text.configure(state="normal")
        self.png_preview_text.delete("1.0", "end")
        if "error" in result:
            self.png_preview_text.insert("1.0", result["error"])
            self.png_drop_frame.configure(border_color="#B3261E")  # error red
        else:
            positive = result.get("positive", "")
            raw = result.get("raw", "")
            log_message(f"PNG Info positive[:{min(80,len(positive))}]: {repr(positive[:80])}")
            if raw != positive:
                log_message(f"PNG Info raw[:{min(80,len(raw))}]: {repr(raw[:80])}")
            self.png_preview_text.insert("1.0", positive if positive else "(情報なし)")
            self.png_drop_frame.configure(border_color=MaterialColors.PRIMARY)  # success
        self.png_preview_text.configure(state="disabled")
        # 内容に応じて高さを自動調整
        self._auto_resize_textbox(self.png_preview_text, min_h=120, max_h=1200)

    def _apply_png_to_prefix(self):
        """PNG Infoのpositiveプロンプトをプレフィックスに適用"""
        result = getattr(self, '_png_info_result', None)
        if not result or "error" in result:
            return
        content = result.get("positive", "").strip()
        if not content:
            return
        existing = self.sd_prefix_text.get("1.0", "end-1c").strip()
        if existing:
            self.sd_prefix_text.delete("1.0", "end")
            self.sd_prefix_text.insert("1.0", f"{existing}, {content}")
        else:
            self.sd_prefix_text.delete("1.0", "end")
            self.sd_prefix_text.insert("1.0", content)
        self._auto_resize_textbox(self.sd_prefix_text, min_h=100, max_h=1200)
        if hasattr(self, 'snackbar'):
            self.snackbar.show("プレフィックスに適用しました", type="success")

    def _apply_png_to_suffix(self):
        """PNG Infoのpositiveプロンプトをサフィックスに適用"""
        result = getattr(self, '_png_info_result', None)
        if not result or "error" in result:
            return
        content = result.get("positive", "").strip()
        if not content:
            return
        existing = self.sd_suffix_text.get("1.0", "end-1c").strip()
        if existing:
            self.sd_suffix_text.delete("1.0", "end")
            self.sd_suffix_text.insert("1.0", f"{existing}, {content}")
        else:
            self.sd_suffix_text.delete("1.0", "end")
            self.sd_suffix_text.insert("1.0", content)
        self._auto_resize_textbox(self.sd_suffix_text, min_h=100, max_h=1200)
        if hasattr(self, 'snackbar'):
            self.snackbar.show("サフィックスに適用しました", type="success")

    def _apply_png_to_negative(self):
        """PNG Infoのnegativeプロンプトをネガティブ品質ベースに適用"""
        result = getattr(self, '_png_info_result', None)
        if not result or "error" in result:
            return
        content = result.get("negative", "").strip()
        if not content:
            # negativeが無い場合はpositiveをネガティブプレフィックスに適用
            content = result.get("positive", "").strip()
            if not content:
                return
            target = self.sd_neg_prefix_text
            label = "ネガティブプレフィックス"
        else:
            target = self.sd_neg_base_text
            label = "ネガティブ品質ベース"
        existing = target.get("1.0", "end-1c").strip()
        if existing:
            target.delete("1.0", "end")
            target.insert("1.0", f"{existing}, {content}")
        else:
            target.delete("1.0", "end")
            target.insert("1.0", content)
        self._auto_resize_textbox(target, min_h=60, max_h=400)
        if hasattr(self, 'snackbar'):
            self.snackbar.show(f"{label}に適用しました", type="success")

    def update_cost_preview(self, event=None):
        """シーン数に基づいてコスト予測を更新"""
        try:
            num_scenes = int(self.scenes_entry.get())
            if num_scenes < 1:
                num_scenes = 1
            elif num_scenes > 500:
                num_scenes = 500

            est = estimate_cost(num_scenes)
            self.cost_preview_label.configure(
                text=f"予想コスト: ${est['estimated_usd']:.4f} (約¥{est['estimated_jpy']:.1f}) | "
                     f"Haiku: ~{est['haiku_tokens']:,}, Sonnet: ~{est['sonnet_tokens']:,}, Opus: ~{est['opus_tokens']:,}"
            )
        except ValueError:
            self.cost_preview_label.configure(
                text="予想コスト: シーン数を入力してください"
            )

    def _build_structure_bar(self, parent):
        """ストーリー構成バーUIを構築"""
        structure_frame = ctk.CTkFrame(parent, fg_color="transparent")
        structure_frame.pack(fill="x", padx=20, pady=(8, 4))

        # ヘッダー行: ラベル + プリセット
        header_row = ctk.CTkFrame(structure_frame, fg_color="transparent")
        header_row.pack(fill="x")
        ctk.CTkLabel(
            header_row, text="ストーリー構成",
            font=ctk.CTkFont(size=13, weight="bold"), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")

        self.structure_preset = ctk.CTkOptionMenu(
            header_row, values=list(STRUCTURE_PRESETS.keys()), height=30,
            font=ctk.CTkFont(size=12), width=200,
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            button_color=MaterialColors.PRIMARY,
            dropdown_fg_color=MaterialColors.SURFACE,
            text_color=MaterialColors.ON_SURFACE,
            dropdown_text_color=MaterialColors.ON_SURFACE,
            command=self._on_structure_preset_changed
        )
        self.structure_preset.pack(side="right")
        self.structure_preset.set("標準バランス (10/80/10)")

        # ビジュアルバー（Canvas）
        bar_frame = ctk.CTkFrame(structure_frame, fg_color="transparent")
        bar_frame.pack(fill="x", pady=(8, 4))
        self.structure_canvas = tk.Canvas(
            bar_frame, height=28, highlightthickness=0,
            bg=MaterialColors.SURFACE_CONTAINER_LOW
        )
        self.structure_canvas.pack(fill="x")
        self.structure_canvas.bind("<Configure>", lambda e: self._update_structure_bar())

        # スライダー行
        slider_frame = ctk.CTkFrame(structure_frame, fg_color="transparent")
        slider_frame.pack(fill="x", pady=(4, 0))

        # プロローグスライダー
        pro_row = ctk.CTkFrame(slider_frame, fg_color="transparent")
        pro_row.pack(fill="x", pady=2)
        ctk.CTkLabel(
            pro_row, text="プロローグ", width=90, anchor="w",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        self.prologue_slider = ctk.CTkSlider(
            pro_row, from_=5, to=30, number_of_steps=5,
            fg_color=MaterialColors.SURFACE_CONTAINER,
            progress_color=MaterialColors.SECONDARY,
            button_color=MaterialColors.PRIMARY,
            button_hover_color=MaterialColors.PRIMARY_CONTAINER,
            command=self._on_structure_slider_changed
        )
        self.prologue_slider.pack(side="left", fill="x", expand=True, padx=(8, 8))
        self.prologue_slider.set(10)
        self.prologue_pct_label = ctk.CTkLabel(
            pro_row, text="10%", width=40,
            font=ctk.CTkFont(size=12, weight="bold"), text_color=MaterialColors.SECONDARY
        )
        self.prologue_pct_label.pack(side="left")

        # エピローグスライダー
        epi_row = ctk.CTkFrame(slider_frame, fg_color="transparent")
        epi_row.pack(fill="x", pady=2)
        ctk.CTkLabel(
            epi_row, text="エピローグ", width=90, anchor="w",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        self.epilogue_slider = ctk.CTkSlider(
            epi_row, from_=5, to=20, number_of_steps=3,
            fg_color=MaterialColors.SURFACE_CONTAINER,
            progress_color=MaterialColors.SECONDARY,
            button_color=MaterialColors.PRIMARY,
            button_hover_color=MaterialColors.PRIMARY_CONTAINER,
            command=self._on_structure_slider_changed
        )
        self.epilogue_slider.pack(side="left", fill="x", expand=True, padx=(8, 8))
        self.epilogue_slider.set(10)
        self.epilogue_pct_label = ctk.CTkLabel(
            epi_row, text="10%", width=40,
            font=ctk.CTkFont(size=12, weight="bold"), text_color=MaterialColors.SECONDARY
        )
        self.epilogue_pct_label.pack(side="left")

        # 本編（自動算出）ラベル
        main_row = ctk.CTkFrame(slider_frame, fg_color="transparent")
        main_row.pack(fill="x", pady=2)
        ctk.CTkLabel(
            main_row, text="→ 本編:", width=90, anchor="w",
            font=ctk.CTkFont(size=12), text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left")
        self.main_pct_label = ctk.CTkLabel(
            main_row, text="自動算出 80%",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=MaterialColors.PRIMARY
        )
        self.main_pct_label.pack(side="left", padx=(8, 0))

    def _update_structure_bar(self):
        """構成バーの再描画"""
        canvas = self.structure_canvas
        w = canvas.winfo_width()
        h = canvas.winfo_height()
        if w < 10:
            return

        canvas.delete("all")

        prologue = int(round(self.prologue_slider.get()))
        epilogue = int(round(self.epilogue_slider.get()))
        main_pct = 100 - prologue - epilogue

        # パーセンテージラベル更新
        self.prologue_pct_label.configure(text=f"{prologue}%")
        self.epilogue_pct_label.configure(text=f"{epilogue}%")
        self.main_pct_label.configure(text=f"自動算出 {main_pct}%")

        # セグメント描画用データ
        segments = [
            (prologue / 100, MaterialColors.SECONDARY, f"プロローグ {prologue}%"),
            (main_pct / 100, MaterialColors.PRIMARY, f"本編 {main_pct}%"),
            (epilogue / 100, MaterialColors.TERTIARY, f"エピローグ {epilogue}%"),
        ]

        r = 6  # 角丸半径
        x = 0
        for i, (ratio, color, label_text) in enumerate(segments):
            seg_w = max(2, ratio * w)
            x1, x2 = x, x + seg_w

            # 左端の角丸
            if i == 0:
                canvas.create_rectangle(x1, 0, x2, h, fill=color, outline="")
                canvas.create_arc(x1, 0, x1 + r * 2, r * 2, start=90, extent=90, fill=color, outline="")
                canvas.create_arc(x1, h - r * 2, x1 + r * 2, h, start=180, extent=90, fill=color, outline="")
            # 右端の角丸
            elif i == len(segments) - 1:
                canvas.create_rectangle(x1, 0, x2, h, fill=color, outline="")
                canvas.create_arc(x2 - r * 2, 0, x2, r * 2, start=0, extent=90, fill=color, outline="")
                canvas.create_arc(x2 - r * 2, h - r * 2, x2, h, start=270, extent=90, fill=color, outline="")
            else:
                canvas.create_rectangle(x1, 0, x2, h, fill=color, outline="")

            # テキストラベル（セグメントが狭すぎなければ）
            mid_x = (x1 + x2) / 2
            if seg_w > 60:
                canvas.create_text(mid_x, h / 2, text=label_text, fill="#FFFFFF",
                                   font=("Noto Sans JP", 9, "bold"))

            x = x2

    def _on_structure_preset_changed(self, value):
        """プリセット選択時"""
        preset = STRUCTURE_PRESETS.get(value)
        if preset is None:
            return
        self.prologue_slider.set(preset["prologue"])
        self.epilogue_slider.set(preset["epilogue"])
        self._update_structure_bar()

    def _on_structure_slider_changed(self, _value=None):
        """スライダー変更時"""
        self._update_structure_bar()
        # プリセットと一致するか確認、しなければ「カスタム」に
        prologue = int(round(self.prologue_slider.get()))
        epilogue = int(round(self.epilogue_slider.get()))
        matched = False
        for name, preset in STRUCTURE_PRESETS.items():
            if preset and preset["prologue"] == prologue and preset["epilogue"] == epilogue:
                self.structure_preset.set(name)
                matched = True
                break
        if not matched:
            self.structure_preset.set("カスタム")

    def _get_story_structure(self) -> dict:
        """現在のストーリー構成比率を取得"""
        prologue = int(round(self.prologue_slider.get()))
        epilogue = int(round(self.epilogue_slider.get()))
        main_pct = 100 - prologue - epilogue
        return {"prologue": prologue, "main": main_pct, "epilogue": epilogue}

    def save_settings(self):
        """設定を保存"""
        theme_jp = self.theme_combo.get()
        self.config_data = {
            "provider": self._provider_var.get(),
            "api_key": self.api_field.get(),
            "concept": self.concept_text.get("1.0", "end-1c"),
            "characters": self._get_characters_text(),
            "characters_fields": self._get_characters_fields(),
            "other_characters": self.other_chars_text.get("1.0", "end-1c") if hasattr(self, "other_chars_text") else "",
            "num_scenes": int(self.scenes_entry.get() or "10"),
            "theme_jp": theme_jp,
            "theme": THEME_OPTIONS.get(theme_jp, ""),
            "work_type": "二次創作",
            "story_structure": {
                "prologue": int(round(self.prologue_slider.get())),
                "epilogue": int(round(self.epilogue_slider.get())),
                "preset": self.structure_preset.get(),
            },
            "concept_preset_category": self.concept_cat_menu.get(),
            "concept_preset_name": self.concept_name_menu.get(),
            "male_preset": self.male_preset_combo.get() if hasattr(self, 'male_preset_combo') else "おまかせ",
            "male_custom": self.male_custom_field.get() if hasattr(self, 'male_custom_field') else "",
            "male_faceless": self.male_faceless_var.get() if hasattr(self, 'male_faceless_var') else True,
            "male_hair_style": self.male_hair_style_combo.get() if hasattr(self, 'male_hair_style_combo') else "おまかせ",
            "male_hair_color": self.male_hair_color_combo.get() if hasattr(self, 'male_hair_color_combo') else "おまかせ",
            "male_skin_color": self.male_skin_color_combo.get() if hasattr(self, 'male_skin_color_combo') else "おまかせ",
            "time_of_day": self.time_of_day_combo.get() if hasattr(self, 'time_of_day_combo') else "おまかせ",
            "location_type": self.location_type_combo.get() if hasattr(self, 'location_type_combo') else "おまかせ",
            "sd_quality_mode": self.sd_quality_mode_var.get() if hasattr(self, 'sd_quality_mode_var') else "auto",
            "sd_quality_custom": (self.sd_quality_custom_entry.get() if self.sd_quality_mode_var.get() == "manual" else "") if hasattr(self, 'sd_quality_custom_entry') else "",
            "sd_prefix_tags": self.sd_prefix_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_prefix_text') else "",
            "sd_suffix_tags": self.sd_suffix_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_suffix_text') else "",
            "sd_neg_base": self.sd_neg_base_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_neg_base_text') else "",
            "sd_neg_prefix": self.sd_neg_prefix_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_neg_prefix_text') else "",
            "sd_neg_suffix": self.sd_neg_suffix_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_neg_suffix_text') else "",
            "quality_priority": self.quality_priority_var.get() if hasattr(self, 'quality_priority_var') else False,
        }
        save_config(self.config_data)
        self.snackbar.show("設定を保存しました", type="success")
        log_message("設定を保存しました")

    def get_current_config(self) -> dict:
        """現在の設定を辞書として取得"""
        theme_jp = self.theme_combo.get()
        return {
            "api_key": self.api_field.get(),
            "concept": self.concept_text.get("1.0", "end-1c"),
            "characters": self._get_characters_text(),
            "characters_fields": self._get_characters_fields(),
            "other_characters": self.other_chars_text.get("1.0", "end-1c") if hasattr(self, "other_chars_text") else "",
            "num_scenes": int(self.scenes_entry.get() or "10"),
            "theme_jp": theme_jp,
            "theme": THEME_OPTIONS.get(theme_jp, ""),
            "work_title": self._work_title_val,
            "char_name": self._char_name_val,
            "work_type": "二次創作",
            "male_preset": self.male_preset_combo.get() if hasattr(self, 'male_preset_combo') else "おまかせ",
            "male_custom": self.male_custom_field.get() if hasattr(self, 'male_custom_field') else "",
            "male_faceless": self.male_faceless_var.get() if hasattr(self, 'male_faceless_var') else True,
            "male_hair_style": self.male_hair_style_combo.get() if hasattr(self, 'male_hair_style_combo') else "おまかせ",
            "male_hair_color": self.male_hair_color_combo.get() if hasattr(self, 'male_hair_color_combo') else "おまかせ",
            "male_skin_color": self.male_skin_color_combo.get() if hasattr(self, 'male_skin_color_combo') else "おまかせ",
            "time_of_day": self.time_of_day_combo.get() if hasattr(self, 'time_of_day_combo') else "おまかせ",
            "location_type": self.location_type_combo.get() if hasattr(self, 'location_type_combo') else "おまかせ",
            "sd_quality_mode": self.sd_quality_mode_var.get() if hasattr(self, 'sd_quality_mode_var') else "auto",
            "sd_quality_custom": (self.sd_quality_custom_entry.get() if self.sd_quality_mode_var.get() == "manual" else "") if hasattr(self, 'sd_quality_custom_entry') else "",
            "sd_prefix_tags": self.sd_prefix_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_prefix_text') else "",
            "sd_suffix_tags": self.sd_suffix_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_suffix_text') else "",
            "sd_neg_base": self.sd_neg_base_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_neg_base_text') else "",
            "sd_neg_prefix": self.sd_neg_prefix_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_neg_prefix_text') else "",
            "sd_neg_suffix": self.sd_neg_suffix_text.get("1.0", "end-1c").strip() if hasattr(self, 'sd_neg_suffix_text') else "",
        }

    def apply_config(self, config: dict):
        """設定を画面に反映"""
        if config.get("api_key"):
            self._set_api_field(config["api_key"])
        if config.get("concept"):
            self._set_concept_text(config["concept"])
        # characters_fields優先、fallbackでテキストパース
        if config.get("characters_fields"):
            self._set_characters_fields(config["characters_fields"])
        elif config.get("characters"):
            self._set_characters_text(config["characters"])
        if hasattr(self, "other_chars_text") and "other_characters" in config:
            self.other_chars_text.delete("1.0", "end")
            self.other_chars_text.insert("1.0", config.get("other_characters", ""))
        if config.get("num_scenes"):
            self.scenes_entry.delete(0, "end")
            self.scenes_entry.insert(0, str(config["num_scenes"]))
        if config.get("theme_jp"):
            self.theme_combo.set(config["theme_jp"])
        if config.get("work_title"):
            self._work_title_val = config["work_title"]
        if config.get("char_name"):
            self._char_name_val = config["char_name"]
        if config.get("story_structure"):
            ss = config["story_structure"]
            self.prologue_slider.set(ss.get("prologue", 10))
            self.epilogue_slider.set(ss.get("epilogue", 10))
            preset_name = ss.get("preset", "標準バランス (10/80/10)")
            if preset_name in STRUCTURE_PRESETS:
                self.structure_preset.set(preset_name)
            self._update_structure_bar()
        # コンセプトプリセット復元
        if config.get("concept_preset_category"):
            cat = config["concept_preset_category"]
            self.concept_cat_menu.set(cat)
            self._on_concept_category_changed(cat)
            if config.get("concept_preset_name"):
                self.concept_name_menu.set(config["concept_preset_name"])
        # 男性キャラ・シーン環境の復元
        if "male_preset" in config and hasattr(self, 'male_preset_combo'):
            self.male_preset_combo.set(config["male_preset"])
        if "male_custom" in config and hasattr(self, 'male_custom_field'):
            _mc = config["male_custom"]
            if _mc:  # 値がある時だけ操作（空の場合はプレースホルダー維持）
                self.male_custom_field.delete(0, "end")
                self.male_custom_field.insert(0, _mc)
        if "male_faceless" in config and hasattr(self, 'male_faceless_var'):
            self.male_faceless_var.set(config["male_faceless"])
        if "male_hair_style" in config and hasattr(self, 'male_hair_style_combo'):
            self.male_hair_style_combo.set(config["male_hair_style"])
        if "male_hair_color" in config and hasattr(self, 'male_hair_color_combo'):
            self.male_hair_color_combo.set(config["male_hair_color"])
        if "male_skin_color" in config and hasattr(self, 'male_skin_color_combo'):
            self.male_skin_color_combo.set(config["male_skin_color"])
        if "time_of_day" in config and hasattr(self, 'time_of_day_combo'):
            self.time_of_day_combo.set(config["time_of_day"])
        if "location_type" in config and hasattr(self, 'location_type_combo'):
            self.location_type_combo.set(config["location_type"])
        # SD設定の復元
        if "sd_quality_mode" in config and hasattr(self, 'sd_quality_mode_var'):
            self.sd_quality_mode_var.set(config["sd_quality_mode"])
            self._on_sd_quality_mode_changed()
        if config.get("sd_quality_custom") and hasattr(self, 'sd_quality_custom_entry'):
            self.sd_quality_custom_entry.configure(state="normal")
            self.sd_quality_custom_entry.delete(0, "end")
            self.sd_quality_custom_entry.insert(0, config["sd_quality_custom"])
            if self.sd_quality_mode_var.get() == "auto":
                self.sd_quality_custom_entry.configure(state="disabled")
        if config.get("sd_prefix_tags") and hasattr(self, 'sd_prefix_text'):
            self.sd_prefix_text.delete("1.0", "end")
            self.sd_prefix_text.insert("1.0", config["sd_prefix_tags"])
            self._auto_resize_textbox(self.sd_prefix_text, 100, 1200)
        if config.get("sd_suffix_tags") and hasattr(self, 'sd_suffix_text'):
            self.sd_suffix_text.delete("1.0", "end")
            self.sd_suffix_text.insert("1.0", config["sd_suffix_tags"])
            self._auto_resize_textbox(self.sd_suffix_text, 100, 1200)
        # ネガティブプロンプト設定
        if config.get("sd_neg_base") and hasattr(self, 'sd_neg_base_text'):
            self.sd_neg_base_text.delete("1.0", "end")
            self.sd_neg_base_text.insert("1.0", config["sd_neg_base"])
            self._auto_resize_textbox(self.sd_neg_base_text, 60, 400)
        if config.get("sd_neg_prefix") and hasattr(self, 'sd_neg_prefix_text'):
            self.sd_neg_prefix_text.delete("1.0", "end")
            self.sd_neg_prefix_text.insert("1.0", config["sd_neg_prefix"])
            self._auto_resize_textbox(self.sd_neg_prefix_text, 60, 400)
        if config.get("sd_neg_suffix") and hasattr(self, 'sd_neg_suffix_text'):
            self.sd_neg_suffix_text.delete("1.0", "end")
            self.sd_neg_suffix_text.insert("1.0", config["sd_neg_suffix"])
            self._auto_resize_textbox(self.sd_neg_suffix_text, 60, 400)
        self.update_cost_preview()

    def refresh_profile_list(self):
        """プロファイル一覧を更新"""
        profiles = ["（新規）"] + get_profile_list()
        self.profile_combo.configure(values=profiles)

    def on_profile_selected(self, choice: str):
        """プロファイル選択時"""
        if choice != "（新規）":
            self.profile_name_entry.delete(0, "end")
            self.profile_name_entry.insert(0, choice)

    def save_current_profile(self):
        """現在の設定をプロファイルとして保存"""
        name = self.profile_name_entry.get().strip()
        if not name:
            self.snackbar.show("プロファイル名を入力してください", type="error")
            return
        
        # 上書き確認
        if name in get_profile_list():
            # 既存プロファイルを上書き
            pass  # 確認ダイアログは省略、直接上書き
        
        config = self.get_current_config()
        save_profile(name, config)
        self.refresh_profile_list()
        self.profile_combo.set(name)
        self.snackbar.show(f"プロファイル '{name}' を保存しました", type="success")

    def load_selected_profile(self):
        """選択したプロファイルを読み込み"""
        name = self.profile_combo.get()
        if name == "（新規）":
            self.snackbar.show("プロファイルを選択してください", type="warning")
            return
        
        config = load_profile(name)
        if config:
            self.apply_config(config)
            self.profile_name_entry.delete(0, "end")
            self.profile_name_entry.insert(0, name)
            self.snackbar.show(f"プロファイル '{name}' を読み込みました", type="success")
            self.log(f"プロファイル読込: {name}")
        else:
            self.snackbar.show(f"プロファイル '{name}' が見つかりません", type="error")

    def copy_selected_profile(self):
        """選択したプロファイルを複製"""
        src_name = self.profile_combo.get()
        if src_name == "（新規）":
            self.snackbar.show("コピー元のプロファイルを選択してください", type="warning")
            return
        
        dst_name = self.profile_name_entry.get().strip()
        if not dst_name:
            dst_name = f"{src_name}_copy"
        
        if dst_name == src_name:
            dst_name = f"{src_name}_copy"
        
        if copy_profile(src_name, dst_name):
            self.refresh_profile_list()
            self.profile_combo.set(dst_name)
            self.profile_name_entry.delete(0, "end")
            self.profile_name_entry.insert(0, dst_name)
            self.snackbar.show(f"'{src_name}' を '{dst_name}' にコピーしました", type="success")
        else:
            self.snackbar.show("コピーに失敗しました", type="error")

    def delete_selected_profile(self):
        """選択したプロファイルを削除"""
        name = self.profile_combo.get()
        if name == "（新規）":
            self.snackbar.show("削除するプロファイルを選択してください", type="warning")
            return
        
        if delete_profile(name):
            self.refresh_profile_list()
            self.profile_combo.set("（新規）")
            self.profile_name_entry.delete(0, "end")
            self.snackbar.show(f"プロファイル '{name}' を削除しました", type="success")
        else:
            self.snackbar.show("削除に失敗しました", type="error")

    def log(self, message: str):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        # see("end")スロットリング: 連続呼び出し時は200ms間隔でスクロール
        _now = time.time()
        if not hasattr(self, '_last_log_scroll') or (_now - self._last_log_scroll) > 0.2:
            self.log_text.see("end")
            self._last_log_scroll = _now
        else:
            # 遅延スクロール（最終行を見せるためにafter予約）
            if not hasattr(self, '_log_scroll_pending') or not self._log_scroll_pending:
                self._log_scroll_pending = True
                def _deferred_scroll():
                    self._log_scroll_pending = False
                    self.log_text.see("end")
                self.after(200, _deferred_scroll)
        log_message(message)

    def update_status(self, message: str):
        # ステータスアイコン自動切替
        if "[ERROR]" in message or "エラー" in message:
            self.status_icon_label.configure(text=Icons.XMARK)
        elif "[OK]" in message or "完了" in message:
            self.status_icon_label.configure(text=Icons.CHECK)
        elif "[WARN]" in message or "停止" in message:
            self.status_icon_label.configure(text=Icons.WARNING)
        elif "開始" in message or "生成中" in message:
            self.status_icon_label.configure(text=Icons.PLAY)
        else:
            self.status_icon_label.configure(text=Icons.CLOCK)
        self.status_label.configure(text=message)
        self.log(message)

        # フェーズインジケーター更新
        self.update_phase_indicator(message)

    def update_phase_indicator(self, message: str):
        """フェーズインジケーターを更新（5段階: 圧縮/あらすじ/分割/シーン生成/品質検証）"""
        import re

        def mark_done(*indices):
            for i in indices:
                if i < len(self.phase_labels):
                    pill, lbl = self.phase_labels[i]
                    pill.configure(fg_color=MaterialColors.SUCCESS)
                    lbl.configure(text_color=MaterialColors.ON_PRIMARY)

        def mark_active(idx):
            if idx < len(self.phase_labels):
                pill, lbl = self.phase_labels[idx]
                pill.configure(fg_color=MaterialColors.PRIMARY)
                lbl.configure(text_color=MaterialColors.ON_PRIMARY)

        def reset_all():
            for pill, lbl in self.phase_labels:
                pill.configure(fg_color=MaterialColors.SURFACE_CONTAINER)
                lbl.configure(text_color=MaterialColors.ON_SURFACE_VARIANT)

        # フェーズ検出（優先順位付き）
        new_phase = None

        if "[DONE]" in message or ("生成完了" in message and "シーン" in message):
            new_phase = "done"
        elif "Phase 5" in message or "品質検証" in message:
            new_phase = 4
        elif "Phase 1" in message and "圧縮" in message:
            new_phase = 0
        elif "[OK]" in message and "圧縮完了" in message:
            new_phase = 1  # Phase 1完了→Phase 2待ち
        elif "Phase 2" in message or "原案作成" in message:
            new_phase = 1
        elif "[OK]" in message and "原案完成" in message:
            new_phase = 2  # Phase 2完了→Phase 3待ち
        elif "Phase 3" in message or "シーン分割" in message:
            new_phase = 2
        elif "[OK]" in message and "分割完成" in message:
            new_phase = 3  # Phase 3完了→Phase 4待ち
        elif re.search(r'シーン \d+/\d+', message):
            new_phase = 3

        # 状態が変わらない場合はプログレスバーのみ更新
        if new_phase is None:
            # シーン進捗のみ更新（フェーズ表示はそのまま維持）
            match = re.search(r'(\d+)/(\d+)', message)
            if match and hasattr(self, '_current_phase') and self._current_phase == 3:
                current, total = int(match.group(1)), int(match.group(2))
                progress = 0.35 + (current / total) * 0.50
                self.progress.set(progress)
            return

        # フェーズ状態を保存
        self._current_phase = new_phase

        # 表示更新
        reset_all()
        if new_phase == "done":
            mark_done(0, 1, 2, 3, 4)
            self.progress.set(1.0)
        elif new_phase == 0:
            mark_active(0)
            self.progress.set(0.05)
        elif new_phase == 1:
            mark_done(0)
            mark_active(1)
            self.progress.set(0.12)
        elif new_phase == 2:
            mark_done(0, 1)
            mark_active(2)
            self.progress.set(0.20)
        elif new_phase == 3:
            mark_done(0, 1, 2)
            mark_active(3)
            match = re.search(r'(\d+)/(\d+)', message)
            if match:
                current, total = int(match.group(1)), int(match.group(2))
                progress = 0.35 + (current / total) * 0.50
                self.progress.set(progress)
            else:
                self.progress.set(0.30)
        elif new_phase == 4:
            mark_done(0, 1, 2, 3)
            mark_active(4)
            self.progress.set(0.90)

    def start_generation(self):
        if self.is_generating:
            return

        # APIキー取得
        api_key = self.api_field.get().strip()
        if not api_key:
                self.snackbar.show("Anthropic APIキーを入力してください", type="error")
                return

        concept = self.concept_text.get("1.0", "end-1c").strip()
        characters = self._get_characters_text().strip()
        other_chars = self.other_chars_text.get("1.0", "end-1c").strip() if hasattr(self, "other_chars_text") else ""

        if not concept:
            self.snackbar.show("コンセプトを入力してください", type="error")
            return

        try:
            num_scenes = int(self.scenes_entry.get())
            if num_scenes < 1 or num_scenes > 500:
                raise ValueError()
        except (ValueError, TypeError):
            self.snackbar.show("シーン数は1〜500の整数で", type="error")
            return

        # ストーリー構成を取得
        story_structure = self._get_story_structure()

        # Auto-save settings
        self.save_settings()

        # アウトラインプレビュー生成（ローカル・API不要）
        theme_jp = self.theme_combo.get()
        theme = THEME_OPTIONS.get(theme_jp, "")
        theme_guide = THEME_GUIDES.get(theme, THEME_GUIDES.get("vanilla", {}))
        theme_name = theme_guide.get("name", "指定なし")

        # 簡易コスト見積もり（ストーリー構成反映）
        pro_pct = story_structure["prologue"] / 100
        epi_pct = story_structure["epilogue"] / 100
        main_pct = story_structure["main"] / 100
        act3_count = max(1, round(num_scenes * main_pct * 0.75))
        low_count = num_scenes - act3_count
        high_count = act3_count
        prep_calls = 2  # あらすじ生成 + シーン分割
        total_api = prep_calls + num_scenes

        # コスト計算
        est_cost_prep = prep_calls * (2000 * 0.25 + 2000 * 1.25) / 1_000_000
        est_cost_low = low_count * (3000 * 0.25 + 2500 * 1.25) / 1_000_000
        est_cost_high = high_count * (3000 * 3.00 + 2500 * 15.00) / 1_000_000
        backend_name = "Claude (Anthropic)"
        low_model = "Haiku"
        high_model = "Sonnet"
        est_total = est_cost_prep + est_cost_low + est_cost_high

        # プレビュー表示
        self.log_text.delete("1.0", "end")
        self.log(f"{'='*50}")
        self.log(f"[INFO]生成プレビュー")
        self.log(f"{'='*50}")
        self.log(f"バックエンド: {backend_name}")
        self.log(f"テーマ: {theme_name}")
        self.log(f"シーン数: {num_scenes}")
        self.log(f"ストーリー構成: プロローグ{story_structure['prologue']}% / 本編{story_structure['main']}% / エピローグ{story_structure['epilogue']}%")
        self.log(f"")
        self.log(f"[STAT]パイプライン:")
        self.log(f"  Step 1: ストーリー原案作成（{low_model}×1）")
        self.log(f"  Step 2: シーン分割（{low_model}×1）")
        self.log(f"  Step 3: シーン生成")
        self.log(f"    Low (1-3): {low_count}シーン → {low_model}")
        self.log(f"    High (4-5): {high_count}シーン → {high_model}")
        self.log(f"")
        self.log(f"[COST]推定コスト: ${est_total:.4f}")
        self.log(f"  準備: ${est_cost_prep:.4f} (あらすじ+分割)")
        self.log(f"  {low_model}: ${est_cost_low:.4f} ({low_count}回)")
        self.log(f"  {high_model}: ${est_cost_high:.4f} ({high_count}回)")
        self.log(f"  合計API呼び出し: {total_api}回")
        self.log(f"{'='*50}")
        self.log(f"")

        self.is_generating = True
        self.stop_requested = False
        self.generate_btn.configure(state="disabled", text="生成中...")
        self.stop_btn.configure(
            state="normal",
            border_color=MaterialColors.ERROR,
            text_color=MaterialColors.ERROR
        )
        self.progress.set(0)

        thread = threading.Thread(
            target=self.run_generation,
            args=(api_key, concept, characters, num_scenes, other_chars, story_structure),
            daemon=True
        )
        thread.start()

    def stop_generation(self):
        if self.is_generating:
            self.stop_requested = True
            self.update_status("[STOP]停止リクエスト送信...")
            self.stop_btn.configure(state="disabled", text="停止中...")

    def run_generation(self, api_key: str, concept: str, characters: str, num_scenes: int, other_chars: str = "", story_structure: dict = None):
        try:
            theme_jp = self.theme_combo.get()
            theme = THEME_OPTIONS.get(theme_jp, "")

            def callback(msg):
                if self.stop_requested:
                    raise InterruptedError("ユーザーによる停止")
                self.after(0, lambda: self.update_status(msg))

            self.after(0, lambda: self.update_status("[START] パイプライン開始... [Claude (Anthropic)]"))

            # その他の登場人物をcharactersに統合
            full_characters = characters
            if other_chars:
                full_characters = f"{characters}\n\n【その他の登場人物】\n{other_chars}"

            # v7.0: 男性設定・時間帯・場所タイプをGUIから取得
            _male_tags = self.get_male_tags() if hasattr(self, 'get_male_tags') else ""
            _time_of_day_jp = self.time_of_day_combo.get() if hasattr(self, 'time_of_day_combo') else "おまかせ"
            _time_tags = TIME_OF_DAY_OPTIONS.get(_time_of_day_jp, "")
            _location_jp = self.location_type_combo.get() if hasattr(self, 'location_type_combo') else "おまかせ"
            _location_type = LOCATION_TYPE_OPTIONS.get(_location_jp, "")

            # v7.3: 男性キャラの日本語説明を取得（LLMプロンプト用）
            _male_desc_jp = ""
            if hasattr(self, 'male_custom_field'):
                _custom = self.male_custom_field.get().strip()
                if _custom:
                    _male_desc_jp = _custom
            if not _male_desc_jp and hasattr(self, 'male_preset_combo'):
                _preset_key = self.male_preset_combo.get()
                if _preset_key != "おまかせ":
                    _male_desc_jp = _preset_key
            # 髪型・髪色・肌色の日本語を追記
            _male_detail_parts = []
            for _attr, _lbl in [
                ('male_hair_style_combo', ''),
                ('male_hair_color_combo', ''),
                ('male_skin_color_combo', '肌:'),
            ]:
                if hasattr(self, _attr):
                    _v = getattr(self, _attr).get()
                    if _v and _v != "おまかせ" and _v != "普通":
                        _male_detail_parts.append(f"{_lbl}{_v}")
            if _male_detail_parts:
                _detail = "、".join(_male_detail_parts)
                _male_desc_jp = f"{_male_desc_jp}、{_detail}" if _male_desc_jp else _detail

            # v7.4: SDプロンプト設定をGUIから取得
            _sd_quality_mode = self.sd_quality_mode_var.get() if hasattr(self, 'sd_quality_mode_var') else "auto"
            _sd_quality_custom = ""
            if _sd_quality_mode == "manual" and hasattr(self, 'sd_quality_custom_entry'):
                _sd_quality_custom = self.sd_quality_custom_entry.get().strip()
                if not _sd_quality_custom:
                    _sd_quality_custom = QUALITY_TAGS_DISABLED  # 空欄→quality tags無し
            _sd_prefix = self.sd_prefix_text.get("1.0", "end-1c").strip().replace("\n", ", ").replace(", , ", ", ") if hasattr(self, 'sd_prefix_text') else ""
            _sd_suffix = self.sd_suffix_text.get("1.0", "end-1c").strip().replace("\n", ", ").replace(", , ", ", ") if hasattr(self, 'sd_suffix_text') else ""

            # ネガティブプロンプト設定をGUIから取得
            _sd_neg_base = self.sd_neg_base_text.get("1.0", "end-1c").strip().replace("\n", ", ").replace(", , ", ", ") if hasattr(self, 'sd_neg_base_text') else ""
            _sd_neg_prefix = self.sd_neg_prefix_text.get("1.0", "end-1c").strip().replace("\n", ", ").replace(", , ", ", ") if hasattr(self, 'sd_neg_prefix_text') else ""
            _sd_neg_suffix = self.sd_neg_suffix_text.get("1.0", "end-1c").strip().replace("\n", ", ").replace(", , ", ", ") if hasattr(self, 'sd_neg_suffix_text') else ""

            _quality_priority = self.quality_priority_var.get() if hasattr(self, 'quality_priority_var') else False
            _faceless_male = self.male_faceless_var.get() if hasattr(self, 'male_faceless_var') else True
            results, cost_tracker, pipeline_metadata = generate_pipeline(
                api_key, concept, full_characters, num_scenes, theme, callback,
                story_structure=story_structure,
                male_tags=_male_tags, time_tags=_time_tags, location_type=_location_type,
                male_description=_male_desc_jp,
                sd_quality_tags=_sd_quality_custom,
                sd_prefix_tags=_sd_prefix,
                sd_suffix_tags=_sd_suffix,
                provider=PROVIDER_CLAUDE,
                quality_priority=_quality_priority,
                faceless_male=_faceless_male,
                sd_neg_base=_sd_neg_base,
                sd_neg_prefix=_sd_neg_prefix,
                sd_neg_suffix=_sd_neg_suffix,
            )

            if self.stop_requested:
                self.after(0, lambda: self.on_stopped())
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_path = EXPORTS_DIR / f"script_{timestamp}.csv"
            json_path = EXPORTS_DIR / f"script_{timestamp}.json"
            xlsx_path = EXPORTS_DIR / f"script_{timestamp}.xlsx"
            sd_path = EXPORTS_DIR / f"sd_prompts_{timestamp}.txt"
            wc_path = EXPORTS_DIR / f"wildcard_{timestamp}.txt"
            dlg_path = EXPORTS_DIR / f"dialogue_{timestamp}.txt"

            export_csv(results, csv_path)
            export_json(results, json_path, metadata=pipeline_metadata)
            export_sd_prompts(results, sd_path)
            export_wildcard(results, wc_path,
                           male_tags=_male_tags, time_tags=_time_tags,
                           location_type=_location_type)
            # ネガティブプロンプト Wildcard出力
            wc_neg_path = EXPORTS_DIR / f"wildcard_negative_{timestamp}.txt"
            export_wildcard_negative(results, wc_neg_path)
            export_dialogue_list(results, dlg_path)

            # Excel出力（openpyxlがある場合）
            excel_ok = export_excel(results, xlsx_path)

            self.after(0, lambda: self.on_complete(results, cost_tracker, csv_path, json_path, xlsx_path if excel_ok else None, pipeline_metadata))

        except InterruptedError:
            # 中断時でも途中結果をエクスポート
            if results:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                partial_json = EXPORTS_DIR / f"script_{timestamp}_partial.json"
                try:
                    export_json(results, partial_json)
                    partial_path = str(partial_json)
                    self.after(0, lambda: self.on_stopped_with_partial(partial_path, len(results)))
                except Exception:
                    self.after(0, lambda: self.on_stopped())
            else:
                self.after(0, lambda: self.on_stopped())
        except Exception as e:
            # エラー時も途中結果があれば保存
            if results:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                partial_json = EXPORTS_DIR / f"script_{timestamp}_error.json"
                try:
                    export_json(results, partial_json)
                except Exception:
                    pass
            self.after(0, lambda: self.on_error(str(e)))

    def reset_buttons(self):
        self.is_generating = False
        self.stop_requested = False
        self.generate_btn.configure(state="normal", text="脚本を生成")
        self.stop_btn.configure(
            state="disabled",
            text="停止",
            border_color=MaterialColors.OUTLINE,
            text_color=MaterialColors.OUTLINE
        )
        # フェーズインジケーターをリセット
        for pill, lbl in self.phase_labels:
            pill.configure(fg_color=MaterialColors.SURFACE_CONTAINER)
            lbl.configure(text_color=MaterialColors.ON_SURFACE_VARIANT)

    def on_complete(self, results, cost_tracker, csv_path, json_path, xlsx_path=None, metadata=None):
        self.reset_buttons()
        self.progress.set(1)
        self.last_results = results  # 再エクスポート用に保持
        self.last_metadata = metadata  # メタデータも保持

        self.cost_label.configure(text=cost_tracker.summary())
        self.update_status(f"[OK]完了! {len(results)}シーン生成")
        self.log(f"[FILE]CSV: {csv_path}")
        self.log(f"[FILE]JSON: {json_path}")
        if xlsx_path:
            self.log(f"[STAT]Excel: {xlsx_path}（折り返し表示対応）")
        self.log(f"[COST]{cost_tracker.summary()}")
        self.snackbar.show(f"{len(results)}シーン生成完了!", type="success")

        # 再エクスポートボタン有効化
        self.export_btn.configure(state="normal")

        # エクスポートフォルダを開くボタンを表示
        self._show_open_folder_btn()

    def _show_open_folder_btn(self):
        """エクスポートフォルダを開くボタンをログ領域の上に表示"""
        if hasattr(self, "_open_folder_btn") and self._open_folder_btn.winfo_exists():
            self._open_folder_btn.destroy()
        self._open_folder_btn = ctk.CTkButton(
            self.log_text.master, text="エクスポートフォルダを開く",
            font=ctk.CTkFont(size=14), height=32,
            fg_color=MaterialColors.SECONDARY_CONTAINER,
            text_color=MaterialColors.ON_SECONDARY_CONTAINER,
            hover_color=MaterialColors.PRIMARY,
            corner_radius=8,
            command=self.open_export_folder
        )
        self._open_folder_btn.pack(pady=(8, 8))

    def open_export_folder(self):
        """エクスポートフォルダをエクスプローラーで開く"""
        import subprocess
        folder = str(EXPORTS_DIR)
        try:
            subprocess.Popen(["explorer", folder])
        except Exception as e:
            self.log(f"フォルダを開けません: {e}")

    def open_export_dialog(self):
        """マルチフォーマットエクスポートダイアログを開く"""
        meta = getattr(self, "last_metadata", None)
        if self.last_results:
            ExportDialog(self, self.last_results, metadata=meta)
        else:
            # last_results が無い場合でもJSONインポートでエクスポート可能
            ExportDialog(self, [], metadata=meta)

    def toggle_presentation_mode(self):
        """Ctrl+Shift+P: APIキー・コスト表示を切替"""
        self._presentation_mode = not self._presentation_mode
        if self._presentation_mode:
            self._api_card.pack_forget()
            self._cost_card.pack_forget()
            self._cost_preview_label.pack_forget()
            self.snackbar.show("プレゼンモード ON（Ctrl+Shift+Pで解除）", type="info")
        else:
            # api_card: contentフレームの先頭に復元
            content = self._api_card.master
            visible = content.pack_slaves()
            if visible:
                self._api_card.pack(fill="x", pady=(0, 16), before=visible[0])
            else:
                self._api_card.pack(fill="x", pady=(0, 16))
            # cost_card: contentフレームの末尾に復元
            self._cost_card.pack(fill="x", pady=(0, 16))
            # cost_preview_label: settings_card内の末尾に復元
            self._cost_preview_label.pack(anchor="w", padx=20, pady=(4, 12))
            self.snackbar.show("プレゼンモード OFF", type="info")

    def _ensure_placeholders(self):
        """CTkEntryのプレースホルダーが初期化後に確実に表示されるようにする"""
        for entry in [
            getattr(self, 'male_custom_field', None),
        ]:
            if entry is not None and not entry.get():
                try:
                    entry._activate_placeholder()
                except Exception:
                    pass

    def on_close(self):
        """ウィンドウ閉じ時の処理（生成中なら確認ダイアログ）"""
        if self.is_generating:
            import tkinter.messagebox as mb
            if mb.askokcancel("確認", "生成中です。停止して終了しますか？"):
                self.stop_requested = True
                self.after(500, self.destroy)
        else:
            self.destroy()

    def on_stopped(self):
        self.reset_buttons()
        self.progress.set(0)
        self.update_status("[STOP]生成を停止しました")
        self.snackbar.show("生成を停止しました", type="warning")

    def on_stopped_with_partial(self, partial_path: str, count: int):
        """中断時に部分結果を保存して通知"""
        self.reset_buttons()
        self.progress.set(0)
        self.update_status(f"[STOP]停止（{count}シーン保存済み）")
        self.log(f"[FILE]途中結果: {partial_path}")
        self.snackbar.show(f"停止（{count}シーン保存済み）", type="warning")

    def on_error(self, error: str):
        self.reset_buttons()
        self.progress.set(0)
        self.update_status(f"[ERROR]エラー: {error}")
        self.snackbar.show(f"エラー: {error[:50]}", type="error")

    def refresh_char_list(self):
        """キャラクター一覧を更新"""
        chars = get_existing_characters()
        values = ["（キャラ選択）"]
        for c in chars:
            values.append(f"{c['name']} ({c['work']})")
        self.char_select_combo.configure(values=values)
        if hasattr(self, '_char_map'):
            pass
        self._char_map = {f"{c['name']} ({c['work']})": c for c in chars}

    def on_char_selected(self, choice: str):
        """キャラ選択時のコールバック"""
        if choice == "（キャラ選択）" or choice not in self._char_map:
            return

        char_info = self._char_map[choice]
        char_id = char_info["char_id"]
        bible_path = CHARACTERS_DIR / f"{char_id}.json"

        if bible_path.exists():
            with open(bible_path, "r", encoding="utf-8") as f:
                bible = json.load(f)

            # キャラ情報を取得
            name = bible.get('character_name', '')
            work = bible.get('work_title', '')
            personality = bible.get('personality_core', {})
            speech = bible.get('speech_pattern', {})
            emotional = bible.get('emotional_speech', {})
            physical = bible.get('physical_description', {})
            tags = bible.get('danbooru_tags', [])

            # 個別フィールドに直接設定
            self._set_characters_fields({
                "name": f"{name}（{work}）" if work else name,
                "personality": personality.get('brief_description', ''),
                "first_person": speech.get('first_person', '私'),
                "endings": ', '.join(speech.get('sentence_endings', [])[:4]),
                "appearance": f"{physical.get('hair', '')}、{physical.get('eyes', '')}",
            })

            # ログに詳細なキャラ設定を出力
            self.log(f"═══ キャラ設定プレビュー: {name} ═══")
            self.log(f"作品: {work}")
            self.log(f"性格: {personality.get('brief_description', '')}")
            self.log(f"特性: {', '.join(personality.get('main_traits', []))}")
            self.log(f"一人称: {speech.get('first_person', '私')}")
            self.log(f"語尾: {', '.join(speech.get('sentence_endings', [])[:5])}")
            self.log(f"照れた時: {emotional.get('when_embarrassed', '')}")
            self.log(f"甘える時: {emotional.get('when_flirty', '')}")
            self.log(f"SDタグ: {', '.join(tags[:8])}...")
            self.log(f"═══════════════════════════════")

            self.snackbar.show(f"{name}を追加（ログに設定詳細）", type="success")

    def refresh_preset_list(self):
        """プリセット一覧を更新"""
        self._all_presets = get_preset_characters()
        self._preset_map = {}
        for p in self._all_presets:
            label = f"【{p.get('work_title', p.get('work', ''))}】{p.get('character_name', p.get('name', ''))}"
            self._preset_map[label] = p

        # Update card title with count
        count = len(self._all_presets)
        if hasattr(self, '_preset_card'):
            self._preset_card.title_label.configure(text=f"二次創作・プリセットキャラ一覧（{count}体収録）")

        # Show all characters immediately
        if hasattr(self, '_category_chips') and self._category_chips:
            self._on_category_chip_click("全て")

    def on_preset_selected(self, choice: str):
        """プリセット選択時（後方互換）"""
        pass

    def load_preset_action(self):
        """プリセット読み込み（後方互換）"""
        pass

    # ======== Preset Tab Methods ========

    def _build_preset_tab(self, parent):
        """プリセットタブUIを構築"""
        # Category chip row
        chip_frame = ctk.CTkFrame(parent, fg_color="transparent")
        chip_frame.pack(fill="x", padx=16, pady=(0, 8))

        categories = ["全て", "ジャンプ", "マガジン", "ラノベ", "アニメ", "ソシャゲ", "ゲーム", "サンデー", "VTuber"]
        self._category_map = {
            "全て": None,
            "ジャンプ": ["ジャンプ", "ジャンプ+"],
            "マガジン": ["マガジン"],
            "ラノベ": ["ラノベ"],
            "アニメ": ["アニメ"],
            "ソシャゲ": ["ソーシャルゲーム"],
            "ゲーム": ["ゲーム"],
            "サンデー": ["サンデー"],
            "VTuber": ["VTuber"],
        }

        for cat in categories:
            chip = MaterialChip(
                chip_frame, text=cat,
                selected=(cat == "全て"),
                chip_type="filter",
                command=lambda c=cat: self._on_category_chip_click(c)
            )
            chip.pack(side="left", padx=(0, 6))
            self._category_chips[cat] = chip

        # Work filter dropdown (optional narrowing)
        filter_row = ctk.CTkFrame(parent, fg_color="transparent")
        filter_row.pack(fill="x", padx=16, pady=(0, 8))

        ctk.CTkLabel(
            filter_row, text="作品で絞り込み:",
            font=ctk.CTkFont(family=FONT_JP, size=13),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left", padx=(0, 6))

        self._work_dropdown = ctk.CTkOptionMenu(
            filter_row, values=["（すべて表示）"],
            command=self._on_work_selected,
            font=ctk.CTkFont(family=FONT_JP, size=14), width=300,
            fg_color=MaterialColors.SURFACE_CONTAINER,
            button_color=MaterialColors.PRIMARY,
            text_color=MaterialColors.ON_SURFACE,
            dropdown_text_color=MaterialColors.ON_SURFACE,
            dropdown_fg_color=MaterialColors.SURFACE
        )
        self._work_dropdown.pack(side="left")

        # Character card scroll area
        self._preset_card_frame = ctk.CTkScrollableFrame(
            parent, fg_color=MaterialColors.SURFACE_CONTAINER_LOWEST,
            height=260, corner_radius=8
        )
        self._preset_card_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        # Placeholder text
        self._preset_placeholder = ctk.CTkLabel(
            self._preset_card_frame,
            text="カテゴリを選択してください",
            font=ctk.CTkFont(family=FONT_JP, size=14),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        )
        self._preset_placeholder.pack(pady=20)

    def _on_category_chip_click(self, category):
        """カテゴリチップ選択→キャラ一覧を即座に表示"""
        # Toggle chips (exclusive selection)
        for cat, chip in self._category_chips.items():
            if cat == category:
                if not chip.selected:
                    chip.toggle()
            else:
                if chip.selected:
                    chip.toggle()
        self._selected_category = category

        # Filter by category
        cat_filters = self._category_map.get(category)
        if cat_filters is None:
            filtered = self._all_presets
        else:
            filtered = [p for p in self._all_presets if p.get("category", "") in cat_filters]

        # Update work dropdown with available works
        seen = set()
        works = []
        for p in filtered:
            wt = p.get("work_title", p.get("work", ""))
            if wt not in seen:
                seen.add(wt)
                works.append(wt)

        values = ["（すべて表示）"] + works
        self._work_dropdown.configure(values=values)
        self._work_dropdown.set("（すべて表示）")

        # Show all characters grouped by work
        self._render_preset_list(filtered)

        # キャラリストを先頭にスクロール
        if self._preset_card_frame:
            try:
                self._preset_card_frame._parent_canvas.yview_moveto(0)
            except Exception:
                pass

    def _on_work_selected(self, work_title):
        """作品選択→キャラカード表示（絞り込み）"""
        cat_filters = self._category_map.get(self._selected_category)

        if work_title == "（すべて表示）":
            if cat_filters is None:
                filtered = self._all_presets
            else:
                filtered = [p for p in self._all_presets if p.get("category", "") in cat_filters]
        else:
            filtered = []
            for p in self._all_presets:
                wt = p.get("work_title", p.get("work", ""))
                cat = p.get("category", "")
                if wt == work_title:
                    if cat_filters is None or cat in cat_filters:
                        filtered.append(p)

        self._render_preset_list(filtered)

    def _render_char_card(self, preset_info):
        """キャラカードを描画（リッチ版）"""
        card = ctk.CTkFrame(
            self._preset_card_frame,
            fg_color=MaterialColors.SURFACE_CONTAINER_LOW,
            corner_radius=0, height=56,
            border_width=1, border_color=MaterialColors.OUTLINE_VARIANT,
        )
        card.pack(fill="x", pady=(0, 6), padx=12)
        card.pack_propagate(False)

        name = preset_info.get("character_name", preset_info.get("name", ""))
        work = preset_info.get("work_title", preset_info.get("work", ""))
        category = preset_info.get("category", "")

        # Left accent bar based on category
        cat_colors = {
            "ジャンプ": "#E85D3A", "ジャンプ+": "#E85D3A",
            "マガジン": "#3A8FE8", "ラノベ": "#8F5FD6",
            "アニメ": "#40B080", "ソーシャルゲーム": "#E8A83A",
            "ゲーム": "#6B8E23", "サンデー": "#FF8C00",
            "VTuber": "#E84F8A",
        }
        accent = cat_colors.get(category, MaterialColors.PRIMARY)

        accent_bar = ctk.CTkFrame(card, fg_color=accent, width=4, corner_radius=0)
        accent_bar.pack(side="left", fill="y", padx=(0, 0), pady=6)

        # Name (bold)
        ctk.CTkLabel(
            card, text=name,
            font=ctk.CTkFont(family=FONT_JP, size=16, weight="bold"),
            text_color=MaterialColors.ON_SURFACE
        ).pack(side="left", padx=(12, 8), pady=8)

        # Work title (smaller, muted)
        ctk.CTkLabel(
            card, text=work,
            font=ctk.CTkFont(family=FONT_JP, size=13),
            text_color=MaterialColors.ON_SURFACE_VARIANT
        ).pack(side="left", padx=(0, 12), pady=8)

        # Load button
        MaterialButton(
            card, text="読み込み", variant="filled_tonal", size="small",
            command=lambda p=preset_info: self._load_preset_direct(p)
        ).pack(side="right", padx=(0, 12), pady=10)

    def _render_preset_list(self, presets):
        """プリセット一覧を作品グループごとに描画"""
        self._clear_preset_cards()

        if not presets:
            self._preset_placeholder.pack(pady=20)
            self._preset_placeholder.configure(text="キャラが見つかりません")
            return

        self._preset_placeholder.pack_forget()

        # Group by work title (dict preserves insertion order in Python 3.7+)
        groups = {}
        for p in presets:
            wt = p.get("work_title", p.get("work", ""))
            if wt not in groups:
                groups[wt] = []
            groups[wt].append(p)

        for work_title, chars in groups.items():
            # Work title header
            header = ctk.CTkFrame(
                self._preset_card_frame, fg_color="transparent", height=28
            )
            header.pack(fill="x", padx=12, pady=(12, 4))
            header.pack_propagate(False)

            ctk.CTkLabel(
                header, text=f"  {work_title}  ({len(chars)})",
                font=ctk.CTkFont(family=FONT_JP, size=14, weight="bold"),
                text_color=MaterialColors.PRIMARY
            ).pack(side="left")

            # Divider line
            divider = ctk.CTkFrame(
                self._preset_card_frame,
                fg_color=MaterialColors.OUTLINE_VARIANT, height=1
            )
            divider.pack(fill="x", padx=16, pady=(0, 6))

            for ch in chars:
                self._render_char_card(ch)

    def _clear_preset_cards(self):
        """プリセットカードをクリア"""
        for widget in self._preset_card_frame.winfo_children():
            if widget != self._preset_placeholder:
                widget.destroy()
        try:
            self._preset_placeholder.pack_forget()
        except:
            pass

    def _setup_nested_scroll(self):
        """ネストされたスクロール領域のスムーズスクロール制御

        1. 全CTkScrollableFrameの内部MouseWheelバインドを無効化
        2. winfo_containingベースでスクロール先を判定
        3. ピクセル単位の慣性アニメーションでスムーズに移動
        4. 内側フレーム端到達時にメインへバブルアップ
        """
        inner_frames = []
        for frame in [
            getattr(self, '_preset_card_frame', None),
        ]:
            if frame:
                inner_frames.append(frame)

        # 全CTkScrollableFrameの内部バインドを無効化
        all_frames = [self.main_container] + inner_frames
        for frame in all_frames:
            try:
                frame.unbind("<MouseWheel>")
            except Exception:
                pass
            try:
                frame._parent_canvas.unbind("<MouseWheel>")
            except Exception:
                pass

        # スムーズスクロール用の状態
        self._scroll_velocity = 0.0
        self._scroll_target_frame = None
        self._scroll_animating = False

        PIXELS_PER_NOTCH = 45
        FRICTION = 0.65
        FRAME_MS = 12
        MIN_VELOCITY = 0.5

        def _find_inner_ancestor(widget):
            """ウィジェットの祖先を辿り、内側CTkScrollableFrameを見つける"""
            w = widget
            depth = 0
            while w is not None and depth < 50:
                for inner in inner_frames:
                    if w is inner:
                        return inner
                try:
                    w = w.master
                except Exception:
                    break
                depth += 1
            return None

        def _can_scroll(frame, direction):
            """フレームがその方向にスクロール可能かチェック"""
            try:
                canvas = frame._parent_canvas
                top, bottom = canvas.yview()
                if direction < 0 and top <= 0.001:
                    return False
                if direction > 0 and bottom >= 0.999:
                    return False
                return True
            except Exception:
                return False

        def _scroll_pixels(frame, pixels):
            """フレームをピクセル単位でスクロール"""
            try:
                canvas = frame._parent_canvas
                scroll_region = canvas.cget("scrollregion")
                if scroll_region:
                    parts = scroll_region.split()
                    total_height = float(parts[3]) - float(parts[1])
                else:
                    total_height = canvas.winfo_height()
                if total_height <= 0:
                    return
                fraction = pixels / total_height
                current = canvas.yview()[0]
                new_pos = max(0.0, min(1.0, current + fraction))
                canvas.yview_moveto(new_pos)
            except Exception:
                pass

        def _animate_scroll():
            """慣性スクロールアニメーション"""
            if not self._scroll_animating:
                return
            if abs(self._scroll_velocity) < MIN_VELOCITY:
                self._scroll_velocity = 0.0
                self._scroll_animating = False
                return

            frame = self._scroll_target_frame
            if frame is None:
                self._scroll_animating = False
                return

            direction = 1 if self._scroll_velocity > 0 else -1

            # 内側フレームの端到達時にメインへバブルアップ
            if frame is not self.main_container and not _can_scroll(frame, direction):
                self._scroll_target_frame = self.main_container
                frame = self.main_container

            _scroll_pixels(frame, self._scroll_velocity)
            self._scroll_velocity *= FRICTION
            self.after(FRAME_MS, _animate_scroll)

        def _on_mousewheel(event):
            """マウスホイールイベント → 慣性スクロール開始"""
            raw_delta = -event.delta / 120.0
            impulse = raw_delta * PIXELS_PER_NOTCH

            try:
                x, y = self.winfo_pointerxy()
                widget = self.winfo_containing(x, y)
            except Exception:
                return "break"

            if widget is None:
                return "break"

            inner = _find_inner_ancestor(widget)
            direction = 1 if impulse > 0 else -1

            if inner is not None and _can_scroll(inner, direction):
                target = inner
            else:
                target = self.main_container

            # 速度加算（連続ホイールで加速、方向転換時はリセット）
            if self._scroll_target_frame is not target:
                self._scroll_velocity = impulse
            elif (self._scroll_velocity > 0) != (impulse > 0):
                self._scroll_velocity = impulse
            else:
                self._scroll_velocity += impulse

            self._scroll_target_frame = target

            if not self._scroll_animating:
                self._scroll_animating = True
                _animate_scroll()

            return "break"

        self.bind_all("<MouseWheel>", _on_mousewheel)


    def _load_preset_direct(self, preset_info):
        """ワンクリックでプリセット読み込み"""
        char_id = preset_info["char_id"]
        try:
            bible, _ = load_preset_character(char_id, callback=lambda msg: self.log(msg))
            self.refresh_char_list()
            name = bible.get("character_name", char_id)
            work = preset_info.get("work_title", "")
            self._work_title_val = work
            self._char_name_val = name
            self.snackbar.show(f"{name}を読み込みました", type="success")
        except Exception as e:
            self.snackbar.show(f"読み込みエラー: {e}", type="error")

    # ======== Custom Character Tab Methods ========

    def save_custom_character(self):
        """オリジナルキャラクターを保存"""
        name = self.custom_name_entry.get().strip()
        if not name:
            self.snackbar.show("キャラ名を入力してください", type="warning")
            return

        # shyness_levelの取得
        shyness_level = int(round(self.shyness_slider.get()))

        # その他の登場人物テキスト取得
        other_chars = ""
        if hasattr(self, "other_chars_text"):
            other_chars = self.other_chars_text.get("1.0", "end-1c").strip()

        bible = build_custom_character_data(
            char_name=name,
            age=self.custom_age_dd.get(),
            relationship=self.custom_rel_dd.get(),
            archetype=self._selected_archetype,
            first_person=self.custom_first_person_dd.get(),
            speech_style=self.custom_speech_dd.get(),
            hair_color=self._selected_hair_color,
            hair_style=self.custom_hair_style_dd.get(),
            body_type=self.custom_body_dd.get(),
            chest=self.custom_chest_dd.get(),
            clothing=self.custom_clothing_dd.get(),
            shyness_level=shyness_level,
            custom_traits=self.custom_traits_entry.get().strip(),
            other_characters=other_chars,
        )

        # char_id生成＆保存
        char_id = generate_char_id("オリジナル", name)
        bible_path = CHARACTERS_DIR / f"{char_id}.json"
        skill_path = CHAR_SKILLS_DIR / f"{char_id}.skill.md"

        with open(bible_path, "w", encoding="utf-8") as f:
            json.dump(bible, f, ensure_ascii=False, indent=2)

        skill_content = generate_character_skill(char_id, bible)
        with open(skill_path, "w", encoding="utf-8") as f:
            f.write(skill_content)

        self.refresh_char_list()
        self.log(f"[OK]オリジナルキャラ保存: {name} ({self._selected_archetype})")
        self.log(f"   性格: {bible['personality_core']['brief_description']}")
        self.log(f"   一人称: {bible['speech_pattern']['first_person']} / 口調: {self.custom_speech_dd.get()}")
        self.log(f"   外見: {bible['physical_description']['hair']}")
        self.snackbar.show(f"{name}を保存しました（API未使用）", type="success")

    def _build_custom_tab(self, parent):
        """オリジナル作成タブUIを構築"""
        custom_scroll = ctk.CTkFrame(
            parent, fg_color="transparent"
        )
        custom_scroll.pack(fill="x")
        self._custom_scroll = custom_scroll

        # Helper for dropdowns
        def add_dropdown(p, label, options, default=None):
            ctk.CTkLabel(p, text=label, font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
                        text_color=MaterialColors.ON_SURFACE_VARIANT).pack(anchor="w", pady=(6,0))
            dd = ctk.CTkOptionMenu(p, values=options, font=ctk.CTkFont(family=FONT_JP, size=14),
                                   width=350, fg_color=MaterialColors.SURFACE_CONTAINER,
                                   button_color=MaterialColors.PRIMARY,
                                   text_color=MaterialColors.ON_SURFACE,
                                   dropdown_text_color=MaterialColors.ON_SURFACE,
                                   dropdown_fg_color=MaterialColors.SURFACE)
            dd.pack(anchor="w", pady=(2, 0))
            if default:
                dd.set(default)
            return dd

        # === Template Quick Start (32種) ===
        tmpl_label = ctk.CTkLabel(custom_scroll, text="テンプレート（ワンクリック雛形）— FANZA売れ筋32種",
                    font=ctk.CTkFont(family=FONT_JP, size=14, weight="bold"),
                    text_color=MaterialColors.PRIMARY)
        tmpl_label.pack(anchor="w", pady=(8, 8))

        templates = {
            # 学園系
            "JKツンデレ": {"age": "JK（女子高生）", "archetype": "ツンデレ", "first_person": "あたし",
                         "speech": "タメ口", "hair_color": "金髪", "hair_style": "ツインテール",
                         "body": "普通", "chest": "大きめ（D-E）", "clothing": "制服（ブレザー）", "shyness": 4},
            "ギャルJK": {"age": "JK（女子高生）", "archetype": "ギャル", "first_person": "ウチ",
                        "speech": "ギャル語", "hair_color": "金髪", "hair_style": "ロングウェーブ",
                        "body": "グラマー", "chest": "大きめ（D-E）", "clothing": "制服（ブレザー）", "shyness": 1},
            "地味子": {"age": "JK（女子高生）", "archetype": "真面目・優等生", "first_person": "私",
                      "speech": "丁寧語", "hair_color": "黒髪", "hair_style": "三つ編み",
                      "body": "小柄・華奢", "chest": "控えめ（A-B）", "clothing": "制服（ブレザー）", "shyness": 5},
            "委員長": {"age": "JK（女子高生）", "archetype": "真面目・優等生", "first_person": "私",
                      "speech": "丁寧語", "hair_color": "黒髪", "hair_style": "ロングストレート",
                      "body": "スレンダー", "chest": "大きめ（D-E）", "clothing": "制服（セーラー服）", "shyness": 4},
            # 純情系
            "甘え妹": {"age": "JK（女子高生）", "archetype": "妹系・甘えん坊", "first_person": "私",
                      "speech": "タメ口", "hair_color": "茶髪", "hair_style": "ツインテール",
                      "body": "小柄・華奢", "chest": "控えめ（A-B）", "clothing": "パジャマ・部屋着", "shyness": 4},
            "後輩マネ": {"age": "JK（女子高生）", "archetype": "元気っ子", "first_person": "私",
                        "speech": "丁寧語", "hair_color": "茶髪", "hair_style": "ポニーテール",
                        "body": "普通", "chest": "普通（C）", "clothing": "体操着・ブルマ", "shyness": 4},
            "メイドさん": {"age": "JD（女子大生）", "archetype": "妹系・甘えん坊", "first_person": "私",
                         "speech": "丁寧語", "hair_color": "黒髪", "hair_style": "ツインテール",
                         "body": "小柄・華奢", "chest": "普通（C）", "clothing": "メイド服", "shyness": 3},
            "巫女さん": {"age": "JD（女子大生）", "archetype": "大和撫子", "first_person": "私",
                        "speech": "古風・時代劇調", "hair_color": "黒髪", "hair_style": "姫カット",
                        "body": "スレンダー", "chest": "普通（C）", "clothing": "巫女服", "shyness": 4},
            # 年上系
            "大人クール": {"age": "OL（20代）", "archetype": "クーデレ", "first_person": "私",
                         "speech": "丁寧語", "hair_color": "黒髪", "hair_style": "ロングストレート",
                         "body": "スレンダー", "chest": "普通（C）", "clothing": "スーツ", "shyness": 2},
            "女教師": {"age": "OL（20代）", "archetype": "真面目・優等生", "first_person": "私",
                      "speech": "敬語（ビジネス）", "hair_color": "黒髪", "hair_style": "ポニーテール",
                      "body": "スレンダー", "chest": "大きめ（D-E）", "clothing": "スーツ", "shyness": 3},
            "ナース": {"age": "OL（20代）", "archetype": "お姉さん系", "first_person": "私",
                      "speech": "丁寧語", "hair_color": "茶髪", "hair_style": "ボブカット",
                      "body": "グラマー", "chest": "巨乳（F以上）", "clothing": "ナース服", "shyness": 3},
            "未亡人": {"age": "お姉さん（30代）", "archetype": "クーデレ", "first_person": "私",
                      "speech": "丁寧語", "hair_color": "黒髪", "hair_style": "ロングストレート",
                      "body": "グラマー", "chest": "大きめ（D-E）", "clothing": "着物・浴衣", "shyness": 4},
            # 個性派
            "お嬢様": {"age": "JD（女子大生）", "archetype": "お嬢様", "first_person": "わたくし",
                      "speech": "お嬢様言葉", "hair_color": "金髪", "hair_style": "ロングウェーブ",
                      "body": "グラマー", "chest": "大きめ（D-E）", "clothing": "ドレス", "shyness": 3},
            "エルフ姫": {"age": "エルフ・長命種", "archetype": "お嬢様", "first_person": "わたくし",
                        "speech": "古風・時代劇調", "hair_color": "銀髪", "hair_style": "ロングストレート",
                        "body": "スレンダー", "chest": "普通（C）", "clothing": "ドレス", "shyness": 3},
            "褐色スポーツ": {"age": "JK（女子高生）", "archetype": "元気っ子", "first_person": "あたし",
                           "speech": "タメ口", "hair_color": "茶髪", "hair_style": "ショートヘア",
                           "body": "筋肉質", "chest": "普通（C）", "clothing": "体操着・ブルマ", "shyness": 2},
            "バニーガール": {"age": "JD（女子大生）", "archetype": "小悪魔", "first_person": "あたし",
                           "speech": "タメ口", "hair_color": "金髪", "hair_style": "ポニーテール",
                           "body": "グラマー", "chest": "巨乳（F以上）", "clothing": "バニーガール", "shyness": 1},
            # NTR/人妻系
            "NTR彼女": {"age": "JD（女子大生）", "archetype": "天然・ドジっ子", "first_person": "私",
                        "speech": "タメ口", "hair_color": "茶髪", "hair_style": "ロングストレート",
                        "body": "普通", "chest": "大きめ（D-E）", "clothing": "私服（清楚系）", "shyness": 4},
            "人妻さん": {"age": "人妻", "archetype": "お姉さん系", "first_person": "私",
                        "speech": "丁寧語", "hair_color": "茶髪", "hair_style": "セミロング",
                        "body": "グラマー", "chest": "大きめ（D-E）", "clothing": "エプロン", "shyness": 4},
            "義母さん": {"age": "お姉さん（30代）", "archetype": "大和撫子", "first_person": "私",
                        "speech": "丁寧語", "hair_color": "黒髪", "hair_style": "ロングストレート",
                        "body": "グラマー", "chest": "巨乳（F以上）", "clothing": "着物・浴衣", "shyness": 4},
            "メスガキ": {"age": "ロリ", "archetype": "小悪魔", "first_person": "あたし",
                        "speech": "タメ口", "hair_color": "ピンク髪", "hair_style": "ツインテール",
                        "body": "小柄・華奢", "chest": "控えめ（A-B）", "clothing": "私服（ギャル系）", "shyness": 1},
            # 異種族系
            "サキュバス": {"age": "エルフ・長命種", "archetype": "サキュバス系", "first_person": "私",
                         "speech": "タメ口", "hair_color": "紫髪", "hair_style": "ロングウェーブ",
                         "body": "グラマー", "chest": "巨乳（F以上）", "clothing": "私服（ギャル系）", "shyness": 1},
            "獣耳メイド": {"age": "JD（女子大生）", "archetype": "妹系・甘えん坊", "first_person": "私",
                          "speech": "丁寧語", "hair_color": "白髪", "hair_style": "ロングストレート",
                          "body": "小柄・華奢", "chest": "普通（C）", "clothing": "メイド服", "shyness": 4},
            "ダークエルフ": {"age": "エルフ・長命種", "archetype": "クーデレ", "first_person": "私",
                           "speech": "古風・時代劇調", "hair_color": "白髪", "hair_style": "ロングストレート",
                           "body": "スレンダー", "chest": "大きめ（D-E）", "clothing": "鎧・アーマー", "shyness": 2},
            "天使堕ち": {"age": "エルフ・長命種", "archetype": "天然・ドジっ子", "first_person": "わたくし",
                        "speech": "丁寧語", "hair_color": "金髪", "hair_style": "ロングストレート",
                        "body": "スレンダー", "chest": "普通（C）", "clothing": "ドレス", "shyness": 5},
            # シチュ特化
            "催眠JK": {"age": "JK（女子高生）", "archetype": "真面目・優等生", "first_person": "私",
                      "speech": "丁寧語", "hair_color": "黒髪", "hair_style": "ロングストレート",
                      "body": "普通", "chest": "大きめ（D-E）", "clothing": "制服（セーラー服）", "shyness": 5},
            "女騎士": {"age": "OL（20代）", "archetype": "真面目・優等生", "first_person": "私",
                      "speech": "古風・時代劇調", "hair_color": "金髪", "hair_style": "ポニーテール",
                      "body": "筋肉質", "chest": "大きめ（D-E）", "clothing": "鎧・アーマー", "shyness": 4},
            "陰キャ同級生": {"age": "JK（女子高生）", "archetype": "陰キャ・オタク", "first_person": "私",
                           "speech": "タメ口", "hair_color": "黒髪", "hair_style": "三つ編み",
                           "body": "小柄・華奢", "chest": "控えめ（A-B）", "clothing": "制服（ブレザー）", "shyness": 5},
            "配信者": {"age": "JD（女子大生）", "archetype": "陰キャ・オタク", "first_person": "私",
                      "speech": "タメ口", "hair_color": "ピンク髪", "hair_style": "ツインテール",
                      "body": "普通", "chest": "普通（C）", "clothing": "パジャマ・部屋着", "shyness": 3},
            # 年齢差系
            "女上司": {"age": "OL（20代）", "archetype": "お姉さん系", "first_person": "私",
                      "speech": "敬語（ビジネス）", "hair_color": "黒髪", "hair_style": "ボブカット",
                      "body": "スレンダー", "chest": "大きめ（D-E）", "clothing": "スーツ", "shyness": 2},
            "ママ友": {"age": "人妻", "archetype": "天然・ドジっ子", "first_person": "私",
                      "speech": "丁寧語", "hair_color": "茶髪", "hair_style": "セミロング",
                      "body": "グラマー", "chest": "巨乳（F以上）", "clothing": "私服（清楚系）", "shyness": 3},
            "若妻先生": {"age": "OL（20代）", "archetype": "大和撫子", "first_person": "私",
                        "speech": "丁寧語", "hair_color": "茶髪", "hair_style": "ポニーテール",
                        "body": "普通", "chest": "大きめ（D-E）", "clothing": "スーツ", "shyness": 4},
            "寮母さん": {"age": "お姉さん（30代）", "archetype": "お姉さん系", "first_person": "私",
                        "speech": "丁寧語", "hair_color": "黒髪", "hair_style": "ロングストレート",
                        "body": "グラマー", "chest": "巨乳（F以上）", "clothing": "エプロン", "shyness": 3},
        }
        self._custom_templates = templates

        # カテゴリ別テンプレートグリッド (8行×4列)
        tmpl_categories = [
            ("学園系", ["JKツンデレ", "ギャルJK", "地味子", "委員長"]),
            ("純情系", ["甘え妹", "後輩マネ", "メイドさん", "巫女さん"]),
            ("年上系", ["大人クール", "女教師", "ナース", "未亡人"]),
            ("個性派", ["お嬢様", "エルフ姫", "褐色スポーツ", "バニーガール"]),
            ("NTR/人妻", ["NTR彼女", "人妻さん", "義母さん", "メスガキ"]),
            ("異種族系", ["サキュバス", "獣耳メイド", "ダークエルフ", "天使堕ち"]),
            ("シチュ特化", ["催眠JK", "女騎士", "陰キャ同級生", "配信者"]),
            ("年齢差系", ["女上司", "ママ友", "若妻先生", "寮母さん"]),
        ]

        tmpl_grid = ctk.CTkFrame(custom_scroll, fg_color="transparent")
        tmpl_grid.pack(fill="x", pady=(0, 12))

        for row_idx, (cat_name, cat_templates) in enumerate(tmpl_categories):
            row_frame = ctk.CTkFrame(
                tmpl_grid, fg_color=MaterialColors.SURFACE_CONTAINER_LOW,
                corner_radius=8
            )
            row_frame.pack(fill="x", pady=(0, 4))
            ctk.CTkLabel(
                row_frame, text=cat_name, width=80,
                font=ctk.CTkFont(family=FONT_JP, size=12, weight="bold"),
                text_color=MaterialColors.ON_SURFACE_VARIANT, anchor="w"
            ).grid(row=0, column=0, padx=(8, 6), pady=4, sticky="w")
            for col_idx, tname in enumerate(cat_templates):
                btn = MaterialButton(
                    row_frame, text=tname, variant="outlined", size="small",
                    width=90,
                    command=lambda t=tname: self._apply_custom_template(t)
                )
                btn.grid(row=0, column=col_idx + 1, padx=(0, 6), pady=4, sticky="w")
                t = templates[tname]
                tip = f"{t['hair_color']}{t['hair_style']} / {t['clothing']} / {t['archetype']} / 恥{t['shyness']}"
                add_tooltip(btn, tip)

        # === 基本情報 Card ===
        basic_card = MaterialCard(custom_scroll, title="基本情報", variant="outlined", collapsible=True, start_collapsed=False)
        basic_card.pack(fill="x", pady=(0, 8))
        bc = basic_card.content_frame

        ctk.CTkLabel(bc, text="キャラ名", font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
                    text_color=MaterialColors.ON_SURFACE_VARIANT).pack(anchor="w", pady=(0,0))
        self.custom_name_entry = ctk.CTkEntry(
            bc, height=36, placeholder_text="例: 佐藤花子",
            font=ctk.CTkFont(family=FONT_JP, size=15), width=350,
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            text_color=MaterialColors.ON_SURFACE,
            placeholder_text_color="#3D3D3D",
            border_width=1, border_color=MaterialColors.OUTLINE_VARIANT
        )
        self.custom_name_entry.pack(anchor="w", pady=(2, 0))

        self.custom_age_dd = add_dropdown(bc, "年齢・外見", AGE_OPTIONS, "JK（女子高生）")
        self.custom_rel_dd = add_dropdown(bc, "主人公との関係", RELATIONSHIP_OPTIONS, "クラスメイト")

        # === 性格・口調 Card ===
        personality_card = MaterialCard(custom_scroll, title="性格・口調", variant="outlined", collapsible=True, start_collapsed=True)
        personality_card.pack(fill="x", pady=(0, 8))
        pc = personality_card.content_frame

        ctk.CTkLabel(pc, text="性格タイプ", font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
                    text_color=MaterialColors.ON_SURFACE_VARIANT).pack(anchor="w", pady=(0, 4))

        # Archetype chip grid (4 cols x 4 rows)
        archetype_grid = ctk.CTkFrame(pc, fg_color="transparent")
        archetype_grid.pack(fill="x", pady=(0, 6))

        for i, arch in enumerate(ARCHETYPE_OPTIONS):
            chip = MaterialChip(
                archetype_grid, text=arch,
                selected=(arch == self._selected_archetype),
                chip_type="filter",
                command=lambda a=arch: self._select_archetype_chip(a)
            )
            row_num = i // 4
            col_num = i % 4
            chip.grid(row=row_num, column=col_num, padx=4, pady=4, sticky="w")
            self._archetype_chips[arch] = chip

        self.custom_first_person_dd = add_dropdown(pc, "一人称", FIRST_PERSON_OPTIONS, "あたし")
        self.custom_speech_dd = add_dropdown(pc, "口調", SPEECH_STYLE_OPTIONS, "タメ口")

        # === 外見 Card ===
        appearance_card = MaterialCard(custom_scroll, title="外見", variant="outlined", collapsible=True, start_collapsed=True)
        appearance_card.pack(fill="x", pady=(0, 8))
        ac = appearance_card.content_frame

        ctk.CTkLabel(ac, text="髪色", font=ctk.CTkFont(family=FONT_JP, size=13, weight="bold"),
                    text_color=MaterialColors.ON_SURFACE_VARIANT).pack(anchor="w", pady=(0, 4))

        # Hair color chips
        hair_color_frame = ctk.CTkFrame(ac, fg_color="transparent")
        hair_color_frame.pack(fill="x", pady=(0, 6))

        for color in HAIR_COLOR_OPTIONS:
            chip = MaterialChip(
                hair_color_frame, text=color,
                selected=(color == self._selected_hair_color),
                chip_type="filter",
                command=lambda c=color: self._select_hair_color_chip(c)
            )
            chip.pack(side="left", padx=(0, 6), pady=4)
            self._hair_color_chips[color] = chip

        self.custom_hair_style_dd = add_dropdown(ac, "髪型", HAIR_STYLE_OPTIONS, "ロングストレート")
        self.custom_body_dd = add_dropdown(ac, "体型", BODY_TYPE_OPTIONS, "普通")
        self.custom_chest_dd = add_dropdown(ac, "胸", CHEST_OPTIONS, "普通（C）")
        self.custom_clothing_dd = add_dropdown(ac, "服装", CLOTHING_OPTIONS, "制服（ブレザー）")

        # === エロシーン設定 Card ===
        ero_card = MaterialCard(custom_scroll, title="エロシーン設定", variant="outlined", collapsible=True, start_collapsed=True)
        ero_card.pack(fill="x", pady=(0, 8))
        ec = ero_card.content_frame

        slider_row = ctk.CTkFrame(ec, fg_color="transparent")
        slider_row.pack(fill="x", pady=(0, 4))

        ctk.CTkLabel(slider_row, text="大胆",
                    font=ctk.CTkFont(family=FONT_JP, size=13),
                    text_color=MaterialColors.ON_SURFACE_VARIANT).pack(side="left", padx=(0, 8))

        self.shyness_slider = ctk.CTkSlider(
            slider_row, from_=1, to=5, number_of_steps=4,
            width=200,
            fg_color=MaterialColors.SURFACE_CONTAINER_HIGH,
            progress_color=MaterialColors.PRIMARY,
            button_color=MaterialColors.PRIMARY,
            button_hover_color=MaterialColors.PRIMARY_VARIANT
        )
        self.shyness_slider.set(3)
        self.shyness_slider.pack(side="left", padx=(0, 8))

        ctk.CTkLabel(slider_row, text="恥ずかしがり",
                    font=ctk.CTkFont(family=FONT_JP, size=13),
                    text_color=MaterialColors.ON_SURFACE_VARIANT).pack(side="left")

        self._shyness_value_label = ctk.CTkLabel(ec, text="恥ずかしがり度: 3",
                    font=ctk.CTkFont(family=FONT_JP, size=13),
                    text_color=MaterialColors.ON_SURFACE_VARIANT)
        self._shyness_value_label.pack(anchor="w")
        self.shyness_slider.configure(command=self._on_shyness_change)

        # === 追加設定 ===
        extra_card = MaterialCard(custom_scroll, title="追加設定（任意）", variant="outlined", collapsible=True, start_collapsed=True)
        extra_card.pack(fill="x", pady=(0, 8))
        xc = extra_card.content_frame

        ctk.CTkLabel(xc, text="追加の性格特性（「、」区切り）",
                    font=ctk.CTkFont(family=FONT_JP, size=13),
                    text_color=MaterialColors.ON_SURFACE_VARIANT).pack(anchor="w", pady=(0,0))
        self.custom_traits_entry = ctk.CTkEntry(
            xc, height=36, placeholder_text="例: 読書好き、猫が好き",
            font=ctk.CTkFont(family=FONT_JP, size=14), width=350,
            fg_color=MaterialColors.SURFACE_CONTAINER, corner_radius=4,
            text_color=MaterialColors.ON_SURFACE,
            placeholder_text_color="#3D3D3D",
            border_width=1, border_color=MaterialColors.OUTLINE_VARIANT
        )
        self.custom_traits_entry.pack(anchor="w", pady=(2, 0))

        # === Live Preview ===
        preview_card = MaterialCard(custom_scroll, title="プレビュー", variant="filled")
        preview_card.pack(fill="x", pady=(0, 8))
        self._custom_preview_label = ctk.CTkLabel(
            preview_card.content_frame,
            text="キャラ名を入力してください",
            font=ctk.CTkFont(family=FONT_JP, size=14),
            text_color=MaterialColors.ON_SURFACE_VARIANT,
            wraplength=380, justify="left"
        )
        self._custom_preview_label.pack(anchor="w")

        # === Save Button ===
        self.custom_save_btn = MaterialButton(
            custom_scroll, text="キャラクターを保存（API不要）",
            variant="filled", command=self.save_custom_character
        )
        self.custom_save_btn.pack(anchor="w", pady=(8, 8))

    def _select_archetype_chip(self, archetype):
        """性格タイプチップの排他選択"""
        self._selected_archetype = archetype
        for arch, chip in self._archetype_chips.items():
            if arch == archetype:
                if not chip.selected:
                    chip.toggle()
            else:
                if chip.selected:
                    chip.toggle()
        self._update_custom_preview()

    def _select_hair_color_chip(self, color):
        """髪色チップの排他選択"""
        self._selected_hair_color = color
        for c, chip in self._hair_color_chips.items():
            if c == color:
                if not chip.selected:
                    chip.toggle()
            else:
                if chip.selected:
                    chip.toggle()
        self._update_custom_preview()

    def _on_shyness_change(self, value):
        """恥ずかしがり度スライダー変更"""
        v = int(round(value))
        labels = {1: "大胆・積極的", 2: "やや積極的", 3: "普通", 4: "恥ずかしがり", 5: "超恥ずかしがり"}
        self._shyness_value_label.configure(text=f"恥ずかしがり度: {v} - {labels.get(v, '')}")
        self._update_custom_preview()

    def _update_custom_preview(self, *args):
        """ライブプレビュー更新"""
        name = self.custom_name_entry.get().strip() if hasattr(self, 'custom_name_entry') else ""
        if not name:
            name = "（未入力）"
        age = self.custom_age_dd.get() if hasattr(self, 'custom_age_dd') else ""
        archetype = self._selected_archetype
        hair_color = self._selected_hair_color
        hair_style = self.custom_hair_style_dd.get() if hasattr(self, 'custom_hair_style_dd') else ""
        chest = self.custom_chest_dd.get() if hasattr(self, 'custom_chest_dd') else ""
        clothing = self.custom_clothing_dd.get() if hasattr(self, 'custom_clothing_dd') else ""
        shyness = int(round(self.shyness_slider.get())) if hasattr(self, 'shyness_slider') else 3

        preview = f"{name} / {age} / {archetype} / {hair_color}{hair_style} / {chest} / {clothing} / 恥度{shyness}"
        if hasattr(self, '_custom_preview_label'):
            self._custom_preview_label.configure(text=preview)

    def _apply_custom_template(self, template_name):
        """テンプレート適用"""
        t = self._custom_templates.get(template_name, {})
        if not t:
            return

        # Set age
        if hasattr(self, 'custom_age_dd'):
            self.custom_age_dd.set(t.get("age", "JK（女子高生）"))
        # Set archetype
        self._select_archetype_chip(t.get("archetype", "ツンデレ"))
        # Set first person
        if hasattr(self, 'custom_first_person_dd'):
            self.custom_first_person_dd.set(t.get("first_person", "私"))
        # Set speech
        if hasattr(self, 'custom_speech_dd'):
            self.custom_speech_dd.set(t.get("speech", "タメ口"))
        # Set hair color
        self._select_hair_color_chip(t.get("hair_color", "黒髪"))
        # Set hair style
        if hasattr(self, 'custom_hair_style_dd'):
            self.custom_hair_style_dd.set(t.get("hair_style", "ロングストレート"))
        # Set body
        if hasattr(self, 'custom_body_dd'):
            self.custom_body_dd.set(t.get("body", "普通"))
        # Set chest
        if hasattr(self, 'custom_chest_dd'):
            self.custom_chest_dd.set(t.get("chest", "普通（C）"))
        # Set clothing
        if hasattr(self, 'custom_clothing_dd'):
            self.custom_clothing_dd.set(t.get("clothing", "制服（ブレザー）"))
        # Set shyness
        if hasattr(self, 'shyness_slider'):
            self.shyness_slider.set(t.get("shyness", 3))
            self._on_shyness_change(t.get("shyness", 3))

        self._update_custom_preview()
        self.snackbar.show(f"テンプレート「{template_name}」を適用", type="info")


if __name__ == "__main__":
    app = App()
    app.mainloop()
